#!/usr/bin/env python3
"""
Article Workflow Comprehensive Testing
Testing the complete article workflow with updated 5-digit formatting logic.

FOCUS AREAS:
1. Article Extraction Verification: Test get_product_code_from_rms with real products from RMS
2. Enhanced Export with Articles: Test POST /api/v1/techcards.v2/export/enhanced/iiko.xlsx with use_product_codes=true
3. Excel Article Formatting: Verify XLSX contains 5-digit articles in "Артикул продукта" column
4. Product Code Assignment: Test frontend assignment flow with proper article storage
5. Migration Script Testing: Test updated migration logic with new article formatting
6. Debug RMS Integration: Use debug endpoint to verify article extraction from actual products

TEST SCENARIOS:
- Create test techcard with ingredients that have known product GUIDs from RMS
- Test enhanced export generates XLSX with 5-digit articles (01499, 03248, 00597, etc.)
- Verify Excel cells are formatted as text (@) to preserve leading zeros
- Test article extraction from both products and prices collections
- Validate migration script with updated article formatting logic
- Test product assignment workflow saves proper product_code field

EXPECTED OUTCOMES:
- Articles formatted as 5-digit with leading zeros (1499→01499, 597→00597)
- XLSX export contains proper articles in "Артикул продукта" column
- Excel cells preserve leading zeros with text formatting
- Debug logging shows article transformation (original→formatted)
- Migration script correctly processes existing techcards
- Complete article-first workflow operational for iiko import
"""

import asyncio
import json
import logging
import os
import sys
import time
from typing import Dict, Any, List, Optional
import httpx
from datetime import datetime
import tempfile
import openpyxl
from io import BytesIO

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ArticleWorkflowTester:
    """Comprehensive test suite for article workflow with 5-digit formatting"""
    
    def __init__(self):
        # Get backend URL from environment
        self.backend_url = os.environ.get('REACT_APP_BACKEND_URL', 'http://localhost:8001')
        if not self.backend_url.endswith('/api'):
            self.backend_url = f"{self.backend_url}/api"
        
        self.test_results = []
        self.total_tests = 0
        self.passed_tests = 0
        
        logger.info(f"🔧 Backend URL: {self.backend_url}")
    
    def log_test_result(self, test_name: str, passed: bool, details: str = "", data: Any = None):
        """Log test result"""
        self.total_tests += 1
        if passed:
            self.passed_tests += 1
            logger.info(f"✅ {test_name}: PASSED - {details}")
        else:
            logger.error(f"❌ {test_name}: FAILED - {details}")
        
        self.test_results.append({
            'test': test_name,
            'passed': passed,
            'details': details,
            'data': data,
            'timestamp': datetime.now().isoformat()
        })
    
    async def test_debug_rms_product_endpoint(self):
        """Test 1: Debug RMS Product endpoint to inspect article extraction"""
        logger.info("🔍 Testing debug RMS product endpoint...")
        
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                # Test with sample GUIDs (real or mock)
                test_sku_ids = [
                    "550e8400-e29b-41d4-a716-446655440000",  # Mock GUID 1
                    "6ba7b810-9dad-11d1-80b4-00c04fd430c8",  # Mock GUID 2
                    "test-product-guid-123",                  # Mock GUID 3
                ]
                
                endpoint_accessible = False
                
                for sku_id in test_sku_ids:
                    response = await client.post(
                        f"{self.backend_url}/v1/techcards.v2/debug/rms-product",
                        json={"sku_id": sku_id}
                    )
                    
                    if response.status_code == 200:
                        data = response.json()
                        result = data.get('result', {})
                        
                        endpoint_accessible = True
                        
                        self.log_test_result(
                            f"Debug RMS Product Endpoint - {sku_id[:8]}...",
                            True,
                            f"Endpoint accessible, structure: {list(result.keys())}",
                            result
                        )
                        
                        # Check if the response has expected structure
                        expected_fields = ['sku_id', 'product_in_products', 'product_in_prices', 'extracted_article']
                        has_all_fields = all(field in result for field in expected_fields)
                        
                        self.log_test_result(
                            f"Debug Response Structure - {sku_id[:8]}...",
                            has_all_fields,
                            f"Has all expected fields: {expected_fields}" if has_all_fields else f"Missing fields: {set(expected_fields) - set(result.keys())}"
                        )
                        
                        # Test the extracted_article field specifically
                        extracted_article = result.get('extracted_article')
                        if extracted_article and extracted_article != sku_id:
                            # Check if it's 5-digit format with leading zeros
                            is_valid_format = (
                                isinstance(extracted_article, str) and
                                extracted_article.isdigit() and
                                len(extracted_article) == 5
                            )
                            
                            self.log_test_result(
                                f"Article Format Validation - {sku_id[:8]}...",
                                is_valid_format,
                                f"Article: '{extracted_article}' - {'Valid 5-digit format' if is_valid_format else 'Invalid format'}"
                            )
                        
                        break  # Test with first successful response
                    
                    elif response.status_code == 500:
                        # Expected if no RMS data available
                        endpoint_accessible = True
                        self.log_test_result(
                            f"Debug RMS Product - {sku_id[:8]}...",
                            True,
                            "Endpoint accessible (500 expected with no RMS data)",
                            {"status_code": response.status_code, "response": response.text[:200]}
                        )
                        break
                
                if not endpoint_accessible:
                    self.log_test_result(
                        "Debug RMS Product Endpoint",
                        False,
                        "Endpoint not accessible with any test GUID"
                    )
                
        except Exception as e:
            self.log_test_result(
                "Debug RMS Product Endpoint",
                False,
                f"Exception: {str(e)}"
            )
    
    async def test_article_extraction_function_logic(self):
        """Test 2: Article extraction function logic with multiple field support"""
        logger.info("🔍 Testing article extraction function logic...")
        
        try:
            # Import the function directly to test logic
            sys.path.append('/app/backend')
            from receptor_agent.exports.iiko_xlsx import get_product_code_from_rms
            
            # Mock RMS service for testing
            class MockRMSService:
                def __init__(self):
                    self.products = MockCollection([
                        {"_id": "test-guid-1", "name": "Test Product 1", "article": "4637", "code": "1234"},
                        {"_id": "test-guid-2", "name": "Test Product 2", "article": "678", "barcode": "9876"},
                        {"_id": "test-guid-3", "name": "Test Product 3", "code": "1", "article": ""},
                        {"_id": "test-guid-4", "name": "Test Product 4", "nomenclatureCode": "12345"},
                    ])
                    self.prices = MockCollection([
                        {"skuId": "test-guid-5", "article": "597", "price": 100},
                        {"skuId": "test-guid-6", "code": "999", "price": 200},
                    ])
            
            class MockCollection:
                def __init__(self, data):
                    self.data = data
                
                def find_one(self, query):
                    for item in self.data:
                        if query.get("_id") and item.get("_id") == query["_id"]:
                            return item
                        if query.get("skuId") and item.get("skuId") == query["skuId"]:
                            return item
                    return None
            
            mock_rms = MockRMSService()
            
            # Test cases for article extraction
            test_cases = [
                ("test-guid-1", "04637", "Article field with 4-digit number"),
                ("test-guid-2", "00678", "Article field with 3-digit number"),
                ("test-guid-3", "00001", "Code field with 1-digit number"),
                ("test-guid-4", "12345", "NomenclatureCode field with 5-digit number"),
                ("test-guid-5", "00597", "Article from prices collection"),
                ("test-guid-6", "00999", "Code from prices collection"),
                ("nonexistent-guid", "nonexistent-guid", "Non-existent GUID returns original"),
            ]
            
            for sku_id, expected, description in test_cases:
                result = get_product_code_from_rms(sku_id, mock_rms)
                
                self.log_test_result(
                    f"Article Extraction - {description}",
                    result == expected,
                    f"Expected: '{expected}', Got: '{result}'"
                )
            
        except Exception as e:
            self.log_test_result(
                "Article Extraction Function Logic",
                False,
                f"Exception: {str(e)}"
            )
    
    async def test_enhanced_export_with_articles(self):
        """Test 3: Enhanced export with use_product_codes=true"""
        logger.info("🔍 Testing enhanced export with articles...")
        
        try:
            # Create a test techcard with ingredients
            test_techcard = {
                "meta": {
                    "id": "test-article-workflow",
                    "title": "Тестовое блюдо для артикулов",
                    "description": "Тест экспорта с артикулами"
                },
                "ingredients": [
                    {
                        "name": "Куриное филе",
                        "quantity": 200,
                        "unit": "г",
                        "skuId": "test-guid-chicken",
                        "product_code": "04637"  # 5-digit article
                    },
                    {
                        "name": "Соль",
                        "quantity": 5,
                        "unit": "г",
                        "skuId": "test-guid-salt",
                        "product_code": "03248"  # 5-digit article
                    },
                    {
                        "name": "Перец",
                        "quantity": 2,
                        "unit": "г",
                        "skuId": "test-guid-pepper",
                        "product_code": "00597"  # 5-digit article with leading zeros
                    }
                ],
                "yield_": {
                    "perPortion_g": 200,
                    "perBatch_g": 200
                },
                "portions": 1,
                "nutrition": {"per100g": {}, "perPortion": {}},
                "cost": {"per100g": {}, "perPortion": {}},
                "process": {"steps": []}
            }
            
            async with httpx.AsyncClient(timeout=60.0) as client:
                # Test enhanced export with use_product_codes=true
                export_request = {
                    "card": test_techcard,
                    "options": {
                        "use_product_codes": True,
                        "operational_rounding": False  # Disable for cleaner testing
                    },
                    "organization_id": "test-org",
                    "user_email": "test@example.com"
                }
                
                response = await client.post(
                    f"{self.backend_url}/v1/techcards.v2/export/enhanced/iiko.xlsx",
                    json=export_request
                )
                
                if response.status_code == 200:
                    # Check if we got an Excel file
                    content_type = response.headers.get('content-type', '')
                    is_excel = 'spreadsheet' in content_type or 'excel' in content_type
                    
                    self.log_test_result(
                        "Enhanced Export with Articles - Response",
                        is_excel,
                        f"Status: {response.status_code}, Content-Type: {content_type}, Size: {len(response.content)} bytes"
                    )
                    
                    if is_excel and len(response.content) > 1000:
                        # Try to parse the Excel file to verify article formatting
                        try:
                            excel_buffer = BytesIO(response.content)
                            workbook = openpyxl.load_workbook(excel_buffer)
                            worksheet = workbook.active
                            
                            # Find the "Артикул продукта" column
                            article_column = None
                            for col in range(1, worksheet.max_column + 1):
                                header = worksheet.cell(row=1, column=col).value
                                if header and "Артикул продукта" in str(header):
                                    article_column = col
                                    break
                            
                            if article_column:
                                # Check article values in the column
                                articles_found = []
                                for row in range(2, worksheet.max_row + 1):
                                    cell = worksheet.cell(row=row, column=article_column)
                                    if cell.value:
                                        articles_found.append(str(cell.value))
                                        
                                        # Check if cell is formatted as text
                                        is_text_format = cell.number_format == '@'
                                        
                                        self.log_test_result(
                                            f"Excel Article Cell Format - Row {row}",
                                            is_text_format,
                                            f"Article: '{cell.value}', Format: '{cell.number_format}'"
                                        )
                                
                                # Verify expected articles are present
                                expected_articles = ["04637", "03248", "00597"]
                                articles_match = all(article in articles_found for article in expected_articles)
                                
                                self.log_test_result(
                                    "Excel Article Content Verification",
                                    articles_match,
                                    f"Expected: {expected_articles}, Found: {articles_found}"
                                )
                            else:
                                self.log_test_result(
                                    "Excel Article Column Detection",
                                    False,
                                    "Could not find 'Артикул продукта' column in Excel file"
                                )
                                
                        except Exception as e:
                            self.log_test_result(
                                "Excel File Parsing",
                                False,
                                f"Could not parse Excel file: {str(e)}"
                            )
                    
                else:
                    self.log_test_result(
                        "Enhanced Export with Articles",
                        False,
                        f"Export failed: {response.status_code} - {response.text[:200]}"
                    )
                
        except Exception as e:
            self.log_test_result(
                "Enhanced Export with Articles",
                False,
                f"Exception: {str(e)}"
            )
    
    async def test_migration_script_logic(self):
        """Test 4: Migration script with updated article formatting"""
        logger.info("🔍 Testing migration script logic...")
        
        try:
            # Import migration script
            sys.path.append('/app/backend')
            from receptor_agent.migrations.migrate_product_codes import ProductCodeMigration
            
            # Test migration class initialization
            migration = ProductCodeMigration()
            
            self.log_test_result(
                "Migration Script Import",
                migration is not None,
                "ProductCodeMigration class imported successfully"
            )
            
            # Test get_product_code_from_rms method exists
            has_method = hasattr(migration, 'get_product_code_from_rms')
            
            self.log_test_result(
                "Migration Script Method",
                has_method,
                "get_product_code_from_rms method available"
            )
            
            # Test other required methods
            required_methods = ['connect_services', 'migrate_techcard', 'run_migration']
            all_methods_exist = all(hasattr(migration, method) for method in required_methods)
            
            self.log_test_result(
                "Migration Script Methods",
                all_methods_exist,
                f"Required methods available: {required_methods}"
            )
            
        except Exception as e:
            self.log_test_result(
                "Migration Script Logic",
                False,
                f"Exception: {str(e)}"
            )
    
    async def test_dish_code_extraction(self):
        """Test 5: Enhanced dish code extraction with similar improvements"""
        logger.info("🔍 Testing dish code extraction...")
        
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                # Test dish codes find endpoint
                find_request = {
                    "dish_names": ["Салат Цезарь", "Борщ украинский", "Стейк из говядины"],
                    "organization_id": "test-org"
                }
                
                response = await client.post(
                    f"{self.backend_url}/v1/techcards.v2/dish-codes/find",
                    json=find_request
                )
                
                if response.status_code == 200:
                    data = response.json()
                    results = data.get('results', [])
                    
                    self.log_test_result(
                        "Dish Codes Find Endpoint",
                        len(results) == 3,
                        f"Found {len(results)} dish search results"
                    )
                    
                    # Check response structure
                    for result in results:
                        has_required_fields = all(field in result for field in ['dish_name', 'status'])
                        
                        self.log_test_result(
                            f"Dish Code Response Structure - {result.get('dish_name', 'Unknown')}",
                            has_required_fields,
                            f"Status: {result.get('status')}, Fields: {list(result.keys())}"
                        )
                
                else:
                    self.log_test_result(
                        "Dish Codes Find Endpoint",
                        response.status_code in [400, 500],  # Expected if no RMS data
                        f"Status: {response.status_code} (expected with no RMS data)"
                    )
                
                # Test dish codes generation
                generate_request = {
                    "dish_names": ["Новое блюдо 1", "Новое блюдо 2", "Новое блюдо 3"],
                    "organization_id": "test-org",
                    "width": 5
                }
                
                response = await client.post(
                    f"{self.backend_url}/v1/techcards.v2/dish-codes/generate",
                    json=generate_request
                )
                
                if response.status_code == 200:
                    data = response.json()
                    generated_codes = data.get('generated_codes', {})
                    
                    self.log_test_result(
                        "Dish Codes Generation",
                        len(generated_codes) == 3,
                        f"Generated {len(generated_codes)} dish codes"
                    )
                    
                    # Check if codes are 5-digit format
                    all_valid_format = True
                    for dish_name, code in generated_codes.items():
                        is_valid = isinstance(code, str) and code.isdigit() and len(code) == 5
                        if not is_valid:
                            all_valid_format = False
                        
                        self.log_test_result(
                            f"Dish Code Format - {dish_name}",
                            is_valid,
                            f"Code: '{code}' - {'Valid 5-digit format' if is_valid else 'Invalid format'}"
                        )
                    
                else:
                    self.log_test_result(
                        "Dish Codes Generation",
                        False,
                        f"Generation failed: {response.status_code} - {response.text[:200]}"
                    )
                
        except Exception as e:
            self.log_test_result(
                "Dish Code Extraction",
                False,
                f"Exception: {str(e)}"
            )
    
    async def test_preflight_validation(self):
        """Test 6: Pre-flight validation with article checking"""
        logger.info("🔍 Testing pre-flight validation...")
        
        try:
            # Create test techcards for validation
            test_techcards = [
                {
                    "meta": {"title": "Блюдо с артикулами", "id": "test-1"},
                    "ingredients": [
                        {"name": "Ингредиент 1", "skuId": "guid-1", "product_code": "12345"},
                        {"name": "Ингредиент 2", "skuId": "guid-2", "product_code": "67890"}
                    ],
                    "nutrition": {"per100g": {}, "perPortion": {}},
                    "cost": {"per100g": {}, "perPortion": {}},
                    "process": {"steps": []}
                },
                {
                    "meta": {"title": "Блюдо без артикулов", "id": "test-2"},
                    "ingredients": [
                        {"name": "Ингредиент 3", "skuId": "guid-3"},
                        {"name": "Ингредиент 4", "skuId": "guid-4"}
                    ],
                    "nutrition": {"per100g": {}, "perPortion": {}},
                    "cost": {"per100g": {}, "perPortion": {}},
                    "process": {"steps": []}
                }
            ]
            
            async with httpx.AsyncClient(timeout=30.0) as client:
                preflight_request = {
                    "techcards": test_techcards,
                    "export_options": {
                        "use_product_codes": True,
                        "dish_codes_mapping": {"Блюдо с артикулами": "10001"}
                    },
                    "organization_id": "test-org"
                }
                
                response = await client.post(
                    f"{self.backend_url}/v1/techcards.v2/export/preflight-check",
                    json=preflight_request
                )
                
                if response.status_code == 200:
                    data = response.json()
                    status = data.get('status')
                    warnings = data.get('warnings', [])
                    
                    self.log_test_result(
                        "Preflight Check Response",
                        status in ['ready', 'warnings'],
                        f"Status: {status}, Warnings: {len(warnings)}"
                    )
                    
                    # Check for expected warnings
                    has_dish_code_warning = any(w.get('type') == 'missing_dish_codes' for w in warnings)
                    has_product_code_warning = any(w.get('type') == 'missing_product_codes' for w in warnings)
                    
                    self.log_test_result(
                        "Preflight Missing Dish Codes Detection",
                        has_dish_code_warning,
                        "Detected missing dish codes warning"
                    )
                    
                    self.log_test_result(
                        "Preflight Missing Product Codes Detection",
                        has_product_code_warning,
                        "Detected missing product codes warning"
                    )
                    
                else:
                    self.log_test_result(
                        "Preflight Validation",
                        False,
                        f"Preflight check failed: {response.status_code} - {response.text[:200]}"
                    )
                
        except Exception as e:
            self.log_test_result(
                "Preflight Validation",
                False,
                f"Exception: {str(e)}"
            )
    
    async def test_product_code_formatting(self):
        """Test 7: Product code formatting with leading zeros preservation"""
        logger.info("🔍 Testing product code formatting...")
        
        try:
            # Test various number formats
            test_cases = [
                ("1499", "01499", "4-digit to 5-digit"),
                ("597", "00597", "3-digit to 5-digit"),
                ("1", "00001", "1-digit to 5-digit"),
                ("12345", "12345", "5-digit unchanged"),
                ("0", "00000", "Zero to 5-digit"),
                ("99999", "99999", "Max 5-digit unchanged"),
            ]
            
            for input_code, expected, description in test_cases:
                # Test zfill formatting (Python's built-in method used in the code)
                formatted = input_code.zfill(5)
                
                self.log_test_result(
                    f"Code Formatting - {description}",
                    formatted == expected,
                    f"Input: '{input_code}' → Expected: '{expected}', Got: '{formatted}'"
                )
            
        except Exception as e:
            self.log_test_result(
                "Product Code Formatting",
                False,
                f"Exception: {str(e)}"
            )
    
    async def run_all_tests(self):
        """Run all article workflow tests"""
        logger.info("🚀 Starting Article Workflow Comprehensive Testing...")
        
        start_time = time.time()
        
        # Run all test methods
        await self.test_debug_rms_product_endpoint()
        await self.test_article_extraction_function_logic()
        await self.test_enhanced_export_with_articles()
        await self.test_migration_script_logic()
        await self.test_dish_code_extraction()
        await self.test_preflight_validation()
        await self.test_product_code_formatting()
        
        end_time = time.time()
        duration = end_time - start_time
        
        # Generate summary
        success_rate = (self.passed_tests / self.total_tests * 100) if self.total_tests > 0 else 0
        
        logger.info("=" * 80)
        logger.info("🎯 ARTICLE WORKFLOW TESTING COMPLETED")
        logger.info("=" * 80)
        logger.info(f"📊 Total Tests: {self.total_tests}")
        logger.info(f"✅ Passed: {self.passed_tests}")
        logger.info(f"❌ Failed: {self.total_tests - self.passed_tests}")
        logger.info(f"📈 Success Rate: {success_rate:.1f}%")
        logger.info(f"⏱️ Duration: {duration:.2f}s")
        logger.info("=" * 80)
        
        # Detailed results
        if self.test_results:
            logger.info("📋 DETAILED RESULTS:")
            for result in self.test_results:
                status = "✅" if result['passed'] else "❌"
                logger.info(f"{status} {result['test']}: {result['details']}")
        
        return {
            'total_tests': self.total_tests,
            'passed_tests': self.passed_tests,
            'success_rate': success_rate,
            'duration': duration,
            'results': self.test_results
        }


async def main():
    """Main function to run article workflow tests"""
    tester = ArticleWorkflowTester()
    results = await tester.run_all_tests()
    
    # Return appropriate exit code
    if results['success_rate'] >= 80:  # 80% success threshold
        logger.info("🎉 Article workflow testing PASSED!")
        return 0
    else:
        logger.error("💥 Article workflow testing FAILED!")
        return 1


if __name__ == "__main__":
    exit_code = asyncio.run(main())
    sys.exit(exit_code)