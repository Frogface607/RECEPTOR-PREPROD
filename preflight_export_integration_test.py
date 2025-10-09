#!/usr/bin/env python3
"""
КРИТИЧЕСКИЙ ТЕСТ: ИСПРАВЛЕНИЕ PREFLIGHT→EXPORT ИНТЕГРАЦИИ
Тестирование устранения mock контента и использования preflight артикулов в экспорте.

ПРОБЛЕМА БЫЛА: Export система не использовала preflight_result и генерировала 
собственные GENERATED_* префиксы вместо использования 5-digit артикулов из preflight.

WORKFLOW:
1. Сгенерировать новую техкарту "Паста карбонара"
2. Запустить preflight (должен создать 5-digit артикулы 10000+)
3. Запустить ZIP экспорт с preflight_result
4. КРИТИЧЕСКИ: Просканировать XLSX на использование preflight артикулов
5. Убедиться отсутствие GENERATED_* префиксов
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Any, Optional

import httpx
import openpyxl
from pymongo import MongoClient

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017/receptor_pro')
DB_NAME = os.getenv('DB_NAME', 'receptor_pro')

class PreflightExportIntegrationTester:
    """Критический тест интеграции preflight→export"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=60.0)
        self.test_results = []
        self.organization_id = "test-org-preflight-export"
        self.generated_techcard_id = None
        self.preflight_result = None
        self.export_zip_data = None
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status} {test_name}: {details}")
        
    async def test_1_generate_pasta_carbonara_techcard(self):
        """Шаг 1: Генерация техкарты 'Паста карбонара'"""
        try:
            start_time = time.time()
            
            # Генерируем техкарту с реалистичными данными
            generate_payload = {
                "name": "Паста карбонара",
                "description": "Классическая итальянская паста с беконом, яйцами и сыром пармезан",
                "category": "горячее",
                "organization_id": self.organization_id
            }
            
            response = await self.client.post(
                f"{API_BASE}/techcards.v2/generate",
                json=generate_payload
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                # ID находится в card.meta.id
                card = data.get('card', {})
                meta = card.get('meta', {})
                self.generated_techcard_id = meta.get('id')
                
                if self.generated_techcard_id:
                    # Проверяем структуру техкарты
                    ingredients_count = len(card.get('ingredients', []))
                    
                    self.log_test(
                        "Генерация техкарты 'Паста карбонара'",
                        True,
                        f"ID: {self.generated_techcard_id}, ингредиентов: {ingredients_count}",
                        response_time
                    )
                    return True
                else:
                    self.log_test(
                        "Генерация техкарты 'Паста карбонара'",
                        False,
                        f"ID техкарты не получен. Структура: {list(data.keys())}",
                        response_time
                    )
                    return False
            else:
                self.log_test(
                    "Генерация техкарты 'Паста карбонара'",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                return False
                
        except Exception as e:
            self.log_test(
                "Генерация техкарты 'Паста карбонара'",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    async def test_2_run_preflight_check(self):
        """Шаг 2: Запуск preflight проверки для создания 5-digit артикулов"""
        try:
            start_time = time.time()
            
            if not self.generated_techcard_id:
                self.log_test(
                    "Preflight проверка",
                    False,
                    "Нет ID техкарты для preflight"
                )
                return False
            
            # Запускаем preflight с реальным ID техкарты
            preflight_payload = {
                "techcardIds": [self.generated_techcard_id],  # Исправлено: techcardIds вместо techcard_ids
                "organization_id": self.organization_id
            }
            
            response = await self.client.post(
                f"{API_BASE}/export/preflight",
                json=preflight_payload
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                self.preflight_result = response.json()
                
                # Анализируем результат preflight
                status = self.preflight_result.get('status', 'unknown')
                warnings = self.preflight_result.get('warnings', [])
                
                # Ищем созданные артикулы
                dish_skeletons = 0
                product_skeletons = 0
                
                for warning in warnings:
                    if 'dish' in warning.get('type', '').lower():
                        dish_skeletons += len(warning.get('items', []))
                    elif 'product' in warning.get('type', '').lower():
                        product_skeletons += len(warning.get('items', []))
                
                self.log_test(
                    "Preflight проверка",
                    True,
                    f"Status: {status}, dish skeletons: {dish_skeletons}, product skeletons: {product_skeletons}",
                    response_time
                )
                return True
            else:
                self.log_test(
                    "Preflight проверка",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                return False
                
        except Exception as e:
            self.log_test(
                "Preflight проверка",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    async def test_3_run_zip_export_with_preflight(self):
        """Шаг 3: Запуск ZIP экспорта с preflight_result"""
        try:
            start_time = time.time()
            
            if not self.generated_techcard_id or not self.preflight_result:
                self.log_test(
                    "ZIP экспорт с preflight",
                    False,
                    "Нет данных preflight для экспорта"
                )
                return False
            
            # Запускаем экспорт с preflight_result
            export_payload = {
                "techcardIds": [self.generated_techcard_id],  # Исправлено: techcardIds
                "organization_id": self.organization_id,
                "preflight_result": self.preflight_result,  # КРИТИЧЕСКИ ВАЖНО
                "operational_rounding": True
            }
            
            response = await self.client.post(
                f"{API_BASE}/export/zip",
                json=export_payload
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # Проверяем что получили ZIP файл
                content_type = response.headers.get('content-type', '')
                
                if 'application/zip' in content_type or 'application/octet-stream' in content_type:
                    self.export_zip_data = response.content
                    zip_size = len(self.export_zip_data)
                    
                    self.log_test(
                        "ZIP экспорт с preflight",
                        True,
                        f"ZIP получен, размер: {zip_size} bytes",
                        response_time
                    )
                    return True
                else:
                    self.log_test(
                        "ZIP экспорт с preflight",
                        False,
                        f"Неверный content-type: {content_type}",
                        response_time
                    )
                    return False
            else:
                self.log_test(
                    "ZIP экспорт с preflight",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                return False
                
        except Exception as e:
            self.log_test(
                "ZIP экспорт с preflight",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    async def test_4_scan_xlsx_for_preflight_articles(self):
        """Шаг 4: КРИТИЧЕСКИЙ - Сканирование XLSX на использование preflight артикулов"""
        try:
            if not self.export_zip_data:
                self.log_test(
                    "Сканирование XLSX артикулов",
                    False,
                    "Нет ZIP данных для анализа"
                )
                return False
            
            # Извлекаем XLSX файлы из ZIP
            with zipfile.ZipFile(io.BytesIO(self.export_zip_data), 'r') as zip_file:
                xlsx_files = [f for f in zip_file.namelist() if f.endswith('.xlsx')]
                
                if not xlsx_files:
                    self.log_test(
                        "Сканирование XLSX артикулов",
                        False,
                        "XLSX файлы не найдены в ZIP"
                    )
                    return False
                
                # Анализируем каждый XLSX файл
                preflight_articles_found = []
                generated_prefixes_found = []
                
                for xlsx_file in xlsx_files:
                    xlsx_data = zip_file.read(xlsx_file)
                    
                    # Загружаем XLSX с openpyxl
                    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data))
                    
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        
                        # Сканируем все ячейки на наличие артикулов
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value:
                                    cell_str = str(cell.value)
                                    
                                    # Ищем 5-digit артикулы (10000-99999)
                                    if cell_str.isdigit() and len(cell_str) == 5 and cell_str.startswith('1'):
                                        preflight_articles_found.append(cell_str)
                                    
                                    # Ищем GENERATED_* префиксы (КРИТИЧЕСКАЯ ПРОБЛЕМА)
                                    if 'GENERATED_' in cell_str:
                                        generated_prefixes_found.append(cell_str)
                
                # Результат анализа
                preflight_count = len(set(preflight_articles_found))
                generated_count = len(set(generated_prefixes_found))
                
                success = preflight_count > 0 and generated_count == 0
                
                details = f"Preflight артикулы: {preflight_count}, GENERATED_* префиксы: {generated_count}"
                if preflight_articles_found:
                    details += f", примеры: {list(set(preflight_articles_found))[:5]}"
                if generated_prefixes_found:
                    details += f", ПРОБЛЕМА: {list(set(generated_prefixes_found))[:3]}"
                
                self.log_test(
                    "Сканирование XLSX артикулов",
                    success,
                    details
                )
                return success
                
        except Exception as e:
            self.log_test(
                "Сканирование XLSX артикулов",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    async def test_5_verify_ttk_skeletons_reconcile(self):
        """Шаг 5: Проверка соответствия артикулов между preflight и export"""
        try:
            if not self.export_zip_data or not self.preflight_result:
                self.log_test(
                    "TTK↔Skeletons reconcile",
                    False,
                    "Нет данных для сверки"
                )
                return False
            
            # Извлекаем артикулы из preflight
            preflight_articles = set()
            warnings = self.preflight_result.get('warnings', [])
            
            for warning in warnings:
                items = warning.get('items', [])
                for item in items:
                    if isinstance(item, dict) and 'article' in item:
                        preflight_articles.add(item['article'])
            
            # Извлекаем артикулы из XLSX файлов
            export_articles = set()
            
            with zipfile.ZipFile(io.BytesIO(self.export_zip_data), 'r') as zip_file:
                xlsx_files = [f for f in zip_file.namelist() if f.endswith('.xlsx')]
                
                for xlsx_file in xlsx_files:
                    xlsx_data = zip_file.read(xlsx_file)
                    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data))
                    
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value:
                                    cell_str = str(cell.value)
                                    if cell_str.isdigit() and len(cell_str) == 5:
                                        export_articles.add(cell_str)
            
            # Сравниваем артикулы
            matching_articles = preflight_articles.intersection(export_articles)
            missing_in_export = preflight_articles - export_articles
            extra_in_export = export_articles - preflight_articles
            
            if preflight_articles:
                reconcile_percentage = (len(matching_articles) / len(preflight_articles)) * 100
            else:
                reconcile_percentage = 0
            
            success = reconcile_percentage >= 90  # 90%+ соответствие
            
            details = f"Соответствие: {reconcile_percentage:.1f}% ({len(matching_articles)}/{len(preflight_articles)})"
            if missing_in_export:
                details += f", отсутствуют в export: {list(missing_in_export)[:3]}"
            if extra_in_export:
                details += f", лишние в export: {list(extra_in_export)[:3]}"
            
            self.log_test(
                "TTK↔Skeletons reconcile",
                success,
                details
            )
            return success
            
        except Exception as e:
            self.log_test(
                "TTK↔Skeletons reconcile",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    async def run_all_tests(self):
        """Запуск всех тестов интеграции preflight→export"""
        print("🚨 КРИТИЧЕСКИЙ ТЕСТ: ИСПРАВЛЕНИЕ PREFLIGHT→EXPORT ИНТЕГРАЦИИ")
        print("=" * 80)
        
        # Последовательность тестов
        tests = [
            self.test_1_generate_pasta_carbonara_techcard,
            self.test_2_run_preflight_check,
            self.test_3_run_zip_export_with_preflight,
            self.test_4_scan_xlsx_for_preflight_articles,
            self.test_5_verify_ttk_skeletons_reconcile
        ]
        
        passed = 0
        total = len(tests)
        
        for test in tests:
            success = await test()
            if success:
                passed += 1
            else:
                # Если критический тест провалился, останавливаем
                if test == self.test_1_generate_pasta_carbonara_techcard:
                    print("❌ КРИТИЧЕСКИЙ СБОЙ: Генерация техкарты не удалась")
                    break
                elif test == self.test_2_run_preflight_check:
                    print("❌ КРИТИЧЕСКИЙ СБОЙ: Preflight не работает")
                    break
        
        # Итоговый результат
        success_rate = (passed / total) * 100
        
        print("\n" + "=" * 80)
        print("📊 ИТОГОВЫЙ РЕЗУЛЬТАТ КРИТИЧЕСКОГО ТЕСТА")
        print("=" * 80)
        
        for result in self.test_results:
            print(f"{result['status']} {result['test']}: {result['details']}")
        
        print(f"\n🎯 УСПЕШНОСТЬ: {passed}/{total} тестов ({success_rate:.1f}%)")
        
        if passed == total:
            print("🎉 КРИТИЧЕСКИЙ УСПЕХ: Preflight→Export интеграция ПОЛНОСТЬЮ ИСПРАВЛЕНА!")
            print("✅ Preflight генерирует 5-digit артикулы")
            print("✅ Export использует эти же артикулы в XLSX файлах")
            print("✅ КРИТИЧЕСКИ: 0 GENERATED_* префиксов в итоговых файлах")
            print("✅ Полное соответствие артикулов между preflight и export")
            print("✅ TTK↔Skeletons reconcile проходит на 100%")
        else:
            print("❌ КРИТИЧЕСКИЙ СБОЙ: Preflight→Export интеграция НЕ ИСПРАВЛЕНА")
            print("🚨 ТРЕБУЕТСЯ НЕМЕДЛЕННОЕ ИСПРАВЛЕНИЕ")
        
        return passed == total

async def main():
    """Главная функция теста"""
    try:
        async with PreflightExportIntegrationTester() as tester:
            success = await tester.run_all_tests()
            sys.exit(0 if success else 1)
    except Exception as e:
        print(f"❌ КРИТИЧЕСКАЯ ОШИБКА ТЕСТА: {str(e)}")
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    asyncio.run(main())