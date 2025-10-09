#!/usr/bin/env python3
"""
FE Export Wizard: Mock Content Validation Test
Critical test to verify that backend changes have eliminated mock content from exported XLSX files.
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import tempfile
import io
from datetime import datetime
from typing import Dict, List, Any, Optional

import httpx
import openpyxl
from openpyxl import load_workbook

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"

class FEExportWizardMockValidationTester:
    """FE Export Wizard Mock Content Validation Testing Suite"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=60.0)
        self.test_results = []
        self.organization_id = "test-org-export-wizard"
        self.generated_techcards = []
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status} {test_name}: {details}")
    
    async def test_01_generate_real_techcards(self):
        """Generate 2 real tech cards for testing"""
        try:
            start_time = time.time()
            
            # Generate first tech card
            dish_name_1 = "Борщ украинский с говядиной"
            response1 = await self.client.post(
                f"{API_BASE}/techcards.v2/generate",
                json={"name": dish_name_1}
            )
            
            if response1.status_code != 200:
                self.log_test("Generate Real TechCard 1", False, 
                            f"Failed to generate first techcard: {response1.status_code} - {response1.text[:200]}")
                return
            
            data1 = response1.json()
            
            # Check if generation was successful
            if data1.get('status') not in ['success', 'draft']:
                self.log_test("Generate Real TechCard 1", False, 
                            f"Generation failed with status: {data1.get('status')} - {data1.get('message', 'No message')}")
                return
            
            # Get the techcard from response
            techcard_1 = data1.get('card')
            if not techcard_1 or not techcard_1.get('meta', {}).get('id'):
                self.log_test("Generate Real TechCard 1", False, 
                            f"No techcard in response: {data1}")
                return
            
            techcard_id_1 = techcard_1.get('meta', {}).get('id')
            
            # Generate second tech card
            dish_name_2 = "Стейк из говядины с картофельным пюре"
            response2 = await self.client.post(
                f"{API_BASE}/techcards.v2/generate",
                json={"name": dish_name_2}
            )
            
            if response2.status_code != 200:
                self.log_test("Generate Real TechCard 2", False, 
                            f"Failed to generate second techcard: {response2.status_code} - {response2.text[:200]}")
                return
            
            data2 = response2.json()
            
            # Check if generation was successful
            if data2.get('status') not in ['success', 'draft']:
                self.log_test("Generate Real TechCard 2", False, 
                            f"Generation failed with status: {data2.get('status')} - {data2.get('message', 'No message')}")
                return
            
            # Get the techcard from response
            techcard_2 = data2.get('card')
            if not techcard_2 or not techcard_2.get('meta', {}).get('id'):
                self.log_test("Generate Real TechCard 2", False, 
                            f"No techcard in response: {data2}")
                return
            
            techcard_id_2 = techcard_2.get('meta', {}).get('id')
            
            # Store generated techcards
            self.generated_techcards = [
                {"id": techcard_id_1, "name": dish_name_1},
                {"id": techcard_id_2, "name": dish_name_2}
            ]
            
            response_time = time.time() - start_time
            self.log_test("Generate Real TechCards", True, 
                        f"Generated 2 real techcards: {dish_name_1} ({techcard_id_1[:8]}...), {dish_name_2} ({techcard_id_2[:8]}...)", 
                        response_time)
            
        except Exception as e:
            self.log_test("Generate Real TechCards", False, f"Exception: {str(e)}")
    
    async def test_02_run_preflight_check(self):
        """Run preflight check to get preflight data"""
        try:
            start_time = time.time()
            
            if not self.generated_techcards:
                self.log_test("Preflight Check", False, "No generated techcards available")
                return
            
            # For the export test, use 'current' which should work with the export system
            # The export system has special handling for 'current' techcard
            techcard_ids = ['current']
            
            response = await self.client.post(
                f"{API_BASE}/export/preflight",
                json={
                    "techcardIds": techcard_ids,
                    "organization_id": self.organization_id
                }
            )
            
            if response.status_code != 200:
                self.log_test("Preflight Check", False, 
                            f"Preflight failed: {response.status_code} - {response.text[:200]}")
                return
            
            data = response.json()
            response_time = time.time() - start_time
            
            # Debug: print actual response structure
            print(f"DEBUG: Preflight response: {json.dumps(data, indent=2)}")
            
            # Validate preflight response structure - adjust based on actual API
            # The API might return different field names
            if "ttkDate" in data:
                # Use the actual field names from the API
                self.preflight_data = data
                self.log_test("Preflight Check", True, 
                            f"Preflight successful - TTK date: {data.get('ttkDate')}, Response keys: {list(data.keys())}", 
                            response_time)
            else:
                # Check what fields are actually present
                self.log_test("Preflight Check", False, 
                            f"Unexpected response structure. Available fields: {list(data.keys())}")
                return
            
        except Exception as e:
            self.log_test("Preflight Check", False, f"Exception: {str(e)}")
    
    async def test_03_export_zip_file(self):
        """Export ZIP file using /export/zip endpoint"""
        try:
            start_time = time.time()
            
            if not hasattr(self, 'preflight_data'):
                self.log_test("ZIP Export", False, "No preflight data available")
                return
            
            # For the export test, use 'current' which should work with the export system
            techcard_ids = ['current']
            
            response = await self.client.post(
                f"{API_BASE}/export/zip",
                json={
                    "techcardIds": techcard_ids,
                    "organization_id": self.organization_id,
                    "preflight_result": self.preflight_data,
                    "operational_rounding": True
                }
            )
            
            if response.status_code != 200:
                self.log_test("ZIP Export", False, 
                            f"ZIP export failed: {response.status_code} - {response.text[:200]}")
                return
            
            # Check if response is actually a ZIP file
            content_type = response.headers.get('content-type', '')
            if 'application/zip' not in content_type and 'application/octet-stream' not in content_type:
                self.log_test("ZIP Export", False, 
                            f"Response is not a ZIP file. Content-Type: {content_type}")
                return
            
            self.zip_content = response.content
            response_time = time.time() - start_time
            
            self.log_test("ZIP Export", True, 
                        f"ZIP export successful - Size: {len(self.zip_content)} bytes", 
                        response_time)
            
        except Exception as e:
            self.log_test("ZIP Export", False, f"Exception: {str(e)}")
    
    async def test_04_extract_and_scan_xlsx_files(self):
        """Extract ZIP and scan XLSX files for mock content"""
        try:
            start_time = time.time()
            
            if not hasattr(self, 'zip_content'):
                self.log_test("XLSX Mock Content Scan", False, "No ZIP content available")
                return
            
            # Extract ZIP file
            with zipfile.ZipFile(io.BytesIO(self.zip_content), 'r') as zip_file:
                file_list = zip_file.namelist()
                xlsx_files = [f for f in file_list if f.endswith('.xlsx')]
                
                if not xlsx_files:
                    self.log_test("XLSX Mock Content Scan", False, "No XLSX files found in ZIP")
                    return
                
                mock_signatures_found = []
                total_cells_scanned = 0
                
                # Define mock signatures to look for
                mock_signatures = [
                    "DISH_MOCK_TECH_CARD",
                    "GENERATED_TEST_INGREDIENT", 
                    "GENERATED_TEST_INGREDIENT_2",
                    "MOCK_DISH",
                    "TEST_INGREDIENT",
                    "DEMO_CONTENT"
                ]
                
                for xlsx_file in xlsx_files:
                    # Extract and load XLSX file
                    xlsx_content = zip_file.read(xlsx_file)
                    
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                        temp_file.write(xlsx_content)
                        temp_file.flush()
                        
                        try:
                            # Load workbook and scan for mock content
                            workbook = load_workbook(temp_file.name, data_only=True)
                            
                            for sheet_name in workbook.sheetnames:
                                sheet = workbook[sheet_name]
                                
                                for row in sheet.iter_rows():
                                    for cell in row:
                                        if cell.value:
                                            cell_value = str(cell.value)
                                            total_cells_scanned += 1
                                            
                                            # Check for mock signatures
                                            for signature in mock_signatures:
                                                if signature in cell_value:
                                                    mock_signatures_found.append({
                                                        "file": xlsx_file,
                                                        "sheet": sheet_name,
                                                        "cell": f"{cell.column_letter}{cell.row}",
                                                        "signature": signature,
                                                        "value": cell_value
                                                    })
                            
                            workbook.close()
                            
                        except Exception as e:
                            self.log_test("XLSX Mock Content Scan", False, 
                                        f"Error reading XLSX file {xlsx_file}: {str(e)}")
                            return
                        finally:
                            os.unlink(temp_file.name)
                
                response_time = time.time() - start_time
                
                if mock_signatures_found:
                    # CRITICAL FAILURE - Mock content detected
                    mock_details = []
                    for mock in mock_signatures_found:
                        mock_details.append(f"{mock['file']}:{mock['sheet']}:{mock['cell']} = '{mock['signature']}'")
                    
                    self.log_test("XLSX Mock Content Scan", False, 
                                f"❌ CRITICAL: {len(mock_signatures_found)} mock signatures found in {len(xlsx_files)} XLSX files. "
                                f"Mock content: {'; '.join(mock_details[:3])}{'...' if len(mock_details) > 3 else ''}", 
                                response_time)
                else:
                    # SUCCESS - No mock content found
                    self.log_test("XLSX Mock Content Scan", True, 
                                f"✅ SUCCESS: No mock signatures found in {len(xlsx_files)} XLSX files. "
                                f"Scanned {total_cells_scanned} cells across all sheets.", 
                                response_time)
            
        except Exception as e:
            self.log_test("XLSX Mock Content Scan", False, f"Exception: {str(e)}")
    
    async def test_05_validate_article_format(self):
        """Validate that article fields contain proper 5-digit codes with leading zeros"""
        try:
            start_time = time.time()
            
            if not hasattr(self, 'zip_content'):
                self.log_test("Article Format Validation", False, "No ZIP content available")
                return
            
            # Extract ZIP file and check article formatting
            with zipfile.ZipFile(io.BytesIO(self.zip_content), 'r') as zip_file:
                file_list = zip_file.namelist()
                xlsx_files = [f for f in file_list if f.endswith('.xlsx')]
                
                article_validation_results = []
                
                for xlsx_file in xlsx_files:
                    xlsx_content = zip_file.read(xlsx_file)
                    
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                        temp_file.write(xlsx_content)
                        temp_file.flush()
                        
                        try:
                            workbook = load_workbook(temp_file.name, data_only=True)
                            
                            for sheet_name in workbook.sheetnames:
                                sheet = workbook[sheet_name]
                                
                                # Look for article columns (typically columns A and C)
                                article_columns = ['A', 'C']  # Common article column positions
                                
                                for col in article_columns:
                                    for row_num in range(2, min(sheet.max_row + 1, 20)):  # Check first 18 data rows
                                        cell = sheet[f"{col}{row_num}"]
                                        if cell.value:
                                            cell_value = str(cell.value)
                                            
                                            # Check if it looks like an article code (5 digits)
                                            if cell_value.isdigit() and len(cell_value) == 5:
                                                article_validation_results.append({
                                                    "file": xlsx_file,
                                                    "sheet": sheet_name,
                                                    "cell": f"{col}{row_num}",
                                                    "value": cell_value,
                                                    "valid_format": True
                                                })
                                            elif cell_value.isdigit() and len(cell_value) < 5:
                                                # Article without leading zeros - potential issue
                                                article_validation_results.append({
                                                    "file": xlsx_file,
                                                    "sheet": sheet_name,
                                                    "cell": f"{col}{row_num}",
                                                    "value": cell_value,
                                                    "valid_format": False,
                                                    "issue": "Missing leading zeros"
                                                })
                            
                            workbook.close()
                            
                        except Exception as e:
                            self.log_test("Article Format Validation", False, 
                                        f"Error reading XLSX file {xlsx_file}: {str(e)}")
                            return
                        finally:
                            os.unlink(temp_file.name)
                
                response_time = time.time() - start_time
                
                # Analyze results
                valid_articles = [r for r in article_validation_results if r.get('valid_format', False)]
                invalid_articles = [r for r in article_validation_results if not r.get('valid_format', True)]
                
                if article_validation_results:
                    if invalid_articles:
                        invalid_details = [f"{a['cell']}='{a['value']}'" for a in invalid_articles[:3]]
                        self.log_test("Article Format Validation", False, 
                                    f"Found {len(invalid_articles)} invalid article formats out of {len(article_validation_results)} total. "
                                    f"Invalid: {'; '.join(invalid_details)}", 
                                    response_time)
                    else:
                        self.log_test("Article Format Validation", True, 
                                    f"All {len(valid_articles)} article codes have proper 5-digit format with leading zeros", 
                                    response_time)
                else:
                    self.log_test("Article Format Validation", True, 
                                f"No article codes found to validate (may be expected for this export type)", 
                                response_time)
            
        except Exception as e:
            self.log_test("Article Format Validation", False, f"Exception: {str(e)}")
    
    async def test_06_validate_real_content(self):
        """Validate that XLSX files contain actual tech card names and ingredient data"""
        try:
            start_time = time.time()
            
            if not hasattr(self, 'zip_content') or not self.generated_techcards:
                self.log_test("Real Content Validation", False, "No ZIP content or techcards available")
                return
            
            # Get expected dish names
            expected_dish_names = [tc["name"] for tc in self.generated_techcards]
            
            # Extract ZIP file and check for real content
            with zipfile.ZipFile(io.BytesIO(self.zip_content), 'r') as zip_file:
                file_list = zip_file.namelist()
                xlsx_files = [f for f in file_list if f.endswith('.xlsx')]
                
                real_content_found = []
                
                for xlsx_file in xlsx_files:
                    xlsx_content = zip_file.read(xlsx_file)
                    
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                        temp_file.write(xlsx_content)
                        temp_file.flush()
                        
                        try:
                            workbook = load_workbook(temp_file.name, data_only=True)
                            
                            for sheet_name in workbook.sheetnames:
                                sheet = workbook[sheet_name]
                                
                                # Scan all cells for expected dish names
                                for row in sheet.iter_rows():
                                    for cell in row:
                                        if cell.value:
                                            cell_value = str(cell.value)
                                            
                                            # Check if any expected dish name is found
                                            for dish_name in expected_dish_names:
                                                if dish_name.lower() in cell_value.lower():
                                                    real_content_found.append({
                                                        "file": xlsx_file,
                                                        "sheet": sheet_name,
                                                        "cell": f"{cell.column_letter}{cell.row}",
                                                        "expected": dish_name,
                                                        "found": cell_value
                                                    })
                            
                            workbook.close()
                            
                        except Exception as e:
                            self.log_test("Real Content Validation", False, 
                                        f"Error reading XLSX file {xlsx_file}: {str(e)}")
                            return
                        finally:
                            os.unlink(temp_file.name)
                
                response_time = time.time() - start_time
                
                if real_content_found:
                    content_details = [f"{r['expected']} found in {r['file']}" for r in real_content_found[:2]]
                    self.log_test("Real Content Validation", True, 
                                f"✅ SUCCESS: Found {len(real_content_found)} instances of real tech card content. "
                                f"Examples: {'; '.join(content_details)}", 
                                response_time)
                else:
                    self.log_test("Real Content Validation", False, 
                                f"❌ CRITICAL: No real tech card content found in XLSX files. "
                                f"Expected dishes: {', '.join(expected_dish_names)}", 
                                response_time)
            
        except Exception as e:
            self.log_test("Real Content Validation", False, f"Exception: {str(e)}")
    
    def generate_summary_report(self):
        """Generate comprehensive test summary"""
        total_tests = len(self.test_results)
        passed_tests = sum(1 for result in self.test_results if result["success"])
        failed_tests = total_tests - passed_tests
        success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
        
        print("\n" + "="*80)
        print("🎯 FE EXPORT WIZARD MOCK CONTENT VALIDATION TEST RESULTS")
        print("="*80)
        print(f"📊 SUMMARY: {passed_tests}/{total_tests} tests passed ({success_rate:.1f}% success rate)")
        print(f"✅ PASSED: {passed_tests}")
        print(f"❌ FAILED: {failed_tests}")
        print("\n📋 DETAILED RESULTS:")
        
        for result in self.test_results:
            print(f"{result['status']} {result['test']}")
            if result['details']:
                print(f"   └─ {result['details']}")
            if result['response_time'] != "N/A":
                print(f"   └─ Response time: {result['response_time']}")
        
        # Critical assessment
        print("\n🎯 CRITICAL ASSESSMENT:")
        
        mock_scan_result = next((r for r in self.test_results if "Mock Content Scan" in r["test"]), None)
        if mock_scan_result:
            if mock_scan_result["success"]:
                print("✅ MOCK CONTENT: No mock signatures detected in exported XLSX files")
            else:
                print("❌ MOCK CONTENT: Mock signatures still present in exported XLSX files")
        
        article_format_result = next((r for r in self.test_results if "Article Format" in r["test"]), None)
        if article_format_result:
            if article_format_result["success"]:
                print("✅ ARTICLE FORMAT: All article codes have proper 5-digit format with leading zeros")
            else:
                print("❌ ARTICLE FORMAT: Some article codes missing leading zeros")
        
        real_content_result = next((r for r in self.test_results if "Real Content" in r["test"]), None)
        if real_content_result:
            if real_content_result["success"]:
                print("✅ REAL CONTENT: Actual tech card data found in exported files")
            else:
                print("❌ REAL CONTENT: No real tech card content detected")
        
        # Overall verdict
        critical_tests = [mock_scan_result, real_content_result]
        critical_passed = sum(1 for test in critical_tests if test and test["success"])
        
        print(f"\n🏆 OVERALL VERDICT:")
        if critical_passed == len([t for t in critical_tests if t]):
            print("✅ SUCCESS: FE Export Wizard is properly configured to use real data without mock content")
        else:
            print("❌ FAILURE: FE Export Wizard still contains mock content or missing real data")
        
        return {
            "total_tests": total_tests,
            "passed_tests": passed_tests,
            "failed_tests": failed_tests,
            "success_rate": success_rate,
            "critical_mock_content_eliminated": mock_scan_result["success"] if mock_scan_result else False,
            "real_content_present": real_content_result["success"] if real_content_result else False
        }

async def main():
    """Main test execution"""
    print("🚀 Starting FE Export Wizard Mock Content Validation Test")
    print(f"🔗 Backend URL: {BACKEND_URL}")
    print(f"🔗 API Base: {API_BASE}")
    
    async with FEExportWizardMockValidationTester() as tester:
        # Execute test sequence
        await tester.test_01_generate_real_techcards()
        await tester.test_02_run_preflight_check()
        await tester.test_03_export_zip_file()
        await tester.test_04_extract_and_scan_xlsx_files()
        await tester.test_05_validate_article_format()
        await tester.test_06_validate_real_content()
        
        # Generate summary
        summary = tester.generate_summary_report()
        
        # Save results to file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        results_file = f"/app/fe_export_wizard_mock_validation_results_{timestamp}.json"
        
        with open(results_file, 'w') as f:
            json.dump({
                "timestamp": timestamp,
                "summary": summary,
                "detailed_results": tester.test_results,
                "generated_techcards": tester.generated_techcards
            }, f, indent=2)
        
        print(f"\n💾 Results saved to: {results_file}")
        
        return summary["critical_mock_content_eliminated"] and summary["real_content_present"]

if __name__ == "__main__":
    try:
        result = asyncio.run(main())
        sys.exit(0 if result else 1)
    except Exception as e:
        print(f"❌ Test execution failed: {str(e)}")
        traceback.print_exc()
        sys.exit(1)