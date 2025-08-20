#!/usr/bin/env python3
"""
Product Skeletons (когда маппинг не нашёлся) Comprehensive Testing
Testing the complete C. Product Skeletons implementation for missing product creation.
"""

import requests
import json
import time
import os
from typing import Dict, List, Any
import openpyxl
from io import BytesIO

# Get backend URL from environment
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://iiko-connect.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"

class ProductSkeletonsTestSuite:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        })
        self.test_results = []
        
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status} {test_name} ({response_time:.3f}s): {details}")
        
    def create_test_techcards(self) -> List[Dict[str, Any]]:
        """Create test techcards with mixed ingredient mapping states"""
        return [
            {
                "meta": {
                    "title": "Салат Цезарь с курицей"
                },
                "ingredients": [
                    {
                        "name": "Куриное филе",
                        "brutto_g": 150,
                        "netto_g": 120,
                        "loss_pct": 20,
                        "unit": "g",
                        "skuId": "guid-chicken-123",  # Has GUID but no product_code
                        "product_code": None
                    },
                    {
                        "name": "Салат Романо",
                        "brutto_g": 100,
                        "netto_g": 90,
                        "loss_pct": 10,
                        "unit": "g",
                        "skuId": None,  # No mapping at all
                        "product_code": None
                    },
                    {
                        "name": "Пармезан",
                        "brutto_g": 50,
                        "netto_g": 50,
                        "loss_pct": 0,
                        "unit": "g",
                        "skuId": "guid-parmesan-456",
                        "product_code": "12345"  # Has product_code - should not be missing
                    },
                    {
                        "name": "Гренки",
                        "brutto_g": 30,
                        "netto_g": 30,
                        "loss_pct": 0,
                        "unit": "g",
                        "skuId": None,
                        "product_code": None
                    },
                    {
                        "name": "Соус Цезарь",
                        "brutto_g": 40,
                        "netto_g": 40,
                        "loss_pct": 0,
                        "unit": "ml",
                        "skuId": "guid-sauce-789",
                        "product_code": "00678"  # Has product_code with leading zeros
                    }
                ],
                "yield": {
                    "perPortion_g": 370,
                    "perBatch_g": 370
                },
                "portions": 1,
                "process": [
                    {
                        "n": 1,
                        "action": "Нарезать куриное филе кубиками",
                        "temp_c": None,
                        "time_min": 5,
                        "equipment": ["нож", "доска"]
                    },
                    {
                        "n": 2,
                        "action": "Промыть и нарезать салат",
                        "temp_c": None,
                        "time_min": 3,
                        "equipment": ["нож"]
                    },
                    {
                        "n": 3,
                        "action": "Смешать ингредиенты с соусом",
                        "temp_c": None,
                        "time_min": 2,
                        "equipment": ["миска"]
                    }
                ],
                "storage": {
                    "conditions": "Хранить в холодильнике при температуре +2...+6°C",
                    "shelfLife_hours": 24,
                    "servingTemp_c": 8
                }
            },
            {
                "meta": {
                    "title": "Борщ украинский"
                },
                "ingredients": [
                    {
                        "name": "Говядина",
                        "brutto_g": 200,
                        "netto_g": 180,
                        "loss_pct": 10,
                        "unit": "g",
                        "skuId": None,  # No mapping
                        "product_code": None
                    },
                    {
                        "name": "Свекла",
                        "brutto_g": 150,
                        "netto_g": 120,
                        "loss_pct": 20,
                        "unit": "g",
                        "skuId": None,
                        "product_code": None
                    },
                    {
                        "name": "Морковь",
                        "brutto_g": 80,
                        "netto_g": 70,
                        "loss_pct": 12.5,
                        "unit": "g",
                        "skuId": "guid-carrot-999",
                        "product_code": None  # Has GUID but no product_code
                    }
                ],
                "yield": {
                    "perPortion_g": 370,
                    "perBatch_g": 370
                },
                "portions": 1,
                "process": [
                    {
                        "n": 1,
                        "action": "Варить говядину в воде",
                        "temp_c": 100,
                        "time_min": 60,
                        "equipment": ["кастрюля", "плита"]
                    },
                    {
                        "n": 2,
                        "action": "Нарезать и добавить овощи",
                        "temp_c": None,
                        "time_min": 10,
                        "equipment": ["нож", "доска"]
                    },
                    {
                        "n": 3,
                        "action": "Варить до готовности",
                        "temp_c": 90,
                        "time_min": 30,
                        "equipment": ["кастрюля"]
                    }
                ],
                "storage": {
                    "conditions": "Хранить в холодильнике при температуре +2...+6°C",
                    "shelfLife_hours": 48,
                    "servingTemp_c": 65
                }
            }
        ]
    
    def test_find_missing_endpoint(self):
        """Test POST /api/v1/techcards.v2/product-codes/find-missing endpoint"""
        test_name = "Missing Product Detection API"
        start_time = time.time()
        
        try:
            test_techcards = self.create_test_techcards()
            
            payload = {
                "techcards": test_techcards,
                "organization_id": "default"
            }
            
            response = self.session.post(
                f"{API_BASE}/techcards.v2/product-codes/find-missing",
                json=payload,
                timeout=30
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ["status", "missing_products", "total_missing", "cards_analyzed"]
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test(test_name, False, f"Missing response fields: {missing_fields}", response_time)
                    return
                
                # Check that correct ingredients are identified as missing
                missing_products = data["missing_products"]
                missing_names = [p["name"] for p in missing_products]
                
                # Expected missing: Куриное филе, Салат Романо, Гренки, Говядина, Свекла, Морковь
                # Not missing: Пармезан (has product_code), Соус Цезарь (has product_code)
                expected_missing = {"Куриное филе", "Салат Романо", "Гренки", "Говядина", "Свекла", "Морковь"}
                actual_missing = set(missing_names)
                
                if expected_missing.issubset(actual_missing):
                    self.log_test(test_name, True, 
                                f"Found {data['total_missing']} missing products correctly: {missing_names}", 
                                response_time)
                else:
                    missing_expected = expected_missing - actual_missing
                    self.log_test(test_name, False, 
                                f"Missing detection incomplete. Expected but not found: {missing_expected}", 
                                response_time)
            else:
                self.log_test(test_name, False, f"HTTP {response.status_code}: {response.text}", response_time)
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(test_name, False, f"Exception: {str(e)}", response_time)
    
    def test_generate_codes_endpoint(self):
        """Test POST /api/v1/techcards.v2/product-codes/generate endpoint"""
        test_name = "Product Code Generation API"
        start_time = time.time()
        
        try:
            ingredient_names = [
                "Куриное филе",
                "Салат Романо", 
                "Гренки",
                "Говядина",
                "Свекла",
                "Морковь"
            ]
            
            payload = {
                "ingredient_names": ingredient_names,
                "start_code": 10000,
                "code_width": 5,
                "organization_id": "default"
            }
            
            response = self.session.post(
                f"{API_BASE}/techcards.v2/product-codes/generate",
                json=payload,
                timeout=30
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ["status", "generated_codes", "codes_generated", "start_code", "code_width"]
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test(test_name, False, f"Missing response fields: {missing_fields}", response_time)
                    return
                
                generated_codes = data["generated_codes"]
                
                # Validate code generation
                if len(generated_codes) != len(ingredient_names):
                    self.log_test(test_name, False, 
                                f"Expected {len(ingredient_names)} codes, got {len(generated_codes)}", 
                                response_time)
                    return
                
                # Check code format (5 digits with leading zeros)
                valid_codes = True
                code_issues = []
                
                for ingredient, code in generated_codes.items():
                    if not isinstance(code, str) or len(code) != 5 or not code.isdigit():
                        valid_codes = False
                        code_issues.append(f"{ingredient}: {code}")
                
                # Check uniqueness
                codes_list = list(generated_codes.values())
                if len(codes_list) != len(set(codes_list)):
                    valid_codes = False
                    code_issues.append("Duplicate codes found")
                
                # Check proper formatting (leading zeros)
                for ingredient, code in generated_codes.items():
                    if code.startswith('0') and len(code) == 5:
                        # Good - has leading zeros
                        continue
                    elif int(code) >= 10000:
                        # Also good - 5 digit number
                        continue
                    else:
                        valid_codes = False
                        code_issues.append(f"Invalid format: {ingredient}: {code}")
                
                if valid_codes:
                    self.log_test(test_name, True, 
                                f"Generated {data['codes_generated']} unique codes with proper formatting", 
                                response_time)
                    # Store codes for export test
                    self.generated_codes = generated_codes
                else:
                    self.log_test(test_name, False, f"Code validation failed: {code_issues}", response_time)
                    
            else:
                self.log_test(test_name, False, f"HTTP {response.status_code}: {response.text}", response_time)
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(test_name, False, f"Exception: {str(e)}", response_time)
    
    def test_export_skeletons_endpoint(self):
        """Test POST /api/v1/techcards.v2/product-skeletons/export endpoint"""
        test_name = "Product Skeletons Export API"
        start_time = time.time()
        
        try:
            # Use codes from previous test or generate new ones
            if not hasattr(self, 'generated_codes'):
                self.generated_codes = {
                    "Куриное филе": "10001",
                    "Салат Романо": "10002",
                    "Гренки": "10003",
                    "Говядина": "10004",
                    "Свекла": "10005",
                    "Морковь": "10006"
                }
            
            missing_ingredients = [
                {"name": "Куриное филе", "unit": "г"},
                {"name": "Салат Романо", "unit": "г"},
                {"name": "Гренки", "unit": "г"},
                {"name": "Говядина", "unit": "г"},
                {"name": "Свекла", "unit": "г"},
                {"name": "Морковь", "unit": "г"}
            ]
            
            payload = {
                "missing_ingredients": missing_ingredients,
                "generated_codes": self.generated_codes
            }
            
            response = self.session.post(
                f"{API_BASE}/techcards.v2/product-skeletons/export",
                json=payload,
                timeout=30
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # Check content type
                content_type = response.headers.get('content-type', '')
                if 'spreadsheet' not in content_type and 'excel' not in content_type:
                    self.log_test(test_name, False, f"Wrong content type: {content_type}", response_time)
                    return
                
                # Check content disposition
                content_disposition = response.headers.get('content-disposition', '')
                if 'Product-Skeletons.xlsx' not in content_disposition:
                    self.log_test(test_name, False, f"Wrong filename in headers: {content_disposition}", response_time)
                    return
                
                # Validate Excel file structure
                try:
                    excel_data = BytesIO(response.content)
                    workbook = openpyxl.load_workbook(excel_data)
                    worksheet = workbook.active
                    
                    # Check headers
                    expected_headers = ["Артикул", "Наименование", "Ед. изм", "Тип", "Группа", "Штрихкод", "Поставщик"]
                    actual_headers = [cell.value for cell in worksheet[1]]
                    
                    if actual_headers != expected_headers:
                        self.log_test(test_name, False, 
                                    f"Wrong headers. Expected: {expected_headers}, Got: {actual_headers}", 
                                    response_time)
                        return
                    
                    # Check data rows
                    data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
                    if len(data_rows) != len(missing_ingredients):
                        self.log_test(test_name, False, 
                                    f"Expected {len(missing_ingredients)} data rows, got {len(data_rows)}", 
                                    response_time)
                        return
                    
                    # Validate data content
                    validation_issues = []
                    for i, row in enumerate(data_rows):
                        ingredient_name = missing_ingredients[i]["name"]
                        expected_code = self.generated_codes[ingredient_name]
                        
                        # Check code (column A)
                        if str(row[0]) != expected_code:
                            validation_issues.append(f"Row {i+2}: Wrong code {row[0]}, expected {expected_code}")
                        
                        # Check name (column B)
                        if row[1] != ingredient_name:
                            validation_issues.append(f"Row {i+2}: Wrong name {row[1]}, expected {ingredient_name}")
                        
                        # Check unit (column C)
                        if row[2] != "г":
                            validation_issues.append(f"Row {i+2}: Wrong unit {row[2]}, expected г")
                        
                        # Check type (column D)
                        if row[3] != "Товар":
                            validation_issues.append(f"Row {i+2}: Wrong type {row[3]}, expected Товар")
                    
                    if validation_issues:
                        self.log_test(test_name, False, f"Data validation failed: {validation_issues[:3]}", response_time)
                    else:
                        file_size = len(response.content)
                        self.log_test(test_name, True, 
                                    f"Valid Product-Skeletons.xlsx created ({file_size} bytes) with correct structure", 
                                    response_time)
                        
                except Exception as excel_error:
                    self.log_test(test_name, False, f"Excel validation error: {str(excel_error)}", response_time)
                    
            else:
                self.log_test(test_name, False, f"HTTP {response.status_code}: {response.text}", response_time)
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(test_name, False, f"Exception: {str(e)}", response_time)
    
    def test_complete_workflow(self):
        """Test complete Product Skeletons workflow: analyze → generate codes → export XLSX"""
        test_name = "Complete Product Skeletons Workflow"
        start_time = time.time()
        
        try:
            # Step 1: Create test techcards
            test_techcards = self.create_test_techcards()
            
            # Step 2: Find missing products
            find_payload = {
                "techcards": test_techcards,
                "organization_id": "default"
            }
            
            find_response = self.session.post(
                f"{API_BASE}/techcards.v2/product-codes/find-missing",
                json=find_payload,
                timeout=30
            )
            
            if find_response.status_code != 200:
                self.log_test(test_name, False, f"Step 1 failed: {find_response.status_code}", time.time() - start_time)
                return
            
            missing_data = find_response.json()
            missing_products = missing_data["missing_products"]
            
            if not missing_products:
                self.log_test(test_name, False, "No missing products found", time.time() - start_time)
                return
            
            # Step 3: Generate codes for missing products
            ingredient_names = [p["name"] for p in missing_products]
            
            generate_payload = {
                "ingredient_names": ingredient_names,
                "start_code": 40000,
                "code_width": 5,
                "organization_id": "default"
            }
            
            generate_response = self.session.post(
                f"{API_BASE}/techcards.v2/product-codes/generate",
                json=generate_payload,
                timeout=30
            )
            
            if generate_response.status_code != 200:
                self.log_test(test_name, False, f"Step 2 failed: {generate_response.status_code}", time.time() - start_time)
                return
            
            codes_data = generate_response.json()
            generated_codes = codes_data["generated_codes"]
            
            # Step 4: Export Product-Skeletons.xlsx
            export_payload = {
                "missing_ingredients": missing_products,
                "generated_codes": generated_codes
            }
            
            export_response = self.session.post(
                f"{API_BASE}/techcards.v2/product-skeletons/export",
                json=export_payload,
                timeout=30
            )
            
            response_time = time.time() - start_time
            
            if export_response.status_code == 200:
                # Validate final Excel file
                excel_data = BytesIO(export_response.content)
                workbook = openpyxl.load_workbook(excel_data)
                worksheet = workbook.active
                
                data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
                
                if len(data_rows) == len(missing_products):
                    file_size = len(export_response.content)
                    self.log_test(test_name, True, 
                                f"Complete workflow successful: {len(missing_products)} missing → {len(generated_codes)} codes → {file_size}B Excel", 
                                response_time)
                else:
                    self.log_test(test_name, False, 
                                f"Row count mismatch: {len(data_rows)} rows vs {len(missing_products)} missing products", 
                                response_time)
            else:
                self.log_test(test_name, False, f"Step 3 failed: {export_response.status_code}", response_time)
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(test_name, False, f"Exception: {str(e)}", response_time)
    
    def run_all_tests(self):
        """Run all Product Skeletons tests"""
        print("🎯 STARTING PRODUCT SKELETONS COMPREHENSIVE TESTING")
        print(f"Backend URL: {BACKEND_URL}")
        print("=" * 80)
        
        # Test individual endpoints
        self.test_find_missing_endpoint()
        self.test_generate_codes_endpoint()
        self.test_export_skeletons_endpoint()
        
        # Test complete workflow
        self.test_complete_workflow()
        
        # Summary
        print("\n" + "=" * 80)
        print("🎯 PRODUCT SKELETONS TESTING SUMMARY")
        print("=" * 80)
        
        passed = sum(1 for result in self.test_results if result["success"])
        total = len(self.test_results)
        success_rate = (passed / total * 100) if total > 0 else 0
        
        print(f"Tests Passed: {passed}/{total} ({success_rate:.1f}%)")
        print()
        
        for result in self.test_results:
            print(f"{result['status']} {result['test']} ({result['response_time']})")
            if result['details']:
                print(f"    {result['details']}")
        
        return passed, total, self.test_results

if __name__ == "__main__":
    test_suite = ProductSkeletonsTestSuite()
    passed, total, results = test_suite.run_all_tests()
    
    # Exit with appropriate code
    exit(0 if passed == total else 1)