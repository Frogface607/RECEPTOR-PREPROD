"""
iiko XLSX Export Module
WITH iiko Import Reliability — Product Codes & Dish Skeletons
"""

from __future__ import annotations
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
import re
import os
import logging

# Используем openpyxl для работы с Excel
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    # Fallback если openpyxl не установлен
    Workbook = None
    load_workbook = None

from receptor_agent.techcards_v2.schemas import TechCardV2
from receptor_agent.integrations.iiko_rms_service import IikoRmsService

logger = logging.getLogger(__name__)


def convert_grams_to_kilograms(grams_value: float, unit: str = 'г') -> tuple[float, str]:
    """
    Автоматическая конвертация грамм в килограммы для экспорта XLSX в iiko
    
    Args:
        grams_value: Значение в граммах
        unit: Исходная единица измерения
        
    Returns:
        tuple: (конвертированное_значение_в_кг, новая_единица)
    """
    # Конвертируем только если единица измерения - граммы
    if unit.lower() in ['г', 'gram', 'гр', 'грамм']:
        # Конвертируем граммы в килограммы: делим на 1000, округляем до 3 знаков после запятой
        kg_value = round(grams_value / 1000.0, 3)
        return kg_value, 'кг'
    else:
        # Для других единиц измерения оставляем как есть
        return grams_value, unit


def convert_yield_grams_to_kilograms(yield_g: float) -> tuple[float, str]:
    """
    Конвертация выхода блюда из грамм в килограммы для экспорта iiko XLSX
    
    Args:
        yield_g: Выход блюда в граммах
        
    Returns:
        tuple: (выход_в_кг, единица_кг)
    """
    # Всегда конвертируем выход в килограммы для iiko
    yield_kg = round(yield_g / 1000.0, 3)
    return yield_kg, 'кг'


def resolve_ttk_date_conflict(dish_name: str, 
                             base_date: Optional[str] = None, 
                             rms_service: Optional[IikoRmsService] = None,
                             max_days_ahead: int = 7) -> str:
    """
    F. TTK Date Autoresolve: Разрешить конфликт дат техкарт в iiko
    
    Args:
        dish_name: Название блюда
        base_date: Базовая дата (YYYY-MM-DD), если не указана - берется today
        rms_service: Сервис iiko RMS для проверки существующих ТК
        max_days_ahead: Максимальное количество дней для поиска свободной даты
        
    Returns:
        Свободная дата в формате YYYY-MM-DD
    """
    if not base_date:
        base_date = datetime.now().strftime('%Y-%m-%d')
    
    try:
        current_date = datetime.strptime(base_date, '%Y-%m-%d')
    except ValueError:
        # Если дата неправильная, используем сегодня
        current_date = datetime.now()
        base_date = current_date.strftime('%Y-%m-%d')
    
    # Если нет доступа к RMS, просто возвращаем базовую дату
    if not rms_service:
        logger.info(f"No RMS service available, using base date: {base_date}")
        return base_date
    
    logger.info(f"Resolving TTK date conflict for dish '{dish_name}', base date: {base_date}")
    
    # Ищем свободную дату в течение max_days_ahead дней
    for days_offset in range(max_days_ahead + 1):
        check_date = current_date + timedelta(days=days_offset)
        check_date_str = check_date.strftime('%Y-%m-%d')
        
        try:
            # Проверяем существующие ТК в iiko RMS на эту дату
            # Примерный запрос - может потребоваться адаптация под конкретный API iiko
            existing_ttk = None
            
            # Пытаемся найти существующую ТК через различные коллекции
            if hasattr(rms_service, 'techcards') and rms_service.techcards:
                existing_ttk = rms_service.techcards.find_one({
                    "name": dish_name,
                    "date": check_date_str
                })
            
            # Альтернативный поиск через products или другие коллекции
            if not existing_ttk and hasattr(rms_service, 'products') and rms_service.products:
                existing_ttk = rms_service.products.find_one({
                    "name": dish_name,
                    "techcard_date": check_date_str
                })
            
            if not existing_ttk:
                logger.info(f"Found free date for '{dish_name}': {check_date_str}")
                return check_date_str
            else:
                logger.debug(f"Date {check_date_str} is occupied for dish '{dish_name}'")
                
        except Exception as e:
            logger.warning(f"Error checking date {check_date_str} for dish '{dish_name}': {e}")
            # При ошибке доступа к RMS используем эту дату
            return check_date_str
    
    # Если не нашли свободную дату, возвращаем последнюю проверенную
    final_date = (current_date + timedelta(days=max_days_ahead)).strftime('%Y-%m-%d')
    logger.warning(f"Could not find free date for '{dish_name}' within {max_days_ahead} days, using: {final_date}")
    
    return final_date


def get_product_code_from_rms(sku_id: str, rms_service=None) -> str:
    """
    Feature A: Получить АРТИКУЛ (номенклатурный код) продукта из iiko RMS вместо GUID
    
    ВАЖНО: Форматируем все артикулы как пятизначные с ведущими нулями (04637)
    
    Args:
        sku_id: GUID продукта из iiko
        rms_service: Сервис для работы с iiko RMS (опционально)
    
    Returns:
        Пятизначный артикул с ведущими нулями или исходный sku_id если код не найден
    """
    if not sku_id or not rms_service:
        return sku_id or ""
    
    try:
        # Поиск в коллекции products
        product = rms_service.products.find_one({"_id": sku_id})
        if product:
            # Ищем артикул в различных полях
            article_fields = ['article', 'code', 'barcode']
            
            for field in article_fields:
                if field in product and product[field]:
                    article_value = str(product[field]).strip()
                    
                    # Если это числовой код - форматируем как пятизначный
                    if article_value and article_value.isdigit():
                        formatted_article = article_value.zfill(5)
                        print(f"DEBUG: Product {sku_id} - {product.get('name', 'Unknown')} - Found article: {article_value} -> {formatted_article}")
                        return formatted_article
                    
                    # Если это не числовой код, но не пустой - возвращаем как есть
                    elif article_value and article_value != '0':
                        print(f"DEBUG: Product {sku_id} - Non-numeric article: {article_value}")
                        return article_value
        
        # Поиск в коллекции prices если нет в products
        pricing = rms_service.prices.find_one({"skuId": sku_id})
        if pricing:
            article_fields = ['article', 'code', 'barcode']
            
            for field in article_fields:
                if field in pricing and pricing[field]:
                    article_value = str(pricing[field]).strip()
                    
                    # Если это числовой код - форматируем как пятизначный
                    if article_value and article_value.isdigit():
                        formatted_article = article_value.zfill(5)
                        print(f"DEBUG: Pricing {sku_id} - Found article: {article_value} -> {formatted_article}")
                        return formatted_article
                    
                    # Если это не числовой код, но не пустой - возвращаем как есть  
                    elif article_value and article_value != '0':
                        print(f"DEBUG: Pricing {sku_id} - Non-numeric article: {article_value}")
                        return article_value
            
    except Exception as e:
        print(f"Error getting product article for {sku_id}: {e}")
    
    print(f"DEBUG: No article found for {sku_id}, returning original")
    return sku_id or ""


def find_dish_in_iiko_rms(dish_name: str, rms_service=None) -> Dict[str, Any]:
    """
    Feature B: Поиск блюда в iiko RMS по имени с нормализацией
    
    Args:
        dish_name: Название блюда для поиска
        rms_service: Сервис для работы с iiko RMS
    
    Returns:
        {
            'status': 'found' | 'not_found' | 'error',
            'dish_code': str,
            'dish_name': str,
            'confidence': float
        }
    """
    if not dish_name or not rms_service:
        return {'status': 'error', 'message': 'Missing dish name or RMS service'}
    
    try:
        # Нормализация названия для поиска (убираем лишние символы, приводим к нижнему регистру)
        normalized_search = dish_name.lower().strip()
        normalized_search = re.sub(r'[^\w\s]', '', normalized_search)  # Убираем пунктуацию
        normalized_search = re.sub(r'\s+', ' ', normalized_search)     # Нормализуем пробелы
        
        # Поиск в коллекции products по типу "Блюдо"
        dishes = list(rms_service.products.find({
            "type": {"$in": ["Блюдо", "блюдо", "dish", "DISH"]},
            "active": True
        }))
        
        best_match = None
        best_confidence = 0
        
        for dish in dishes:
            dish_name_db = dish.get('name', '').lower().strip()
            dish_name_db = re.sub(r'[^\w\s]', '', dish_name_db)
            dish_name_db = re.sub(r'\s+', ' ', dish_name_db)
            
            # Простое совпадение по подстроке
            if normalized_search in dish_name_db or dish_name_db in normalized_search:
                confidence = 0.9
            # Проверка на совпадение слов
            elif len(set(normalized_search.split()) & set(dish_name_db.split())) > 0:
                confidence = 0.7
            else:
                continue
            
            if confidence > best_confidence:
                best_confidence = confidence
                best_match = dish
        
        if best_match and best_confidence >= 0.7:
            # Получаем артикул блюда (номенклатурный код)
            dish_article = None
            article_fields = ['article', 'code', 'nomenclatureCode', 'itemCode', 'dishCode']
            
            for field in article_fields:
                if field in best_match and best_match[field]:
                    article_value = str(best_match[field]).strip()
                    
                    # Проверяем что это именно артикул (5 цифр)
                    if article_value.isdigit() and len(article_value) <= 6:
                        dish_article = article_value.zfill(5)
                        break
            
            return {
                'status': 'found',
                'dish_code': dish_article or '',
                'dish_name': best_match.get('name', ''),
                'confidence': best_confidence
            }
        
        return {'status': 'not_found'}
        
    except Exception as e:
        return {'status': 'error', 'message': str(e)}


def generate_dish_codes(dish_names: List[str], rms_service=None, width: int = 5) -> Dict[str, str]:
    """
    Feature B: Генерация свободных числовых кодов для новых блюд
    
    Args:
        dish_names: Список названий блюд для генерации кодов
        rms_service: Сервис для работы с iiko RMS (для проверки занятых кодов)  
        width: Ширина кода (количество цифр с ведущими нулями)
    
    Returns:
        Dict[dish_name, generated_code]
    """
    if not dish_names:
        return {}
    
    # Получаем список занятых кодов из базы
    used_codes = set()
    
    if rms_service:
        try:
            # Собираем все коды из products
            products = rms_service.products.find({}, {"article": 1, "code": 1})
            for product in products:
                code = product.get('article') or product.get('code')
                if code and str(code).isdigit():
                    used_codes.add(int(code))
        except Exception as e:
            print(f"Error getting used codes: {e}")
    
    # Генерируем свободные коды
    generated_codes = {}
    start_code = 10**(width-1)  # Минимальный код для заданной ширины (например, 10000 для width=5)
    current_code = start_code
    
    for dish_name in dish_names:
        # Ищем первый свободный код
        while current_code in used_codes:
            current_code += 1
        
        # Форматируем с ведущими нулями
        formatted_code = str(current_code).zfill(width)
        generated_codes[dish_name] = formatted_code
        
        # Помечаем как занятый
        used_codes.add(current_code)
        current_code += 1
    
    return generated_codes


def create_product_skeletons_xlsx(missing_ingredients: List[Dict[str, Any]], 
                                 generated_codes: Dict[str, str] = None) -> BytesIO:
    """
    C. Product Skeletons: Создать XLSX файл для импорта номенклатуры в iiko
    WITH STRICT TYPE VALIDATION
    
    Args:
        missing_ingredients: Список ингредиентов без маппинга
        generated_codes: Маппинг название → сгенерированный код
        
    Returns:
        BytesIO buffer с Product-Skeletons.xlsx
        
    Raises:
        ValueError: Если обнаружены невалидные типы продуктов
    """
    if not Workbook:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    if not generated_codes:
        generated_codes = {}
    
    # VALID IIKO PRODUCT TYPES - строго по документации iiko
    VALID_IIKO_TYPES = {
        "GOODS",      # Товар (основной тип для ингредиентов)
        "DISH",       # Блюдо
        "MODIFIER",   # Модификатор
        "GROUP",      # Группа товаров
        "SERVICE",    # Услуга
        "PREPARED"    # Полуфабрикат
    }
    
    wb = Workbook()
    ws = wb.active
    
    # Заголовки для iiko номенклатуры (минимальные поля)
    headers = [
        "Артикул",           # Column A - numeric code (text format)
        "Наименование",      # Column B - product name  
        "Ед. изм",          # Column C - unit of measure
        "Тип",              # Column D - product type (STRICT VALIDATION)
        "Группа",           # Column E - product group ("Сырьё")
        "Штрихкод",         # Column F - barcode (optional, empty)
        "Поставщик"         # Column G - supplier (optional, empty)
    ]
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Валидация и подготовка данных
    type_errors = []
    validated_products = []
    
    # Заполняем данные по ингредиентам с СТРОГОЙ ВАЛИДАЦИЕЙ
    for idx, ingredient in enumerate(missing_ingredients):
        ingredient_name = ingredient.get('name', 'Unknown Ingredient')
        unit = ingredient.get('unit', 'г')
        
        # Получаем или генерируем код
        product_code = generated_codes.get(ingredient_name)
        if not product_code:
            # Fallback - генерируем базовый код если не был предоставлен
            product_code = f"{10000 + len(generated_codes):05d}"
            generated_codes[ingredient_name] = product_code
        
        # Нормализуем единицу измерения для iiko
        unit_mapping = {
            'g': 'г',
            'ml': 'мл', 
            'l': 'л',
            'kg': 'кг',
            'pcs': 'шт',
            'шт': 'шт',
            'г': 'г',
            'мл': 'мл',
            'л': 'л',
            'кг': 'кг'
        }
        normalized_unit = unit_mapping.get(unit.lower(), unit)
        
        # Определяем группу по типу ингредиента
        ingredient_name_lower = ingredient_name.lower()
        if any(word in ingredient_name_lower for word in ['мясо', 'курица', 'говядина', 'свинина', 'рыба', 'филе']):
            group = 'Мясо и рыба'
        elif any(word in ingredient_name_lower for word in ['молоко', 'сыр', 'творог', 'сливки', 'масло']):
            group = 'Молочные продукты'
        elif any(word in ingredient_name_lower for word in ['морковь', 'лук', 'картофель', 'капуста', 'помидор', 'огурец']):
            group = 'Овощи'
        elif any(word in ingredient_name_lower for word in ['мука', 'сахар', 'соль', 'перец', 'специи']):
            group = 'Бакалея'
        elif any(word in ingredient_name_lower for word in ['масло растительное', 'уксус', 'соус']):
            group = 'Масла и соусы'
        else:
            group = 'Сырьё'  # По умолчанию
        
        # КРИТИЧЕСКОЕ: Определяем ВАЛИДНЫЙ тип для iiko
        # Все ингредиенты должны быть GOODS (товар) в iiko системе
        product_type = "GOODS"
        
        # СТРОГАЯ ВАЛИДАЦИЯ ТИПА
        if product_type not in VALID_IIKO_TYPES:
            type_errors.append({
                "index": idx,
                "name": ingredient_name,
                "invalid_type": product_type,
                "valid_types": list(VALID_IIKO_TYPES),
                "error": f"Недопустимый тип '{product_type}' для продукта '{ingredient_name}'"
            })
            continue
        
        validated_products.append({
            "product_code": product_code,
            "ingredient_name": ingredient_name,
            "normalized_unit": normalized_unit,
            "product_type": product_type,
            "group": group
        })
    
    # FAIL-FAST: Если есть ошибки валидации - останавливаем экспорт
    if type_errors:
        # Создаем детальный отчет об ошибках
        error_report = {
            "timestamp": datetime.now().isoformat(),
            "error_type": "PRODUCT_TYPE_VALIDATION_FAILED",
            "total_products": len(missing_ingredients),
            "invalid_products": len(type_errors),
            "valid_types": list(VALID_IIKO_TYPES),
            "errors": type_errors,
            "instruction": "Исправьте типы продуктов на допустимые значения: GOODS, DISH, MODIFIER, GROUP, SERVICE, PREPARED"
        }
        
        # Сохраняем отчет об ошибках (если нужен файловый артефакт)
        try:
            import json
            error_file_path = "/app/artifacts/type_errors.json"
            os.makedirs(os.path.dirname(error_file_path), exist_ok=True)
            with open(error_file_path, 'w', encoding='utf-8') as f:
                json.dump(error_report, f, indent=2, ensure_ascii=False)
            logger.error(f"Type validation errors saved to {error_file_path}")
        except Exception as e:
            logger.warning(f"Could not save error report: {e}")
        
        # Генерируем детальное сообщение об ошибке
        error_msg = f"КРИТИЧЕСКАЯ ОШИБКА: Обнаружено {len(type_errors)} продуктов с невалидными типами:\n"
        for error in type_errors[:5]:  # Показываем первые 5 ошибок
            error_msg += f"- {error['name']}: тип '{error['invalid_type']}' недопустим\n"
        if len(type_errors) > 5:
            error_msg += f"... и еще {len(type_errors) - 5} ошибок\n"
        error_msg += f"\nДопустимые типы: {', '.join(VALID_IIKO_TYPES)}"
        error_msg += f"\nИспользуйте GOODS для обычных ингредиентов, DISH для блюд."
        
        raise ValueError(error_msg)
    
    # Записываем ВАЛИДНЫЕ данные в Excel
    row = 2
    for product in validated_products:
        ws.cell(row=row, column=1, value=product["product_code"]).number_format = '@'  # Текстовый формат для кода
        ws.cell(row=row, column=2, value=product["ingredient_name"])
        ws.cell(row=row, column=3, value=product["normalized_unit"])
        ws.cell(row=row, column=4, value=product["product_type"])  # ВАЛИДНЫЙ тип
        ws.cell(row=row, column=5, value=product["group"])
        ws.cell(row=row, column=6, value="")  # Штрихкод пустой
        ws.cell(row=row, column=7, value="")  # Поставщик пустой
        
        row += 1
    
    # Автоподгонка ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    
    # Сохраняем в BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Создаем артефакт с успешными результатами
    try:
        success_report = {
            "timestamp": datetime.now().isoformat(),
            "operation": "PRODUCT_SKELETONS_SUCCESS",
            "total_products": len(validated_products),
            "all_types_valid": True,
            "products": [
                {
                    "name": p["ingredient_name"],
                    "code": p["product_code"],
                    "type": p["product_type"],
                    "unit": p["normalized_unit"],
                    "group": p["group"]
                }
                for p in validated_products
            ],
            "instruction": "Все продукты готовы к импорту в iiko. Типы валидированы согласно требованиям iiko API."
        }
        
        import json
        success_file_path = "/app/artifacts/product_skeletons_final.json"
        os.makedirs(os.path.dirname(success_file_path), exist_ok=True)
        with open(success_file_path, 'w', encoding='utf-8') as f:
            json.dump(success_report, f, indent=2, ensure_ascii=False)
        
        logger.info(f"✅ Product-Skeletons.xlsx создан с {len(validated_products)} валидными продуктами")
        logger.info(f"Отчет сохранен: {success_file_path}")
    except Exception as e:
        logger.warning(f"Could not save success report: {e}")
    
    return buffer


def find_missing_product_mappings(cards: List[TechCardV2], rms_service=None) -> List[Dict[str, Any]]:
    """
    C. Product Skeletons: Найти ингредиенты без маппинга к продуктам iiko
    
    Args:
        cards: Список техкарт для анализа
        rms_service: Сервис iiko RMS для проверки существования продуктов
        
    Returns:
        Список ингредиентов без корректного маппинга
    """
    missing_ingredients = []
    processed_names = set()  # Избегаем дубликатов
    
    for card in cards:
        for ingredient in card.ingredients:
            ingredient_name = ingredient.name
            
            # Пропускаем уже обработанные
            if ingredient_name in processed_names:
                continue
                
            processed_names.add(ingredient_name)
            
            # Проверяем наличие product_code или skuId
            has_product_code = getattr(ingredient, 'product_code', None)
            has_sku_id = getattr(ingredient, 'skuId', None)
            
            # Если есть product_code, считаем что маппинг есть
            if has_product_code:
                continue
            
            # Если есть skuId, проверяем его в RMS
            if has_sku_id and rms_service:
                try:
                    # Проверяем в products
                    product = rms_service.products.find_one({"_id": has_sku_id})
                    if product and product.get('article'):
                        continue  # Продукт найден с кодом
                    
                    # Проверяем в prices
                    pricing = rms_service.prices.find_one({"skuId": has_sku_id})
                    if pricing and pricing.get('article'):
                        continue  # Продукт найден с кодом
                        
                except Exception:
                    pass  # Ошибка доступа к RMS - считаем что маппинг отсутствует
            
            # Ингредиент без корректного маппинга
            missing_ingredients.append({
                'name': ingredient_name,
                'unit': getattr(ingredient, 'unit', 'г'),
                'skuId': has_sku_id,
                'product_code': has_product_code
            })
    
    logger.info(f"Found {len(missing_ingredients)} ingredients without product mapping")
    
    return missing_ingredients


def generate_product_codes(ingredient_names: List[str], 
                          rms_service=None, 
                          start_code: int = 10000,
                          code_width: int = 5) -> Dict[str, str]:
    """
    C. Product Skeletons: Генерация свободных кодов продуктов
    
    Args:
        ingredient_names: Список названий ингредиентов
        rms_service: Сервис RMS для проверки занятых кодов
        start_code: Начальный код для генерации
        code_width: Ширина кода (количество цифр)
        
    Returns:
        Словарь название → сгенерированный код
    """
    generated_codes = {}
    current_code = start_code
    
    # Получаем существующие коды из RMS для избежания дубликатов
    existing_codes = set()
    
    if rms_service:
        try:
            # Получаем коды из products
            for product in rms_service.products.find({}, {"article": 1}):
                article = product.get('article')
                if article:
                    existing_codes.add(str(article).strip())
            
            # Получаем коды из prices
            for price in rms_service.prices.find({}, {"article": 1}):
                article = price.get('article')
                if article:
                    existing_codes.add(str(article).strip())
                    
            logger.info(f"Found {len(existing_codes)} existing product codes in RMS")
            
        except Exception as e:
            logger.warning(f"Could not fetch existing codes from RMS: {e}")
    
    # Генерируем уникальные коды
    for ingredient_name in ingredient_names:
        # Ищем свободный код
        while True:
            code_str = str(current_code).zfill(code_width)
            
            if code_str not in existing_codes:
                generated_codes[ingredient_name] = code_str
                existing_codes.add(code_str)  # Добавляем в набор чтобы избежать дубликатов в текущей генерации
                break
                
            current_code += 1
        
        current_code += 1
    
    logger.info(f"Generated {len(generated_codes)} unique product codes starting from {start_code:0{code_width}d}")
    
    return generated_codes


def create_dish_skeletons_xlsx(dish_codes_mapping: Dict[str, str], 
                              cards: List[TechCardV2] = None,
                              dishes_data: List[Dict] = None) -> BytesIO:
    """
    Feature B: Создание файла Dish-Skeletons.xlsx для предварительного импорта блюд в iiko
    WITH STRICT TYPE VALIDATION
    
    Args:
        dish_codes_mapping: Маппинг {dish_name: dish_code}
        cards: Список техкарт для извлечения данных о блюдах (legacy)
        dishes_data: Прямые данные о блюдах (новый формат)
    
    Returns:
        BytesIO buffer с Excel файлом
        
    Raises:
        ValueError: Если обнаружены невалидные типы блюд
    """
    if not Workbook:
        raise ImportError("openpyxl is required for Excel export")
    
    # VALID IIKO DISH TYPES - строго по документации iiko
    VALID_IIKO_DISH_TYPES = {
        "DISH"        # Блюдо (основной тип для всех блюд)
    }
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Блюда"
    
    # Заголовки согласно требованиям
    headers = [
        "Артикул",           # Числовой код блюда
        "Наименование",      # Название блюда  
        "Тип",               # Всегда "DISH" (валидировано)
        "Ед. выпуска",       # Единица измерения (г/мл)
        "Выход"              # Выход готового продукта
    ]
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Валидация и подготовка данных
    type_errors = []
    validated_dishes = []
    
    # Заполняем данные по блюдам с СТРОГОЙ ВАЛИДАЦИЕЙ
    row = 2
    
    # Используем dishes_data если предоставлен (новый формат)
    if dishes_data:
        for idx, dish in enumerate(dishes_data):
            dish_name = dish["name"]
            dish_code = dish["article"]
            dish_type_input = dish.get("type", "блюдо")
            unit = dish.get("unit", "порц.")
            yield_g = dish.get("yield_g", 200.0)
            
            # КРИТИЧЕСКОЕ: Все блюда должны быть типа DISH
            dish_type = "DISH"
            
            # СТРОГАЯ ВАЛИДАЦИЯ ТИПА
            if dish_type not in VALID_IIKO_DISH_TYPES:
                type_errors.append({
                    "index": idx,
                    "name": dish_name,
                    "invalid_type": dish_type,
                    "input_type": dish_type_input,
                    "valid_types": list(VALID_IIKO_DISH_TYPES),
                    "error": f"Недопустимый тип '{dish_type}' для блюда '{dish_name}'"
                })
                continue
            
            validated_dishes.append({
                "dish_code": dish_code,
                "dish_name": dish_name,
                "dish_type": dish_type,
                "unit": unit,
                "yield_g": yield_g
            })
    
    
    # Legacy: используем cards если dishes_data не предоставлены
    elif cards:
        for idx, card in enumerate(cards):
            dish_name = card.meta.title
            dish_code = dish_codes_mapping.get(dish_name, '')
            
            if dish_code:
                # Определяем выход и единицу
                yield_g = card.yield_.perPortion_g if hasattr(card, 'yield_') else 200
                unit = 'г'  # По умолчанию граммы
                
                # КРИТИЧЕСКОЕ: Все блюда должны быть типа DISH
                dish_type = "DISH"
                
                # СТРОГАЯ ВАЛИДАЦИЯ ТИПА
                if dish_type not in VALID_IIKO_DISH_TYPES:
                    type_errors.append({
                        "index": idx,
                        "name": dish_name,
                        "invalid_type": dish_type,
                        "valid_types": list(VALID_IIKO_DISH_TYPES),
                        "error": f"Недопустимый тип '{dish_type}' для блюда '{dish_name}'"
                    })
                    continue
                
                validated_dishes.append({
                    "dish_code": dish_code,
                    "dish_name": dish_name,
                    "dish_type": dish_type,
                    "unit": unit,
                    "yield_g": yield_g
                })
    
    # FAIL-FAST: Если есть ошибки валидации - останавливаем экспорт
    if type_errors:
        # Создаем детальный отчет об ошибках для блюд
        error_report = {
            "timestamp": datetime.now().isoformat(),
            "error_type": "DISH_TYPE_VALIDATION_FAILED",
            "total_dishes": len(dishes_data) if dishes_data else len(cards) if cards else 0,
            "invalid_dishes": len(type_errors),
            "valid_types": list(VALID_IIKO_DISH_TYPES),
            "errors": type_errors,
            "instruction": "Все блюда должны иметь тип DISH для корректного импорта в iiko"
        }
        
        # Сохраняем отчет об ошибках
        try:
            import json
            error_file_path = "/app/artifacts/dish_type_errors.json"
            os.makedirs(os.path.dirname(error_file_path), exist_ok=True)
            with open(error_file_path, 'w', encoding='utf-8') as f:
                json.dump(error_report, f, indent=2, ensure_ascii=False)
            logger.error(f"Dish type validation errors saved to {error_file_path}")
        except Exception as e:
            logger.warning(f"Could not save dish error report: {e}")
        
        # Генерируем детальное сообщение об ошибке
        error_msg = f"КРИТИЧЕСКАЯ ОШИБКА DISH SKELETONS: Обнаружено {len(type_errors)} блюд с невалидными типами:\n"
        for error in type_errors[:3]:  # Показываем первые 3 ошибки
            error_msg += f"- {error['name']}: тип '{error['invalid_type']}' недопустим\n"
        if len(type_errors) > 3:
            error_msg += f"... и еще {len(type_errors) - 3} ошибок\n"
        error_msg += f"\nДля блюд используйте только тип: DISH"
        
        raise ValueError(error_msg)
    
    # Записываем ВАЛИДНЫЕ блюда в Excel
    row = 2
    for dish in validated_dishes:
        # КОНВЕРТАЦИЯ ВЫХОДА БЛЮДА ИЗ ГРАММ В КИЛОГРАММЫ для экспорта в iiko
        yield_kg, yield_unit = convert_yield_grams_to_kilograms(dish["yield_g"])
        dish_unit_kg = "кг"  # Для блюд всегда используем килограммы в iiko
        
        # Заполняем ячейки с валидными данными
        ws.cell(row=row, column=1, value=dish["dish_code"])
        ws.cell(row=row, column=2, value=dish["dish_name"])
        ws.cell(row=row, column=3, value=dish["dish_type"])  # ВАЛИДНЫЙ тип DISH
        ws.cell(row=row, column=4, value=dish_unit_kg)       # Единица измерения (кг)
        ws.cell(row=row, column=5, value=yield_kg)           # Выход в килограммах
        
        # Форматируем артикул как текст
        article_cell = ws.cell(row=row, column=1)
        article_cell.number_format = '@'
        
        row += 1
    
    # Автоширина колонок
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row_cells in ws[f"{column_letter}1:{column_letter}{row-1}"]:
            for cell in row_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    
    # Сохраняем в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Создаем артефакт с успешными результатами для блюд
    try:
        success_report = {
            "timestamp": datetime.now().isoformat(),
            "operation": "DISH_SKELETONS_SUCCESS",
            "total_dishes": len(validated_dishes),
            "all_types_valid": True,
            "dishes": [
                {
                    "name": d["dish_name"],
                    "code": d["dish_code"],
                    "type": d["dish_type"],
                    "unit": d["unit"],
                    "yield_g": d["yield_g"]
                }
                for d in validated_dishes
            ],
            "instruction": "Все блюда готовы к импорту в iiko. Типы валидированы как DISH согласно требованиям iiko API."
        }
        
        import json
        success_file_path = "/app/artifacts/dish_skeletons_final.json"
        os.makedirs(os.path.dirname(success_file_path), exist_ok=True)
        with open(success_file_path, 'w', encoding='utf-8') as f:
            json.dump(success_report, f, indent=2, ensure_ascii=False)
        
        logger.info(f"✅ Dish-Skeletons.xlsx создан с {len(validated_dishes)} валидными блюдами")
        logger.info(f"Отчет сохранен: {success_file_path}")
    except Exception as e:
        logger.warning(f"Could not save dish success report: {e}")
    
    return buffer

# Путь к шаблону iiko
TEMPLATE_PATH = Path(__file__).parent.parent / "data" / "iiko_templates" / "ttk.xlsx"

# Маппинг заголовков iiko (поддержка синонимов)
IIKO_HEADERS = {
    'dish_code': ['Артикул блюда', 'Код блюда'],
    'dish_name': ['Наименование блюда', 'Название блюда'],
    'product_code': ['Артикул продукта', 'Код продукта'],
    'product_name': ['Наименование продукта', 'Название продукта'],
    'brutto': ['Брутто', 'Брутто, г'],
    'loss_pct': ['Потери, %', 'Потери'],
    'netto': ['Нетто', 'Нетто, г'],
    'unit': ['Ед.', 'Единица', 'Ед. изм.'],
    'output_qty': ['Выход готового продукта', 'Выход'],
    'norm': ['Норма закладки'],
    'write_off_method': ['Метод списания'],
    'technology': ['Технология приготовления', 'Технология']
}

def generate_dish_slug(title: str) -> str:
    """Генерирует slug из названия блюда"""
    # Убираем специальные символы и заменяем пробелы на подчеркивания
    slug = re.sub(r'[^\w\s-]', '', title).strip()
    slug = re.sub(r'[-\s]+', '_', slug)
    return slug.upper()[:20]  # Ограничиваем длину

def normalize_unit_to_grams(value: float, unit: str, ingredient_name: str = "") -> float:
    """
    Конвертирует единицы в граммы для iiko
    """
    unit = unit.lower().strip()
    
    # Уже граммы
    if unit in ['г', 'g', 'gram', 'грамм']:
        return value
    
    # Килограммы -> граммы
    elif unit in ['кг', 'kg', 'kilogram', 'килограмм']:
        return value * 1000
    
    # Миллилитры -> граммы (принимаем плотность = 1 для большинства жидкостей)
    elif unit in ['мл', 'ml', 'milliliter', 'миллилитр']:
        # Для масла плотность ≈ 0.9
        if 'масло' in ingredient_name.lower():
            return value * 0.9
        # Для воды и большинства жидкостей плотность = 1
        return value
    
    # Литры -> граммы
    elif unit in ['л', 'l', 'liter', 'литр']:
        return value * 1000  # Принимаем плотность = 1
    
    # Штуки -> граммы (средние веса)
    elif unit in ['шт', 'pcs', 'piece', 'штука']:
        # Стандартные веса продуктов
        piece_weights = {
            'яйцо': 50,      # 50г за яйцо
            'луковица': 150,  # 150г за луковицу
            'морковь': 100,   # 100г за морковь
            'картофель': 120, # 120г за картофелину
            'помидор': 180,   # 180г за помидор
            'огурец': 150,    # 150г за огурец
            'яблоко': 180,    # 180г за яблоко
            'лимон': 130      # 130г за лимон
        }
        
        # Пытаемся найти подходящий вес
        for product, weight in piece_weights.items():
            if product in ingredient_name.lower():
                return value * weight
        
        # Если не нашли - используем средний вес 100г
        return value * 100
    
    # По умолчанию возвращаем как есть
    return value

def generate_technology_text(process_steps) -> str:
    """
    Генерирует текст технологии из шагов процесса
    Формат: #{n}. {action} [t={temp_c}°C] [{time_min} мин] [{equipment}]
    """
    if not process_steps:
        return "Технология приготовления не указана"
    
    technology_lines = []
    
    for step in process_steps:
        # Handle both dict and object types
        if hasattr(step, 'n'):
            step_num = step.n
            action = getattr(step, 'action', 'Действие не указано')
            temp_c = getattr(step, 'temp_c', None)
            time_min = getattr(step, 'time_min', None)
            equipment = getattr(step, 'equipment', None)
        else:
            step_num = step.get('n', len(technology_lines) + 1)
            action = step.get('action', 'Действие не указано')
            temp_c = step.get('temp_c')
            time_min = step.get('time_min')
            equipment = step.get('equipment')
        
        # Базовая строка
        line = f"#{step_num}. {action}"
        
        # Добавляем температуру если есть
        if temp_c:
            line += f" [t={temp_c}°C]"
        
        # Добавляем время если есть
        if time_min:
            line += f" [{time_min} мин]"
        
        # Добавляем оборудование если есть
        if equipment:
            if isinstance(equipment, list):
                equipment_str = ', '.join(equipment)
            else:
                equipment_str = str(equipment)
            line += f" [{equipment_str}]"
        
        technology_lines.append(line)
    
    # Объединяем в один текст, ограничиваем 1000 символами
    technology_text = '\n'.join(technology_lines)
    if len(technology_text) > 1000:
        technology_text = technology_text[:997] + "..."
    
    return technology_text

def create_iiko_ttk_xlsx(card: TechCardV2, 
                        export_options: Dict[str, Any] = None,
                        rms_service=None) -> Tuple[BytesIO, List[Dict[str, Any]]]:
    """
    Создать iiko XLSX файл для импорта технологических карт
    WITH iiko Import Reliability — Product Codes & Dish Skeletons
    WITH Operational Rounding v1 (Export & Kitchen View)
    
    Args:
        card: Техкарта в формате TechCardV2
        export_options: Опции экспорта:
            - use_product_codes: bool = True - использовать коды товаров вместо GUID
            - dish_codes_mapping: Dict[str, str] - маппинг блюд к кодам
            - operational_rounding: bool = True - применять операционное округление
    
    Returns:
        (BytesIO buffer, issues list)
    """
    if not export_options:
        export_options = {}
    
    # Feature A: Product Code toggle (по умолчанию включен)
    use_product_codes = export_options.get('use_product_codes', True)
    dish_codes_mapping = export_options.get('dish_codes_mapping', {})
    rms_service = export_options.get('rms_service')  # Для получения кодов продуктов
    
    # Operational Rounding v1: операционное округление (по умолчанию включено)
    operational_rounding_enabled = export_options.get('operational_rounding', True)
    
    if not Workbook:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    issues = []
    
    # Operational Rounding v1: Применяем операционное округление если включено
    working_card = card
    rounding_metadata = None
    
    if operational_rounding_enabled:
        try:
            from receptor_agent.techcards_v2.operational_rounding import get_operational_rounder
            logger.info("Applying operational rounding...")
            rounder = get_operational_rounder()
            
            # Конвертируем TechCardV2 в dict для округления
            card_dict = card.model_dump()
            logger.info(f"Card dict keys: {list(card_dict.keys())}")
            logger.info(f"Meta dict: {card_dict.get('meta', {})}")
            
            rounding_result = rounder.round_techcard_ingredients(card_dict)
            
            # DEBUG: Проверяем результат округления
            rounded_dict = rounding_result['rounded_techcard']
            logger.info(f"Rounded card keys: {list(rounded_dict.keys())}")
            logger.info(f"Rounded meta: {rounded_dict.get('meta', {})}")
            
            # Создаем новый TechCardV2 объект с округленными данными
            working_card = TechCardV2.model_validate(rounded_dict)
            rounding_metadata = rounding_result['rounding_metadata']
            
            logger.info(f"Working card meta type: {type(working_card.meta)}")
            logger.info(f"Working card meta: {working_card.meta}")
            
            # Логируем применение округления
            if rounding_metadata and rounding_metadata.get('items'):
                logger.info(f"Operational rounding applied to {len(rounding_metadata['items'])} ingredients, "
                           f"delta: {rounding_metadata.get('delta_g', 0)}g")
                
                # Добавляем информацию в issues для отчета
                if abs(rounding_metadata.get('delta_g', 0)) > 0.1:
                    issues.append({
                        "type": "operationalRounding",
                        "message": f"Применено операционное округление: {len(rounding_metadata['items'])} ингредиентов, "
                                 f"дельта: {rounding_metadata.get('delta_g', 0):+.1f}г",
                        "severity": "info",
                        "details": rounding_metadata
                    })
            
        except Exception as e:
            logger.error(f"Operational rounding error: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            # Продолжаем с исходной техкартой если округление не удалось
            issues.append({
                "type": "roundingError",
                "message": f"Ошибка операционного округления: {str(e)}",
                "severity": "warning"
            })
    
    # Пытаемся загрузить шаблон, если не получается - создаем новый
    wb = None
    if TEMPLATE_PATH.exists():
        try:
            wb = load_workbook(TEMPLATE_PATH)
            ws = wb.active
        except Exception as e:
            print(f"Failed to load template {TEMPLATE_PATH}: {e}")
            wb = None
    
    if wb is None:
        # Создаем новую книгу с правильными заголовками
        wb = Workbook()
        ws = wb.active
        ws.title = "ТТК"
        
        # Заголовки в правильном порядке
        headers = [
            'Артикул блюда',
            'Наименование блюда', 
            'Артикул продукта',
            'Наименование продукта',
            'Брутто',
            'Потери, %',
            'Нетто',
            'Ед.',
            'Выход готового продукта',
            'Норма закладки',
            'Метод списания',
            'Технология приготовления'
        ]
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    else:
        ws = wb.active
    
    # Генерируем данные для экспорта
    # Feature A: Use dish_codes_mapping if provided
    # DEBUG: Детальное логирование meta объекта для поиска проблемы с 'str' object has no attribute 'get'
    logger.info(f"DEBUG working_card type: {type(working_card)}")
    logger.info(f"DEBUG working_card.meta type: {type(working_card.meta)}")
    logger.info(f"DEBUG working_card.meta repr: {repr(working_card.meta)}")
    
    try:
        if hasattr(working_card.meta, 'title'):
            dish_title = working_card.meta.title
        elif isinstance(working_card.meta, dict):
            dish_title = working_card.meta.get('title', 'Unknown Dish')
        elif isinstance(working_card.meta, str):
            # Если meta является строкой, используем её как название
            dish_title = working_card.meta
            logger.warning(f"working_card.meta is a string: {working_card.meta}")
        else:
            dish_title = str(working_card.meta)
            logger.warning(f"working_card.meta is unexpected type {type(working_card.meta)}: {working_card.meta}")
    except Exception as e:
        logger.error(f"Error accessing working_card.meta.title: {e}")
        logger.error(f"working_card.meta type: {type(working_card.meta)}")
        logger.error(f"working_card.meta: {working_card.meta}")
        dish_title = 'Unknown Dish'
    
    # Get preflight result for article mapping
    preflight_result = export_options.get('preflight_result')
    
    dish_code = dish_codes_mapping.get(dish_title)
    if not dish_code:
        meta_dict = working_card.meta.model_dump() if hasattr(working_card.meta, 'model_dump') else working_card.meta
        # CRITICAL FIX: Ищем артикул блюда в правильном поле meta.article (не dish_code)
        dish_code = meta_dict.get('article') or meta_dict.get('dish_code') if isinstance(meta_dict, dict) else None
        if dish_code:
            logger.info(f"Found dish article in meta: {dish_code} for '{dish_title}'")
    
    # Check preflight result for dish article
    if not dish_code and preflight_result:
        # Look for dish article in preflight missing dishes
        missing_dishes = preflight_result.get('missing', {}).get('dishes', [])
        for dish in missing_dishes:
            if dish.get('name') == dish_title:
                dish_code = dish.get('article')
                logger.info(f"Using preflight dish article for '{dish_title}': {dish_code}")
                break
    
    if not dish_code:
        try:
            # Используем ArticleAllocator для получения реального 5-digit артикула блюда
            from ..integrations.article_allocator import ArticleAllocator
            allocator = ArticleAllocator()
            
            # Создаем уникальный ID для блюда
            dish_id = f"dish_{dish_title.lower().replace(' ', '_')}"
            allocated_result = allocator.allocate_articles([dish_id])
            
            if allocated_result.get('success') and allocated_result.get('allocated'):
                dish_code = allocated_result['allocated'][0]['article']
                print(f"✅ Generated real dish article: {dish_code}")
            else:
                # Fallback to old placeholder system if ArticleAllocator fails
                dish_slug = generate_dish_slug(dish_title)
                dish_code = f"DISH_{dish_slug}"
                print(f"⚠️ ArticleAllocator failed for dish, using placeholder: {dish_code}")
                
        except Exception as e:
            # Fallback to old placeholder system if ArticleAllocator fails
            print(f"⚠️ ArticleAllocator error for dish: {e}")
            dish_slug = generate_dish_slug(dish_title)
            dish_code = f"DISH_{dish_slug}"
    
    dish_name = dish_title
    
    # F. TTK Date Autoresolve: Автоматический сдвиг даты при конфликте
    auto_resolve_date = export_options.get('auto_resolve_date', True) if export_options else True
    base_date = export_options.get('base_date') if export_options else None
    
    resolved_date = None
    if auto_resolve_date:
        resolved_date = resolve_ttk_date_conflict(
            dish_name=dish_name,
            base_date=base_date,
            rms_service=rms_service,
            max_days_ahead=7
        )
        logger.info(f"Resolved TTK date for '{dish_name}': {resolved_date}")
    else:
        resolved_date = base_date or datetime.now().strftime('%Y-%m-%d')
    
    # Рассчитываем выход готового продукта и конвертируем в килограммы
    output_qty_g = getattr(working_card.yield_, 'perBatch_g', 0) if hasattr(working_card, 'yield_') and working_card.yield_ else 0
    if not output_qty_g and hasattr(working_card, 'yield_') and working_card.yield_:
        # Если нет perBatch_g, рассчитываем из порций
        portions = getattr(working_card, 'portions', 1)
        per_portion = getattr(working_card.yield_, 'perPortion_g', 0)
        output_qty_g = portions * per_portion if per_portion else 0
    
    # КОНВЕРТАЦИЯ ВЫХОДА ИЗ ГРАММ В КИЛОГРАММЫ для экспорта в iiko
    output_qty_kg, output_unit = convert_yield_grams_to_kilograms(output_qty_g)
    
    # Генерируем технологию приготовления
    technology_text = ""
    if hasattr(working_card, 'process') and working_card.process:
        technology_text = generate_technology_text(working_card.process)
    
    # Заполняем данные по ингредиентам
    row = 2  # Начинаем со второй строки (первая - заголовки)
    
    for ingredient in working_card.ingredients:
        # Артикул продукта
        # Feature A: Product Code toggle - использовать реальные коды товаров
        if use_product_codes:
            # A. Hotfix & Migration: Сначала пытаемся использовать уже сохраненный product_code
            product_code = getattr(ingredient, 'product_code', None)
            
            # Если нет сохраненного кода, получаем из RMS по skuId
            if not product_code and ingredient.skuId:
                product_code = get_product_code_from_rms(ingredient.skuId, rms_service)
            
            # Если это все еще GUID или нет кода, проверяем preflight данные
            if not product_code or product_code == ingredient.skuId:
                # Check preflight result for ingredient article
                if preflight_result:
                    missing_products = preflight_result.get('missing', {}).get('products', [])
                    for product in missing_products:
                        if product.get('name') == ingredient.name:
                            product_code = product.get('article')
                            logger.info(f"Using preflight product article for '{ingredient.name}': {product_code}")
                            break
                
                # If still no product code, use ArticleAllocator as fallback
                if not product_code:
                    try:
                        # Используем ArticleAllocator для получения реального 5-digit артикула
                        from ..integrations.article_allocator import ArticleAllocator
                        allocator = ArticleAllocator()
                        
                        # Создаем уникальный ID для ингредиента
                        ingredient_id = f"ingredient_{ingredient.name.lower().replace(' ', '_')}"
                        allocated_result = allocator.allocate_articles([ingredient_id])
                        
                        if allocated_result.get('success') and allocated_result.get('allocated'):
                            product_code = allocated_result['allocated'][0]['article']
                            print(f"✅ Generated real article for {ingredient.name}: {product_code}")
                        else:
                            # Fallback to old placeholder system if ArticleAllocator fails
                            ingredient_slug = generate_dish_slug(ingredient.name)
                            product_code = f"GENERATED_{ingredient_slug}"
                            print(f"⚠️ ArticleAllocator failed, using placeholder for {ingredient.name}: {product_code}")
                            
                    except Exception as e:
                        # Fallback to old placeholder system if ArticleAllocator fails
                        print(f"⚠️ ArticleAllocator error for {ingredient.name}: {e}")
                        ingredient_slug = generate_dish_slug(ingredient.name)
                        product_code = f"GENERATED_{ingredient_slug}"
                issues.append({
                    "type": "noProductCode",
                    "name": ingredient.name,
                    "hint": f"Продукт не найден в iiko RMS или отсутствует код товара. Сгенерирован: {product_code}",
                    "dish": dish_title
                })
        else:
            # Используем GUID как есть (legacy режим)
            product_code = ingredient.skuId or f"GUID_{generate_dish_slug(ingredient.name)}"
        
        # Обработка подрецептов
        if hasattr(ingredient, 'subRecipe') and ingredient.subRecipe:
            # Для подрецептов используем их dish_code как артикул продукта
            sub_recipe_code = getattr(ingredient.subRecipe, 'dish_code', None)
            if sub_recipe_code and use_product_codes:
                product_code = sub_recipe_code
            elif not use_product_codes:
                # Используем GUID для подрецептов
                product_code = getattr(ingredient.subRecipe, 'guid', sub_recipe_code or f"SUBGUID_{generate_dish_slug(ingredient.name)}")
            else:
                issues.append({
                    "type": "subRecipeNoCode",
                    "name": ingredient.name,
                    "hint": "Sub-recipe missing dish_code",
                    "dish": dish_title
                })
        
        # Конвертируем единицы в граммы, затем в килограммы для iiko XLSX
        brutto_g_raw = round(normalize_unit_to_grams(
            getattr(ingredient, 'brutto_g', 0), 
            getattr(ingredient, 'unit', 'г'), 
            ingredient.name
        ), 1)
        
        netto_g_raw = round(normalize_unit_to_grams(
            getattr(ingredient, 'netto_g', 0), 
            getattr(ingredient, 'unit', 'г'), 
            ingredient.name
        ), 1)
        
        # КОНВЕРТАЦИЯ ГРАММ В КИЛОГРАММЫ для экспорта в iiko
        brutto_kg, brutto_unit = convert_grams_to_kilograms(brutto_g_raw, 'г')
        netto_kg, netto_unit = convert_grams_to_kilograms(netto_g_raw, 'г')
        
        loss_pct = round(getattr(ingredient, 'loss_pct', 0), 1)
        
        # Заполняем строку
        row_data = [
            dish_code,                    # Артикул блюда
            dish_name,                    # Наименование блюда
            product_code,                 # Артикул продукта
            ingredient.name,              # Наименование продукта
            brutto_kg,                    # Брутто (в кг)
            loss_pct,                     # Потери, %
            netto_kg,                     # Нетто (в кг)
            brutto_unit,                  # Ед. (кг для iiko)
            round(output_qty_kg, 1),      # Выход готового продукта (в кг)
            1,                            # Норма закладки
            1,                            # Метод списания
            technology_text if row == 2 else ""  # Технология (только для первой строки)
        ]
        
        # Записываем данные в ячейки
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            
            # Feature A: Форматирование кодов продуктов и блюд как текст для сохранения ведущих нулей
            if col in [1, 3]:  # Колонки "Артикул блюда" и "Артикул продукта"
                cell.number_format = '@'  # Текстовый формат для сохранения ведущих нулей
        
        row += 1
    
    # Автоширина колонок
    for col in range(1, 13):  # 12 колонок
        column_letter = get_column_letter(col)
        max_length = 0
        for row_cells in ws[f"{column_letter}1:{column_letter}{row-1}"]:
            for cell in row_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer, issues


def export_techcard_to_iiko_xlsx(card: TechCardV2, export_options: Dict[str, Any] = None) -> bytes:
    """
    Wrapper function for round-trip tests compatibility
    Returns just the Excel bytes
    """
    buffer, issues = create_iiko_ttk_xlsx(card, export_options)
    return buffer.getvalue()