#!/usr/bin/env python3
"""
IK-04/01 Backend XLSX Parser Testing
Comprehensive testing of iiko XLSX import functionality as specified in review request
"""

import os
import sys
import json
import time
import tempfile
import requests
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl import Workbook

# Add backend path for imports
sys.path.append('/app/backend')

# Configuration
BACKEND_URL = os.environ.get('REACT_APP_BACKEND_URL', 'http://localhost:8001')
API_BASE = f"{BACKEND_URL}/api"

class IK04XlsxParserTester:
    """Comprehensive tester for IK-04/01 XLSX Parser functionality"""
    
    def __init__(self):
        self.test_results = []
        self.total_tests = 0
        self.passed_tests = 0
        
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0):
        """Log test result"""
        self.total_tests += 1
        if success:
            self.passed_tests += 1
            
        result = {
            "test": test_name,
            "success": success,
            "details": details,
            "response_time_ms": round(response_time * 1000, 2),
            "timestamp": datetime.now().isoformat()
        }
        self.test_results.append(result)
        
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"{status} {test_name}: {details}")
        
    def create_test_xlsx_file(self, filename: str, data_type: str = "basic") -> bytes:
        """Create test XLSX files with different structures"""
        wb = Workbook()
        ws = wb.active
        
        if data_type == "basic":
            # Basic tech card structure with RU headers
            ws['A1'] = 'Название блюда'
            ws['B1'] = 'Говядина тушеная с овощами'
            
            ws['A3'] = 'Артикул продукта'
            ws['B3'] = 'Наименование продукта'
            ws['C3'] = 'Брутто'
            ws['D3'] = 'Потери %'
            ws['E3'] = 'Нетто'
            ws['F3'] = 'Ед. изм.'
            
            # Ingredients data
            ingredients = [
                ['BEEF001', 'Говядина для тушения', 500, 15, 425, 'г'],
                ['VEG002', 'Лук репчатый', 100, 10, 90, 'г'],
                ['VEG003', 'Морковь', 80, 5, 76, 'г'],
                ['OIL001', 'Масло растительное', 30, 0, 30, 'мл'],
                ['SPICE001', 'Соль поваренная', 5, 0, 5, 'г']
            ]
            
            for i, ingredient in enumerate(ingredients, 4):
                for j, value in enumerate(ingredient):
                    ws.cell(row=i, column=j+1, value=value)
                    
        elif data_type == "english":
            # English headers
            ws['A1'] = 'Dish Name'
            ws['B1'] = 'Beef Stew with Vegetables'
            
            ws['A3'] = 'Product Code'
            ws['B3'] = 'Product Name'
            ws['C3'] = 'Gross'
            ws['D3'] = 'Loss %'
            ws['E3'] = 'Net'
            ws['F3'] = 'Unit'
            
            ingredients = [
                ['BEEF001', 'Beef for stewing', 500, 15, 425, 'g'],
                ['VEG002', 'Onion', 100, 10, 90, 'g'],
                ['VEG003', 'Carrot', 80, 5, 76, 'g']
            ]
            
            for i, ingredient in enumerate(ingredients, 4):
                for j, value in enumerate(ingredient):
                    ws.cell(row=i, column=j+1, value=value)
                    
        elif data_type == "units_test":
            # Test different units and conversions
            ws['A1'] = 'Название блюда'
            ws['B1'] = 'Тест единиц измерения'
            
            ws['A3'] = 'Наименование продукта'
            ws['B3'] = 'Брутто'
            ws['C3'] = 'Нетто'
            ws['D3'] = 'Ед. изм.'
            
            ingredients = [
                ['Говядина', 1.5, 1.275, 'кг'],  # kg to g conversion
                ['Молоко', 0.5, 0.5, 'л'],       # l to ml conversion
                ['Масло растительное', 50, 50, 'мл'],  # ml to g with density
                ['Яйца куриные', 2, 2, 'шт'],    # pieces to weight
                ['Соль', 10, 10, 'г']            # grams (no conversion)
            ]
            
            for i, ingredient in enumerate(ingredients, 4):
                for j, value in enumerate(ingredient):
                    ws.cell(row=i, column=j+1, value=value)
                    
        elif data_type == "technology":
            # Include technology description
            ws['A1'] = 'Название блюда'
            ws['B1'] = 'Борщ украинский'
            
            ws['A2'] = 'Технология приготовления'
            ws['B2'] = 'Мясо нарезать кусочками и обжарить 15 мин при 180°C. Овощи нашинковать и добавить к мясу. Тушить 45 мин при 160°C. Подавать горячим.'
            
            ws['A4'] = 'Наименование продукта'
            ws['B4'] = 'Брутто'
            ws['C4'] = 'Нетто'
            ws['D4'] = 'Ед. изм.'
            
            ingredients = [
                ['Говядина', 400, 340, 'г'],
                ['Капуста белокочанная', 200, 180, 'г'],
                ['Свекла', 100, 85, 'г']
            ]
            
            for i, ingredient in enumerate(ingredients, 5):
                for j, value in enumerate(ingredient):
                    ws.cell(row=i, column=j+1, value=value)
                    
        elif data_type == "empty":
            # Empty file with just headers
            ws['A1'] = 'Наименование продукта'
            ws['B1'] = 'Брутто'
            ws['C1'] = 'Нетто'
            
        elif data_type == "large":
            # Large file with ~200 rows for performance testing
            ws['A1'] = 'Название блюда'
            ws['B1'] = 'Большое блюдо для тестирования'
            
            ws['A3'] = 'Наименование продукта'
            ws['B3'] = 'Брутто'
            ws['C3'] = 'Нетто'
            ws['D3'] = 'Ед. изм.'
            
            # Generate 200 ingredients
            for i in range(200):
                ws.cell(row=i+4, column=1, value=f'Ингредиент {i+1}')
                ws.cell(row=i+4, column=2, value=100 + i)
                ws.cell(row=i+4, column=3, value=90 + i)
                ws.cell(row=i+4, column=4, value='г')
        
        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def test_endpoint_availability(self):
        """Test 1: API Endpoint Availability"""
        print("\n🔍 Testing API Endpoint Availability...")
        
        # Test endpoint exists and accepts multipart/form-data
        try:
            # Create a simple test file
            test_file_bytes = self.create_test_xlsx_file("test.xlsx", "basic")
            
            start_time = time.time()
            response = requests.post(
                f"{API_BASE}/v1/iiko/import/ttk.xlsx",
                files={'file': ('test.xlsx', test_file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=30
            )
            response_time = time.time() - start_time
            
            if response.status_code in [200, 400]:  # 400 is OK for validation errors
                self.log_test(
                    "Endpoint Availability", 
                    True, 
                    f"Endpoint accessible, status: {response.status_code}",
                    response_time
                )
            else:
                self.log_test(
                    "Endpoint Availability", 
                    False, 
                    f"Unexpected status: {response.status_code}, response: {response.text[:200]}"
                )
                
        except Exception as e:
            self.log_test("Endpoint Availability", False, f"Request failed: {str(e)}")
    
    def test_file_validation(self):
        """Test 2: File Type Validation"""
        print("\n🔍 Testing File Type Validation...")
        
        # Test .xlsx file acceptance
        try:
            test_file_bytes = self.create_test_xlsx_file("test.xlsx", "basic")
            
            start_time = time.time()
            response = requests.post(
                f"{API_BASE}/v1/iiko/import/ttk.xlsx",
                files={'file': ('test.xlsx', test_file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=30
            )
            response_time = time.time() - start_time
            
            if response.status_code in [200, 400]:
                self.log_test(
                    "XLSX File Acceptance", 
                    True, 
                    f"XLSX file accepted, status: {response.status_code}",
                    response_time
                )
            else:
                self.log_test("XLSX File Acceptance", False, f"XLSX rejected: {response.status_code}")
                
        except Exception as e:
            self.log_test("XLSX File Acceptance", False, f"Error: {str(e)}")
        
        # Test invalid file type rejection
        try:
            start_time = time.time()
            response = requests.post(
                f"{API_BASE}/v1/iiko/import/ttk.xlsx",
                files={'file': ('test.txt', b'invalid content', 'text/plain')},
                timeout=30
            )
            response_time = time.time() - start_time
            
            if response.status_code == 400:
                self.log_test(
                    "Invalid File Type Rejection", 
                    True, 
                    "Invalid file type properly rejected",
                    response_time
                )
            else:
                self.log_test("Invalid File Type Rejection", False, f"Should reject invalid files: {response.status_code}")
                
        except Exception as e:
            self.log_test("Invalid File Type Rejection", False, f"Error: {str(e)}")
        
        # Test empty file handling
        try:
            start_time = time.time()
            response = requests.post(
                f"{API_BASE}/v1/iiko/import/ttk.xlsx",
                files={'file': ('empty.xlsx', b'', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=30
            )
            response_time = time.time() - start_time
            
            if response.status_code == 400:
                self.log_test(
                    "Empty File Handling", 
                    True, 
                    "Empty file properly rejected",
                    response_time
                )
            else:
                self.log_test("Empty File Handling", False, f"Should reject empty files: {response.status_code}")
                
        except Exception as e:
            self.log_test("Empty File Handling", False, f"Error: {str(e)}")
    
    def test_parser_core_functionality(self):
        """Test 3: XLSX Parser Core Functionality"""
        print("\n🔍 Testing XLSX Parser Core Functionality...")
        
        # Test basic parsing with RU headers
        try:
            test_file_bytes = self.create_test_xlsx_file("basic_ru.xlsx", "basic")
            
            start_time = time.time()
            response = requests.post(
                f"{API_BASE}/v1/iiko/import/ttk.xlsx",
                files={'file': ('basic_ru.xlsx', test_file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=30
            )
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                techcard = data.get('techcard', {})
                
                # Validate basic structure
                has_title = bool(techcard.get('meta', {}).get('title'))
                has_ingredients = len(techcard.get('ingredients', [])) > 0
                has_process = len(techcard.get('process', [])) > 0
                
                success = has_title and has_ingredients and has_process
                details = f"Title: {has_title}, Ingredients: {len(techcard.get('ingredients', []))}, Process: {len(techcard.get('process', []))}"
                
                self.log_test(
                    "Basic RU Headers Parsing", 
                    success, 
                    details,
                    response_time
                )
            else:
                self.log_test("Basic RU Headers Parsing", False, f"Failed: {response.status_code}, {response.text[:200]}")
                
        except Exception as e:
            self.log_test("Basic RU Headers Parsing", False, f"Error: {str(e)}")
        
        # Test English headers support
        try:
            test_file_bytes = self.create_test_xlsx_file("basic_en.xlsx", "english")
            
            start_time = time.time()
            response = requests.post(
                f"{API_BASE}/v1/iiko/import/ttk.xlsx",
                files={'file': ('basic_en.xlsx', test_file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=30
            )
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                techcard = data.get('techcard', {})
                has_ingredients = len(techcard.get('ingredients', [])) > 0
                
                self.log_test(
                    "English Headers Support", 
                    has_ingredients, 
                    f"Parsed {len(techcard.get('ingredients', []))} ingredients from EN headers",
                    response_time
                )
            else:
                self.log_test("English Headers Support", False, f"Failed: {response.status_code}")
                
        except Exception as e:
            self.log_test("English Headers Support", False, f"Error: {str(e)}")
    
    def test_unit_normalization(self):
        """Test 4: Unit Normalization"""
        print("\n🔍 Testing Unit Normalization...")
        
        try:
            test_file_bytes = self.create_test_xlsx_file("units_test.xlsx", "units_test")
            
            start_time = time.time()
            response = requests.post(
                f"{API_BASE}/v1/iiko/import/ttk.xlsx",
                files={'file': ('units_test.xlsx', test_file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=30
            )
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                techcard = data.get('techcard', {})
                ingredients = techcard.get('ingredients', [])
                
                # Check unit conversions
                conversions_found = []
                for ing in ingredients:
                    name = ing.get('name', '')
                    unit = ing.get('unit', '')
                    brutto = ing.get('brutto_g', 0)
                    
                    if 'говядина' in name.lower() and unit == 'g' and brutto > 1000:
                        conversions_found.append('kg→g')
                    elif 'молоко' in name.lower() and unit == 'ml' and brutto > 400:
                        conversions_found.append('l→ml')
                    elif 'масло' in name.lower() and unit == 'g' and brutto > 40:
                        conversions_found.append('ml→g')
                    elif 'яйца' in name.lower() and unit == 'g' and brutto > 50:
                        conversions_found.append('pcs→g')
                
                success = len(conversions_found) >= 3
                self.log_test(
                    "Unit Normalization", 
                    success, 
                    f"Conversions detected: {', '.join(conversions_found)}",
                    response_time
                )
            else:
                self.log_test("Unit Normalization", False, f"Failed: {response.status_code}")
                
        except Exception as e:
            self.log_test("Unit Normalization", False, f"Error: {str(e)}")
    
    def test_techcard_structure_validation(self):
        """Test 5: TechCardV2 Structure Validation"""
        print("\n🔍 Testing TechCardV2 Structure Validation...")
        
        try:
            test_file_bytes = self.create_test_xlsx_file("structure_test.xlsx", "basic")
            
            start_time = time.time()
            response = requests.post(
                f"{API_BASE}/v1/iiko/import/ttk.xlsx",
                files={'file': ('structure_test.xlsx', test_file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=30
            )
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                techcard = data.get('techcard', {})
                
                # Validate TechCardV2 structure
                required_fields = ['meta', 'portions', 'yield', 'ingredients', 'process', 'storage', 'nutrition', 'cost']
                missing_fields = [field for field in required_fields if field not in techcard]
                
                # Check SKU preservation
                ingredients = techcard.get('ingredients', [])
                sku_preserved = any(ing.get('skuId') for ing in ingredients)
                
                # Check meta information
                meta = techcard.get('meta', {})
                has_meta = bool(meta.get('title')) and bool(meta.get('version'))
                
                success = len(missing_fields) == 0 and has_meta
                details = f"Missing fields: {missing_fields}, SKU preserved: {sku_preserved}, Meta valid: {has_meta}"
                
                self.log_test(
                    "TechCardV2 Structure", 
                    success, 
                    details,
                    response_time
                )
            else:
                self.log_test("TechCardV2 Structure", False, f"Failed: {response.status_code}")
                
        except Exception as e:
            self.log_test("TechCardV2 Structure", False, f"Error: {str(e)}")
    
    def test_technology_parsing(self):
        """Test 6: Technology/Process Steps Parsing"""
        print("\n🔍 Testing Technology Parsing...")
        
        try:
            test_file_bytes = self.create_test_xlsx_file("technology_test.xlsx", "technology")
            
            start_time = time.time()
            response = requests.post(
                f"{API_BASE}/v1/iiko/import/ttk.xlsx",
                files={'file': ('technology_test.xlsx', test_file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=30
            )
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                techcard = data.get('techcard', {})
                process_steps = techcard.get('process', [])
                
                # Check if technology was parsed into steps
                has_steps = len(process_steps) > 0
                has_time_temp = any(step.get('time_min') or step.get('temp_c') for step in process_steps)
                
                success = has_steps
                details = f"Steps: {len(process_steps)}, Time/Temp extracted: {has_time_temp}"
                
                self.log_test(
                    "Technology Parsing", 
                    success, 
                    details,
                    response_time
                )
            else:
                self.log_test("Technology Parsing", False, f"Failed: {response.status_code}")
                
        except Exception as e:
            self.log_test("Technology Parsing", False, f"Error: {str(e)}")
    
    def test_error_handling(self):
        """Test 7: Error Handling and Issues Generation"""
        print("\n🔍 Testing Error Handling...")
        
        # Test empty file handling
        try:
            test_file_bytes = self.create_test_xlsx_file("empty_test.xlsx", "empty")
            
            start_time = time.time()
            response = requests.post(
                f"{API_BASE}/v1/iiko/import/ttk.xlsx",
                files={'file': ('empty_test.xlsx', test_file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=30
            )
            response_time = time.time() - start_time
            
            if response.status_code in [200, 400]:
                data = response.json()
                issues = data.get('issues', [])
                
                # Should have issues for empty/missing data
                has_issues = len(issues) > 0
                issue_types = [issue.get('code') for issue in issues]
                
                self.log_test(
                    "Empty File Issues", 
                    has_issues, 
                    f"Issues generated: {len(issues)}, Types: {issue_types[:3]}",
                    response_time
                )
            else:
                self.log_test("Empty File Issues", False, f"Unexpected status: {response.status_code}")
                
        except Exception as e:
            self.log_test("Empty File Issues", False, f"Error: {str(e)}")
        
        # Test graceful handling of parsing errors
        try:
            # Create corrupted XLSX-like content
            corrupted_content = b"Not a real XLSX file content"
            
            start_time = time.time()
            response = requests.post(
                f"{API_BASE}/v1/iiko/import/ttk.xlsx",
                files={'file': ('corrupted.xlsx', corrupted_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=30
            )
            response_time = time.time() - start_time
            
            # Should return 400 with error message, not crash
            if response.status_code == 400:
                self.log_test(
                    "Corrupted File Handling", 
                    True, 
                    "Corrupted file handled gracefully with 400 error",
                    response_time
                )
            else:
                self.log_test("Corrupted File Handling", False, f"Unexpected handling: {response.status_code}")
                
        except Exception as e:
            self.log_test("Corrupted File Handling", False, f"Error: {str(e)}")
    
    def test_recalculation_integration(self):
        """Test 8: Integration with Recalculation"""
        print("\n🔍 Testing Recalculation Integration...")
        
        try:
            test_file_bytes = self.create_test_xlsx_file("recalc_test.xlsx", "basic")
            
            start_time = time.time()
            response = requests.post(
                f"{API_BASE}/v1/iiko/import/ttk.xlsx",
                files={'file': ('recalc_test.xlsx', test_file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=30
            )
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                techcard = data.get('techcard', {})
                
                # Check if cost and nutrition were calculated
                cost = techcard.get('cost', {})
                nutrition = techcard.get('nutrition', {})
                cost_meta = techcard.get('costMeta', {})
                nutrition_meta = techcard.get('nutritionMeta', {})
                
                has_cost_calc = cost.get('rawCost') is not None or cost_meta.get('source') != 'calculation_needed'
                has_nutrition_calc = nutrition.get('per100g', {}).get('kcal', 0) > 0 or nutrition_meta.get('source') != 'calculation_needed'
                
                # Check SKU preservation through recalc
                ingredients = techcard.get('ingredients', [])
                skus_preserved = sum(1 for ing in ingredients if ing.get('skuId'))
                
                success = has_cost_calc or has_nutrition_calc
                details = f"Cost calc: {has_cost_calc}, Nutrition calc: {has_nutrition_calc}, SKUs preserved: {skus_preserved}"
                
                self.log_test(
                    "Recalculation Integration", 
                    success, 
                    details,
                    response_time
                )
            else:
                self.log_test("Recalculation Integration", False, f"Failed: {response.status_code}")
                
        except Exception as e:
            self.log_test("Recalculation Integration", False, f"Error: {str(e)}")
    
    def test_performance(self):
        """Test 9: Performance Testing"""
        print("\n🔍 Testing Performance...")
        
        # Test with medium-sized file
        try:
            test_file_bytes = self.create_test_xlsx_file("large_test.xlsx", "large")
            
            start_time = time.time()
            response = requests.post(
                f"{API_BASE}/v1/iiko/import/ttk.xlsx",
                files={'file': ('large_test.xlsx', test_file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=60  # Extended timeout for large file
            )
            response_time = time.time() - start_time
            
            # Performance target: ≤3 seconds for average file
            performance_ok = response_time <= 3.0
            
            if response.status_code == 200:
                data = response.json()
                meta = data.get('meta', {})
                parsed_rows = meta.get('parsed_rows', 0)
                
                success = performance_ok and parsed_rows > 100
                details = f"Time: {response_time:.2f}s (target ≤3s), Rows: {parsed_rows}"
                
                self.log_test(
                    "Performance Test", 
                    success, 
                    details,
                    response_time
                )
            else:
                self.log_test("Performance Test", False, f"Failed: {response.status_code}, Time: {response_time:.2f}s")
                
        except Exception as e:
            self.log_test("Performance Test", False, f"Error: {str(e)}")
    
    def test_yield_calculation(self):
        """Test 10: Yield and Portions Calculation"""
        print("\n🔍 Testing Yield Calculation...")
        
        try:
            test_file_bytes = self.create_test_xlsx_file("yield_test.xlsx", "basic")
            
            start_time = time.time()
            response = requests.post(
                f"{API_BASE}/v1/iiko/import/ttk.xlsx",
                files={'file': ('yield_test.xlsx', test_file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=30
            )
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                techcard = data.get('techcard', {})
                
                # Check yield calculation
                yield_info = techcard.get('yield', {})
                portions = techcard.get('portions', 0)
                
                has_yield = yield_info.get('perBatch_g', 0) > 0
                has_portions = portions > 0
                per_portion = yield_info.get('perPortion_g', 0)
                
                # Validate yield logic
                batch_weight = yield_info.get('perBatch_g', 0)
                calculated_per_portion = batch_weight / portions if portions > 0 else 0
                yield_logic_ok = abs(per_portion - calculated_per_portion) < 1.0  # Allow 1g tolerance
                
                success = has_yield and has_portions and yield_logic_ok
                details = f"Batch: {batch_weight}g, Portions: {portions}, Per portion: {per_portion}g"
                
                self.log_test(
                    "Yield Calculation", 
                    success, 
                    details,
                    response_time
                )
            else:
                self.log_test("Yield Calculation", False, f"Failed: {response.status_code}")
                
        except Exception as e:
            self.log_test("Yield Calculation", False, f"Error: {str(e)}")
    
    def run_comprehensive_tests(self):
        """Run all IK-04/01 tests"""
        print("🚀 Starting IK-04/01 Backend XLSX Parser Comprehensive Testing")
        print(f"Backend URL: {BACKEND_URL}")
        print("=" * 80)
        
        # Run all test categories
        self.test_endpoint_availability()
        self.test_file_validation()
        self.test_parser_core_functionality()
        self.test_unit_normalization()
        self.test_techcard_structure_validation()
        self.test_technology_parsing()
        self.test_error_handling()
        self.test_recalculation_integration()
        self.test_performance()
        self.test_yield_calculation()
        
        # Summary
        print("\n" + "=" * 80)
        print("🎯 IK-04/01 XLSX Parser Testing Summary")
        print("=" * 80)
        
        success_rate = (self.passed_tests / self.total_tests * 100) if self.total_tests > 0 else 0
        print(f"Total Tests: {self.total_tests}")
        print(f"Passed: {self.passed_tests}")
        print(f"Failed: {self.total_tests - self.passed_tests}")
        print(f"Success Rate: {success_rate:.1f}%")
        
        # Performance summary
        response_times = [r['response_time_ms'] for r in self.test_results if r['response_time_ms'] > 0]
        if response_times:
            avg_time = sum(response_times) / len(response_times)
            max_time = max(response_times)
            print(f"Average Response Time: {avg_time:.0f}ms")
            print(f"Max Response Time: {max_time:.0f}ms")
        
        # Critical issues
        failed_tests = [r for r in self.test_results if not r['success']]
        if failed_tests:
            print(f"\n❌ Failed Tests ({len(failed_tests)}):")
            for test in failed_tests:
                print(f"  • {test['test']}: {test['details']}")
        
        # Success indicators
        critical_tests = [
            "Endpoint Availability",
            "XLSX File Acceptance", 
            "Basic RU Headers Parsing",
            "TechCardV2 Structure",
            "Performance Test"
        ]
        
        critical_passed = sum(1 for r in self.test_results if r['test'] in critical_tests and r['success'])
        critical_total = len([r for r in self.test_results if r['test'] in critical_tests])
        
        print(f"\n🎯 Critical Tests: {critical_passed}/{critical_total} passed")
        
        if success_rate >= 80 and critical_passed >= 4:
            print("✅ IK-04/01 XLSX Parser is FUNCTIONAL and ready for production")
        elif success_rate >= 60:
            print("⚠️ IK-04/01 XLSX Parser has MINOR ISSUES but core functionality works")
        else:
            print("❌ IK-04/01 XLSX Parser has CRITICAL ISSUES requiring fixes")
        
        return success_rate >= 80

if __name__ == "__main__":
    tester = IK04XlsxParserTester()
    success = tester.run_comprehensive_tests()
    
    # Exit with appropriate code
    sys.exit(0 if success else 1)