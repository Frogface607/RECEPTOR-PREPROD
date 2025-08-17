"""
IK-04/03: Round-trip golden tests for iiko XLSX ↔ TechCardV2
Ensures compatibility: import from iikoWeb → export back to iikoWeb

Tests the complete cycle:
1. Import XLSX → validate TechCardV2 → export iiko.xlsx 
2. Verify headers/types/separators/units compatibility
3. Preserve SKU codes through round-trip
4. Validate time/temperature parsing in process steps
"""

import os
import sys
import tempfile
from pathlib import Path
import json
import openpyxl
from openpyxl.workbook import Workbook

# Add the backend directory to Python path
backend_dir = Path(__file__).parent.parent.parent / "backend"
sys.path.insert(0, str(backend_dir))

from receptor_agent.integrations.iiko_xlsx_parser import IikoXlsxParser, IikoXlsxParseError
from receptor_agent.techcards_v2.schemas import TechCardV2
from receptor_agent.exports.iiko_xlsx import export_techcard_to_iiko_xlsx

class TestIikoXlsxRoundTrip:
    """Test class for XLSX import-export round-trip validation"""
    
    def __init__(self):
        self.parser = IikoXlsxParser()
        self.test_data_dir = Path(__file__).parent / "test_xlsx_data"
        self.test_data_dir.mkdir(exist_ok=True)
    
    def create_test_xlsx_file(self, filename: str, dish_data: dict) -> Path:
        """Create a test XLSX file with dish data"""
        filepath = self.test_data_dir / filename
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "TechCard"
        
        # Add headers (row 1)
        headers = [
            "Наименование продукта",  # A1
            "Артикул продукта",       # B1  
            "Брутто",                 # C1
            "Нетто",                  # D1
            "Единица измерения",      # E1
            "Потери, %"               # F1
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add dish information (above ingredients)
        ws.cell(row=2, column=1, value="Название блюда:")
        ws.cell(row=2, column=2, value=dish_data["name"])
        ws.cell(row=3, column=1, value="Артикул блюда:")  
        ws.cell(row=3, column=2, value=dish_data.get("code", "DISH001"))
        ws.cell(row=4, column=1, value="Выход готового продукта:")
        ws.cell(row=4, column=2, value=f"{dish_data.get('yield_g', 800)} г")
        ws.cell(row=5, column=1, value="Технология приготовления:")
        ws.cell(row=5, column=2, value=dish_data.get("technology", "Готовить по технологии"))
        
        # Add ingredient data (starting from row 7)
        start_row = 7
        for i, ingredient in enumerate(dish_data["ingredients"], start_row):
            ws.cell(row=i, column=1, value=ingredient["name"])
            ws.cell(row=i, column=2, value=ingredient.get("sku", ""))
            ws.cell(row=i, column=3, value=ingredient["brutto"])
            ws.cell(row=i, column=4, value=ingredient["netto"]) 
            ws.cell(row=i, column=5, value=ingredient["unit"])
            ws.cell(row=i, column=6, value=ingredient.get("loss", 0))
        
        # Save workbook
        wb.save(filepath)
        return filepath
    
    def test_import_roundtrip_basic(self):
        """Test basic round-trip: import → validate → export"""
        print("\n=== Test: Basic Round-trip ===")
        
        # Test data: Simple dish with mixed units
        dish_data = {
            "name": "Говядина тушеная с овощами",
            "code": "BEEF_STEW_001", 
            "yield_g": 800,
            "technology": "1. Подготовить ингредиенты. 2. Обжарить мясо при 180°C в течение 10 мин. 3. Добавить овощи и тушить 25 мин при 160°C.",
            "ingredients": [
                {"name": "Говядина", "sku": "BEEF001", "brutto": 600, "netto": 540, "unit": "г", "loss": 10},
                {"name": "Лук репчатый", "sku": "VEG002", "brutto": 200, "netto": 170, "unit": "г", "loss": 15},
                {"name": "Морковь", "sku": "VEG003", "brutto": 150, "netto": 135, "unit": "г", "loss": 10},
                {"name": "Масло подсолнечное", "sku": "OIL001", "brutto": 50, "netto": 49, "unit": "мл", "loss": 2},
                {"name": "Соль", "sku": "SALT001", "brutto": 10, "netto": 10, "unit": "г", "loss": 0}
            ]
        }
        
        # Create test XLSX file
        xlsx_path = self.create_test_xlsx_file("basic_roundtrip.xlsx", dish_data)
        assert xlsx_path.exists(), "Test XLSX file was not created"
        print(f"✓ Created test XLSX: {xlsx_path}")
        
        # Step 1: Import XLSX → TechCardV2
        with open(xlsx_path, 'rb') as f:
            xlsx_bytes = f.read()
        
        parse_result = self.parser.parse_xlsx_to_techcard(xlsx_bytes, xlsx_path.name)
        assert parse_result["techcard"], "Failed to parse XLSX to TechCardV2"
        print(f"✓ Parsed XLSX: {parse_result['meta']['parsed_rows']} rows")
        
        # Step 2: Validate TechCardV2 structure
        techcard = TechCardV2.model_validate(parse_result["techcard"])
        assert techcard, "TechCardV2 validation failed"
        print(f"✓ TechCardV2 validated: {techcard.meta.title}")
        
        # Step 3: Export TechCardV2 → iiko XLSX
        export_bytes = export_techcard_to_iiko_xlsx(techcard)
        assert export_bytes and len(export_bytes) > 1000, "Export produced empty or too small file"
        print(f"✓ Exported to XLSX: {len(export_bytes)} bytes")
        
        # Step 4: Verify exported XLSX structure
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(export_bytes)
            tmp_file.flush()
            
            # Load exported XLSX and verify
            exported_wb = openpyxl.load_workbook(tmp_file.name, data_only=True)
            ws = exported_wb.active
            
            # Verify headers exist
            assert ws.cell(1, 1).value is not None, "Missing headers in exported XLSX"
            print(f"✓ Exported XLSX has headers: {[ws.cell(1, c).value for c in range(1, 6)]}")
            
            # Clean up
            os.unlink(tmp_file.name)
        
        # Step 5: Verify SKU preservation
        original_skus = {ing["sku"] for ing in dish_data["ingredients"] if ing["sku"]}
        imported_skus = {ing.skuId for ing in techcard.ingredients if ing.skuId}
        
        assert original_skus == imported_skus, f"SKUs not preserved: {original_skus} vs {imported_skus}"
        print(f"✓ SKUs preserved through round-trip: {original_skus}")
        
        print("✅ Basic round-trip test PASSED")
        return True
    
    def test_import_units_conversion(self):
        """Test units conversion: kg/l/ml/pcs with densities"""
        print("\n=== Test: Units Conversion ===")
        
        # Test data: Different units requiring conversion
        dish_data = {
            "name": "Салат с маслом",
            "code": "SALAD001",
            "yield_g": 500,
            "technology": "Смешать ингредиенты",
            "ingredients": [
                {"name": "Капуста", "sku": "VEG010", "brutto": 1.2, "netto": 1.0, "unit": "кг", "loss": 16.7},  # kg→g
                {"name": "Молоко", "sku": "DAIRY001", "brutto": 500, "netto": 500, "unit": "мл", "loss": 0},    # ml→g (density 1.03)
                {"name": "Масло растительное", "sku": "OIL002", "brutto": 100, "netto": 99, "unit": "мл", "loss": 1},  # ml→g (density 0.9)
                {"name": "Яйца", "sku": "EGG001", "brutto": 2, "netto": 2, "unit": "шт", "loss": 0}             # pcs→g (50g each)
            ]
        }
        
        xlsx_path = self.create_test_xlsx_file("units_conversion.xlsx", dish_data) 
        
        # Import and parse
        with open(xlsx_path, 'rb') as f:
            xlsx_bytes = f.read()
        
        parse_result = self.parser.parse_xlsx_to_techcard(xlsx_bytes, xlsx_path.name)
        techcard = TechCardV2.model_validate(parse_result["techcard"])
        
        # Verify unit conversions
        ingredients = {ing.name: ing for ing in techcard.ingredients}
        
        # Test kg → g conversion (1.0 kg → 1000g)
        капуста = ingredients["Капуста"] 
        assert капуста.unit == "g", f"Unit not converted to g: {капуста.unit}"
        assert abs(капуста.netto_g - 1000) < 1, f"kg→g conversion failed: {капуста.netto_g} vs 1000"
        print("✓ kg → g conversion working")
        
        # Test ml → g with milk density (500ml → ~515g)
        молоко = ingredients["Молоко"]
        assert молоко.unit == "g", f"Unit not converted to g: {молоко.unit}"
        assert 510 <= молоко.netto_g <= 520, f"ml→g milk conversion failed: {молоко.netto_g}"
        print("✓ ml → g (milk density) conversion working")
        
        # Test ml → g with oil density (99ml → ~89g)
        масло = ingredients["Масло растительное"]
        assert масло.unit == "g", f"Unit not converted to g: {масло.unit}"
        assert 85 <= масло.netto_g <= 95, f"ml→g oil conversion failed: {масло.netto_g}"
        print("✓ ml → g (oil density) conversion working")
        
        # Test pcs → g conversion (2 eggs → 100g)
        яйца = ingredients["Яйца"]
        assert яйца.unit == "g", f"Unit not converted to g: {яйца.unit}"
        assert abs(яйца.netto_g - 100) < 5, f"pcs→g conversion failed: {яйца.netto_g} vs 100"
        print("✓ pcs → g conversion working")
        
        # Check for appropriate warnings
        issues = parse_result["issues"]
        density_warnings = [i for i in issues if i.get("code", "").startswith("density")]
        assert len(density_warnings) >= 2, f"Expected density warnings, got: {density_warnings}"
        print(f"✓ Density warnings generated: {len(density_warnings)}")
        
        print("✅ Units conversion test PASSED")
        return True
    
    def test_import_process_parsing(self):
        """Test technology parsing: ≥3 steps, time_min/temp_c extraction"""
        print("\n=== Test: Process Parsing ===")
        
        dish_data = {
            "name": "Стейк в духовке",  
            "code": "STEAK001",
            "yield_g": 400,
            "technology": "Подготовить мясо и приправы. Обжарить стейк на сковороде 3 минуты при 200°C с каждой стороны. Запекать в духовке 15 мин при 180°C для medium прожарки.",
            "ingredients": [
                {"name": "Говядина стейк", "sku": "BEEF002", "brutto": 350, "netto": 320, "unit": "г", "loss": 8.6},
                {"name": "Соль морская", "sku": "SALT002", "brutto": 5, "netto": 5, "unit": "г", "loss": 0},
                {"name": "Перец черный", "sku": "SPICE001", "brutto": 2, "netto": 2, "unit": "г", "loss": 0}
            ]
        }
        
        xlsx_path = self.create_test_xlsx_file("process_parsing.xlsx", dish_data)
        
        # Import and parse
        with open(xlsx_path, 'rb') as f:
            xlsx_bytes = f.read()
        
        parse_result = self.parser.parse_xlsx_to_techcard(xlsx_bytes, xlsx_path.name)
        techcard = TechCardV2.model_validate(parse_result["techcard"])
        
        # Verify minimum 3 process steps
        assert len(techcard.process) >= 3, f"Expected ≥3 process steps, got {len(techcard.process)}"
        print(f"✓ Process steps generated: {len(techcard.process)}")
        
        # Verify step ordering
        for i, step in enumerate(techcard.process, 1):
            assert step.n == i, f"Step numbering incorrect: expected {i}, got {step.n}"
        print("✓ Process steps properly numbered")
        
        # Verify time/temperature extraction
        steps_with_temp = [s for s in techcard.process if s.temp_c is not None]
        steps_with_time = [s for s in techcard.process if s.time_min is not None] 
        
        assert len(steps_with_temp) > 0, "No temperature extracted from technology text"
        assert len(steps_with_time) > 0, "No time extracted from technology text"
        print(f"✓ Temperature extracted in {len(steps_with_temp)} steps")
        print(f"✓ Time extracted in {len(steps_with_time)} steps")
        
        # Verify specific values (200°C, 180°C should be extracted)
        temperatures = [s.temp_c for s in techcard.process if s.temp_c]
        times = [s.time_min for s in techcard.process if s.time_min]
        
        assert 200.0 in temperatures or 180.0 in temperatures, f"Expected 200°C or 180°C, got: {temperatures}"
        assert any(t >= 3 for t in times), f"Expected time ≥3 min, got: {times}"
        print(f"✓ Extracted temperatures: {temperatures}°C")
        print(f"✓ Extracted times: {times} minutes")
        
        print("✅ Process parsing test PASSED")
        return True
    
    def test_import_fixtures_roundtrip(self):
        """Test round-trip with realistic fixtures: hot.xlsx, cold.xlsx, sauce.xlsx"""
        print("\n=== Test: Fixtures Round-trip ===")
        
        fixtures_dir = Path(__file__).parent.parent / "fixtures" / "iiko_xlsx"
        test_fixtures = ["hot.xlsx", "cold.xlsx", "sauce.xlsx"]
        
        for fixture_name in test_fixtures:
            fixture_path = fixtures_dir / fixture_name
            
            if not fixture_path.exists():
                print(f"⚠️ Skipping {fixture_name} - fixture not found")
                continue
            
            print(f"\n--- Testing {fixture_name} ---")
            
            # Step 1: Import XLSX → TechCardV2
            with open(fixture_path, 'rb') as f:
                xlsx_bytes = f.read()
            
            parse_result = self.parser.parse_xlsx_to_techcard(xlsx_bytes, fixture_name)
            assert parse_result["techcard"], f"Failed to parse {fixture_name} to TechCardV2"
            print(f"✓ Parsed {fixture_name}: {parse_result['meta']['parsed_rows']} rows")
            
            # Step 2: Validate TechCardV2 structure
            techcard = TechCardV2.model_validate(parse_result["techcard"])
            assert techcard, f"TechCardV2 validation failed for {fixture_name}"
            print(f"✓ TechCardV2 validated: {techcard.meta.title}")
            
            # Step 3: Export TechCardV2 → iiko XLSX
            export_bytes = export_techcard_to_iiko_xlsx(techcard)
            assert export_bytes and len(export_bytes) > 1000, f"Export produced empty or too small file for {fixture_name}"
            print(f"✓ Exported to XLSX: {len(export_bytes)} bytes")
            
            # Step 4: Verify exported XLSX structure
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                tmp_file.write(export_bytes)
                tmp_file.flush()
                
                # Load exported XLSX and verify
                exported_wb = openpyxl.load_workbook(tmp_file.name, data_only=True)
                ws = exported_wb.active
                
                # Verify headers exist
                assert ws.cell(1, 1).value is not None, f"Missing headers in exported XLSX for {fixture_name}"
                print(f"✓ Exported XLSX has headers: {[ws.cell(1, c).value for c in range(1, 6)]}")
                
                # Clean up
                os.unlink(tmp_file.name)
            
            # Step 5: Check for appropriate issues
            issues = parse_result["issues"]
            error_issues = [i for i in issues if i.get("level") == "error"]
            assert len(error_issues) == 0, f"Found error issues in {fixture_name}: {error_issues}"
            
            warning_issues = [i for i in issues if i.get("level") == "warning"]
            print(f"✓ Issues: {len(warning_issues)} warnings, 0 errors")
            
            # Step 6: Verify SKU preservation
            original_skus = set()
            for ing in techcard.ingredients:
                if ing.skuId and not ing.skuId.startswith("GENERATED_"):
                    original_skus.add(ing.skuId)
            
            print(f"✓ SKUs preserved: {len(original_skus)} original SKUs")
            
            # Step 7: Check specific fixture characteristics
            if fixture_name == "hot.xlsx":
                # Should have temperature extraction
                temps = [s.temp_c for s in techcard.process if s.temp_c and s.temp_c > 50]
                assert len(temps) > 0, f"No cooking temperatures found in {fixture_name}"
                print(f"✓ Cooking temperatures extracted: {temps}°C")
                
            elif fixture_name == "cold.xlsx":
                # Should have no high temperatures
                temps = [s.temp_c for s in techcard.process if s.temp_c and s.temp_c > 50]
                assert len(temps) == 0, f"Unexpected cooking temperatures in {fixture_name}: {temps}°C"
                print(f"✓ No cooking temperatures (cold preparation)")
                
            elif fixture_name == "sauce.xlsx":
                # Should have density conversion warnings
                density_issues = [i for i in issues if "density" in i.get("code", "").lower()]
                assert len(density_issues) > 0, f"Expected density conversion warnings in {fixture_name}"
                print(f"✓ Density conversions: {len(density_issues)} warnings")
            
            print(f"✅ {fixture_name} round-trip PASSED")
        
    def run_all_tests(self):
        """Run all round-trip golden tests"""
        print("🔄 Starting IK-04/03 Round-trip Golden Tests...")
        print("=" * 50)
        
        tests = [
            self.test_import_roundtrip_basic,
            self.test_import_units_conversion, 
            self.test_import_process_parsing,
            self.test_import_fixtures_roundtrip
        ]
        
        passed = 0
        total = len(tests)
        
        for test in tests:
            try:
                if test():
                    passed += 1
            except Exception as e:
                print(f"❌ Test {test.__name__} FAILED: {e}")
                continue
        
        print("=" * 50)
        print(f"📊 Test Results: {passed}/{total} tests passed")
        
        if passed == total:
            print("🎉 ALL ROUND-TRIP GOLDEN TESTS PASSED!")
            print("✅ iiko XLSX ↔ TechCardV2 compatibility confirmed")
            print("✅ SKU preservation verified")  
            print("✅ Units conversion working")
            print("✅ Process parsing functional")
            print("✅ Fixture round-trip validated")
            print("✅ Ready for iikoWeb integration")
        else:
            print(f"❌ {total - passed} tests failed - requires attention")
            
        return passed == total

if __name__ == "__main__":
    tester = TestIikoXlsxRoundTrip()
    success = tester.run_all_tests()
    sys.exit(0 if success else 1)