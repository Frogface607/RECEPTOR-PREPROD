"""
IK-04/03: Create test fixtures for round-trip testing
Creates hot.xlsx, cold.xlsx, sauce.xlsx with real iiko data
"""

import openpyxl
from pathlib import Path


def create_hot_xlsx():
    """Create hot.xlsx fixture with cooking temperatures"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ТТК"
    
    # Headers
    headers = [
        "Наименование продукта",  # A1
        "Артикул продукта",       # B1  
        "Брутто",                 # C1
        "Нетто",                  # D1
        "Единица измерения",      # E1
        "Потери, %"               # F1
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Dish metadata
    ws.cell(row=2, column=1, value="Название блюда:")
    ws.cell(row=2, column=2, value="Говядина жареная с луком")
    ws.cell(row=3, column=1, value="Артикул блюда:")  
    ws.cell(row=3, column=2, value="HOT_BEEF_001")
    ws.cell(row=4, column=1, value="Выход готового продукта:")
    ws.cell(row=4, column=2, value="350 г")
    ws.cell(row=5, column=1, value="Технология приготовления:")
    ws.cell(row=5, column=2, value="Подготовить мясо и овощи. Нарезать говядину кусочками 2x2 см. Разогреть сковороду до 180°C с маслом. Обжарить мясо 8 минут до золотистой корочки. Добавить лук и жарить еще 5 минут при 160°C. Довести до готовности за 3 мин при 140°C.")
    
    # Ingredients
    ingredients = [
        {"name": "Говядина", "sku": "MEAT_BEEF_001", "brutto": 400, "netto": 350, "unit": "г", "loss": 12.5},
        {"name": "Лук репчатый", "sku": "VEG_ONION_002", "brutto": 80, "netto": 70, "unit": "г", "loss": 12.5},
        {"name": "Масло подсолнечное", "sku": "OIL_SUNFLOWER_001", "brutto": 20, "netto": 19, "unit": "мл", "loss": 5.0},
        {"name": "Соль поваренная", "sku": "SPICE_SALT_001", "brutto": 3, "netto": 3, "unit": "г", "loss": 0},
        {"name": "Перец черный молотый", "sku": "SPICE_PEPPER_001", "brutto": 1, "netto": 1, "unit": "г", "loss": 0}
    ]
    
    start_row = 7
    for i, ingredient in enumerate(ingredients, start_row):
        ws.cell(row=i, column=1, value=ingredient["name"])
        ws.cell(row=i, column=2, value=ingredient["sku"])
        ws.cell(row=i, column=3, value=ingredient["brutto"])
        ws.cell(row=i, column=4, value=ingredient["netto"]) 
        ws.cell(row=i, column=5, value=ingredient["unit"])
        ws.cell(row=i, column=6, value=ingredient["loss"])
    
    filepath = Path(__file__).parent / "iiko_xlsx" / "hot.xlsx"
    wb.save(filepath)
    print(f"✓ Created {filepath}")


def create_cold_xlsx():
    """Create cold.xlsx fixture without thermal processing"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ТТК"
    
    # Headers
    headers = [
        "Наименование продукта",  # A1
        "Артикул продукта",       # B1  
        "Брутто",                 # C1
        "Нетто",                  # D1
        "Единица измерения",      # E1
        "Потери, %"               # F1
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Dish metadata
    ws.cell(row=2, column=1, value="Название блюда:")
    ws.cell(row=2, column=2, value="Салат греческий")
    ws.cell(row=3, column=1, value="Артикул блюда:")  
    ws.cell(row=3, column=2, value="COLD_SALAD_001")
    ws.cell(row=4, column=1, value="Выход готового продукта:")
    ws.cell(row=4, column=2, value="250 г")
    ws.cell(row=5, column=1, value="Технология приготовления:")
    ws.cell(row=5, column=2, value="Промыть и обсушить овощи. Нарезать помидоры дольками, огурцы кружочками, лук полукольцами. Сыр нарезать кубиками. Смешать все ингредиенты в салатнице. Заправить маслом и лимонным соком. Добавить специи по вкусу.")
    
    # Ingredients
    ingredients = [
        {"name": "Помидоры", "sku": "VEG_TOMATO_001", "brutto": 100, "netto": 90, "unit": "г", "loss": 10.0},
        {"name": "Огурцы свежие", "sku": "VEG_CUCUMBER_001", "brutto": 80, "netto": 75, "unit": "г", "loss": 6.25},
        {"name": "Лук красный", "sku": "VEG_REDONION_001", "brutto": 40, "netto": 35, "unit": "г", "loss": 12.5},
        {"name": "Сыр фета", "sku": "DAIRY_FETA_001", "brutto": 50, "netto": 50, "unit": "г", "loss": 0},
        {"name": "Маслины", "sku": "VEG_OLIVES_001", "brutto": 30, "netto": 25, "unit": "г", "loss": 16.7},
        {"name": "Масло оливковое", "sku": "OIL_OLIVE_001", "brutto": 15, "netto": 15, "unit": "мл", "loss": 0},
        {"name": "Лимонный сок", "sku": "JUICE_LEMON_001", "brutto": 10, "netto": 10, "unit": "мл", "loss": 0}
    ]
    
    start_row = 7
    for i, ingredient in enumerate(ingredients, start_row):
        ws.cell(row=i, column=1, value=ingredient["name"])
        ws.cell(row=i, column=2, value=ingredient["sku"])
        ws.cell(row=i, column=3, value=ingredient["brutto"])
        ws.cell(row=i, column=4, value=ingredient["netto"]) 
        ws.cell(row=i, column=5, value=ingredient["unit"])
        ws.cell(row=i, column=6, value=ingredient["loss"])
    
    filepath = Path(__file__).parent / "iiko_xlsx" / "cold.xlsx"
    wb.save(filepath)
    print(f"✓ Created {filepath}")


def create_sauce_xlsx():
    """Create sauce.xlsx fixture with viscous units ml→g conversions"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ТТК"
    
    # Headers
    headers = [
        "Наименование продукта",  # A1
        "Артикул продукта",       # B1  
        "Брутто",                 # C1
        "Нетто",                  # D1
        "Единица измерения",      # E1
        "Потери, %"               # F1
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Dish metadata
    ws.cell(row=2, column=1, value="Название блюда:")
    ws.cell(row=2, column=2, value="Соус сливочно-медовый")
    ws.cell(row=3, column=1, value="Артикул блюда:")  
    ws.cell(row=3, column=2, value="SAUCE_HONEY_001")
    ws.cell(row=4, column=1, value="Выход готового продукта:")
    ws.cell(row=4, column=2, value="200 мл")
    ws.cell(row=5, column=1, value="Технология приготовления:")
    ws.cell(row=5, column=2, value="Разогреть сливки до 60°C на медленном огне 3 минуты. Добавить мёд и перемешать. Варить 5 минут при 70°C помешивая. Добавить сливочное масло и взбить венчиком 2 минуты. Довести до кипения за 1 минуту при 95°C. Остудить до комнатной температуры.")
    
    # Ingredients with various viscous units
    ingredients = [
        {"name": "Сливки жирные", "sku": "DAIRY_CREAM_001", "brutto": 150, "netto": 150, "unit": "мл", "loss": 0},  # ml→g with cream density
        {"name": "Мёд натуральный", "sku": "SWEET_HONEY_001", "brutto": 40, "netto": 40, "unit": "мл", "loss": 0},  # ml→g with honey density  
        {"name": "Масло сливочное", "sku": "DAIRY_BUTTER_001", "brutto": 20, "netto": 20, "unit": "г", "loss": 0},
        {"name": "Ванильный экстракт", "sku": "EXTRACT_VANILLA_001", "brutto": 2, "netto": 2, "unit": "мл", "loss": 0},
        {"name": "Сироп кленовый", "sku": "SWEET_SYRUP_001", "brutto": 10, "netto": 10, "unit": "мл", "loss": 0}  # ml→g with syrup density
    ]
    
    start_row = 7
    for i, ingredient in enumerate(ingredients, start_row):
        ws.cell(row=i, column=1, value=ingredient["name"])
        ws.cell(row=i, column=2, value=ingredient["sku"])
        ws.cell(row=i, column=3, value=ingredient["brutto"])
        ws.cell(row=i, column=4, value=ingredient["netto"]) 
        ws.cell(row=i, column=5, value=ingredient["unit"])
        ws.cell(row=i, column=6, value=ingredient["loss"])
    
    filepath = Path(__file__).parent / "iiko_xlsx" / "sauce.xlsx"
    wb.save(filepath)
    print(f"✓ Created {filepath}")


if __name__ == "__main__":
    print("🔄 Creating IK-04/03 test fixtures...")
    print("=" * 50)
    
    create_hot_xlsx()
    create_cold_xlsx() 
    create_sauce_xlsx()
    
    print("=" * 50)
    print("🎉 All fixtures created successfully!")
    print("✅ hot.xlsx - Thermal processing with temperatures")
    print("✅ cold.xlsx - Cold preparation without thermal processing")  
    print("✅ sauce.xlsx - Viscous units with ml→g density conversions")