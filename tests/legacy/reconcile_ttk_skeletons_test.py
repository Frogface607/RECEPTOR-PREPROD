#!/usr/bin/env python3
"""
FINAL RECONCILE TTK↔SKELETONS COMPREHENSIVE TESTING
Critical verification of complete reconcile workflow with no GENERATED_*, proper Dish-Skeletons writer, and alt/legacy UI removal.
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Any, Optional

import httpx
import openpyxl
from pymongo import MongoClient

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017/receptor_pro')
DB_NAME = os.getenv('DB_NAME', 'receptor_pro')

class ReconcileTTKSkeletonsTest:
    """Comprehensive Reconcile TTK↔Skeletons Testing Suite"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=60.0)
        self.test_results = []
        self.organization_id = "test-org-reconcile"
        self.generated_techcard_id = None
        self.artifacts = {}
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status}: {test_name} ({response_time:.3f}s) - {details}")
    
    async def step_1_generate_real_techcard(self):
        """STEP 1: Generate real tech card to replace REPLACE_TECHCARD_ID_1"""
        print("\n🎯 STEP 1: Generating Real Tech Card")
        
        try:
            start_time = time.time()
            
            # Generate a real tech card with Russian dish name
            payload = {
                "name": "Борщ украинский с говядиной",
                "cuisine": "русская",
                "equipment": [],
                "budget": None,
                "dietary": []
            }
            
            response = await self.client.post(f"{API_BASE}/techcards.v2/generate", json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                print(f"DEBUG: Response data keys: {list(data.keys())}")
                print(f"DEBUG: Status: {data.get('status')}")
                
                if data.get("status") in ["success", "draft"] and data.get("card"):
                    techcard = data["card"]
                    print(f"DEBUG: TechCard keys: {list(techcard.keys()) if techcard else 'None'}")
                    
                    # Try to get ID from meta field
                    meta = techcard.get("meta", {})
                    self.generated_techcard_id = meta.get("id") or techcard.get("id")
                    print(f"DEBUG: TechCard ID: {self.generated_techcard_id}")
                    
                    # Validate tech card structure
                    required_fields = ["meta", "ingredients"]
                    has_required = all(field in techcard for field in required_fields)
                    
                    ingredients_count = len(techcard.get("ingredients", []))
                    
                    if has_required and self.generated_techcard_id and ingredients_count > 0:
                        self.log_test("Generate Real TechCard", True,
                                    f"ID: {self.generated_techcard_id}, Ingredients: {ingredients_count}", response_time)
                        
                        # Save to artifacts
                        self.artifacts["generated_techcard"] = {
                            "id": self.generated_techcard_id,
                            "name": techcard.get("name"),
                            "ingredients_count": ingredients_count,
                            "generation_time": response_time
                        }
                        
                        return True
                    else:
                        self.log_test("Generate Real TechCard", False,
                                    f"Invalid structure: required={has_required}, id={bool(self.generated_techcard_id)}, ingredients={ingredients_count}", response_time)
                else:
                    self.log_test("Generate Real TechCard", False,
                                f"Generation failed: {data.get('error', 'Unknown error')}", response_time)
            else:
                self.log_test("Generate Real TechCard", False,
                            f"HTTP {response.status_code}: {response.text[:200]}", response_time)
                
        except Exception as e:
            self.log_test("Generate Real TechCard", False, f"Exception: {str(e)}", 0.0)
        
        return False
    
    async def step_2_run_preflight(self):
        """STEP 2: Run preflight check and save artifacts"""
        print("\n🎯 STEP 2: Running Preflight Check")
        
        if not self.generated_techcard_id:
            self.log_test("Preflight Check", False, "No generated tech card ID available", 0.0)
            return False
        
        try:
            start_time = time.time()
            
            payload = {
                "techcardIds": [self.generated_techcard_id],
                "organization_id": self.organization_id
            }
            
            response = await self.client.post(f"{API_BASE}/export/preflight", json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate preflight response structure
                required_fields = ["status", "ttkDate", "missing", "generated", "counts"]
                missing_fields = [field for field in required_fields if field not in data]
                
                if not missing_fields:
                    # Extract key metrics
                    ttk_date = data.get("ttkDate")
                    missing_dishes = data.get("missing", {}).get("dishes", [])
                    missing_products = data.get("missing", {}).get("products", [])
                    dish_count = data.get("counts", {}).get("dishSkeletons", 0)
                    product_count = data.get("counts", {}).get("productSkeletons", 0)
                    
                    # Save to artifacts/preflight.json
                    self.artifacts["preflight"] = {
                        "status": data.get("status"),
                        "ttkDate": ttk_date,
                        "counts": {
                            "dishSkeletons": dish_count,
                            "productSkeletons": product_count
                        },
                        "missing": {
                            "dishes": len(missing_dishes),
                            "products": len(missing_products)
                        },
                        "response_time": response_time
                    }
                    
                    details = f"TTK Date: {ttk_date}, Dish Skeletons: {dish_count}, Product Skeletons: {product_count}"
                    self.log_test("Preflight Check", True, details, response_time)
                    
                    # Store full preflight result for next step
                    self.preflight_result = data
                    return True
                else:
                    self.log_test("Preflight Check", False,
                                f"Missing fields: {missing_fields}", response_time)
            else:
                self.log_test("Preflight Check", False,
                            f"HTTP {response.status_code}: {response.text[:200]}", response_time)
                
        except Exception as e:
            self.log_test("Preflight Check", False, f"Exception: {str(e)}", 0.0)
        
        return False
    
    async def step_3_zip_export(self):
        """STEP 3: Call /export/zip and save zipUrl and composition"""
        print("\n🎯 STEP 3: ZIP Export and Analysis")
        
        if not hasattr(self, 'preflight_result'):
            self.log_test("ZIP Export", False, "No preflight result available", 0.0)
            return False
        
        try:
            start_time = time.time()
            
            payload = {
                "techcardIds": [self.generated_techcard_id],
                "operational_rounding": True,
                "organization_id": self.organization_id,
                "preflight_result": self.preflight_result
            }
            
            response = await self.client.post(f"{API_BASE}/export/zip", json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # Check if response is a ZIP file
                content_type = response.headers.get('content-type', '')
                content_length = len(response.content)
                
                is_zip = (content_type == 'application/zip' or 
                         response.content.startswith(b'PK'))
                
                if is_zip and content_length > 0:
                    # Unpack and analyze ZIP contents
                    zip_analysis = await self._analyze_zip_contents(response.content)
                    
                    # Save to artifacts
                    self.artifacts["zip_export"] = {
                        "size_bytes": content_length,
                        "files": zip_analysis.get("files", []),
                        "file_count": len(zip_analysis.get("files", [])),
                        "response_time": response_time
                    }
                    
                    # Store ZIP content for next steps
                    self.zip_content = response.content
                    
                    details = f"ZIP: {content_length} bytes, Files: {len(zip_analysis.get('files', []))}"
                    self.log_test("ZIP Export", True, details, response_time)
                    return True
                else:
                    self.log_test("ZIP Export", False,
                                f"Invalid ZIP: content-type={content_type}, size={content_length}", response_time)
            else:
                self.log_test("ZIP Export", False,
                            f"HTTP {response.status_code}: {response.text[:200]}", response_time)
                
        except Exception as e:
            self.log_test("ZIP Export", False, f"Exception: {str(e)}", 0.0)
        
        return False
    
    async def step_4_reconcile_products(self):
        """STEP 4: Collect map from Product-Skeletons.xlsx and check for GENERATED_* in TTK"""
        print("\n🎯 STEP 4: Reconcile Products - Check for GENERATED_*")
        
        if not hasattr(self, 'zip_content'):
            self.log_test("Reconcile Products", False, "No ZIP content available", 0.0)
            return False
        
        try:
            start_time = time.time()
            
            # Extract and analyze XLSX files from ZIP
            with zipfile.ZipFile(io.BytesIO(self.zip_content), 'r') as zip_file:
                file_names = zip_file.namelist()
                
                # Look for Product-Skeletons.xlsx and iiko_TTK.xlsx
                product_skeletons_found = False
                ttk_file_found = False
                generated_found = []
                
                for file_name in file_names:
                    if 'Product-Skeletons.xlsx' in file_name:
                        product_skeletons_found = True
                        
                        # Analyze Product-Skeletons.xlsx
                        with zip_file.open(file_name) as xlsx_file:
                            workbook = openpyxl.load_workbook(xlsx_file)
                            sheet = workbook.active
                            
                            # Check for article mapping
                            product_articles = []
                            for row in sheet.iter_rows(min_row=2, values_only=True):
                                if row and row[0]:  # Article column
                                    article = str(row[0])
                                    product_articles.append(article)
                            
                            self.artifacts["product_skeletons"] = {
                                "found": True,
                                "articles_count": len(product_articles),
                                "sample_articles": product_articles[:5]
                            }
                    
                    elif 'iiko_TTK.xlsx' in file_name:
                        ttk_file_found = True
                        
                        # Analyze iiko_TTK.xlsx for GENERATED_* content
                        with zip_file.open(file_name) as xlsx_file:
                            workbook = openpyxl.load_workbook(xlsx_file)
                            sheet = workbook.active
                            
                            # Check all cells for GENERATED_* content
                            for row in sheet.iter_rows(values_only=True):
                                for cell_value in row:
                                    if cell_value and isinstance(cell_value, str):
                                        if 'GENERATED_' in cell_value:
                                            generated_found.append(cell_value)
                
                response_time = time.time() - start_time
                
                # Critical validation: NO GENERATED_* should be found
                no_generated = len(generated_found) == 0
                
                if no_generated:
                    details = f"✅ NO GENERATED_* found in TTK, Product Skeletons: {product_skeletons_found}"
                    self.log_test("Reconcile Products - No GENERATED_*", True, details, response_time)
                else:
                    details = f"❌ FOUND GENERATED_*: {generated_found[:3]}"
                    self.log_test("Reconcile Products - No GENERATED_*", False, details, response_time)
                
                # Save reconciliation results
                self.artifacts["reconcile_products"] = {
                    "product_skeletons_found": product_skeletons_found,
                    "ttk_file_found": ttk_file_found,
                    "generated_content_found": generated_found,
                    "no_generated_content": no_generated
                }
                
                return no_generated
                
        except Exception as e:
            self.log_test("Reconcile Products", False, f"Exception: {str(e)}", 0.0)
        
        return False
    
    async def step_5_ensure_dish_skeleton(self):
        """STEP 5: Check that Dish-Skeletons.xlsx is filled when missing>0"""
        print("\n🎯 STEP 5: Ensure Dish-Skeletons Writer")
        
        if not hasattr(self, 'zip_content') or not hasattr(self, 'preflight_result'):
            self.log_test("Dish-Skeletons Writer", False, "Missing ZIP content or preflight result", 0.0)
            return False
        
        try:
            start_time = time.time()
            
            # Check if we have missing dishes from preflight
            missing_dishes = self.preflight_result.get("missing", {}).get("dishes", [])
            missing_count = len(missing_dishes)
            
            # Extract and check Dish-Skeletons.xlsx
            with zipfile.ZipFile(io.BytesIO(self.zip_content), 'r') as zip_file:
                file_names = zip_file.namelist()
                
                dish_skeletons_found = False
                dish_rows_count = 0
                
                for file_name in file_names:
                    if 'Dish-Skeletons.xlsx' in file_name:
                        dish_skeletons_found = True
                        
                        # Analyze Dish-Skeletons.xlsx content
                        with zip_file.open(file_name) as xlsx_file:
                            workbook = openpyxl.load_workbook(xlsx_file)
                            sheet = workbook.active
                            
                            # Count data rows (excluding header)
                            for row in sheet.iter_rows(min_row=2, values_only=True):
                                if row and any(cell for cell in row):  # Non-empty row
                                    dish_rows_count += 1
                
                response_time = time.time() - start_time
                
                # Validation logic: If missing>0, Dish-Skeletons.xlsx should contain rows
                if missing_count > 0:
                    skeleton_properly_filled = dish_skeletons_found and dish_rows_count > 0
                    
                    if skeleton_properly_filled:
                        details = f"✅ Missing: {missing_count}, Dish-Skeletons rows: {dish_rows_count}"
                        self.log_test("Dish-Skeletons Writer", True, details, response_time)
                    else:
                        details = f"❌ Missing: {missing_count}, but Dish-Skeletons empty or not found"
                        self.log_test("Dish-Skeletons Writer", False, details, response_time)
                        return False
                else:
                    # No missing dishes, so Dish-Skeletons may be empty (acceptable)
                    details = f"✅ No missing dishes, Dish-Skeletons: {dish_rows_count} rows"
                    self.log_test("Dish-Skeletons Writer", True, details, response_time)
                
                # Save results
                self.artifacts["dish_skeletons"] = {
                    "missing_dishes_count": missing_count,
                    "file_found": dish_skeletons_found,
                    "rows_count": dish_rows_count,
                    "properly_filled": missing_count == 0 or (dish_skeletons_found and dish_rows_count > 0)
                }
                
                return True
                
        except Exception as e:
            self.log_test("Dish-Skeletons Writer", False, f"Exception: {str(e)}", 0.0)
        
        return False
    
    async def step_6_excel_invariants(self):
        """STEP 6: Check articles as strings (@) with leading zeros"""
        print("\n🎯 STEP 6: Excel Invariants - Article Format Validation")
        
        if not hasattr(self, 'zip_content'):
            self.log_test("Excel Invariants", False, "No ZIP content available", 0.0)
            return False
        
        try:
            start_time = time.time()
            
            # Check all XLSX files for proper article formatting
            with zipfile.ZipFile(io.BytesIO(self.zip_content), 'r') as zip_file:
                file_names = zip_file.namelist()
                
                format_results = {}
                
                for file_name in file_names:
                    if file_name.endswith('.xlsx'):
                        with zip_file.open(file_name) as xlsx_file:
                            workbook = openpyxl.load_workbook(xlsx_file)
                            sheet = workbook.active
                            
                            # Check article columns (typically columns A and C)
                            article_cells_checked = 0
                            properly_formatted = 0
                            
                            for row in sheet.iter_rows(min_row=2, max_col=5, values_only=False):
                                for col_idx, cell in enumerate(row):
                                    if col_idx in [0, 2] and cell.value:  # Article columns
                                        article_cells_checked += 1
                                        
                                        # Check if cell is formatted as text (@)
                                        cell_format = cell.number_format
                                        is_text_format = '@' in cell_format or cell_format == '@'
                                        
                                        # Check if value has proper 5-digit format
                                        value = str(cell.value)
                                        is_5_digit = len(value) == 5 and value.isdigit()
                                        
                                        if is_text_format and is_5_digit:
                                            properly_formatted += 1
                            
                            format_results[file_name] = {
                                "cells_checked": article_cells_checked,
                                "properly_formatted": properly_formatted,
                                "format_compliance": properly_formatted / max(article_cells_checked, 1)
                            }
                
                response_time = time.time() - start_time
                
                # Calculate overall compliance
                total_checked = sum(r["cells_checked"] for r in format_results.values())
                total_formatted = sum(r["properly_formatted"] for r in format_results.values())
                overall_compliance = total_formatted / max(total_checked, 1)
                
                # Validation: Should have high compliance (>80%)
                invariants_met = overall_compliance >= 0.8
                
                details = f"Compliance: {overall_compliance:.1%} ({total_formatted}/{total_checked} cells)"
                self.log_test("Excel Invariants", invariants_met, details, response_time)
                
                # Save results
                self.artifacts["excel_invariants"] = {
                    "files_analyzed": list(format_results.keys()),
                    "total_cells_checked": total_checked,
                    "properly_formatted": total_formatted,
                    "compliance_rate": overall_compliance,
                    "meets_requirements": invariants_met
                }
                
                return invariants_met
                
        except Exception as e:
            self.log_test("Excel Invariants", False, f"Exception: {str(e)}", 0.0)
        
        return False
    
    async def _analyze_zip_contents(self, zip_content: bytes) -> Dict[str, Any]:
        """Analyze ZIP file contents"""
        try:
            with zipfile.ZipFile(io.BytesIO(zip_content), 'r') as zip_file:
                files = []
                for file_info in zip_file.filelist:
                    files.append({
                        "name": file_info.filename,
                        "size": file_info.file_size,
                        "compressed_size": file_info.compress_size
                    })
                
                return {
                    "files": files,
                    "total_files": len(files)
                }
        except Exception as e:
            return {"error": str(e), "files": []}
    
    async def validate_acceptance_criteria(self):
        """Validate all acceptance criteria"""
        print("\n🎯 FINAL VALIDATION: Acceptance Criteria")
        
        criteria_results = []
        
        # Criteria 1: No GENERATED_* in iiko_TTK.xlsx
        reconcile_result = self.artifacts.get("reconcile_products", {})
        no_generated = reconcile_result.get("no_generated_content", False)
        criteria_results.append(("No GENERATED_* in TTK", no_generated))
        
        # Criteria 2: Dish-Skeletons.xlsx contains row when missing>0
        dish_result = self.artifacts.get("dish_skeletons", {})
        dish_properly_filled = dish_result.get("properly_filled", False)
        criteria_results.append(("Dish-Skeletons filled when missing>0", dish_properly_filled))
        
        # Criteria 3: Excel invariants met
        excel_result = self.artifacts.get("excel_invariants", {})
        excel_invariants_met = excel_result.get("meets_requirements", False)
        criteria_results.append(("Excel invariants (@ format with zeros)", excel_invariants_met))
        
        # Overall success
        all_criteria_met = all(result for _, result in criteria_results)
        
        # Log results
        for criteria, result in criteria_results:
            status = "✅" if result else "❌"
            print(f"{status} {criteria}")
        
        self.log_test("FINAL ACCEPTANCE CRITERIA", all_criteria_met,
                    f"Met: {sum(1 for _, r in criteria_results if r)}/{len(criteria_results)}", 0.0)
        
        return all_criteria_met
    
    async def save_artifacts(self):
        """Save all artifacts to files"""
        try:
            # Save artifacts to JSON files
            with open('/app/artifacts_preflight.json', 'w') as f:
                json.dump(self.artifacts.get("preflight", {}), f, indent=2)
            
            with open('/app/artifacts_reconcile_summary.json', 'w') as f:
                json.dump({
                    "generated_techcard": self.artifacts.get("generated_techcard", {}),
                    "preflight": self.artifacts.get("preflight", {}),
                    "zip_export": self.artifacts.get("zip_export", {}),
                    "reconcile_products": self.artifacts.get("reconcile_products", {}),
                    "dish_skeletons": self.artifacts.get("dish_skeletons", {}),
                    "excel_invariants": self.artifacts.get("excel_invariants", {})
                }, f, indent=2)
            
            print("✅ Artifacts saved to /app/artifacts_*.json")
            
        except Exception as e:
            print(f"⚠️ Failed to save artifacts: {str(e)}")
    
    def print_summary(self):
        """Print comprehensive test summary"""
        print("\n" + "="*80)
        print("🎯 RECONCILE TTK↔SKELETONS FINAL TEST SUMMARY")
        print("="*80)
        
        total_tests = len(self.test_results)
        passed_tests = sum(1 for result in self.test_results if result["success"])
        success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
        
        print(f"📊 OVERALL RESULTS: {passed_tests}/{total_tests} tests passed ({success_rate:.1f}% success rate)")
        print()
        
        # Group results by workflow step
        workflow_steps = {
            "STEP 1: Generate Real TechCard": [],
            "STEP 2: Preflight Check": [],
            "STEP 3: ZIP Export": [],
            "STEP 4: Reconcile Products": [],
            "STEP 5: Dish-Skeletons Writer": [],
            "STEP 6: Excel Invariants": [],
            "FINAL VALIDATION": []
        }
        
        for result in self.test_results:
            test_name = result["test"]
            assigned = False
            
            for step_name in workflow_steps.keys():
                if any(keyword in test_name for keyword in step_name.split()[1:]):
                    workflow_steps[step_name].append(result)
                    assigned = True
                    break
            
            if not assigned:
                workflow_steps["FINAL VALIDATION"].append(result)
        
        # Print results by step
        for step_name, results in workflow_steps.items():
            if results:
                print(f"📋 {step_name}:")
                for result in results:
                    print(f"  {result['status']}: {result['test']} - {result['details']}")
                print()
        
        # Print critical findings
        print("🔍 CRITICAL FINDINGS:")
        
        # Check for GENERATED_* content
        reconcile_result = self.artifacts.get("reconcile_products", {})
        if reconcile_result.get("no_generated_content"):
            print("  ✅ NO GENERATED_* content found in exported XLSX files")
        else:
            generated_content = reconcile_result.get("generated_content_found", [])
            print(f"  ❌ GENERATED_* content found: {generated_content}")
        
        # Check dish skeletons
        dish_result = self.artifacts.get("dish_skeletons", {})
        if dish_result.get("properly_filled"):
            print("  ✅ Dish-Skeletons.xlsx properly filled when missing dishes detected")
        else:
            print("  ❌ Dish-Skeletons.xlsx not properly filled")
        
        # Check Excel formatting
        excel_result = self.artifacts.get("excel_invariants", {})
        compliance = excel_result.get("compliance_rate", 0)
        print(f"  📊 Excel article formatting compliance: {compliance:.1%}")
        
        print("\n" + "="*80)

async def main():
    """Main test execution"""
    print("🚀 Starting RECONCILE TTK↔SKELETONS Final Testing")
    print("="*80)
    
    async with ReconcileTTKSkeletonsTest() as tester:
        # Execute workflow steps
        step1_success = await tester.step_1_generate_real_techcard()
        if not step1_success:
            print("❌ CRITICAL: Failed to generate real tech card. Cannot proceed.")
            return
        
        step2_success = await tester.step_2_run_preflight()
        if not step2_success:
            print("❌ CRITICAL: Preflight check failed. Cannot proceed.")
            return
        
        step3_success = await tester.step_3_zip_export()
        if not step3_success:
            print("❌ CRITICAL: ZIP export failed. Cannot proceed.")
            return
        
        # Continue with analysis steps
        await tester.step_4_reconcile_products()
        await tester.step_5_ensure_dish_skeleton()
        await tester.step_6_excel_invariants()
        
        # Final validation
        await tester.validate_acceptance_criteria()
        
        # Save artifacts and print summary
        await tester.save_artifacts()
        tester.print_summary()

if __name__ == "__main__":
    asyncio.run(main())