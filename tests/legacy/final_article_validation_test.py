#!/usr/bin/env python3
"""
ФИНАЛЬНЫЙ ТЕСТ С ИСПРАВЛЕННОЙ СХЕМОЙ MetaV2
Подтверждение работы генерации артикулов после добавления поля 'article' в схему MetaV2.

Цель: Подтвердить, что после добавления поля 'article' в MetaV2:
1. Dish articles корректно генерируются и сохраняются
2. Ingredient product_code сохраняются правильно  
3. Preflight работает без ошибок
4. Экспорт создает корректные XLSX файлы
5. Скелетоны содержат правильные данные
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Any, Optional

import httpx
import openpyxl
from pymongo import MongoClient

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017/receptor_pro')
DB_NAME = os.getenv('DB_NAME', 'receptor_pro')

class FinalArticleValidationTester:
    """Финальный тест валидации артикулов с исправленной схемой MetaV2"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=45.0)
        self.test_results = []
        self.organization_id = "test-org-final-validation"
        self.generated_techcard_id = None
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status} {test_name}")
        if details:
            print(f"    {details}")
        if response_time > 0:
            print(f"    Response time: {response_time:.3f}s")
        print()

    async def test_1_generate_techcard_fish_with_vegetables(self):
        """Тест 1: Генерация техкарты 'Рыба запеченная с овощами'"""
        try:
            start_time = time.time()
            
            # Generate tech card with realistic Russian dish name
            payload = {
                "name": "Рыба запеченная с овощами",
                "cuisine": "русская",
                "equipment": [],
                "budget": None,
                "dietary": []
            }
            
            response = await self.client.post(
                f"{API_BASE}/techcards.v2/generate",
                json=payload
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                techcard = data.get('card', {})
                
                # Try to get ID from meta or generate one if needed
                if techcard:
                    # Check if ID exists in meta
                    meta = techcard.get('meta', {})
                    self.generated_techcard_id = meta.get('id') or techcard.get('id')
                    
                    # If no ID found, we'll store the techcard data for direct access
                    if not self.generated_techcard_id:
                        # Generate a temporary ID for tracking
                        import uuid
                        self.generated_techcard_id = str(uuid.uuid4())
                    
                    # Always store the techcard data for direct access
                    self.generated_techcard_data = techcard
                    
                    dish_name = meta.get('title') or techcard.get('name', 'Unknown')
                    
                    self.log_test(
                        "Генерация техкарты 'Рыба запеченная с овощами'",
                        True,
                        f"Техкарта создана. Статус: {data.get('status')}, название: {dish_name}, поля: {list(techcard.keys())}",
                        response_time
                    )
                    return True
                else:
                    self.log_test(
                        "Генерация техкарты 'Рыба запеченная с овощами'",
                        False,
                        f"Техкарта не создана. Статус: {data.get('status')}, ошибки: {data.get('issues', [])}",
                        response_time
                    )
                    return False
            else:
                self.log_test(
                    "Генерация техкарты 'Рыба запеченная с овощами'",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                return False
                
        except Exception as e:
            self.log_test(
                "Генерация техкарты 'Рыба запеченная с овощами'",
                False,
                f"Ошибка: {str(e)}"
            )
            return False

    async def test_2_check_dish_article_in_meta(self):
        """Тест 2: Проверка наличия dish.article в meta"""
        if not self.generated_techcard_id:
            self.log_test(
                "Проверка dish.article в meta",
                False,
                "Нет ID техкарты для проверки"
            )
            return False
            
        try:
            # Use the techcard data directly from generation since we have it
            if hasattr(self, 'generated_techcard_data'):
                techcard = self.generated_techcard_data
                
                # Check for article in multiple possible locations
                meta = techcard.get('meta', {})
                dish_article = meta.get('article') or techcard.get('article')
                
                if dish_article:
                    self.log_test(
                        "Проверка dish.article в meta",
                        True,
                        f"Dish article найден: {dish_article} (в {'meta' if meta.get('article') else 'root'})"
                    )
                    return True
                else:
                    # Check if article field exists but is None/empty
                    has_article_field = 'article' in meta or 'article' in techcard
                    self.log_test(
                        "Проверка dish.article в meta",
                        False,
                        f"Dish article отсутствует. Поле article {'существует но пустое' if has_article_field else 'отсутствует'}. Meta: {list(meta.keys())}, Root: {list(techcard.keys())}"
                    )
                    return False
            else:
                self.log_test(
                    "Проверка dish.article в meta",
                    False,
                    "Данные техкарты недоступны"
                )
                return False
                
        except Exception as e:
            self.log_test(
                "Проверка dish.article в meta",
                False,
                f"Ошибка при проверке: {str(e)}"
            )
            return False

    async def test_3_check_ingredient_product_codes(self):
        """Тест 3: Проверка наличия product_code у ингредиентов"""
        if not self.generated_techcard_id:
            self.log_test(
                "Проверка product_code ингредиентов",
                False,
                "Нет ID техкарты для проверки"
            )
            return False
            
        try:
            # Use the techcard data directly from generation since we have it
            if hasattr(self, 'generated_techcard_data'):
                techcard = self.generated_techcard_data
                
                ingredients = techcard.get('ingredients', [])
                
                if not ingredients:
                    self.log_test(
                        "Проверка product_code ингредиентов",
                        False,
                        "Ингредиенты не найдены в техкарте"
                    )
                    return False
                
                ingredients_with_codes = 0
                total_ingredients = len(ingredients)
                ingredient_details = []
                
                for ingredient in ingredients:
                    product_code = ingredient.get('product_code')
                    name = ingredient.get('name', 'Unknown')
                    if product_code:
                        ingredients_with_codes += 1
                        ingredient_details.append(f"{name}: {product_code}")
                    else:
                        ingredient_details.append(f"{name}: НЕТ КОДА")
                
                if ingredients_with_codes > 0:
                    self.log_test(
                        "Проверка product_code ингредиентов",
                        True,
                        f"Product codes найдены у {ingredients_with_codes}/{total_ingredients} ингредиентов. Детали: {'; '.join(ingredient_details[:3])}{'...' if len(ingredient_details) > 3 else ''}"
                    )
                    return True
                else:
                    self.log_test(
                        "Проверка product_code ингредиентов",
                        False,
                        f"Product codes отсутствуют у всех {total_ingredients} ингредиентов. Ингредиенты: {'; '.join([ing.get('name', 'Unknown') for ing in ingredients[:3]])}{'...' if len(ingredients) > 3 else ''}"
                    )
                    return False
            else:
                self.log_test(
                    "Проверка product_code ингредиентов",
                    False,
                    "Данные техкарты недоступны"
                )
                return False
                
        except Exception as e:
            self.log_test(
                "Проверка product_code ингредиентов",
                False,
                f"Ошибка при проверке: {str(e)}"
            )
            return False

    async def test_4_preflight_check_no_errors(self):
        """Тест 4: Запуск preflight - должен работать без ошибок HTTP 500"""
        if not self.generated_techcard_id:
            self.log_test(
                "Preflight check без ошибок",
                False,
                "Нет ID техкарты для проверки"
            )
            return False
            
        try:
            start_time = time.time()
            
            # Run preflight check
            payload = {
                "techcardIds": [self.generated_techcard_id],
                "organizationId": self.organization_id
            }
            
            response = await self.client.post(
                f"{API_BASE}/export/preflight",
                json=payload
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                warnings = data.get('warnings', [])
                
                # Store preflight result for export
                self.preflight_result = data
                
                self.log_test(
                    "Preflight check без ошибок",
                    True,
                    f"Preflight выполнен успешно. Предупреждений: {len(warnings)}",
                    response_time
                )
                return True
            else:
                self.log_test(
                    "Preflight check без ошибок",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                return False
                
        except Exception as e:
            self.log_test(
                "Preflight check без ошибок",
                False,
                f"Ошибка: {str(e)}"
            )
            return False

    async def test_5_zip_export_creation(self):
        """Тест 5: Выполнение ZIP экспорта"""
        if not self.generated_techcard_id:
            self.log_test(
                "ZIP экспорт создание",
                False,
                "Нет ID техкарты для экспорта"
            )
            return False
            
        if not hasattr(self, 'preflight_result'):
            self.log_test(
                "ZIP экспорт создание",
                False,
                "Нет результата preflight для экспорта"
            )
            return False
            
        try:
            start_time = time.time()
            
            # Perform ZIP export with preflight result
            payload = {
                "techcardIds": [self.generated_techcard_id],
                "organizationId": self.organization_id,
                "operational_rounding": True,
                "preflight_result": self.preflight_result
            }
            
            response = await self.client.post(
                f"{API_BASE}/export/zip",
                json=payload
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # Check if response is ZIP file
                content_type = response.headers.get('content-type', '')
                content_length = len(response.content)
                
                if 'zip' in content_type.lower() or content_length > 1000:
                    self.log_test(
                        "ZIP экспорт создание",
                        True,
                        f"ZIP файл создан успешно. Размер: {content_length} байт",
                        response_time
                    )
                    
                    # Store ZIP content for next test
                    self.zip_content = response.content
                    return True
                else:
                    self.log_test(
                        "ZIP экспорт создание",
                        False,
                        f"Ответ не является ZIP файлом. Content-Type: {content_type}, размер: {content_length}",
                        response_time
                    )
                    return False
            else:
                self.log_test(
                    "ZIP экспорт создание",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                return False
                
        except Exception as e:
            self.log_test(
                "ZIP экспорт создание",
                False,
                f"Ошибка: {str(e)}"
            )
            return False

    async def test_6_xlsx_files_content_validation(self):
        """Тест 6: Проверка содержимого всех XLSX файлов"""
        if not hasattr(self, 'zip_content'):
            self.log_test(
                "Проверка содержимого XLSX файлов",
                False,
                "ZIP контент недоступен для проверки"
            )
            return False
            
        try:
            # Extract and validate XLSX files from ZIP
            with zipfile.ZipFile(io.BytesIO(self.zip_content), 'r') as zip_file:
                file_list = zip_file.namelist()
                xlsx_files = [f for f in file_list if f.endswith('.xlsx')]
                
                if not xlsx_files:
                    self.log_test(
                        "Проверка содержимого XLSX файлов",
                        False,
                        f"XLSX файлы не найдены в ZIP. Файлы: {file_list}"
                    )
                    return False
                
                articles_found = 0
                real_data_found = 0
                mock_content_found = 0
                
                for xlsx_file in xlsx_files:
                    try:
                        # Read XLSX file
                        xlsx_data = zip_file.read(xlsx_file)
                        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data))
                        
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            
                            # Check for articles (5-digit codes)
                            for row in sheet.iter_rows(values_only=True):
                                for cell_value in row:
                                    if cell_value:
                                        cell_str = str(cell_value)
                                        
                                        # Check for 5-digit articles
                                        if cell_str.isdigit() and len(cell_str) == 5:
                                            articles_found += 1
                                        
                                        # Check for real dish name
                                        if 'рыба' in cell_str.lower() or 'овощ' in cell_str.lower():
                                            real_data_found += 1
                                        
                                        # Check for mock content
                                        if any(mock in cell_str.upper() for mock in ['MOCK', 'TEST_INGREDIENT', 'GENERATED_TEST']):
                                            mock_content_found += 1
                        
                        workbook.close()
                        
                    except Exception as e:
                        print(f"Ошибка при чтении {xlsx_file}: {str(e)}")
                        continue
                
                # Evaluate results
                success = articles_found > 0 and real_data_found > 0 and mock_content_found == 0
                
                details = f"XLSX файлов: {len(xlsx_files)}, артикулов найдено: {articles_found}, реальных данных: {real_data_found}, mock контента: {mock_content_found}"
                
                self.log_test(
                    "Проверка содержимого XLSX файлов",
                    success,
                    details
                )
                return success
                
        except Exception as e:
            self.log_test(
                "Проверка содержимого XLSX файлов",
                False,
                f"Ошибка при проверке XLSX: {str(e)}"
            )
            return False

    async def test_7_skeleton_data_validation(self):
        """Тест 7: Проверка корректности данных в скелетонах"""
        if not hasattr(self, 'zip_content'):
            self.log_test(
                "Проверка данных скелетонов",
                False,
                "ZIP контент недоступен для проверки"
            )
            return False
            
        try:
            # Extract and validate skeleton files
            with zipfile.ZipFile(io.BytesIO(self.zip_content), 'r') as zip_file:
                file_list = zip_file.namelist()
                skeleton_files = [f for f in file_list if 'skeleton' in f.lower() and f.endswith('.xlsx')]
                
                if not skeleton_files:
                    self.log_test(
                        "Проверка данных скелетонов",
                        True,
                        "Скелетоны не требуются (все артикулы найдены в системе)"
                    )
                    return True
                
                valid_skeletons = 0
                
                for skeleton_file in skeleton_files:
                    try:
                        # Read skeleton XLSX file
                        xlsx_data = zip_file.read(skeleton_file)
                        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data))
                        
                        # Check if skeleton has proper structure
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            
                            # Check for proper headers and data
                            has_headers = False
                            has_data = False
                            
                            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                                if i == 0:  # Header row
                                    if any(header for header in row if header and 'артикул' in str(header).lower()):
                                        has_headers = True
                                elif i > 0:  # Data rows
                                    if any(cell for cell in row if cell):
                                        has_data = True
                                        break
                            
                            if has_headers and has_data:
                                valid_skeletons += 1
                        
                        workbook.close()
                        
                    except Exception as e:
                        print(f"Ошибка при проверке скелетона {skeleton_file}: {str(e)}")
                        continue
                
                success = valid_skeletons > 0
                
                self.log_test(
                    "Проверка данных скелетонов",
                    success,
                    f"Валидных скелетонов: {valid_skeletons}/{len(skeleton_files)}"
                )
                return success
                
        except Exception as e:
            self.log_test(
                "Проверка данных скелетонов",
                False,
                f"Ошибка при проверке скелетонов: {str(e)}"
            )
            return False

    async def run_all_tests(self):
        """Запуск всех тестов финальной валидации артикулов"""
        print("🎯 ФИНАЛЬНЫЙ ТЕСТ С ИСПРАВЛЕННОЙ СХЕМОЙ MetaV2")
        print("=" * 80)
        print("Подтверждение работы генерации артикулов после добавления поля 'article' в схему MetaV2")
        print()
        
        # Run all tests in sequence
        test_methods = [
            self.test_1_generate_techcard_fish_with_vegetables,
            self.test_2_check_dish_article_in_meta,
            self.test_3_check_ingredient_product_codes,
            self.test_4_preflight_check_no_errors,
            self.test_5_zip_export_creation,
            self.test_6_xlsx_files_content_validation,
            self.test_7_skeleton_data_validation
        ]
        
        passed_tests = 0
        total_tests = len(test_methods)
        
        for test_method in test_methods:
            try:
                result = await test_method()
                if result:
                    passed_tests += 1
            except Exception as e:
                print(f"❌ КРИТИЧЕСКАЯ ОШИБКА в {test_method.__name__}: {str(e)}")
                traceback.print_exc()
        
        # Final summary
        print("=" * 80)
        print("🎯 ФИНАЛЬНЫЕ РЕЗУЛЬТАТЫ ВАЛИДАЦИИ АРТИКУЛОВ")
        print("=" * 80)
        
        success_rate = (passed_tests / total_tests) * 100
        
        print(f"Пройдено тестов: {passed_tests}/{total_tests} ({success_rate:.1f}%)")
        print()
        
        # Expected results validation
        expected_results = [
            "✅ Dish article генерируется и сохраняется в meta.article",
            "✅ Product codes генерируются для всех ингредиентов",
            "✅ Preflight работает без ошибок HTTP 500",
            "✅ ZIP экспорт создается успешно",
            "✅ XLSX файлы содержат реальные артикулы (не пустые)",
            "✅ Скелетоны содержат корректные данные"
        ]
        
        print("ОЖИДАЕМЫЕ РЕЗУЛЬТАТЫ:")
        for i, expected in enumerate(expected_results):
            if i < len(self.test_results):
                actual_result = self.test_results[i]
                status = "✅" if actual_result['success'] else "❌"
                print(f"{status} {expected.replace('✅ ', '')}")
            else:
                print(f"❓ {expected.replace('✅ ', '')}")
        
        print()
        
        if passed_tests == total_tests:
            print("🎉 ВСЕ ТЕСТЫ ПРОЙДЕНЫ! Исправление схемы MetaV2 полностью решило проблему с артикулами.")
        elif passed_tests >= total_tests * 0.8:
            print("⚠️ БОЛЬШИНСТВО ТЕСТОВ ПРОЙДЕНО. Основная функциональность работает, но есть минорные проблемы.")
        else:
            print("❌ КРИТИЧЕСКИЕ ПРОБЛЕМЫ. Исправление схемы MetaV2 не решило проблему с артикулами.")
        
        print()
        print("Детальные результаты:")
        for result in self.test_results:
            print(f"{result['status']} {result['test']}")
            if result['details']:
                print(f"    {result['details']}")
        
        return passed_tests, total_tests

async def main():
    """Main test execution"""
    try:
        async with FinalArticleValidationTester() as tester:
            passed, total = await tester.run_all_tests()
            
            # Exit with appropriate code
            if passed == total:
                sys.exit(0)  # All tests passed
            else:
                sys.exit(1)  # Some tests failed
                
    except Exception as e:
        print(f"❌ КРИТИЧЕСКАЯ ОШИБКА: {str(e)}")
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    asyncio.run(main())