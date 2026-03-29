#!/usr/bin/env python3
"""
ФИНАЛЬНЫЙ ИНТЕГРАЦИОННЫЙ ТЕСТ СИСТЕМЫ
Проверка полного workflow: генерация → артикулы → экспорт с реальными данными
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import io
import re
from datetime import datetime
from typing import Dict, List, Any, Optional

import httpx
import openpyxl
from pymongo import MongoClient

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017/receptor_pro')
DB_NAME = os.getenv('DB_NAME', 'receptor_pro')

class FinalIntegrationTester:
    """Финальный интеграционный тест всей системы"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=60.0)
        self.test_results = []
        self.organization_id = "test-org-final"
        self.generated_techcard_id = None
        self.generated_techcard_data = None
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status} {test_name}: {details}")
        return success

    async def test_1_generate_tom_yam_techcard(self) -> bool:
        """1. Генерация техкарты 'Суп Том Ям' с артикулами"""
        start_time = time.time()
        
        try:
            # Generate tech card for "Суп Том Ям"
            generate_url = f"{API_BASE}/techcards.v2/generate"
            
            payload = {
                "name": "Суп Том Ям",
                "description": "Острый тайский суп с креветками, грибами и лемонграссом"
            }
            
            response = await self.client.post(generate_url, json=payload)
            response_time = time.time() - start_time
            
            if response.status_code != 200:
                return self.log_test(
                    "Generate Tom Yam TechCard", 
                    False, 
                    f"HTTP {response.status_code}: {response.text[:200]}", 
                    response_time
                )
            
            data = response.json()
            
            # Extract tech card ID and data from the correct response structure
            if 'card' in data:
                self.generated_techcard_data = data['card']
                if 'meta' in self.generated_techcard_data:
                    self.generated_techcard_id = self.generated_techcard_data['meta'].get('id')
            else:
                return self.log_test(
                    "Generate Tom Yam TechCard", 
                    False, 
                    "No card found in response", 
                    response_time
                )
            
            if not self.generated_techcard_id:
                return self.log_test(
                    "Generate Tom Yam TechCard", 
                    False, 
                    "No techcard ID found in response", 
                    response_time
                )
            
            # Check if dish has article
            dish_article = None
            if isinstance(self.generated_techcard_data, dict):
                dish_article = (
                    self.generated_techcard_data.get('article') or 
                    self.generated_techcard_data.get('meta', {}).get('article')
                )
            
            # Check ingredients for product codes
            ingredients_with_codes = 0
            total_ingredients = 0
            
            if 'ingredients' in self.generated_techcard_data:
                ingredients = self.generated_techcard_data['ingredients']
                if isinstance(ingredients, list):
                    total_ingredients = len(ingredients)
                    for ingredient in ingredients:
                        if isinstance(ingredient, dict) and ingredient.get('product_code'):
                            ingredients_with_codes += 1
            
            success_details = f"Generated TechCard ID: {self.generated_techcard_id}, "
            success_details += f"Dish article: {dish_article}, "
            success_details += f"Ingredients with codes: {ingredients_with_codes}/{total_ingredients}"
            
            return self.log_test(
                "Generate Tom Yam TechCard", 
                True, 
                success_details, 
                response_time
            )
            
        except Exception as e:
            return self.log_test(
                "Generate Tom Yam TechCard", 
                False, 
                f"Exception: {str(e)}", 
                time.time() - start_time
            )

    async def test_2_verify_articles_in_techcard(self) -> bool:
        """2. Проверка артикулов в сгенерированной техкарте"""
        
        if not self.generated_techcard_data:
            return self.log_test(
                "Verify Articles in TechCard", 
                False, 
                "No generated techcard data available"
            )
        
        try:
            # Check dish article
            dish_article = None
            if isinstance(self.generated_techcard_data, dict):
                dish_article = (
                    self.generated_techcard_data.get('article') or 
                    self.generated_techcard_data.get('dish', {}).get('article') or
                    self.generated_techcard_data.get('meta', {}).get('article')
                )
            
            # Check ingredient product codes
            ingredients_analysis = {
                'total': 0,
                'with_codes': 0,
                'codes_found': []
            }
            
            if 'ingredients' in self.generated_techcard_data:
                ingredients = self.generated_techcard_data['ingredients']
                if isinstance(ingredients, list):
                    ingredients_analysis['total'] = len(ingredients)
                    
                    for ingredient in ingredients:
                        if isinstance(ingredient, dict):
                            product_code = ingredient.get('product_code') or ingredient.get('skuId')
                            if product_code:
                                ingredients_analysis['with_codes'] += 1
                                ingredients_analysis['codes_found'].append(product_code)
            
            # Validate 5-digit format for codes
            valid_5digit_codes = 0
            for code in ingredients_analysis['codes_found']:
                if isinstance(code, str) and len(code) == 5 and code.isdigit():
                    valid_5digit_codes += 1
            
            success = (
                dish_article is not None and 
                ingredients_analysis['with_codes'] > 0 and
                valid_5digit_codes > 0
            )
            
            details = f"Dish article: {dish_article}, "
            details += f"Ingredients with codes: {ingredients_analysis['with_codes']}/{ingredients_analysis['total']}, "
            details += f"Valid 5-digit codes: {valid_5digit_codes}, "
            details += f"Sample codes: {ingredients_analysis['codes_found'][:3]}"
            
            return self.log_test(
                "Verify Articles in TechCard", 
                success, 
                details
            )
            
        except Exception as e:
            return self.log_test(
                "Verify Articles in TechCard", 
                False, 
                f"Exception: {str(e)}"
            )

    async def test_3_verify_db_persistence(self) -> bool:
        """3. Проверка сохранения в БД с артикулами"""
        
        if not self.generated_techcard_id:
            return self.log_test(
                "Verify DB Persistence", 
                False, 
                "No generated techcard ID available"
            )
        
        try:
            # Connect to MongoDB
            client = MongoClient(MONGO_URL)
            db = client[DB_NAME]
            
            # Find the techcard in database
            techcard_doc = db.techcards_v2.find_one({"id": self.generated_techcard_id})
            
            if not techcard_doc:
                client.close()
                return self.log_test(
                    "Verify DB Persistence", 
                    False, 
                    f"TechCard {self.generated_techcard_id} not found in database"
                )
            
            # Check articles in database document
            db_dish_article = None
            if 'article' in techcard_doc:
                db_dish_article = techcard_doc['article']
            elif 'dish' in techcard_doc and isinstance(techcard_doc['dish'], dict):
                db_dish_article = techcard_doc['dish'].get('article')
            elif 'meta' in techcard_doc and isinstance(techcard_doc['meta'], dict):
                db_dish_article = techcard_doc['meta'].get('article')
            
            # Check ingredient codes in database
            db_ingredients_with_codes = 0
            db_total_ingredients = 0
            db_codes_found = []
            
            if 'ingredients' in techcard_doc and isinstance(techcard_doc['ingredients'], list):
                db_total_ingredients = len(techcard_doc['ingredients'])
                
                for ingredient in techcard_doc['ingredients']:
                    if isinstance(ingredient, dict):
                        product_code = ingredient.get('product_code') or ingredient.get('skuId')
                        if product_code:
                            db_ingredients_with_codes += 1
                            db_codes_found.append(product_code)
            
            client.close()
            
            success = (
                db_dish_article is not None and 
                db_ingredients_with_codes > 0
            )
            
            details = f"DB Dish article: {db_dish_article}, "
            details += f"DB Ingredients with codes: {db_ingredients_with_codes}/{db_total_ingredients}, "
            details += f"DB Sample codes: {db_codes_found[:3]}"
            
            return self.log_test(
                "Verify DB Persistence", 
                success, 
                details
            )
            
        except Exception as e:
            return self.log_test(
                "Verify DB Persistence", 
                False, 
                f"Exception: {str(e)}"
            )

    async def test_4_run_zip_export_with_real_id(self) -> bool:
        """4. Запуск ZIP export с реальным ID техкарты"""
        
        if not self.generated_techcard_id:
            return self.log_test(
                "Run ZIP Export with Real ID", 
                False, 
                "No generated techcard ID available"
            )
        
        start_time = time.time()
        
        try:
            # First run preflight check
            preflight_url = f"{API_BASE}/export/preflight"
            preflight_payload = {
                "techcardIds": [self.generated_techcard_id]
            }
            
            preflight_response = await self.client.post(preflight_url, json=preflight_payload)
            
            if preflight_response.status_code != 200:
                return self.log_test(
                    "Run ZIP Export with Real ID", 
                    False, 
                    f"Preflight failed: HTTP {preflight_response.status_code}: {preflight_response.text[:200]}", 
                    time.time() - start_time
                )
            
            preflight_data = preflight_response.json()
            
            # Now run ZIP export
            export_url = f"{API_BASE}/export/zip"
            export_payload = {
                "techcardIds": [self.generated_techcard_id]
            }
            
            export_response = await self.client.post(export_url, json=export_payload)
            response_time = time.time() - start_time
            
            if export_response.status_code != 200:
                return self.log_test(
                    "Run ZIP Export with Real ID", 
                    False, 
                    f"Export failed: HTTP {export_response.status_code}: {export_response.text[:200]}", 
                    response_time
                )
            
            # Check if response is ZIP file
            content_type = export_response.headers.get('content-type', '')
            zip_size = len(export_response.content)
            
            if zip_size < 100:
                return self.log_test(
                    "Run ZIP Export with Real ID", 
                    False, 
                    f"Invalid response: content-type={content_type}, size={zip_size}", 
                    response_time
                )
            
            # Store ZIP content for next test
            self.exported_zip_content = export_response.content
            self.preflight_data = preflight_data
            
            return self.log_test(
                "Run ZIP Export with Real ID", 
                True, 
                f"ZIP export successful: {zip_size} bytes, content-type: {content_type}", 
                response_time
            )
            
        except Exception as e:
            return self.log_test(
                "Run ZIP Export with Real ID", 
                False, 
                f"Exception: {str(e)}", 
                time.time() - start_time
            )

    async def test_5_analyze_xlsx_content(self) -> bool:
        """5. Анализ содержимого XLSX файлов"""
        
        if not hasattr(self, 'exported_zip_content'):
            return self.log_test(
                "Analyze XLSX Content", 
                False, 
                "No exported ZIP content available"
            )
        
        try:
            # Extract ZIP file
            zip_buffer = io.BytesIO(self.exported_zip_content)
            
            with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
                file_list = zip_file.namelist()
                xlsx_files = [f for f in file_list if f.endswith('.xlsx')]
                
                if not xlsx_files:
                    return self.log_test(
                        "Analyze XLSX Content", 
                        False, 
                        f"No XLSX files found in ZIP. Files: {file_list}"
                    )
                
                analysis_results = {
                    'files_analyzed': len(xlsx_files),
                    'real_articles_found': 0,
                    'mock_content_found': 0,
                    'total_articles': 0,
                    'files_details': {}
                }
                
                # Analyze each XLSX file
                for xlsx_file in xlsx_files:
                    file_analysis = await self._analyze_xlsx_file(zip_file, xlsx_file)
                    analysis_results['files_details'][xlsx_file] = file_analysis
                    
                    analysis_results['real_articles_found'] += file_analysis.get('real_articles', 0)
                    analysis_results['mock_content_found'] += file_analysis.get('mock_signatures', 0)
                    analysis_results['total_articles'] += file_analysis.get('total_articles', 0)
                
                # Check success criteria
                success = (
                    analysis_results['real_articles_found'] > 0 and
                    analysis_results['mock_content_found'] == 0 and
                    analysis_results['total_articles'] > 0
                )
                
                details = f"Files: {analysis_results['files_analyzed']}, "
                details += f"Real articles: {analysis_results['real_articles_found']}, "
                details += f"Mock content: {analysis_results['mock_content_found']}, "
                details += f"Total articles: {analysis_results['total_articles']}"
                
                return self.log_test(
                    "Analyze XLSX Content", 
                    success, 
                    details
                )
                
        except Exception as e:
            return self.log_test(
                "Analyze XLSX Content", 
                False, 
                f"Exception: {str(e)}"
            )

    async def _analyze_xlsx_file(self, zip_file, xlsx_filename: str) -> Dict[str, Any]:
        """Analyze individual XLSX file for articles and mock content"""
        
        try:
            # Extract XLSX file
            xlsx_data = zip_file.read(xlsx_filename)
            xlsx_buffer = io.BytesIO(xlsx_data)
            
            # Load workbook
            workbook = openpyxl.load_workbook(xlsx_buffer)
            
            analysis = {
                'filename': xlsx_filename,
                'sheets': len(workbook.sheetnames),
                'real_articles': 0,
                'mock_signatures': 0,
                'total_articles': 0,
                'sample_articles': [],
                'mock_content_found': []
            }
            
            # Mock content patterns to detect
            mock_patterns = [
                'DISH_MOCK_TECH_CARD',
                'GENERATED_TEST_INGREDIENT',
                'TEST_INGREDIENT',
                'MOCK_',
                'TEST_',
                'DEMO_'
            ]
            
            # Article patterns (5-digit codes)
            article_pattern = re.compile(r'\b\d{5}\b')
            
            # Analyze each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                for row in sheet.iter_rows(values_only=True):
                    for cell_value in row:
                        if cell_value is None:
                            continue
                            
                        cell_str = str(cell_value)
                        
                        # Check for mock content
                        for mock_pattern in mock_patterns:
                            if mock_pattern in cell_str:
                                analysis['mock_signatures'] += 1
                                analysis['mock_content_found'].append(f"{sheet_name}: {cell_str[:50]}")
                        
                        # Check for real 5-digit articles
                        articles_in_cell = article_pattern.findall(cell_str)
                        for article in articles_in_cell:
                            analysis['total_articles'] += 1
                            
                            # Consider it "real" if it's not in a mock context
                            is_real_article = True
                            for mock_pattern in mock_patterns:
                                if mock_pattern in cell_str:
                                    is_real_article = False
                                    break
                            
                            if is_real_article:
                                analysis['real_articles'] += 1
                                if len(analysis['sample_articles']) < 5:
                                    analysis['sample_articles'].append(article)
            
            workbook.close()
            return analysis
            
        except Exception as e:
            return {
                'filename': xlsx_filename,
                'error': str(e),
                'real_articles': 0,
                'mock_signatures': 0,
                'total_articles': 0
            }

    async def test_6_verify_5digit_articles(self) -> bool:
        """6. Убедиться в наличии реальных 5-digit артикулов"""
        
        if not hasattr(self, 'exported_zip_content'):
            return self.log_test(
                "Verify 5-Digit Articles", 
                False, 
                "No exported ZIP content available"
            )
        
        try:
            # Extract and analyze ZIP for 5-digit articles
            zip_buffer = io.BytesIO(self.exported_zip_content)
            
            five_digit_articles = set()
            valid_format_count = 0
            
            with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
                xlsx_files = [f for f in zip_file.namelist() if f.endswith('.xlsx')]
                
                for xlsx_file in xlsx_files:
                    xlsx_data = zip_file.read(xlsx_file)
                    xlsx_buffer = io.BytesIO(xlsx_data)
                    
                    workbook = openpyxl.load_workbook(xlsx_buffer)
                    
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        
                        for row in sheet.iter_rows(values_only=True):
                            for cell_value in row:
                                if cell_value is None:
                                    continue
                                
                                cell_str = str(cell_value).strip()
                                
                                # Check for 5-digit pattern
                                if len(cell_str) == 5 and cell_str.isdigit():
                                    five_digit_articles.add(cell_str)
                                    valid_format_count += 1
                    
                    workbook.close()
            
            # Success criteria: at least 3 unique 5-digit articles
            success = len(five_digit_articles) >= 3 and valid_format_count >= 5
            
            details = f"Unique 5-digit articles: {len(five_digit_articles)}, "
            details += f"Total valid format occurrences: {valid_format_count}, "
            details += f"Sample articles: {list(five_digit_articles)[:5]}"
            
            return self.log_test(
                "Verify 5-Digit Articles", 
                success, 
                details
            )
            
        except Exception as e:
            return self.log_test(
                "Verify 5-Digit Articles", 
                False, 
                f"Exception: {str(e)}"
            )

    async def run_all_tests(self):
        """Run all integration tests"""
        print("🎯 ФИНАЛЬНЫЙ ИНТЕГРАЦИОННЫЙ ТЕСТ СИСТЕМЫ")
        print("=" * 60)
        
        # Run tests in sequence
        test_methods = [
            self.test_1_generate_tom_yam_techcard,
            self.test_2_verify_articles_in_techcard,
            self.test_3_verify_db_persistence,
            self.test_4_run_zip_export_with_real_id,
            self.test_5_analyze_xlsx_content,
            self.test_6_verify_5digit_articles
        ]
        
        passed_tests = 0
        total_tests = len(test_methods)
        
        for test_method in test_methods:
            try:
                success = await test_method()
                if success:
                    passed_tests += 1
            except Exception as e:
                print(f"❌ CRITICAL ERROR in {test_method.__name__}: {str(e)}")
                traceback.print_exc()
        
        # Final summary
        print("\n" + "=" * 60)
        print("🎯 ФИНАЛЬНЫЙ РЕЗУЛЬТАТ ИНТЕГРАЦИОННОГО ТЕСТА")
        print("=" * 60)
        
        success_rate = (passed_tests / total_tests) * 100
        
        print(f"Пройдено тестов: {passed_tests}/{total_tests} ({success_rate:.1f}%)")
        
        if success_rate >= 80:
            print("🎉 СИСТЕМА ГОТОВА К ПРОДАКШЕНУ!")
        elif success_rate >= 60:
            print("⚠️ СИСТЕМА ЧАСТИЧНО РАБОТАЕТ - ТРЕБУЮТСЯ ДОРАБОТКИ")
        else:
            print("🚨 КРИТИЧЕСКИЕ ПРОБЛЕМЫ - СИСТЕМА НЕ ГОТОВА")
        
        # Detailed results
        print("\nДетальные результаты:")
        for result in self.test_results:
            print(f"{result['status']} {result['test']}: {result['details']}")
        
        return success_rate >= 80

async def main():
    """Main test execution"""
    async with FinalIntegrationTester() as tester:
        success = await tester.run_all_tests()
        return success

if __name__ == "__main__":
    try:
        result = asyncio.run(main())
        sys.exit(0 if result else 1)
    except KeyboardInterrupt:
        print("\n🛑 Test interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"🚨 CRITICAL ERROR: {str(e)}")
        traceback.print_exc()
        sys.exit(1)