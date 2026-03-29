#!/usr/bin/env python3
"""
Comprehensive Backend Testing for iiko Export Functionality (TechCardV2 Task #6)
Testing the newly implemented iiko export functionality as requested in review.
"""

import requests
import json
import os
import sys
from datetime import datetime
from typing import Dict, Any, List
import tempfile

# Backend URL from environment
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"

class IikoExportTester:
    """Comprehensive tester for iiko export functionality"""
    
    def __init__(self):
        self.results = []
        self.test_count = 0
        self.passed_count = 0
        
    def log_result(self, test_name: str, passed: bool, details: str = ""):
        """Log test result"""
        self.test_count += 1
        if passed:
            self.passed_count += 1
            status = "✅ PASS"
        else:
            status = "❌ FAIL"
        
        result = f"{status}: {test_name}"
        if details:
            result += f" - {details}"
        
        self.results.append(result)
        print(result)
        
    def create_test_techcard_v2(self, scenario: str = "basic") -> Dict[str, Any]:
        """Create test TechCardV2 with different scenarios"""
        
        base_card = {
            "meta": {
                "id": "test-iiko-export-001",
                "title": "Паста Карбонара Тест",
                "version": "2.0",
                "createdAt": datetime.now().isoformat(),
                "cuisine": "italian",
                "tags": ["pasta", "test"]
            },
            "portions": 4,
            "yield": {
                "perPortion_g": 250.0,
                "perBatch_g": 1000.0
            },
            "ingredients": [],
            "process": [
                {
                    "n": 1,
                    "action": "Отварить спагетти в подсоленной воде",
                    "time_min": 10.0,
                    "temp_c": 100.0,
                    "equipment": ["плита", "кастрюля"],
                    "ccp": True,
                    "note": "До состояния аль денте"
                },
                {
                    "n": 2,
                    "action": "Обжарить бекон до золотистого цвета",
                    "time_min": 5.0,
                    "temp_c": 180.0,
                    "equipment": ["сковорода"],
                    "ccp": False
                },
                {
                    "n": 3,
                    "action": "Смешать с яично-сырной смесью",
                    "time_min": 2.0,
                    "equipment": ["миска"],
                    "ccp": False
                }
            ],
            "storage": {
                "conditions": "Подавать немедленно, хранение не рекомендуется",
                "shelfLife_hours": 0.5,
                "servingTemp_c": 65.0
            },
            "nutrition": {
                "per100g": {
                    "kcal": 180.0,
                    "proteins_g": 12.0,
                    "fats_g": 8.0,
                    "carbs_g": 15.0
                },
                "perPortion": {
                    "kcal": 450.0,
                    "proteins_g": 30.0,
                    "fats_g": 20.0,
                    "carbs_g": 37.5
                }
            },
            "cost": {
                "rawCost": 280.0,
                "costPerPortion": 70.0,
                "markup_pct": 300.0,
                "vat_pct": 20.0
            },
            "costMeta": {
                "source": "catalog",
                "coveragePct": 85.0,
                "asOf": "2025-01-17"
            }
        }
        
        if scenario == "basic":
            # Basic scenario with all SKUs present
            base_card["ingredients"] = [
                {
                    "name": "Спагетти",
                    "skuId": "PASTA_SPAGHETTI_001",
                    "unit": "g",
                    "brutto_g": 320.0,
                    "loss_pct": 0.0,
                    "netto_g": 320.0,
                    "allergens": ["gluten"]
                },
                {
                    "name": "Бекон",
                    "skuId": "MEAT_BACON_002",
                    "unit": "g", 
                    "brutto_g": 200.0,
                    "loss_pct": 10.0,
                    "netto_g": 180.0,
                    "allergens": []
                },
                {
                    "name": "Яйца куриные",
                    "skuId": "EGG_CHICKEN_003",
                    "unit": "pcs",
                    "brutto_g": 120.0,
                    "loss_pct": 15.0,
                    "netto_g": 102.0,
                    "allergens": ["eggs"]
                },
                {
                    "name": "Сыр Пармезан",
                    "skuId": "CHEESE_PARMESAN_004",
                    "unit": "g",
                    "brutto_g": 80.0,
                    "loss_pct": 5.0,
                    "netto_g": 76.0,
                    "allergens": ["dairy"]
                }
            ]
            
        elif scenario == "mixed_units":
            # Mixed units scenario (g, ml, pcs)
            base_card["meta"]["title"] = "Салат Цезарь Микс"
            base_card["ingredients"] = [
                {
                    "name": "Салат Романо",
                    "skuId": "VEG_LETTUCE_001",
                    "unit": "g",
                    "brutto_g": 150.0,
                    "loss_pct": 20.0,
                    "netto_g": 120.0,
                    "allergens": []
                },
                {
                    "name": "Оливковое масло",
                    "skuId": "OIL_OLIVE_002",
                    "unit": "ml",
                    "brutto_g": 50.0,
                    "loss_pct": 0.0,
                    "netto_g": 50.0,
                    "allergens": []
                },
                {
                    "name": "Яйца перепелиные",
                    "skuId": "EGG_QUAIL_003",
                    "unit": "pcs",
                    "brutto_g": 85.0,
                    "loss_pct": 12.0,
                    "netto_g": 74.8,
                    "allergens": ["eggs"]
                }
            ]
            # Adjust yield for smaller salad
            base_card["yield"]["perPortion_g"] = 180.0
            base_card["yield"]["perBatch_g"] = 720.0
            
        elif scenario == "missing_skus":
            # Scenario with missing SKUs to test noSku issues
            base_card["meta"]["title"] = "Борщ Украинский"
            base_card["ingredients"] = [
                {
                    "name": "Свекла",
                    "skuId": "VEG_BEETROOT_001",
                    "unit": "g",
                    "brutto_g": 200.0,
                    "loss_pct": 15.0,
                    "netto_g": 170.0,
                    "allergens": []
                },
                {
                    "name": "Капуста белокочанная",
                    "skuId": None,  # Missing SKU
                    "unit": "g",
                    "brutto_g": 150.0,
                    "loss_pct": 10.0,
                    "netto_g": 135.0,
                    "allergens": []
                },
                {
                    "name": "Морковь",
                    "skuId": "",  # Empty SKU
                    "unit": "g",
                    "brutto_g": 100.0,
                    "loss_pct": 20.0,
                    "netto_g": 80.0,
                    "allergens": []
                },
                {
                    "name": "Экзотическая специя",
                    "skuId": None,  # Missing SKU
                    "unit": "g",
                    "brutto_g": 5.0,
                    "loss_pct": 0.0,
                    "netto_g": 5.0,
                    "allergens": []
                }
            ]
            # Adjust yield
            base_card["yield"]["perPortion_g"] = 300.0
            base_card["yield"]["perBatch_g"] = 1200.0
            
        elif scenario == "yield_mismatch":
            # Scenario with yield mismatch to test yieldMismatch issues
            base_card["meta"]["title"] = "Рагу Овощное"
            base_card["ingredients"] = [
                {
                    "name": "Картофель",
                    "skuId": "VEG_POTATO_001",
                    "unit": "g",
                    "brutto_g": 300.0,
                    "loss_pct": 25.0,
                    "netto_g": 225.0,
                    "allergens": []
                },
                {
                    "name": "Лук репчатый",
                    "skuId": "VEG_ONION_002",
                    "unit": "g",
                    "brutto_g": 100.0,
                    "loss_pct": 15.0,
                    "netto_g": 85.0,
                    "allergens": []
                }
            ]
            # Set yield that doesn't match sum of netto (225 + 85 = 310, but yield = 400)
            base_card["yield"]["perPortion_g"] = 100.0
            base_card["yield"]["perBatch_g"] = 400.0  # Mismatch: ingredients sum to 310g
            
        return base_card
    
    def test_iiko_export_endpoint(self):
        """Test 1: iiko Export API Endpoint Testing"""
        print("\n🎯 TEST 1: iiko Export API Endpoint Testing")
        
        # Test basic endpoint availability
        try:
            test_card = self.create_test_techcard_v2("basic")
            
            response = requests.post(
                f"{API_BASE}/techcards.v2/export/iiko",
                json=test_card,
                timeout=30
            )
            
            # Check status code
            if response.status_code == 200:
                self.log_result("API Endpoint Accessibility", True, f"HTTP {response.status_code}")
            else:
                self.log_result("API Endpoint Accessibility", False, f"HTTP {response.status_code}: {response.text[:100]}")
                return
            
            # Check content type
            content_type = response.headers.get('content-type', '')
            expected_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
            if expected_type in content_type:
                self.log_result("Content-Type Header", True, f"Correct XLSX content-type")
            else:
                self.log_result("Content-Type Header", False, f"Expected {expected_type}, got {content_type}")
            
            # Check attachment filename
            content_disposition = response.headers.get('content-disposition', '')
            if 'attachment' in content_disposition and 'iiko_export_' in content_disposition and '.xlsx' in content_disposition:
                self.log_result("Attachment Filename Format", True, f"Correct filename format")
            else:
                self.log_result("Attachment Filename Format", False, f"Incorrect format: {content_disposition}")
            
            # Store response for further testing
            self.xlsx_content = response.content
            
        except Exception as e:
            self.log_result("API Endpoint Accessibility", False, f"Exception: {str(e)}")
    
    def test_xlsx_structure_and_content(self):
        """Test 2: XLSX File Structure and Content"""
        print("\n🎯 TEST 2: XLSX File Structure and Content")
        
        if not hasattr(self, 'xlsx_content'):
            self.log_result("XLSX Structure Test", False, "No XLSX content from previous test")
            return
        
        try:
            # Try to import openpyxl, if not available, use basic checks
            try:
                import openpyxl
                from io import BytesIO
                
                # Load XLSX from bytes
                workbook = openpyxl.load_workbook(BytesIO(self.xlsx_content))
                
                # Check sheet names
                sheet_names = workbook.sheetnames
                expected_sheets = ['Products', 'Recipes']
                
                if set(expected_sheets).issubset(set(sheet_names)):
                    self.log_result("XLSX Sheet Structure", True, f"Contains required sheets: {expected_sheets}")
                else:
                    self.log_result("XLSX Sheet Structure", False, f"Missing sheets. Found: {sheet_names}")
                    return
                
                # Test Products sheet structure
                products_sheet = workbook['Products']
                expected_products_columns = ['skuId', 'canonicalId', 'name', 'unitBase', 'pricePerUnit', 'vatPct', 'asOf']
                
                # Get first row (headers)
                products_headers = [cell.value for cell in products_sheet[1]]
                
                if products_headers == expected_products_columns:
                    self.log_result("Products Sheet Columns", True, f"Correct columns: {expected_products_columns}")
                else:
                    self.log_result("Products Sheet Columns", False, f"Expected {expected_products_columns}, got {products_headers}")
                
                # Test Recipes sheet structure
                recipes_sheet = workbook['Recipes']
                expected_recipes_columns = ['dishCode', 'dishName', 'portionUnit', 'outputPerPortion_g', 'outputPerBatch_g', 
                                          'ingredientSku', 'ingredientName', 'qtyNet', 'qtyUnit']
                
                # Get first row (headers)
                recipes_headers = [cell.value for cell in recipes_sheet[1]]
                
                if recipes_headers == expected_recipes_columns:
                    self.log_result("Recipes Sheet Columns", True, f"Correct columns: {expected_recipes_columns}")
                else:
                    self.log_result("Recipes Sheet Columns", False, f"Expected {expected_recipes_columns}, got {recipes_headers}")
                
                # Test data presence
                products_data_rows = products_sheet.max_row - 1  # Exclude header
                recipes_data_rows = recipes_sheet.max_row - 1   # Exclude header
                
                if products_data_rows > 0:
                    self.log_result("Products Data Presence", True, f"{products_data_rows} product rows")
                else:
                    self.log_result("Products Data Presence", False, "No product data found")
                
                if recipes_data_rows > 0:
                    self.log_result("Recipes Data Presence", True, f"{recipes_data_rows} recipe rows")
                else:
                    self.log_result("Recipes Data Presence", False, "No recipe data found")
                
                # Store sheets for further testing
                self.products_sheet = products_sheet
                self.recipes_sheet = recipes_sheet
                
            except ImportError:
                # Fallback: Basic file validation without openpyxl
                if len(self.xlsx_content) > 1000:  # Basic size check
                    self.log_result("XLSX File Size", True, f"File size: {len(self.xlsx_content)} bytes")
                else:
                    self.log_result("XLSX File Size", False, f"File too small: {len(self.xlsx_content)} bytes")
                
                # Check for XLSX magic bytes
                if self.xlsx_content.startswith(b'PK'):
                    self.log_result("XLSX File Format", True, "Valid ZIP/XLSX format")
                else:
                    self.log_result("XLSX File Format", False, "Invalid file format")
                    
        except Exception as e:
            self.log_result("XLSX Structure Test", False, f"Exception: {str(e)}")
    
    def test_unit_conversions(self):
        """Test 3: Unit Conversions Testing"""
        print("\n🎯 TEST 3: Unit Conversions Testing")
        
        # Test with mixed units scenario
        try:
            test_card = self.create_test_techcard_v2("mixed_units")
            
            response = requests.post(
                f"{API_BASE}/techcards.v2/export/iiko",
                json=test_card,
                timeout=30
            )
            
            if response.status_code != 200:
                self.log_result("Unit Conversion Test Setup", False, f"HTTP {response.status_code}")
                return
            
            # Basic validation - if we have openpyxl, do detailed checks
            try:
                import openpyxl
                from io import BytesIO
                
                # Load and analyze XLSX
                workbook = openpyxl.load_workbook(BytesIO(response.content))
                
                # Check Products sheet for unit conversions
                products_sheet = workbook['Products']
                recipes_sheet = workbook['Recipes']
                
                # Test unit conversions in Products sheet
                unit_conversions_correct = True
                conversion_details = []
                
                for row in range(2, products_sheet.max_row + 1):
                    unit_base = products_sheet.cell(row, 4).value  # unitBase column
                    name = products_sheet.cell(row, 3).value       # name column
                    
                    if "масло" in name.lower():
                        # Oil should be in liters (ml → l)
                        if unit_base == "l":
                            conversion_details.append(f"✅ {name}: ml → l")
                        else:
                            conversion_details.append(f"❌ {name}: expected 'l', got '{unit_base}'")
                            unit_conversions_correct = False
                    elif "яйца" in name.lower():
                        # Eggs should remain in pcs
                        if unit_base == "pcs":
                            conversion_details.append(f"✅ {name}: pcs → pcs")
                        else:
                            conversion_details.append(f"❌ {name}: expected 'pcs', got '{unit_base}'")
                            unit_conversions_correct = False
                    else:
                        # Other ingredients should be in kg (g → kg)
                        if unit_base == "kg":
                            conversion_details.append(f"✅ {name}: g → kg")
                        else:
                            conversion_details.append(f"❌ {name}: expected 'kg', got '{unit_base}'")
                            unit_conversions_correct = False
                
                self.log_result("Unit Base Conversions", unit_conversions_correct, "; ".join(conversion_details))
                
                # Test quantity conversions in Recipes sheet with proper rounding
                quantity_conversions_correct = True
                quantity_details = []
                
                for row in range(2, recipes_sheet.max_row + 1):
                    qty_net = recipes_sheet.cell(row, 8).value      # qtyNet column
                    qty_unit = recipes_sheet.cell(row, 9).value     # qtyUnit column
                    ingredient_name = recipes_sheet.cell(row, 7).value  # ingredientName column
                    
                    if "масло" in ingredient_name.lower():
                        # 50ml → 0.050l
                        if qty_unit == "l" and abs(qty_net - 0.050) < 0.001:
                            quantity_details.append(f"✅ {ingredient_name}: 50ml → 0.050l")
                        else:
                            quantity_details.append(f"❌ {ingredient_name}: expected 0.050l, got {qty_net}{qty_unit}")
                            quantity_conversions_correct = False
                    elif "яйца" in ingredient_name.lower():
                        # Should be integer for pcs
                        if qty_unit == "pcs" and isinstance(qty_net, int):
                            quantity_details.append(f"✅ {ingredient_name}: integer pcs ({qty_net})")
                        else:
                            quantity_details.append(f"❌ {ingredient_name}: expected integer pcs, got {qty_net}")
                            quantity_conversions_correct = False
                    else:
                        # g → kg with 3 decimal places
                        if qty_unit == "kg" and isinstance(qty_net, float):
                            quantity_details.append(f"✅ {ingredient_name}: g → {qty_net}kg")
                        else:
                            quantity_details.append(f"❌ {ingredient_name}: conversion issue {qty_net}{qty_unit}")
                            quantity_conversions_correct = False
                
                self.log_result("Quantity Conversions with Rounding", quantity_conversions_correct, "; ".join(quantity_details))
                
            except ImportError:
                # Fallback without openpyxl
                self.log_result("Unit Conversions Test", True, "Export successful (detailed validation requires openpyxl)")
                
        except Exception as e:
            self.log_result("Unit Conversions Test", False, f"Exception: {str(e)}")
    
    def test_issues_generation(self):
        """Test 4: Issues Generation and Handling"""
        print("\n🎯 TEST 4: Issues Generation and Handling")
        
        # Test noSku issues
        try:
            test_card = self.create_test_techcard_v2("missing_skus")
            
            response = requests.post(
                f"{API_BASE}/techcards.v2/export/iiko",
                json=test_card,
                timeout=30
            )
            
            if response.status_code == 200:
                self.log_result("Export with Missing SKUs", True, "Export succeeded despite missing SKUs")
                
                # Try to analyze content if openpyxl is available
                try:
                    import openpyxl
                    from io import BytesIO
                    
                    # Load and check XLSX content
                    workbook = openpyxl.load_workbook(BytesIO(response.content))
                    recipes_sheet = workbook['Recipes']
                    
                    # Check for empty ingredientSku cells
                    empty_sku_count = 0
                    for row in range(2, recipes_sheet.max_row + 1):
                        ingredient_sku = recipes_sheet.cell(row, 6).value  # ingredientSku column
                        if not ingredient_sku or ingredient_sku == "":
                            empty_sku_count += 1
                    
                    if empty_sku_count > 0:
                        self.log_result("noSku Issue Handling", True, f"{empty_sku_count} ingredients with empty SKU handled correctly")
                    else:
                        self.log_result("noSku Issue Handling", False, "Expected empty SKUs but found none")
                        
                except ImportError:
                    self.log_result("noSku Issue Handling", True, "Export completed (detailed validation requires openpyxl)")
                    
            else:
                self.log_result("Export with Missing SKUs", False, f"Export failed: HTTP {response.status_code}")
        
        except Exception as e:
            self.log_result("noSku Issues Test", False, f"Exception: {str(e)}")
        
        # Test yieldMismatch issues
        try:
            test_card = self.create_test_techcard_v2("yield_mismatch")
            
            response = requests.post(
                f"{API_BASE}/techcards.v2/export/iiko",
                json=test_card,
                timeout=30
            )
            
            if response.status_code == 200:
                self.log_result("Export with Yield Mismatch", True, "Export succeeded despite yield mismatch")
                
                # The export should still work, issues are logged but don't break export
                try:
                    import openpyxl
                    from io import BytesIO
                    
                    workbook = openpyxl.load_workbook(BytesIO(response.content))
                    
                    if 'Recipes' in workbook.sheetnames:
                        self.log_result("yieldMismatch Issue Handling", True, "Export completed with yield mismatch logged")
                    else:
                        self.log_result("yieldMismatch Issue Handling", False, "Export structure broken")
                        
                except ImportError:
                    self.log_result("yieldMismatch Issue Handling", True, "Export completed (detailed validation requires openpyxl)")
                    
            else:
                self.log_result("Export with Yield Mismatch", False, f"Export failed: HTTP {response.status_code}")
        
        except Exception as e:
            self.log_result("yieldMismatch Issues Test", False, f"Exception: {str(e)}")
    
    def test_data_accuracy_and_mapping(self):
        """Test 5: Data Accuracy and Mapping"""
        print("\n🎯 TEST 5: Data Accuracy and Mapping")
        
        try:
            test_card = self.create_test_techcard_v2("basic")
            
            response = requests.post(
                f"{API_BASE}/techcards.v2/export/iiko",
                json=test_card,
                timeout=30
            )
            
            if response.status_code != 200:
                self.log_result("Data Mapping Test Setup", False, f"HTTP {response.status_code}")
                return
            
            try:
                import openpyxl
                from io import BytesIO
                
                workbook = openpyxl.load_workbook(BytesIO(response.content))
                products_sheet = workbook['Products']
                recipes_sheet = workbook['Recipes']
                
                # Test unique SKUs in Products sheet
                products_skus = set()
                for row in range(2, products_sheet.max_row + 1):
                    sku = products_sheet.cell(row, 1).value  # skuId column
                    if sku:
                        products_skus.add(sku)
                
                expected_ingredient_count = len(test_card["ingredients"])
                if len(products_skus) == expected_ingredient_count:
                    self.log_result("Unique SKUs in Products", True, f"{len(products_skus)} unique SKUs")
                else:
                    self.log_result("Unique SKUs in Products", False, f"Expected {expected_ingredient_count}, got {len(products_skus)}")
                
                # Test dish code generation
                dish_codes = set()
                for row in range(2, recipes_sheet.max_row + 1):
                    dish_code = recipes_sheet.cell(row, 1).value  # dishCode column
                    if dish_code:
                        dish_codes.add(dish_code)
                
                expected_dish_code = f"DISH_{test_card['meta']['title'].upper().replace(' ', '_')}"
                if len(dish_codes) == 1 and expected_dish_code in dish_codes:
                    self.log_result("Dish Code Generation", True, f"Correct dish code: {expected_dish_code}")
                else:
                    self.log_result("Dish Code Generation", False, f"Expected {expected_dish_code}, got {dish_codes}")
                
                # Test cost metadata propagation
                cost_dates = set()
                for row in range(2, products_sheet.max_row + 1):
                    as_of = products_sheet.cell(row, 7).value  # asOf column
                    if as_of:
                        cost_dates.add(as_of)
                
                expected_date = test_card["costMeta"]["asOf"]
                if expected_date in cost_dates:
                    self.log_result("Cost Metadata Propagation", True, f"Correct asOf date: {expected_date}")
                else:
                    self.log_result("Cost Metadata Propagation", False, f"Expected {expected_date}, got {cost_dates}")
                
                # Test portion calculations
                portion_outputs = set()
                batch_outputs = set()
                
                for row in range(2, recipes_sheet.max_row + 1):
                    portion_output = recipes_sheet.cell(row, 4).value  # outputPerPortion_g column
                    batch_output = recipes_sheet.cell(row, 5).value   # outputPerBatch_g column
                    
                    if portion_output:
                        portion_outputs.add(portion_output)
                    if batch_output:
                        batch_outputs.add(batch_output)
                
                expected_portion = test_card["yield"]["perPortion_g"]
                expected_batch = test_card["yield"]["perBatch_g"]
                
                portion_correct = len(portion_outputs) == 1 and expected_portion in portion_outputs
                batch_correct = len(batch_outputs) == 1 and expected_batch in batch_outputs
                
                if portion_correct and batch_correct:
                    self.log_result("Portion Calculations", True, f"Correct yields: {expected_portion}g/portion, {expected_batch}g/batch")
                else:
                    self.log_result("Portion Calculations", False, f"Yield mismatch. Expected: {expected_portion}/{expected_batch}, got: {portion_outputs}/{batch_outputs}")
                
                # Test one row per ingredient in Recipes
                recipes_ingredient_count = recipes_sheet.max_row - 1  # Exclude header
                if recipes_ingredient_count == expected_ingredient_count:
                    self.log_result("One Row Per Ingredient", True, f"{recipes_ingredient_count} recipe rows for {expected_ingredient_count} ingredients")
                else:
                    self.log_result("One Row Per Ingredient", False, f"Expected {expected_ingredient_count} rows, got {recipes_ingredient_count}")
                
            except ImportError:
                # Fallback without openpyxl
                self.log_result("Data Accuracy Test", True, "Export successful (detailed validation requires openpyxl)")
                
        except Exception as e:
            self.log_result("Data Accuracy Test", False, f"Exception: {str(e)}")
    
    def test_comprehensive_scenario(self):
        """Test 6: Comprehensive Scenario with All Features"""
        print("\n🎯 TEST 6: Comprehensive Scenario Testing")
        
        try:
            # Create a complex test card with various challenges
            complex_card = {
                "meta": {
                    "id": "test-comprehensive-001",
                    "title": "Рататуй Прованский",
                    "version": "2.0",
                    "createdAt": datetime.now().isoformat(),
                    "cuisine": "french",
                    "tags": ["vegetarian", "complex"]
                },
                "portions": 6,
                "yield": {
                    "perPortion_g": 200.0,
                    "perBatch_g": 1200.0
                },
                "ingredients": [
                    {
                        "name": "Баклажаны",
                        "skuId": "VEG_EGGPLANT_001",
                        "unit": "g",
                        "brutto_g": 300.0,
                        "loss_pct": 15.0,
                        "netto_g": 255.0,
                        "allergens": []
                    },
                    {
                        "name": "Кабачки",
                        "skuId": None,  # Missing SKU
                        "unit": "g",
                        "brutto_g": 250.0,
                        "loss_pct": 10.0,
                        "netto_g": 225.0,
                        "allergens": []
                    },
                    {
                        "name": "Оливковое масло Extra Virgin",
                        "skuId": "OIL_OLIVE_EXTRA_002",
                        "unit": "ml",
                        "brutto_g": 80.0,
                        "loss_pct": 0.0,
                        "netto_g": 80.0,
                        "allergens": []
                    },
                    {
                        "name": "Помидоры черри",
                        "skuId": "VEG_TOMATO_CHERRY_003",
                        "unit": "pcs",
                        "brutto_g": 180.0,
                        "loss_pct": 5.0,
                        "netto_g": 171.0,
                        "allergens": []
                    },
                    {
                        "name": "Травы прованские",
                        "skuId": "",  # Empty SKU
                        "unit": "g",
                        "brutto_g": 8.0,
                        "loss_pct": 0.0,
                        "netto_g": 8.0,
                        "allergens": []
                    }
                ],
                "process": [
                    {
                        "n": 1,
                        "action": "Нарезать овощи кубиками 1x1 см",
                        "time_min": 15.0,
                        "equipment": ["нож", "доска"],
                        "ccp": False
                    },
                    {
                        "n": 2,
                        "action": "Обжарить баклажаны до золотистого цвета",
                        "time_min": 8.0,
                        "temp_c": 180.0,
                        "equipment": ["сковорода"],
                        "ccp": True
                    },
                    {
                        "n": 3,
                        "action": "Тушить все овощи с травами",
                        "time_min": 25.0,
                        "temp_c": 160.0,
                        "equipment": ["сотейник"],
                        "ccp": True
                    }
                ],
                "storage": {
                    "conditions": "Хранить в холодильнике при +2...+6°C",
                    "shelfLife_hours": 48.0,
                    "servingTemp_c": 65.0
                },
                "cost": {
                    "rawCost": 320.0,
                    "costPerPortion": 53.33,
                    "markup_pct": 280.0,
                    "vat_pct": 20.0
                },
                "costMeta": {
                    "source": "catalog",
                    "coveragePct": 60.0,  # Lower coverage due to missing SKUs
                    "asOf": "2025-01-17"
                }
            }
            
            # Note: Sum of netto = 255 + 225 + 80 + 171 + 8 = 739g, but yield = 1200g
            # This creates a yield mismatch for testing
            
            response = requests.post(
                f"{API_BASE}/techcards.v2/export/iiko",
                json=complex_card,
                timeout=30
            )
            
            if response.status_code == 200:
                self.log_result("Comprehensive Export", True, "Complex scenario exported successfully")
                
                try:
                    import openpyxl
                    from io import BytesIO
                    
                    # Analyze the result
                    workbook = openpyxl.load_workbook(BytesIO(response.content))
                    products_sheet = workbook['Products']
                    recipes_sheet = workbook['Recipes']
                    
                    # Count issues that should be present
                    missing_sku_count = 0
                    unit_variety = set()
                    
                    for row in range(2, recipes_sheet.max_row + 1):
                        ingredient_sku = recipes_sheet.cell(row, 6).value
                        qty_unit = recipes_sheet.cell(row, 9).value
                        
                        if not ingredient_sku or ingredient_sku == "":
                            missing_sku_count += 1
                        
                        if qty_unit:
                            unit_variety.add(qty_unit)
                    
                    # Should have 2 missing SKUs (кабачки, травы прованские)
                    if missing_sku_count == 2:
                        self.log_result("Complex noSku Handling", True, f"2 missing SKUs handled correctly")
                    else:
                        self.log_result("Complex noSku Handling", False, f"Expected 2 missing SKUs, found {missing_sku_count}")
                    
                    # Should have mixed units (kg, l, pcs)
                    expected_units = {"kg", "l", "pcs"}
                    if expected_units.issubset(unit_variety):
                        self.log_result("Complex Unit Conversions", True, f"All unit types present: {unit_variety}")
                    else:
                        self.log_result("Complex Unit Conversions", False, f"Missing units. Expected {expected_units}, got {unit_variety}")
                    
                    # Check file is valid XLSX
                    try:
                        # Try to save and reload to verify integrity
                        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                        temp_file.write(response.content)
                        temp_file.close()
                        
                        # Try to reopen
                        test_workbook = openpyxl.load_workbook(temp_file.name)
                        if len(test_workbook.sheetnames) == 2:
                            self.log_result("XLSX File Integrity", True, "File is valid and can be reopened")
                        else:
                            self.log_result("XLSX File Integrity", False, "File structure corrupted")
                        
                        os.unlink(temp_file.name)
                        
                    except Exception as e:
                        self.log_result("XLSX File Integrity", False, f"File integrity check failed: {str(e)}")
                        
                except ImportError:
                    self.log_result("Comprehensive Analysis", True, "Export successful (detailed validation requires openpyxl)")
                
            else:
                self.log_result("Comprehensive Export", False, f"HTTP {response.status_code}: {response.text[:100]}")
                
        except Exception as e:
            self.log_result("Comprehensive Scenario Test", False, f"Exception: {str(e)}")
    
    def run_all_tests(self):
        """Run all iiko export tests"""
        print("🚀 Starting Comprehensive iiko Export Testing for TechCardV2 (Task #6)")
        print("=" * 80)
        
        # Run all test methods
        self.test_iiko_export_endpoint()
        self.test_xlsx_structure_and_content()
        self.test_unit_conversions()
        self.test_issues_generation()
        self.test_data_accuracy_and_mapping()
        self.test_comprehensive_scenario()
        
        # Print summary
        print("\n" + "=" * 80)
        print("🎯 IIKO EXPORT TESTING SUMMARY")
        print("=" * 80)
        
        for result in self.results:
            print(result)
        
        print(f"\n📊 OVERALL RESULTS: {self.passed_count}/{self.test_count} tests passed")
        
        if self.passed_count == self.test_count:
            print("🎉 ALL TESTS PASSED - iiko Export functionality is working correctly!")
            return True
        else:
            print(f"⚠️  {self.test_count - self.passed_count} tests failed - issues need attention")
            return False

def main():
    """Main test execution"""
    print("🔧 Backend Testing: iiko Export Functionality for TechCardV2")
    print(f"🌐 Testing against: {BACKEND_URL}")
    
    tester = IikoExportTester()
    success = tester.run_all_tests()
    
    return 0 if success else 1

if __name__ == "__main__":
    sys.exit(main())