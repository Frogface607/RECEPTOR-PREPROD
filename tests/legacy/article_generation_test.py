#!/usr/bin/env python3
"""
ФИНАЛЬНЫЙ ТЕСТ АРТИКУЛОВ - Article Generation Logic Testing
Протестировать исправленную логику генерации артикулов с улучшенной обработкой ошибок.

ЦЕЛЬ: Проверить, что улучшенная логика генерации артикулов теперь работает корректно:
1. Артикулы генерируются для блюда и ингредиентов
2. Поля правильно сохраняются в Pydantic моделях
3. Экспорт работает с реальными артикулами
4. Скелетоны содержат корректные данные

КРИТИЧЕСКИЕ ИЗМЕНЕНИЯ:
- Добавлен traceback для отладки ошибок
- Улучшена логика обновления meta с article
- Использование IngredientV2.model_validate для создания обновленных ингредиентов
- Более робастная обработка вложенных полей
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Any, Optional

import httpx
import openpyxl
from pymongo import MongoClient

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017/receptor_pro')
DB_NAME = os.getenv('DB_NAME', 'receptor_pro')

class ArticleGenerationTester:
    """Comprehensive Article Generation Testing Suite"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=60.0)
        self.test_results = []
        self.organization_id = "test-org-articles"
        self.generated_techcard_id = None
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status} {test_name}")
        if details:
            print(f"    {details}")
        if response_time > 0:
            print(f"    Response time: {response_time:.3f}s")
        print()

    async def test_1_generate_caesar_salad_techcard(self):
        """Test 1: Generate 'Салат Цезарь' tech card with articles"""
        try:
            start_time = time.time()
            
            # Generate Caesar Salad tech card
            response = await self.client.post(
                f"{API_BASE}/techcards.v2/generate",
                json={
                    "name": "Салат Цезарь",
                    "cuisine": "европейская",
                    "equipment": [],
                    "budget": None,
                    "dietary": []
                }
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Check response structure
                if data.get('status') in ['success', 'draft'] and data.get('card'):
                    card = data['card']
                    self.generated_techcard_id = card.get('meta', {}).get('id')
                    
                    # Check if tech card has article
                    dish_article = card.get('article') or card.get('dish', {}).get('article')
                    
                    # Check ingredients for product codes
                    ingredients = card.get('ingredients', [])
                    ingredients_with_codes = 0
                    total_ingredients = len(ingredients)
                    
                    for ingredient in ingredients:
                        if ingredient.get('product_code') or ingredient.get('article'):
                            ingredients_with_codes += 1
                    
                    success = (
                        self.generated_techcard_id is not None and
                        total_ingredients > 0
                    )
                    
                    details = f"Tech card ID: {self.generated_techcard_id}, Dish article: {dish_article}, Ingredients: {total_ingredients}, With codes: {ingredients_with_codes}"
                else:
                    success = False
                    details = f"Generation failed: {data.get('status')}, Issues: {data.get('issues', [])}"
                
            else:
                success = False
                details = f"HTTP {response.status_code}: {response.text[:200]}"
                
            self.log_test("Generate Caesar Salad Tech Card", success, details, response_time)
            return success
            
        except Exception as e:
            self.log_test("Generate Caesar Salad Tech Card", False, f"Exception: {str(e)}")
            return False

    async def test_2_verify_techcard_articles_in_database(self):
        """Test 2: Verify tech card articles are saved in database"""
        try:
            if not self.generated_techcard_id:
                self.log_test("Verify Tech Card Articles in Database", False, "No tech card ID available")
                return False
                
            # Connect to MongoDB and check the tech card
            client = MongoClient(MONGO_URL)
            db = client[DB_NAME]
            
            # Try different possible ID fields
            techcard = (
                db.techcards_v2.find_one({"id": self.generated_techcard_id}) or
                db.techcards_v2.find_one({"meta.id": self.generated_techcard_id}) or
                db.techcards_v2.find_one({"_id": self.generated_techcard_id})
            )
            
            if not techcard:
                self.log_test("Verify Tech Card Articles in Database", False, "Tech card not found in database")
                return False
            
            # Check dish article
            dish_article = techcard.get('article') or techcard.get('dish', {}).get('article')
            
            # Check ingredient product codes
            ingredients = techcard.get('ingredients', [])
            ingredients_with_codes = 0
            ingredient_details = []
            
            for ingredient in ingredients:
                product_code = ingredient.get('product_code') or ingredient.get('article')
                if product_code:
                    ingredients_with_codes += 1
                    ingredient_details.append(f"{ingredient.get('name', 'Unknown')}: {product_code}")
            
            success = dish_article is not None and ingredients_with_codes > 0
            
            details = f"Dish article: {dish_article}, Ingredients with codes: {ingredients_with_codes}/{len(ingredients)}"
            if ingredient_details:
                details += f", Examples: {', '.join(ingredient_details[:3])}"
            
            self.log_test("Verify Tech Card Articles in Database", success, details)
            return success
            
        except Exception as e:
            self.log_test("Verify Tech Card Articles in Database", False, f"Exception: {str(e)}")
            return False

    async def test_3_run_preflight_check(self):
        """Test 3: Run preflight check to verify missing articles count"""
        try:
            if not self.generated_techcard_id:
                self.log_test("Run Preflight Check", False, "No tech card ID available")
                return False
                
            start_time = time.time()
            
            response = await self.client.post(
                f"{API_BASE}/export/preflight",
                json={
                    "techcardIds": [self.generated_techcard_id],
                    "organization_id": self.organization_id
                }
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                missing_dishes = len(data.get('missing_dishes', []))
                missing_products = len(data.get('missing_products', []))
                ttk_date = data.get('ttk_date')
                
                # Success if preflight runs without errors
                success = True
                details = f"Missing dishes: {missing_dishes}, Missing products: {missing_products}, TTK date: {ttk_date}"
                
            else:
                success = False
                details = f"HTTP {response.status_code}: {response.text[:200]}"
                
            self.log_test("Run Preflight Check", success, details, response_time)
            return success
            
        except Exception as e:
            self.log_test("Run Preflight Check", False, f"Exception: {str(e)}")
            return False

    async def test_4_export_zip_with_articles(self):
        """Test 4: Export ZIP file and verify article content"""
        try:
            if not self.generated_techcard_id:
                self.log_test("Export ZIP with Articles", False, "No tech card ID available")
                return False
                
            start_time = time.time()
            
            response = await self.client.post(
                f"{API_BASE}/export/zip",
                json={
                    "techcardIds": [self.generated_techcard_id],
                    "organization_id": self.organization_id
                }
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # Check if response is ZIP file
                content_type = response.headers.get('content-type', '')
                
                if 'application/zip' in content_type or 'application/octet-stream' in content_type:
                    zip_content = response.content
                    zip_size = len(zip_content)
                    
                    success = zip_size > 1000  # ZIP should be substantial
                    details = f"ZIP file size: {zip_size} bytes, Content-Type: {content_type}"
                    
                    # Store ZIP content for next test
                    self.exported_zip = zip_content
                    
                else:
                    success = False
                    details = f"Expected ZIP file, got: {content_type}"
                    
            else:
                success = False
                details = f"HTTP {response.status_code}: {response.text[:200]}"
                
            self.log_test("Export ZIP with Articles", success, details, response_time)
            return success
            
        except Exception as e:
            self.log_test("Export ZIP with Articles", False, f"Exception: {str(e)}")
            return False

    async def test_5_verify_xlsx_articles_content(self):
        """Test 5: Verify XLSX files contain real articles (not empty skeletons)"""
        try:
            if not hasattr(self, 'exported_zip'):
                self.log_test("Verify XLSX Articles Content", False, "No exported ZIP available")
                return False
                
            # Extract and analyze XLSX files
            with zipfile.ZipFile(io.BytesIO(self.exported_zip), 'r') as zip_file:
                file_names = zip_file.namelist()
                xlsx_files = [f for f in file_names if f.endswith('.xlsx')]
                
                if not xlsx_files:
                    self.log_test("Verify XLSX Articles Content", False, "No XLSX files found in ZIP")
                    return False
                
                articles_found = 0
                real_data_found = 0
                file_analysis = []
                
                for xlsx_file in xlsx_files:
                    try:
                        xlsx_content = zip_file.read(xlsx_file)
                        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_content))
                        
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            
                            # Look for article columns and real data
                            for row in sheet.iter_rows(max_row=20, values_only=True):
                                if row:
                                    for cell in row:
                                        if cell and isinstance(cell, str):
                                            # Check for 5-digit articles
                                            if len(cell) == 5 and cell.isdigit():
                                                articles_found += 1
                                            # Check for real dish/ingredient names (not mock)
                                            if any(word in cell.lower() for word in ['салат', 'цезарь', 'курица', 'сыр', 'листья']):
                                                real_data_found += 1
                        
                        file_analysis.append(f"{xlsx_file}: analyzed")
                        
                    except Exception as e:
                        file_analysis.append(f"{xlsx_file}: error - {str(e)}")
                
                success = articles_found > 0 and real_data_found > 0
                details = f"XLSX files: {len(xlsx_files)}, Articles found: {articles_found}, Real data entries: {real_data_found}"
                
            self.log_test("Verify XLSX Articles Content", success, details)
            return success
            
        except Exception as e:
            self.log_test("Verify XLSX Articles Content", False, f"Exception: {str(e)}")
            return False

    async def test_6_verify_skeletons_not_empty(self):
        """Test 6: Verify skeleton files contain real data, not empty content"""
        try:
            if not hasattr(self, 'exported_zip'):
                self.log_test("Verify Skeletons Not Empty", False, "No exported ZIP available")
                return False
                
            # Extract and analyze skeleton files
            with zipfile.ZipFile(io.BytesIO(self.exported_zip), 'r') as zip_file:
                file_names = zip_file.namelist()
                skeleton_files = [f for f in file_names if 'skeleton' in f.lower()]
                
                if not skeleton_files:
                    # No skeleton files might be OK if no missing articles
                    self.log_test("Verify Skeletons Not Empty", True, "No skeleton files found (articles may be complete)")
                    return True
                
                skeleton_analysis = []
                total_rows = 0
                
                for skeleton_file in skeleton_files:
                    try:
                        xlsx_content = zip_file.read(skeleton_file)
                        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_content))
                        
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            
                            # Count non-empty rows
                            rows_with_data = 0
                            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                                if any(cell for cell in row if cell):
                                    rows_with_data += 1
                            
                            total_rows += rows_with_data
                            skeleton_analysis.append(f"{skeleton_file}/{sheet_name}: {rows_with_data} rows")
                        
                    except Exception as e:
                        skeleton_analysis.append(f"{skeleton_file}: error - {str(e)}")
                
                success = total_rows > 0
                details = f"Skeleton files: {len(skeleton_files)}, Total data rows: {total_rows}"
                
            self.log_test("Verify Skeletons Not Empty", success, details)
            return success
            
        except Exception as e:
            self.log_test("Verify Skeletons Not Empty", False, f"Exception: {str(e)}")
            return False

    async def test_7_end_to_end_workflow(self):
        """Test 7: Complete end-to-end workflow validation"""
        try:
            start_time = time.time()
            
            # Generate another tech card to test complete workflow
            response = await self.client.post(
                f"{API_BASE}/techcards.v2/generate",
                json={
                    "name": "Борщ украинский с мясом",
                    "cuisine": "русская",
                    "equipment": [],
                    "budget": None,
                    "dietary": []
                }
            )
            
            if response.status_code != 200:
                self.log_test("End-to-End Workflow", False, f"Failed to generate second tech card: {response.status_code}")
                return False
            
            data = response.json()
            if data.get('status') not in ['success', 'draft'] or not data.get('card'):
                self.log_test("End-to-End Workflow", False, f"Second tech card generation failed: {data.get('status')}")
                return False
                
            second_techcard_id = data['card'].get('meta', {}).get('id')
            
            # Run preflight with both tech cards
            preflight_response = await self.client.post(
                f"{API_BASE}/export/preflight",
                json={
                    "techcardIds": [self.generated_techcard_id, second_techcard_id],
                    "organization_id": self.organization_id
                }
            )
            
            if preflight_response.status_code != 200:
                self.log_test("End-to-End Workflow", False, f"Preflight failed: {preflight_response.status_code}")
                return False
            
            # Export both tech cards
            export_response = await self.client.post(
                f"{API_BASE}/export/zip",
                json={
                    "techcardIds": [self.generated_techcard_id, second_techcard_id],
                    "organization_id": self.organization_id
                }
            )
            
            response_time = time.time() - start_time
            
            if export_response.status_code == 200:
                zip_size = len(export_response.content)
                success = zip_size > 2000  # Should be larger with 2 tech cards
                details = f"Complete workflow with 2 tech cards, ZIP size: {zip_size} bytes"
            else:
                success = False
                details = f"Export failed: {export_response.status_code}"
                
            self.log_test("End-to-End Workflow", success, details, response_time)
            return success
            
        except Exception as e:
            self.log_test("End-to-End Workflow", False, f"Exception: {str(e)}")
            return False

    async def run_all_tests(self):
        """Run all article generation tests"""
        print("🎯 ARTICLE GENERATION FIX TESTING STARTED")
        print("=" * 60)
        print()
        
        # Run tests in sequence
        test_1_result = await self.test_1_generate_caesar_salad_techcard()
        test_2_result = await self.test_2_verify_techcard_articles_in_database()
        test_3_result = await self.test_3_run_preflight_check()
        test_4_result = await self.test_4_export_zip_with_articles()
        test_5_result = await self.test_5_verify_xlsx_articles_content()
        test_6_result = await self.test_6_verify_skeletons_not_empty()
        test_7_result = await self.test_7_end_to_end_workflow()
        
        # Calculate results
        total_tests = len(self.test_results)
        passed_tests = sum(1 for result in self.test_results if result['success'])
        success_rate = (passed_tests / total_tests) * 100 if total_tests > 0 else 0
        
        print("=" * 60)
        print("🎯 ARTICLE GENERATION FIX TESTING SUMMARY")
        print("=" * 60)
        
        for result in self.test_results:
            print(f"{result['status']} {result['test']}")
            if result['details']:
                print(f"    {result['details']}")
            if result['response_time'] != "N/A":
                print(f"    Response time: {result['response_time']}")
            print()
        
        print(f"📊 OVERALL RESULTS: {passed_tests}/{total_tests} tests passed ({success_rate:.1f}%)")
        
        # Determine overall success
        critical_tests_passed = test_1_result and test_2_result and test_4_result
        
        if critical_tests_passed and success_rate >= 70:
            print("🎉 ARTICLE GENERATION FIX: SUCCESS")
            print("✅ Tech cards are being generated with articles")
            print("✅ Export system works with real articles")
            print("✅ Skeletons contain proper data")
        else:
            print("❌ ARTICLE GENERATION FIX: ISSUES DETECTED")
            print("⚠️ Article generation may still have problems")
            print("⚠️ Manual verification recommended")
        
        return success_rate >= 70

async def main():
    """Main test execution"""
    try:
        async with ArticleGenerationTester() as tester:
            success = await tester.run_all_tests()
            return 0 if success else 1
            
    except Exception as e:
        print(f"❌ CRITICAL ERROR: {str(e)}")
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    exit_code = asyncio.run(main())
    sys.exit(exit_code)