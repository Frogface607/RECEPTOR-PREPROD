#!/usr/bin/env python3
"""
Skeleton + Export Workflow Comprehensive Testing
Тестирование полной цепочки: скелетоны + экспорт техкарты

ЦЕЛЬ: Протестировать правильный workflow экспорта согласно логике системы:
1. Сгенерировать техкарту "Простой салат"
2. Получить скелетоны для создания блюда в iiko
3. Экспортировать техкарту после создания скелетонов
4. Проверить содержимое скелетонов на отсутствие mock данных
5. Убедиться что скелетоны содержат реальные данные техкарты
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Any, Optional

import httpx
import openpyxl
from pymongo import MongoClient

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017/receptor_pro')
DB_NAME = os.getenv('DB_NAME', 'receptor_pro')

class SkeletonExportTester:
    """Comprehensive Skeleton + Export Workflow Testing Suite"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=45.0)
        self.test_results = []
        self.organization_id = "test-org-skeleton"
        self.generated_techcard_id = None
        self.generated_techcard_name = "Простой салат"
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status} {test_name}")
        if details:
            print(f"    {details}")
        if response_time > 0:
            print(f"    Response time: {response_time:.3f}s")
        print()

    async def test_1_generate_techcard(self) -> bool:
        """Тест 1: Генерация техкарты 'Простой салат'"""
        try:
            start_time = time.time()
            
            # Generate tech card using the v2 endpoint
            endpoint = f"{API_BASE}/techcards.v2/generate"
            
            payload = {
                "name": self.generated_techcard_name,
                "cuisine": "европейская",
                "equipment": [],
                "budget": None,
                "dietary": []
            }
            
            response = await self.client.post(endpoint, json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Extract tech card ID from response
                if 'card' in data and data['card'] and 'meta' in data['card'] and 'id' in data['card']['meta']:
                    self.generated_techcard_id = data['card']['meta']['id']
                    techcard_name = data['card']['meta'].get('title', 'Unknown')
                    ingredients_count = len(data['card'].get('ingredients', []))
                    status = data.get('status', 'unknown')
                    
                    self.log_test(
                        "Генерация техкарты 'Простой салат'",
                        True,
                        f"Техкарта создана: ID={self.generated_techcard_id}, название='{techcard_name}', ингредиентов={ingredients_count}, статус={status}",
                        response_time
                    )
                    return True
                else:
                    self.log_test(
                        "Генерация техкарты 'Простой салат'",
                        False,
                        f"Техкарта создана, но ID не найден в ответе: {data}",
                        response_time
                    )
                    return False
            else:
                self.log_test(
                    "Генерация техкарты 'Простой салат'",
                    False,
                    f"HTTP {response.status_code}: {response.text}",
                    response_time
                )
                return False
                
        except Exception as e:
            self.log_test(
                "Генерация техкарты 'Простой салат'",
                False,
                f"Ошибка: {str(e)}"
            )
            return False

    async def test_2_preflight_check(self) -> Dict[str, Any]:
        """Тест 2: Preflight проверка для получения скелетонов"""
        try:
            start_time = time.time()
            
            if not self.generated_techcard_id:
                self.log_test(
                    "Preflight проверка",
                    False,
                    "Нет ID техкарты для проверки"
                )
                return {}
            
            # Run preflight check
            endpoint = f"{API_BASE}/export/preflight"
            
            payload = {
                "techcardIds": [self.generated_techcard_id],
                "organization_id": self.organization_id
            }
            
            response = await self.client.post(endpoint, json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Extract preflight results
                ttk_date = data.get('ttk_date', 'Unknown')
                dish_skeletons = data.get('dish_skeletons_count', 0)
                product_skeletons = data.get('product_skeletons_count', 0)
                warnings = data.get('warnings', [])
                
                self.log_test(
                    "Preflight проверка",
                    True,
                    f"TTK дата: {ttk_date}, скелетонов блюд: {dish_skeletons}, скелетонов продуктов: {product_skeletons}, предупреждений: {len(warnings)}",
                    response_time
                )
                return data
            else:
                self.log_test(
                    "Preflight проверка",
                    False,
                    f"HTTP {response.status_code}: {response.text}",
                    response_time
                )
                return {}
                
        except Exception as e:
            self.log_test(
                "Preflight проверка",
                False,
                f"Ошибка: {str(e)}"
            )
            return {}

    async def test_3_zip_export(self, preflight_result: Dict[str, Any] = None) -> Optional[bytes]:
        """Тест 3: ZIP экспорт для получения файлов скелетонов"""
        try:
            start_time = time.time()
            
            if not self.generated_techcard_id:
                self.log_test(
                    "ZIP экспорт",
                    False,
                    "Нет ID техкарты для экспорта"
                )
                return None
            
            # Export ZIP with skeletons
            endpoint = f"{API_BASE}/export/zip"
            
            payload = {
                "techcardIds": [self.generated_techcard_id],
                "organization_id": self.organization_id,
                "operational_rounding": True
            }
            
            # Include preflight result if provided (proper workflow)
            if preflight_result:
                payload["preflight_result"] = preflight_result
            
            response = await self.client.post(endpoint, json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # Check if response is ZIP file
                content_type = response.headers.get('content-type', '')
                zip_content = response.content
                zip_size = len(zip_content)
                
                if 'application/zip' in content_type or zip_size > 1000:
                    self.log_test(
                        "ZIP экспорт",
                        True,
                        f"ZIP файл получен: размер={zip_size} байт, content-type='{content_type}'",
                        response_time
                    )
                    return zip_content
                else:
                    self.log_test(
                        "ZIP экспорт",
                        False,
                        f"Ответ не является ZIP файлом: content-type='{content_type}', размер={zip_size}",
                        response_time
                    )
                    return None
            else:
                self.log_test(
                    "ZIP экспорт",
                    False,
                    f"HTTP {response.status_code}: {response.text}",
                    response_time
                )
                return None
                
        except Exception as e:
            self.log_test(
                "ZIP экспорт",
                False,
                f"Ошибка: {str(e)}"
            )
            return None

    async def test_4_skeleton_content_validation(self, zip_content: bytes) -> bool:
        """Тест 4: Проверка содержимого скелетонов на отсутствие mock данных"""
        try:
            if not zip_content:
                self.log_test(
                    "Проверка содержимого скелетонов",
                    False,
                    "Нет ZIP контента для проверки"
                )
                return False
            
            # Extract ZIP and check files
            with zipfile.ZipFile(io.BytesIO(zip_content), 'r') as zip_file:
                file_list = zip_file.namelist()
                
                # Look for skeleton files
                dish_skeleton_file = None
                product_skeleton_file = None
                ttk_file = None
                
                for filename in file_list:
                    if 'Dish-Skeletons.xlsx' in filename:
                        dish_skeleton_file = filename
                    elif 'Product-Skeletons.xlsx' in filename:
                        product_skeleton_file = filename
                    elif 'iiko_TTK.xlsx' in filename:
                        ttk_file = filename
                
                mock_signatures_found = []
                real_data_found = []
                
                # Check Dish-Skeletons.xlsx
                if dish_skeleton_file:
                    dish_content = zip_file.read(dish_skeleton_file)
                    dish_wb = openpyxl.load_workbook(io.BytesIO(dish_content))
                    dish_ws = dish_wb.active
                    
                    # Check for mock content
                    for row in dish_ws.iter_rows(values_only=True):
                        for cell_value in row:
                            if cell_value and isinstance(cell_value, str):
                                cell_lower = cell_value.lower()
                                if any(mock_sig in cell_lower for mock_sig in ['mock', 'test_', 'generated_test', 'dish_mock']):
                                    mock_signatures_found.append(f"Dish-Skeletons: {cell_value}")
                                elif self.generated_techcard_name.lower() in cell_lower:
                                    real_data_found.append(f"Dish-Skeletons: найдено '{self.generated_techcard_name}'")
                
                # Check Product-Skeletons.xlsx
                if product_skeleton_file:
                    product_content = zip_file.read(product_skeleton_file)
                    product_wb = openpyxl.load_workbook(io.BytesIO(product_content))
                    product_ws = product_wb.active
                    
                    # Check for mock content and real ingredients
                    for row in product_ws.iter_rows(values_only=True):
                        for cell_value in row:
                            if cell_value and isinstance(cell_value, str):
                                cell_lower = cell_value.lower()
                                if any(mock_sig in cell_lower for mock_sig in ['mock', 'test_', 'generated_test']):
                                    mock_signatures_found.append(f"Product-Skeletons: {cell_value}")
                                elif any(ingredient in cell_lower for ingredient in ['салат', 'помидор', 'огурец', 'лук', 'масло']):
                                    real_data_found.append(f"Product-Skeletons: найден ингредиент '{cell_value}'")
                
                # Summary
                files_found = len(file_list)
                has_dish_skeletons = dish_skeleton_file is not None
                has_product_skeletons = product_skeleton_file is not None
                has_ttk = ttk_file is not None
                mock_count = len(mock_signatures_found)
                real_count = len(real_data_found)
                
                success = mock_count == 0 and real_count > 0
                
                details = f"Файлов в ZIP: {files_found}, Dish-Skeletons: {'✅' if has_dish_skeletons else '❌'}, Product-Skeletons: {'✅' if has_product_skeletons else '❌'}, TTK: {'✅' if has_ttk else '❌'}"
                details += f"\nMock подписей найдено: {mock_count}"
                details += f"\nРеальных данных найдено: {real_count}"
                
                if mock_signatures_found:
                    details += f"\nMock подписи: {mock_signatures_found[:3]}"  # Show first 3
                if real_data_found:
                    details += f"\nРеальные данные: {real_data_found[:3]}"  # Show first 3
                
                self.log_test(
                    "Проверка содержимого скелетонов",
                    success,
                    details
                )
                return success
                
        except Exception as e:
            self.log_test(
                "Проверка содержимого скелетонов",
                False,
                f"Ошибка при проверке: {str(e)}"
            )
            return False

    async def test_5_techcard_database_verification(self) -> bool:
        """Тест 5: Проверка сохранения техкарты в базе данных"""
        try:
            if not self.generated_techcard_id:
                self.log_test(
                    "Проверка техкарты в БД",
                    False,
                    "Нет ID техкарты для проверки"
                )
                return False
            
            # Connect to MongoDB and check tech card
            client = MongoClient(MONGO_URL)
            db = client[DB_NAME]
            
            # Try different collection names
            collections_to_try = ['techcards_v2', 'techcards', 'tech_cards']
            techcard = None
            
            for collection_name in collections_to_try:
                collection = db[collection_name]
                # Try different ID field names
                for id_field in ['meta.id', 'id', '_id']:
                    if id_field == 'meta.id':
                        techcard = collection.find_one({"meta.id": self.generated_techcard_id})
                    else:
                        techcard = collection.find_one({id_field: self.generated_techcard_id})
                    
                    if techcard:
                        break
                if techcard:
                    break
            
            if techcard:
                # Handle different document structures
                if 'meta' in techcard and 'title' in techcard['meta']:
                    name = techcard['meta']['title']
                else:
                    name = techcard.get('name', 'Unknown')
                
                ingredients = techcard.get('ingredients', [])
                ingredients_count = len(ingredients)
                
                # Check for articles in ingredients
                articles_found = 0
                for ingredient in ingredients:
                    if ingredient.get('product_code') or ingredient.get('article'):
                        articles_found += 1
                
                success = ingredients_count > 0
                
                self.log_test(
                    "Проверка техкарты в БД",
                    success,
                    f"Техкарта найдена: название='{name}', ингредиентов={ingredients_count}, артикулов={articles_found}",
                )
                return success
            else:
                self.log_test(
                    "Проверка техкарты в БД",
                    False,
                    f"Техкарта с ID {self.generated_techcard_id} не найдена в БД"
                )
                return False
                
        except Exception as e:
            self.log_test(
                "Проверка техкарты в БД",
                False,
                f"Ошибка подключения к БД: {str(e)}"
            )
            return False

    async def test_6_article_allocation_workflow(self) -> bool:
        """Тест 6: Проверка работы системы выделения артикулов"""
        try:
            start_time = time.time()
            
            # Test article allocation endpoint
            endpoint = f"{API_BASE}/articles/allocate"
            
            payload = {
                "entity_ids": [f"dish_{self.generated_techcard_id}", f"product_1_{self.generated_techcard_id}"],
                "article_types": ["dish", "product"],
                "organization_id": self.organization_id
            }
            
            response = await self.client.post(endpoint, json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                allocated_articles = data.get('allocated_articles', [])
                
                # Check article format (should be 5-digit)
                valid_articles = 0
                for article in allocated_articles:
                    article_code = article.get('article', '')
                    if len(article_code) == 5 and article_code.isdigit():
                        valid_articles += 1
                
                success = len(allocated_articles) > 0 and valid_articles == len(allocated_articles)
                
                self.log_test(
                    "Система выделения артикулов",
                    success,
                    f"Выделено артикулов: {len(allocated_articles)}, валидных 5-значных: {valid_articles}",
                    response_time
                )
                return success
            else:
                self.log_test(
                    "Система выделения артикулов",
                    False,
                    f"HTTP {response.status_code}: {response.text}",
                    response_time
                )
                return False
                
        except Exception as e:
            self.log_test(
                "Система выделения артикулов",
                False,
                f"Ошибка: {str(e)}"
            )
            return False

    async def run_comprehensive_test(self):
        """Запуск полного комплексного теста"""
        print("🎯 НАЧАЛО КОМПЛЕКСНОГО ТЕСТИРОВАНИЯ СКЕЛЕТОНОВ + ЭКСПОРТ")
        print("=" * 80)
        print(f"Backend URL: {BACKEND_URL}")
        print(f"Техкарта для тестирования: '{self.generated_techcard_name}'")
        print(f"Organization ID: {self.organization_id}")
        print()
        
        # Test sequence
        test_sequence = [
            ("1. Генерация техкарты", "test_1_generate_techcard"),
            ("2. Preflight проверка", "test_2_preflight_check"),
            ("3. ZIP экспорт", "test_3_zip_export"),
            ("4. Проверка содержимого скелетонов", "test_4_skeleton_content_validation"),
            ("5. Проверка техкарты в БД", "test_5_techcard_database_verification"),
            ("6. Система выделения артикулов", "test_6_article_allocation_workflow")
        ]
        
        zip_content = None
        
        # Execute tests in sequence
        preflight_result = {}
        zip_content = None
        
        for i, (test_name, test_method) in enumerate(test_sequence):
            if i == 0:  # Generate techcard
                success = await getattr(self, test_method)()
                if not success:
                    print("❌ Критическая ошибка: не удалось сгенерировать техкарту")
                    break
            elif i == 1:  # Preflight check
                preflight_result = await getattr(self, test_method)()
                if not preflight_result:
                    print("⚠️ Preflight проверка не прошла, но продолжаем тестирование")
            elif i == 2:  # ZIP export
                zip_content = await getattr(self, test_method)(preflight_result)
                if not zip_content:
                    print("❌ Критическая ошибка: не удалось получить ZIP файл")
                    break
            elif i == 3:  # Skeleton content validation
                if zip_content:
                    await getattr(self, test_method)(zip_content)
                else:
                    self.log_test("Проверка содержимого скелетонов", False, "Нет ZIP контента")
            else:  # Other tests
                await getattr(self, test_method)()
        
        # Generate summary
        await self.generate_test_summary()

    async def generate_test_summary(self):
        """Генерация итогового отчета"""
        print("\n" + "=" * 80)
        print("📊 ИТОГОВЫЙ ОТЧЕТ ТЕСТИРОВАНИЯ")
        print("=" * 80)
        
        total_tests = len(self.test_results)
        passed_tests = sum(1 for result in self.test_results if result['success'])
        failed_tests = total_tests - passed_tests
        success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
        
        print(f"Всего тестов: {total_tests}")
        print(f"Пройдено: {passed_tests} ✅")
        print(f"Провалено: {failed_tests} ❌")
        print(f"Процент успеха: {success_rate:.1f}%")
        print()
        
        # Detailed results
        print("ДЕТАЛЬНЫЕ РЕЗУЛЬТАТЫ:")
        print("-" * 40)
        for result in self.test_results:
            print(f"{result['status']} {result['test']}")
            if result['details']:
                print(f"    {result['details']}")
            if result['response_time'] != "N/A":
                print(f"    Время ответа: {result['response_time']}")
            print()
        
        # Critical findings
        critical_issues = []
        success_points = []
        
        for result in self.test_results:
            if not result['success']:
                critical_issues.append(result['test'])
            else:
                success_points.append(result['test'])
        
        if critical_issues:
            print("🚨 КРИТИЧЕСКИЕ ПРОБЛЕМЫ:")
            for issue in critical_issues:
                print(f"  ❌ {issue}")
            print()
        
        if success_points:
            print("✅ УСПЕШНЫЕ КОМПОНЕНТЫ:")
            for success in success_points:
                print(f"  ✅ {success}")
            print()
        
        # Final verdict
        if success_rate >= 80:
            print("🎉 ОБЩИЙ РЕЗУЛЬТАТ: УСПЕШНО")
            print("Система скелетонов + экспорт работает корректно")
        elif success_rate >= 60:
            print("⚠️ ОБЩИЙ РЕЗУЛЬТАТ: ЧАСТИЧНО УСПЕШНО")
            print("Основная функциональность работает, есть незначительные проблемы")
        else:
            print("❌ ОБЩИЙ РЕЗУЛЬТАТ: ТРЕБУЕТСЯ ИСПРАВЛЕНИЕ")
            print("Критические проблемы в системе скелетонов + экспорт")
        
        print("=" * 80)

async def main():
    """Main test execution"""
    try:
        async with SkeletonExportTester() as tester:
            await tester.run_comprehensive_test()
    except KeyboardInterrupt:
        print("\n⚠️ Тестирование прервано пользователем")
    except Exception as e:
        print(f"\n❌ Критическая ошибка тестирования: {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    asyncio.run(main())