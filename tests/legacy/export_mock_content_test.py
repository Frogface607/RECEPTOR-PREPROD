#!/usr/bin/env python3
"""
Export System Mock Content Validation Test
Critical verification that real tech card data is now being used instead of mock content in exported XLSX files.
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Any, Optional

import httpx
import openpyxl
from pymongo import MongoClient

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017/receptor_pro')
DB_NAME = os.getenv('DB_NAME', 'receptor_pro')

class ExportMockContentTester:
    """Comprehensive Export Mock Content Validation Testing Suite"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=60.0)
        self.test_results = []
        self.organization_id = "test-org-export-validation"
        self.generated_tech_card_ids = []
        self.test_artifacts = {}
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status}: {test_name} ({response_time:.3f}s) - {details}")
    
    async def test_generate_real_tech_cards(self):
        """Test 1: Generate real tech cards and extract IDs"""
        print("\n🎯 Test 1: Generate Real Tech Cards and Extract IDs")
        
        # Generate 2 tech cards with realistic dish names
        test_dishes = [
            {
                "name": "Борщ украинский с говядиной",
                "description": "Традиционный украинский борщ с говядиной и овощами",
                "category": "Супы"
            },
            {
                "name": "Стейк из говядины с картофельным пюре",
                "description": "Сочный стейк из говядины с гарниром из картофельного пюре",
                "category": "Горячие блюда"
            }
        ]
        
        for i, dish_info in enumerate(test_dishes, 1):
            try:
                start_time = time.time()
                
                payload = {
                    "name": dish_info["name"],
                    "description": dish_info["description"],
                    "category": dish_info["category"],
                    "portions": 1,
                    "cuisine": "Русская",
                    "difficulty": "medium"
                }
                
                response = await self.client.post(f"{API_BASE}/techcards.v2/generate", json=payload)
                response_time = time.time() - start_time
                
                if response.status_code == 200:
                    data = response.json()
                    
                    if data.get("status") in ["success", "draft"] and data.get("card"):
                        tech_card = data["card"]
                        tech_card_id = tech_card.get("meta", {}).get("id")
                        
                        if tech_card_id:
                            self.generated_tech_card_ids.append(tech_card_id)
                            self.test_artifacts[f"gen_run_{i}"] = {
                                "id": tech_card_id,
                                "name": dish_info["name"],
                                "status": data.get("status"),
                                "ingredients_count": len(tech_card.get("ingredients", [])),
                                "response_time": response_time
                            }
                            
                            self.log_test(f"Generate Tech Card {i}", True,
                                        f"Generated '{dish_info['name']}' with ID: {tech_card_id}", response_time)
                        else:
                            self.log_test(f"Generate Tech Card {i}", False,
                                        f"No ID found in generated tech card", response_time)
                    else:
                        self.log_test(f"Generate Tech Card {i}", False,
                                    f"Generation failed: {data.get('status', 'unknown')}", response_time)
                else:
                    self.log_test(f"Generate Tech Card {i}", False,
                                f"HTTP {response.status_code}: {response.text[:200]}", response_time)
                    
            except Exception as e:
                self.log_test(f"Generate Tech Card {i}", False, f"Exception: {str(e)}", 0.0)
        
        # Summary of generation
        success_count = len(self.generated_tech_card_ids)
        self.log_test("Tech Card Generation Summary", success_count >= 1,
                    f"Generated {success_count}/2 tech cards with IDs: {self.generated_tech_card_ids}", 0.0)
    
    async def test_preflight_with_real_ids(self):
        """Test 2: Run preflight with real tech card IDs"""
        print("\n🎯 Test 2: Run Preflight with Real Tech Card IDs")
        
        if not self.generated_tech_card_ids:
            self.log_test("Preflight with Real IDs", False, "No tech card IDs available", 0.0)
            return
        
        try:
            start_time = time.time()
            
            payload = {
                "techcardIds": self.generated_tech_card_ids,
                "organization_id": self.organization_id
            }
            
            response = await self.client.post(f"{API_BASE}/export/preflight", json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ["status", "ttkDate", "missing", "generated", "counts"]
                missing_fields = [field for field in required_fields if field not in data]
                
                if not missing_fields:
                    self.test_artifacts["preflight_result"] = data
                    
                    ttk_date = data.get("ttkDate")
                    dish_skeletons = data.get("counts", {}).get("dishSkeletons", 0)
                    product_skeletons = data.get("counts", {}).get("productSkeletons", 0)
                    
                    details = f"TTK Date: {ttk_date}, Dish Skeletons: {dish_skeletons}, Product Skeletons: {product_skeletons}"
                    self.log_test("Preflight with Real IDs", True, details, response_time)
                else:
                    self.log_test("Preflight with Real IDs", False,
                                f"Missing fields: {missing_fields}", response_time)
            else:
                self.log_test("Preflight with Real IDs", False,
                            f"HTTP {response.status_code}: {response.text[:200]}", response_time)
                
        except Exception as e:
            self.log_test("Preflight with Real IDs", False, f"Exception: {str(e)}", 0.0)
    
    async def test_zip_export_with_real_ids(self):
        """Test 3: Run ZIP export with real tech card IDs"""
        print("\n🎯 Test 3: Run ZIP Export with Real Tech Card IDs")
        
        if not self.generated_tech_card_ids or "preflight_result" not in self.test_artifacts:
            self.log_test("ZIP Export with Real IDs", False, "Prerequisites not met", 0.0)
            return
        
        try:
            start_time = time.time()
            
            payload = {
                "techcardIds": self.generated_tech_card_ids,
                "operational_rounding": True,
                "organization_id": self.organization_id,
                "preflight_result": self.test_artifacts["preflight_result"]
            }
            
            response = await self.client.post(f"{API_BASE}/export/zip", json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # Check if response is a ZIP file
                content_type = response.headers.get('content-type', '')
                content_length = len(response.content)
                
                is_zip = (content_type == 'application/zip' or 
                         response.content.startswith(b'PK'))
                
                if is_zip and content_length > 0:
                    self.test_artifacts["export_zip"] = response.content
                    self.test_artifacts["zip_size"] = content_length
                    
                    self.log_test("ZIP Export with Real IDs", True,
                                f"ZIP file generated: {content_length} bytes", response_time)
                else:
                    self.log_test("ZIP Export with Real IDs", False,
                                f"Invalid ZIP: content-type={content_type}, size={content_length}", response_time)
            else:
                self.log_test("ZIP Export with Real IDs", False,
                            f"HTTP {response.status_code}: {response.text[:200]}", response_time)
                
        except Exception as e:
            self.log_test("ZIP Export with Real IDs", False, f"Exception: {str(e)}", 0.0)
    
    async def test_xlsx_content_validation(self):
        """Test 4: Validate XLSX files for mock content"""
        print("\n🎯 Test 4: Validate XLSX Files for Mock Content")
        
        if "export_zip" not in self.test_artifacts:
            self.log_test("XLSX Content Validation", False, "No ZIP file available", 0.0)
            return
        
        try:
            start_time = time.time()
            
            # Extract ZIP file
            zip_buffer = io.BytesIO(self.test_artifacts["export_zip"])
            
            mock_signatures_found = []
            real_content_found = []
            xlsx_files_processed = 0
            
            with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
                for file_name in zip_file.namelist():
                    if file_name.endswith('.xlsx'):
                        xlsx_files_processed += 1
                        print(f"  📄 Processing {file_name}...")
                        
                        # Extract and read XLSX file
                        xlsx_data = zip_file.read(file_name)
                        xlsx_buffer = io.BytesIO(xlsx_data)
                        
                        try:
                            workbook = openpyxl.load_workbook(xlsx_buffer)
                            
                            for sheet_name in workbook.sheetnames:
                                sheet = workbook[sheet_name]
                                
                                # Scan all cells for mock content
                                for row in sheet.iter_rows():
                                    for cell in row:
                                        if cell.value and isinstance(cell.value, str):
                                            cell_value = str(cell.value).strip()
                                            
                                            # Check for mock signatures
                                            mock_patterns = [
                                                "DISH_MOCK_TECH_CARD",
                                                "GENERATED_TEST_INGREDIENT",
                                                "TEST_INGREDIENT",
                                                "Mock Tech Card",
                                                "Test Ingredient"
                                            ]
                                            
                                            for pattern in mock_patterns:
                                                if pattern in cell_value:
                                                    mock_signatures_found.append({
                                                        "file": file_name,
                                                        "sheet": sheet_name,
                                                        "cell": cell.coordinate,
                                                        "pattern": pattern,
                                                        "value": cell_value[:50]
                                                    })
                                            
                                            # Check for real content (our generated dish names)
                                            for artifact in self.test_artifacts.values():
                                                if isinstance(artifact, dict) and "name" in artifact:
                                                    if artifact["name"] in cell_value:
                                                        real_content_found.append({
                                                            "file": file_name,
                                                            "sheet": sheet_name,
                                                            "cell": cell.coordinate,
                                                            "dish_name": artifact["name"],
                                                            "value": cell_value[:50]
                                                        })
                            
                            workbook.close()
                            
                        except Exception as xlsx_error:
                            print(f"    ⚠️ Error reading {file_name}: {xlsx_error}")
            
            response_time = time.time() - start_time
            
            # Store results
            self.test_artifacts["xlsx_validation"] = {
                "files_processed": xlsx_files_processed,
                "mock_signatures": mock_signatures_found,
                "real_content": real_content_found
            }
            
            # Determine test result
            has_mock_content = len(mock_signatures_found) > 0
            has_real_content = len(real_content_found) > 0
            
            if has_mock_content:
                details = f"❌ MOCK CONTENT DETECTED: Found {len(mock_signatures_found)} mock signatures in {xlsx_files_processed} XLSX files"
                for signature in mock_signatures_found[:3]:  # Show first 3
                    details += f"\n    • {signature['file']} {signature['cell']}: {signature['pattern']}"
                self.log_test("XLSX Mock Content Check", False, details, response_time)
            else:
                details = f"✅ NO MOCK CONTENT: Scanned {xlsx_files_processed} XLSX files, no mock signatures found"
                if has_real_content:
                    details += f", found {len(real_content_found)} real content matches"
                self.log_test("XLSX Mock Content Check", True, details, response_time)
            
            # Separate test for real content presence
            if has_real_content:
                details = f"Found real dish names in {len(real_content_found)} locations"
                for content in real_content_found[:2]:  # Show first 2
                    details += f"\n    • {content['file']} {content['cell']}: {content['dish_name']}"
                self.log_test("XLSX Real Content Check", True, details, response_time)
            else:
                self.log_test("XLSX Real Content Check", False,
                            f"No real dish names found in exported XLSX files", response_time)
                
        except Exception as e:
            self.log_test("XLSX Content Validation", False, f"Exception: {str(e)}", 0.0)
    
    async def test_article_formatting_validation(self):
        """Test 5: Validate article formatting (5-digit with leading zeros)"""
        print("\n🎯 Test 5: Validate Article Formatting")
        
        if "export_zip" not in self.test_artifacts:
            self.log_test("Article Formatting Validation", False, "No ZIP file available", 0.0)
            return
        
        try:
            start_time = time.time()
            
            # Extract ZIP file
            zip_buffer = io.BytesIO(self.test_artifacts["export_zip"])
            
            articles_found = []
            formatting_issues = []
            
            with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
                for file_name in zip_file.namelist():
                    if file_name.endswith('.xlsx'):
                        print(f"  📄 Checking articles in {file_name}...")
                        
                        # Extract and read XLSX file
                        xlsx_data = zip_file.read(file_name)
                        xlsx_buffer = io.BytesIO(xlsx_data)
                        
                        try:
                            workbook = openpyxl.load_workbook(xlsx_buffer)
                            
                            for sheet_name in workbook.sheetnames:
                                sheet = workbook[sheet_name]
                                
                                # Look for article columns (typically columns A and C)
                                for col_idx in [1, 3]:  # A and C columns
                                    for row_idx in range(2, min(20, sheet.max_row + 1)):  # Skip header, check first 18 rows
                                        cell = sheet.cell(row=row_idx, column=col_idx)
                                        
                                        if cell.value and isinstance(cell.value, str):
                                            cell_value = str(cell.value).strip()
                                            
                                            # Check if it looks like an article (5 digits)
                                            if cell_value.isdigit() and len(cell_value) == 5:
                                                articles_found.append({
                                                    "file": file_name,
                                                    "sheet": sheet_name,
                                                    "cell": cell.coordinate,
                                                    "article": cell_value,
                                                    "format_type": cell.data_type
                                                })
                                            elif cell_value.isdigit() and len(cell_value) != 5:
                                                formatting_issues.append({
                                                    "file": file_name,
                                                    "sheet": sheet_name,
                                                    "cell": cell.coordinate,
                                                    "article": cell_value,
                                                    "issue": f"Wrong length: {len(cell_value)} digits"
                                                })
                            
                            workbook.close()
                            
                        except Exception as xlsx_error:
                            print(f"    ⚠️ Error reading {file_name}: {xlsx_error}")
            
            response_time = time.time() - start_time
            
            # Store results
            self.test_artifacts["article_validation"] = {
                "articles_found": articles_found,
                "formatting_issues": formatting_issues
            }
            
            # Determine test result
            valid_articles = len(articles_found)
            issues_count = len(formatting_issues)
            
            if valid_articles > 0 and issues_count == 0:
                details = f"Found {valid_articles} properly formatted 5-digit articles"
                for article in articles_found[:3]:  # Show first 3
                    details += f"\n    • {article['file']} {article['cell']}: {article['article']}"
                self.log_test("Article Formatting Validation", True, details, response_time)
            elif valid_articles > 0 and issues_count > 0:
                details = f"Found {valid_articles} valid articles but {issues_count} formatting issues"
                self.log_test("Article Formatting Validation", False, details, response_time)
            else:
                details = f"No properly formatted articles found, {issues_count} issues detected"
                self.log_test("Article Formatting Validation", False, details, response_time)
                
        except Exception as e:
            self.log_test("Article Formatting Validation", False, f"Exception: {str(e)}", 0.0)
    
    async def test_database_persistence_check(self):
        """Test 6: Verify tech cards are persisted in database"""
        print("\n🎯 Test 6: Verify Tech Cards Database Persistence")
        
        if not self.generated_tech_card_ids:
            self.log_test("Database Persistence Check", False, "No tech card IDs to check", 0.0)
            return
        
        try:
            start_time = time.time()
            
            # Connect to MongoDB
            client = MongoClient(MONGO_URL)
            db = client[DB_NAME.strip('"')]
            techcards_collection = db.techcards_v2
            
            found_cards = []
            missing_cards = []
            
            for tech_card_id in self.generated_tech_card_ids:
                doc = techcards_collection.find_one({"_id": tech_card_id})
                if doc:
                    found_cards.append({
                        "id": tech_card_id,
                        "name": doc.get("meta", {}).get("title", "Unknown"),
                        "ingredients_count": len(doc.get("ingredients", [])),
                        "created_at": doc.get("created_at")
                    })
                else:
                    missing_cards.append(tech_card_id)
            
            client.close()
            response_time = time.time() - start_time
            
            # Store results
            self.test_artifacts["database_check"] = {
                "found_cards": found_cards,
                "missing_cards": missing_cards
            }
            
            # Determine test result
            found_count = len(found_cards)
            missing_count = len(missing_cards)
            
            if found_count > 0 and missing_count == 0:
                details = f"All {found_count} tech cards found in database"
                for card in found_cards:
                    details += f"\n    • {card['id']}: {card['name']} ({card['ingredients_count']} ingredients)"
                self.log_test("Database Persistence Check", True, details, response_time)
            elif found_count > 0 and missing_count > 0:
                details = f"Found {found_count} cards, missing {missing_count} cards"
                self.log_test("Database Persistence Check", False, details, response_time)
            else:
                details = f"No tech cards found in database"
                self.log_test("Database Persistence Check", False, details, response_time)
                
        except Exception as e:
            self.log_test("Database Persistence Check", False, f"Exception: {str(e)}", 0.0)
    
    def save_test_artifacts(self):
        """Save test artifacts to JSON files"""
        try:
            # Save generation runs
            if any(key.startswith("gen_run_") for key in self.test_artifacts.keys()):
                gen_runs = {k: v for k, v in self.test_artifacts.items() if k.startswith("gen_run_")}
                with open("/app/gen_runs.json", "w", encoding="utf-8") as f:
                    json.dump(gen_runs, f, indent=2, ensure_ascii=False, default=str)
            
            # Save preflight result
            if "preflight_result" in self.test_artifacts:
                with open("/app/preflight.json", "w", encoding="utf-8") as f:
                    json.dump(self.test_artifacts["preflight_result"], f, indent=2, ensure_ascii=False, default=str)
            
            # Save XLSX validation results
            if "xlsx_validation" in self.test_artifacts:
                with open("/app/xlsx_checks.json", "w", encoding="utf-8") as f:
                    json.dump(self.test_artifacts["xlsx_validation"], f, indent=2, ensure_ascii=False, default=str)
            
            # Save article validation results
            if "article_validation" in self.test_artifacts:
                with open("/app/article_checks.json", "w", encoding="utf-8") as f:
                    json.dump(self.test_artifacts["article_validation"], f, indent=2, ensure_ascii=False, default=str)
            
            # Save database check results
            if "database_check" in self.test_artifacts:
                with open("/app/database_check.json", "w", encoding="utf-8") as f:
                    json.dump(self.test_artifacts["database_check"], f, indent=2, ensure_ascii=False, default=str)
            
            print("📁 Test artifacts saved to JSON files")
            
        except Exception as e:
            print(f"⚠️ Failed to save test artifacts: {e}")
    
    def print_summary(self):
        """Print comprehensive test summary"""
        print("\n" + "="*80)
        print("🎯 EXPORT SYSTEM MOCK CONTENT VALIDATION SUMMARY")
        print("="*80)
        
        total_tests = len(self.test_results)
        passed_tests = sum(1 for result in self.test_results if result["success"])
        failed_tests = total_tests - passed_tests
        success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
        
        print(f"\n📊 OVERALL RESULTS:")
        print(f"   Total Tests: {total_tests}")
        print(f"   Passed: {passed_tests} ✅")
        print(f"   Failed: {failed_tests} ❌")
        print(f"   Success Rate: {success_rate:.1f}%")
        
        # Critical findings
        print(f"\n🔍 CRITICAL FINDINGS:")
        
        # Mock content check
        xlsx_validation = self.test_artifacts.get("xlsx_validation", {})
        mock_signatures = xlsx_validation.get("mock_signatures", [])
        real_content = xlsx_validation.get("real_content", [])
        
        if mock_signatures:
            print(f"   ❌ MOCK CONTENT DETECTED: {len(mock_signatures)} mock signatures found")
            for signature in mock_signatures[:3]:
                print(f"      • {signature['file']} {signature['cell']}: {signature['pattern']}")
        else:
            print(f"   ✅ NO MOCK CONTENT: Export system using real data")
        
        if real_content:
            print(f"   ✅ REAL CONTENT FOUND: {len(real_content)} real dish names in exports")
        else:
            print(f"   ❌ NO REAL CONTENT: Generated dish names not found in exports")
        
        # Database persistence
        database_check = self.test_artifacts.get("database_check", {})
        found_cards = database_check.get("found_cards", [])
        
        if found_cards:
            print(f"   ✅ DATABASE PERSISTENCE: {len(found_cards)} tech cards saved to database")
        else:
            print(f"   ❌ DATABASE PERSISTENCE: Tech cards not saved to database")
        
        # Article formatting
        article_validation = self.test_artifacts.get("article_validation", {})
        articles_found = article_validation.get("articles_found", [])
        
        if articles_found:
            print(f"   ✅ ARTICLE FORMATTING: {len(articles_found)} properly formatted 5-digit articles")
        else:
            print(f"   ❌ ARTICLE FORMATTING: No properly formatted articles found")
        
        if failed_tests > 0:
            print(f"\n❌ FAILED TESTS:")
            for result in self.test_results:
                if not result["success"]:
                    print(f"   • {result['test']}: {result['details']}")
        
        print(f"\n✅ PASSED TESTS:")
        for result in self.test_results:
            if result["success"]:
                print(f"   • {result['test']}: {result['details']}")
        
        # Final verdict
        mock_content_eliminated = len(mock_signatures) == 0
        real_content_present = len(real_content) > 0
        database_working = len(found_cards) > 0
        
        print(f"\n🎯 FINAL VERDICT:")
        if mock_content_eliminated and real_content_present and database_working:
            print(f"   ✅ SUCCESS: Mock content issue has been COMPLETELY RESOLVED")
            print(f"   ✅ Export system is using real tech card data from database")
            print(f"   ✅ Generated tech cards are properly persisted and exported")
        else:
            print(f"   ❌ FAILURE: Mock content issue has NOT been resolved")
            print(f"   ❌ Export system still using mock/demo content")
            print(f"   ❌ Immediate action required to fix export pipeline")
        
        return mock_content_eliminated and real_content_present and database_working


async def main():
    """Main test execution"""
    print("🚀 Starting Export System Mock Content Validation Testing")
    print(f"Backend URL: {BACKEND_URL}")
    print(f"MongoDB URL: {MONGO_URL}")
    print(f"Database: {DB_NAME}")
    
    async with ExportMockContentTester() as tester:
        # Execute all test suites in sequence
        await tester.test_generate_real_tech_cards()
        await tester.test_preflight_with_real_ids()
        await tester.test_zip_export_with_real_ids()
        await tester.test_xlsx_content_validation()
        await tester.test_article_formatting_validation()
        await tester.test_database_persistence_check()
        
        # Save test artifacts
        tester.save_test_artifacts()
        
        # Print comprehensive summary
        success = tester.print_summary()
        
        if success:
            print(f"\n🎉 EXPORT MOCK CONTENT VALIDATION COMPLETED SUCCESSFULLY!")
            print(f"Real tech card data is now being used instead of mock content.")
        else:
            print(f"\n⚠️  EXPORT MOCK CONTENT VALIDATION FAILED")
            print(f"Mock content issue has NOT been resolved - immediate action required.")
        
        return success


if __name__ == "__main__":
    try:
        success = asyncio.run(main())
        sys.exit(0 if success else 1)
    except KeyboardInterrupt:
        print("\n⚠️ Testing interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Testing failed with exception: {e}")
        traceback.print_exc()
        sys.exit(1)