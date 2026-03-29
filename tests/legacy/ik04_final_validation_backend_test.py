#!/usr/bin/env python3
"""
IK-04/01 Backend XLSX Parser - FINAL VALIDATION TESTING
Проведи финальную валидацию исправлений для IK-04/01 парсера после фиксов схемы TechCardV2.

Focus Areas:
1. Schema Validation Fix Verification
2. Recalculation Integration Testing  
3. Complete Workflow Testing
4. Performance and Edge Cases
5. Response Structure Validation
6. Round-trip Validation Preparation
"""

import requests
import json
import time
import io
import openpyxl
from openpyxl.styles import Font, Alignment
import os
from datetime import datetime

# Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api"

class IK04FinalValidationTester:
    def __init__(self):
        self.results = []
        self.test_count = 0
        self.passed_count = 0
        
    def log_result(self, test_name: str, success: bool, details: str = "", response_data: dict = None):
        """Log test result"""
        self.test_count += 1
        if success:
            self.passed_count += 1
            
        result = {
            "test": test_name,
            "success": success,
            "details": details,
            "response_data": response_data
        }
        self.results.append(result)
        
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"{status} {test_name}")
        if details:
            print(f"    {details}")
        if not success and response_data:
            print(f"    Response: {json.dumps(response_data, indent=2, ensure_ascii=False)[:200]}...")
        print()

    def create_test_xlsx_basic(self) -> bytes:
        """Create basic test XLSX file with minimal data"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ТТК"
        
        # Headers (Russian)
        headers = [
            "Артикул блюда", "Наименование блюда", "Артикул продукта", 
            "Наименование продукта", "Брутто", "Потери %", "Нетто", 
            "Ед.", "Выход готового продукта", "Норма закладки", 
            "Метод списания", "Технология приготовления"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Test dish data
        dish_data = [
            ["DISH_BEEF_001", "Говядина тушеная с овощами", "BEEF001", "Говядина для тушения", 500, 15, 425, "г", 623, 1, "ASSEMBLE", "Нарезать мясо кубиками. Обжарить на сковороде 10 минут при 180°C. Добавить овощи и тушить 30 минут."],
            ["", "", "VEG002", "Лук репчатый", 100, 10, 90, "г", "", "", "", ""],
            ["", "", "VEG003", "Морковь", 80, 5, 76, "г", "", "", "", ""],
            ["", "", "OIL001", "Масло растительное", 30, 0, 30, "мл", "", "", "", ""],
            ["", "", "SPICE001", "Соль поваренная", 8, 0, 8, "г", "", "", "", ""]
        ]
        
        for row_idx, row_data in enumerate(dish_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def create_test_xlsx_comprehensive(self) -> bytes:
        """Create comprehensive test XLSX with full dataset"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ТТК"
        
        # Headers
        headers = [
            "Артикул блюда", "Наименование блюда", "Артикул продукта", 
            "Наименование продукта", "Брутто", "Потери %", "Нетто", 
            "Ед.", "Выход готового продукта", "Норма закладки", 
            "Метод списания", "Технология приготовления"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Comprehensive dish data with different units and complex technology
        dish_data = [
            ["DISH_COMPLEX_001", "Говядина тушеная с овощами", "BEEF001", "Говядина для тушения", 1.5, 15, 1.275, "кг", 2500, 4, "ASSEMBLE", "1. Подготовить мясо: нарезать кубиками 3x3 см, промыть холодной водой. 2. Обжарить мясо на растительном масле при температуре 180°C в течение 10-12 минут до золотистой корочки. 3. Добавить нарезанный лук и морковь, тушить 5 минут. 4. Залить горячей водой, добавить специи, тушить под крышкой 45-60 минут при температуре 160°C. 5. За 10 минут до готовности добавить соль и перец по вкусу. 6. Подавать горячим с гарниром."],
            ["", "", "VEG002", "Лук репчатый", 300, 10, 270, "г", "", "", "", ""],
            ["", "", "VEG003", "Морковь", 250, 5, 237.5, "г", "", "", "", ""],
            ["", "", "OIL001", "Масло растительное", 0.05, 0, 0.05, "л", "", "", "", ""],
            ["", "", "SPICE001", "Соль поваренная", 15, 0, 15, "г", "", "", "", ""],
            ["", "", "SPICE002", "Перец черный молотый", 3, 0, 3, "г", "", "", "", ""],
            ["", "", "WATER001", "Вода питьевая", 500, 0, 500, "мл", "", "", "", ""],
            ["", "", "EGG001", "Яйца куриные", 2, 0, 2, "шт", "", "", "", ""]
        ]
        
        for row_idx, row_data in enumerate(dish_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def create_test_xlsx_edge_case_minimal(self) -> bytes:
        """Create minimal edge case XLSX"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Minimal headers
        headers = ["Наименование продукта", "Брутто", "Нетто", "Ед."]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Single ingredient
        ws.cell(row=2, column=1, value="Мука пшеничная")
        ws.cell(row=2, column=2, value=100)
        ws.cell(row=2, column=3, value=100)
        ws.cell(row=2, column=4, value="г")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def test_1_schema_validation_fix_verification(self):
        """1. Schema Validation Fix Verification"""
        print("🎯 TEST 1: Schema Validation Fix Verification")
        
        # Test with basic XLSX file
        xlsx_data = self.create_test_xlsx_basic()
        
        try:
            files = {'file': ('test_basic.xlsx', xlsx_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(f"{API_BASE}/v1/iiko/import/ttk.xlsx", files=files, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                
                # Check TechCardV2 validation passes
                validation_success = data.get('status') in ['success', 'draft']
                
                # Check meta.code field is removed (no extra_forbidden error)
                validation_issues = [issue for issue in data.get('issues', []) if 'extra_forbidden' in issue.get('msg', '').lower()]
                no_extra_forbidden = len(validation_issues) == 0
                
                # Check minimum 3 process steps generated
                techcard = data.get('techcard', {})
                process_steps = techcard.get('process', [])
                min_3_steps = len(process_steps) >= 3
                
                # Check valid enum values for sources
                nutrition_meta = techcard.get('nutritionMeta', {})
                cost_meta = techcard.get('costMeta', {})
                valid_nutrition_source = nutrition_meta.get('source') in ['bootstrap', 'catalog', 'usda', 'llm', 'csv', 'none']
                valid_cost_source = cost_meta.get('source') in ['catalog', 'csv', 'llm', 'none']
                
                all_checks_pass = validation_success and no_extra_forbidden and min_3_steps and valid_nutrition_source and valid_cost_source
                
                details = f"Status: {data.get('status')}, Process steps: {len(process_steps)}, Nutrition source: {nutrition_meta.get('source')}, Cost source: {cost_meta.get('source')}"
                if validation_issues:
                    details += f", Extra forbidden errors: {len(validation_issues)}"
                
                self.log_result("Schema Validation Fix", all_checks_pass, details, data)
                
            else:
                self.log_result("Schema Validation Fix", False, f"HTTP {response.status_code}: {response.text[:200]}")
                
        except Exception as e:
            self.log_result("Schema Validation Fix", False, f"Exception: {str(e)}")

    def test_2_recalculation_integration(self):
        """2. Recalculation Integration Testing"""
        print("🎯 TEST 2: Recalculation Integration Testing")
        
        xlsx_data = self.create_test_xlsx_basic()
        
        try:
            files = {'file': ('test_recalc.xlsx', xlsx_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(f"{API_BASE}/v1/iiko/import/ttk.xlsx", files=files, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                techcard = data.get('techcard', {})
                
                # Check recalculation occurred
                cost_data = techcard.get('cost', {})
                nutrition_data = techcard.get('nutrition', {})
                
                # Check cost calculator worked
                cost_calculated = cost_data.get('rawCost') is not None or cost_data.get('costPerPortion') is not None
                
                # Check nutrition calculator worked  
                nutrition_calculated = nutrition_data.get('per100g', {}).get('kcal', 0) > 0
                
                # Check SKU preservation through recalc
                ingredients = techcard.get('ingredients', [])
                skus_preserved = any(ing.get('skuId') for ing in ingredients)
                
                # Check final TechCardV2 structure
                has_required_fields = all(field in techcard for field in ['meta', 'ingredients', 'process', 'cost', 'nutrition'])
                
                all_checks_pass = cost_calculated and has_required_fields and skus_preserved
                
                details = f"Cost calculated: {cost_calculated}, Nutrition calculated: {nutrition_calculated}, SKUs preserved: {skus_preserved}, Required fields: {has_required_fields}"
                
                self.log_result("Recalculation Integration", all_checks_pass, details, data)
                
            else:
                self.log_result("Recalculation Integration", False, f"HTTP {response.status_code}: {response.text[:200]}")
                
        except Exception as e:
            self.log_result("Recalculation Integration", False, f"Exception: {str(e)}")

    def test_3_complete_workflow_testing(self):
        """3. Complete Workflow Testing"""
        print("🎯 TEST 3: Complete Workflow Testing")
        
        xlsx_data = self.create_test_xlsx_comprehensive()
        
        try:
            start_time = time.time()
            files = {'file': ('test_comprehensive.xlsx', xlsx_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(f"{API_BASE}/v1/iiko/import/ttk.xlsx", files=files, timeout=30)
            processing_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                techcard = data.get('techcard', {})
                
                # Check workflow completion without errors
                workflow_success = data.get('status') == 'success'
                
                # Check dish name preserved
                dish_name = techcard.get('meta', {}).get('title', '')
                correct_dish_name = 'говядина тушеная с овощами' in dish_name.lower()
                
                # Check ingredients with different units processed
                ingredients = techcard.get('ingredients', [])
                has_beef = any('говядина' in ing.get('name', '').lower() for ing in ingredients)
                has_vegetables = any('лук' in ing.get('name', '').lower() or 'морковь' in ing.get('name', '').lower() for ing in ingredients)
                has_oil = any('масло' in ing.get('name', '').lower() for ing in ingredients)
                
                # Check unit conversions worked
                unit_conversions_ok = all(ing.get('unit') in ['g', 'ml', 'pcs'] for ing in ingredients)
                
                # Check technology processing
                process_steps = techcard.get('process', [])
                technology_processed = len(process_steps) >= 3
                has_temperature = any(step.get('temp_c') for step in process_steps)
                has_time = any(step.get('time_min') for step in process_steps)
                
                # Check performance (≤3 seconds)
                performance_ok = processing_time <= 3.0
                
                all_checks_pass = (workflow_success and correct_dish_name and has_beef and 
                                 has_vegetables and has_oil and unit_conversions_ok and 
                                 technology_processed and performance_ok)
                
                details = f"Status: {data.get('status')}, Processing time: {processing_time:.2f}s, Ingredients: {len(ingredients)}, Process steps: {len(process_steps)}, Temperature extracted: {has_temperature}, Time extracted: {has_time}"
                
                self.log_result("Complete Workflow", all_checks_pass, details, data)
                
            else:
                self.log_result("Complete Workflow", False, f"HTTP {response.status_code}: {response.text[:200]}")
                
        except Exception as e:
            self.log_result("Complete Workflow", False, f"Exception: {str(e)}")

    def test_4_performance_and_edge_cases(self):
        """4. Performance and Edge Cases"""
        print("🎯 TEST 4: Performance and Edge Cases")
        
        # Test 4a: Performance with large file
        try:
            # Create large XLSX (50+ ingredients)
            wb = openpyxl.Workbook()
            ws = wb.active
            
            headers = ["Наименование продукта", "Брутто", "Нетто", "Ед."]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add 50 ingredients
            for i in range(50):
                ws.cell(row=i+2, column=1, value=f"Ингредиент {i+1}")
                ws.cell(row=i+2, column=2, value=100 + i)
                ws.cell(row=i+2, column=3, value=95 + i)
                ws.cell(row=i+2, column=4, value="г")
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            large_xlsx = buffer.getvalue()
            
            start_time = time.time()
            files = {'file': ('test_large.xlsx', large_xlsx, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(f"{API_BASE}/v1/iiko/import/ttk.xlsx", files=files, timeout=30)
            processing_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                performance_ok = processing_time <= 3.0
                ingredients_count = len(data.get('techcard', {}).get('ingredients', []))
                
                self.log_result("Performance Test (50 ingredients)", performance_ok, 
                              f"Processing time: {processing_time:.2f}s, Ingredients processed: {ingredients_count}")
            else:
                self.log_result("Performance Test (50 ingredients)", False, f"HTTP {response.status_code}")
                
        except Exception as e:
            self.log_result("Performance Test (50 ingredients)", False, f"Exception: {str(e)}")
        
        # Test 4b: Edge case - minimal data
        try:
            minimal_xlsx = self.create_test_xlsx_edge_case_minimal()
            
            files = {'file': ('test_minimal.xlsx', minimal_xlsx, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(f"{API_BASE}/v1/iiko/import/ttk.xlsx", files=files, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                minimal_success = data.get('status') in ['success', 'draft']
                has_ingredients = len(data.get('techcard', {}).get('ingredients', [])) > 0
                
                self.log_result("Edge Case - Minimal Data", minimal_success and has_ingredients, 
                              f"Status: {data.get('status')}, Ingredients: {len(data.get('techcard', {}).get('ingredients', []))}")
            else:
                self.log_result("Edge Case - Minimal Data", False, f"HTTP {response.status_code}")
                
        except Exception as e:
            self.log_result("Edge Case - Minimal Data", False, f"Exception: {str(e)}")
        
        # Test 4c: Edge case - unknown units
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            
            headers = ["Наименование продукта", "Брутто", "Нетто", "Ед."]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add ingredient with unknown unit
            ws.cell(row=2, column=1, value="Специя экзотическая")
            ws.cell(row=2, column=2, value=50)
            ws.cell(row=2, column=3, value=50)
            ws.cell(row=2, column=4, value="пинч")  # Unknown unit
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            unknown_unit_xlsx = buffer.getvalue()
            
            files = {'file': ('test_unknown_unit.xlsx', unknown_unit_xlsx, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(f"{API_BASE}/v1/iiko/import/ttk.xlsx", files=files, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                handles_unknown_unit = data.get('status') in ['success', 'draft']
                has_unit_warning = any('unitUnknown' in issue.get('code', '') for issue in data.get('issues', []))
                
                self.log_result("Edge Case - Unknown Units", handles_unknown_unit, 
                              f"Status: {data.get('status')}, Unit warning: {has_unit_warning}")
            else:
                self.log_result("Edge Case - Unknown Units", False, f"HTTP {response.status_code}")
                
        except Exception as e:
            self.log_result("Edge Case - Unknown Units", False, f"Exception: {str(e)}")

    def test_5_response_structure_validation(self):
        """5. Response Structure Validation"""
        print("🎯 TEST 5: Response Structure Validation")
        
        xlsx_data = self.create_test_xlsx_basic()
        
        try:
            files = {'file': ('test_structure.xlsx', xlsx_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(f"{API_BASE}/v1/iiko/import/ttk.xlsx", files=files, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                
                # Check required response fields
                has_status = 'status' in data
                has_techcard = 'techcard' in data
                has_issues = 'issues' in data
                has_meta = 'meta' in data
                
                # Check TechCardV2 required fields
                techcard = data.get('techcard', {})
                required_tc_fields = ['meta', 'portions', 'yield', 'ingredients', 'process', 'storage', 'nutrition', 'cost']
                has_all_tc_fields = all(field in techcard for field in required_tc_fields)
                
                # Check meta structure
                meta = data.get('meta', {})
                has_parsed_rows = 'parsed_rows' in meta
                has_filename = 'filename' in meta
                has_source = 'source' in meta
                
                # Check issues classification
                issues = data.get('issues', [])
                issues_properly_classified = all(
                    issue.get('level') in ['info', 'warning', 'error'] and 
                    'code' in issue and 'msg' in issue 
                    for issue in issues
                )
                
                # Check dish code preservation (in printNotes as workaround)
                print_notes = techcard.get('printNotes', [])
                dish_code_preserved = any('артикул блюда' in note.lower() for note in print_notes)
                
                all_checks_pass = (has_status and has_techcard and has_issues and has_meta and 
                                 has_all_tc_fields and has_parsed_rows and has_filename and 
                                 has_source and issues_properly_classified)
                
                details = f"Response fields: {has_status and has_techcard and has_issues and has_meta}, TechCard fields: {has_all_tc_fields}, Meta fields: {has_parsed_rows and has_filename and has_source}, Issues classified: {issues_properly_classified}, Dish code preserved: {dish_code_preserved}"
                
                self.log_result("Response Structure Validation", all_checks_pass, details, data)
                
            else:
                self.log_result("Response Structure Validation", False, f"HTTP {response.status_code}: {response.text[:200]}")
                
        except Exception as e:
            self.log_result("Response Structure Validation", False, f"Exception: {str(e)}")

    def test_6_round_trip_validation_prep(self):
        """6. Round-trip Validation Preparation"""
        print("🎯 TEST 6: Round-trip Validation Preparation")
        
        xlsx_data = self.create_test_xlsx_basic()
        
        try:
            # Step 1: Import XLSX to TechCardV2
            files = {'file': ('test_roundtrip.xlsx', xlsx_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            import_response = requests.post(f"{API_BASE}/v1/iiko/import/ttk.xlsx", files=files, timeout=30)
            
            if import_response.status_code == 200:
                import_data = import_response.json()
                
                if import_data.get('status') == 'success':
                    techcard = import_data.get('techcard', {})
                    
                    # Step 2: Try to export the imported TechCardV2 back to XLSX
                    try:
                        export_response = requests.post(
                            f"{API_BASE}/v1/techcards.v2/export/iiko.xlsx",
                            json=techcard,
                            timeout=30
                        )
                        
                        if export_response.status_code == 200:
                            # Check export response
                            content_type = export_response.headers.get('content-type', '')
                            is_xlsx = 'spreadsheet' in content_type
                            has_content = len(export_response.content) > 0
                            
                            # Check key data preservation
                            original_ingredients = techcard.get('ingredients', [])
                            original_dish_name = techcard.get('meta', {}).get('title', '')
                            
                            round_trip_success = is_xlsx and has_content
                            
                            details = f"Import status: {import_data.get('status')}, Export content-type: {content_type}, Export size: {len(export_response.content)} bytes, Ingredients: {len(original_ingredients)}, Dish: {original_dish_name}"
                            
                            self.log_result("Round-trip Validation Prep", round_trip_success, details)
                            
                        else:
                            self.log_result("Round-trip Validation Prep", False, f"Export failed: HTTP {export_response.status_code}")
                            
                    except Exception as export_error:
                        self.log_result("Round-trip Validation Prep", False, f"Export exception: {str(export_error)}")
                        
                else:
                    self.log_result("Round-trip Validation Prep", False, f"Import failed with status: {import_data.get('status')}")
                    
            else:
                self.log_result("Round-trip Validation Prep", False, f"Import HTTP {import_response.status_code}")
                
        except Exception as e:
            self.log_result("Round-trip Validation Prep", False, f"Exception: {str(e)}")

    def run_all_tests(self):
        """Run all final validation tests"""
        print("🎯 IK-04/01 BACKEND XLSX PARSER - FINAL VALIDATION TESTING")
        print("=" * 80)
        print()
        
        # Run all test categories
        self.test_1_schema_validation_fix_verification()
        self.test_2_recalculation_integration()
        self.test_3_complete_workflow_testing()
        self.test_4_performance_and_edge_cases()
        self.test_5_response_structure_validation()
        self.test_6_round_trip_validation_prep()
        
        # Summary
        print("=" * 80)
        print(f"🎯 FINAL VALIDATION RESULTS: {self.passed_count}/{self.test_count} tests passed")
        
        if self.passed_count == self.test_count:
            print("✅ ALL TESTS PASSED - IK-04/01 готов к продакшену!")
        else:
            print("❌ SOME TESTS FAILED - требуются дополнительные исправления")
            
        print()
        
        # Detailed results
        for result in self.results:
            status = "✅" if result["success"] else "❌"
            print(f"{status} {result['test']}: {result['details']}")
        
        return self.passed_count == self.test_count

if __name__ == "__main__":
    tester = IK04FinalValidationTester()
    success = tester.run_all_tests()
    exit(0 if success else 1)