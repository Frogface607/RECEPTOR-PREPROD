#!/usr/bin/env python3
"""
Excel File Validation Test for iiko Import Reliability
Validates the actual Excel file structure and content
"""

import requests
import json
import os
import io
import zipfile
from openpyxl import load_workbook

# Backend URL from environment
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'http://localhost:8001')
API_BASE = f"{BACKEND_URL}/api/v1"

def create_test_techcard(title: str) -> dict:
    """Create a test techcard"""
    ingredients = [
        {
            "name": "Куриное филе",
            "brutto_g": 150.0,
            "netto_g": 120.0,
            "loss_pct": 20.0,
            "unit": "g",
            "skuId": "test-chicken-sku-001"
        },
        {
            "name": "Картофель",
            "brutto_g": 200.0,
            "netto_g": 180.0,
            "loss_pct": 10.0,
            "unit": "g",
            "skuId": "test-potato-sku-002"
        }
    ]
    
    total_netto = sum(ing["netto_g"] for ing in ingredients)
    
    return {
        "meta": {
            "title": title,
            "id": f"test-{title.lower().replace(' ', '-')}",
            "version": "2.0",
            "createdAt": "2025-01-27T10:00:00Z"
        },
        "ingredients": ingredients,
        "yield": {
            "perPortion_g": total_netto,
            "perBatch_g": total_netto
        },
        "portions": 1,
        "process": [
            {
                "n": 1,
                "action": "Нарезать куриное филе кубиками",
                "time_min": 5,
                "equipment": ["нож", "доска"]
            },
            {
                "n": 2,
                "action": "Обжарить курицу на сковороде",
                "temp_c": 180,
                "time_min": 10,
                "equipment": ["сковорода"]
            },
            {
                "n": 3,
                "action": "Добавить овощи и тушить",
                "temp_c": 160,
                "time_min": 15,
                "equipment": ["сковорода"]
            }
        ],
        "storage": {
            "conditions": "Хранить в холодильнике при температуре +2...+6°C",
            "shelfLife_hours": 24.0,
            "servingTemp_c": 65.0
        },
        "nutrition": {
            "per100g": {
                "kcal": 150.0,
                "proteins_g": 20.0,
                "fats_g": 5.0,
                "carbs_g": 10.0
            },
            "perPortion": {
                "kcal": 150.0 * total_netto / 100,
                "proteins_g": 20.0 * total_netto / 100,
                "fats_g": 5.0 * total_netto / 100,
                "carbs_g": 10.0 * total_netto / 100
            }
        },
        "cost": {
            "rawCost": 100.0,
            "costPerPortion": 100.0,
            "markup_pct": 200.0,
            "vat_pct": 20.0
        }
    }

def test_excel_file_validation():
    """Test Excel file structure and content validation"""
    print("🔍 TESTING EXCEL FILE VALIDATION")
    
    # Create test data
    techcards = [
        create_test_techcard("Тестовое блюдо с кодами")
    ]
    
    # Generate dish codes first
    dish_names = [card["meta"]["title"] for card in techcards]
    generate_payload = {
        "dish_names": dish_names,
        "organization_id": "default",
        "width": 5
    }
    
    session = requests.Session()
    generate_response = session.post(
        f"{API_BASE}/techcards.v2/dish-codes/generate",
        json=generate_payload,
        timeout=10
    )
    
    if generate_response.status_code != 200:
        print(f"❌ Failed to generate dish codes: {generate_response.status_code}")
        return
    
    dish_codes_mapping = generate_response.json().get("generated_codes", {})
    print(f"✅ Generated dish codes: {dish_codes_mapping}")
    
    # Test dual export
    export_payload = {
        "techcards": techcards,
        "export_options": {
            "use_product_codes": True,
            "dish_codes_mapping": dish_codes_mapping
        },
        "organization_id": "default",
        "user_email": "test@example.com"
    }
    
    export_response = session.post(
        f"{API_BASE}/techcards.v2/export/enhanced-dual/iiko.xlsx",
        json=export_payload,
        timeout=20
    )
    
    if export_response.status_code != 200:
        print(f"❌ Export failed: {export_response.status_code}")
        return
    
    print("✅ Export successful, validating ZIP contents...")
    
    # Validate ZIP contents
    zip_buffer = io.BytesIO(export_response.content)
    with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
        file_list = zip_file.namelist()
        print(f"📁 ZIP contains files: {file_list}")
        
        # Find and validate Dish-Skeletons.xlsx
        skeleton_file = None
        ttk_file = None
        
        for f in file_list:
            if 'Dish-Skeletons' in f:
                skeleton_file = f
            elif 'iiko_TTK' in f:
                ttk_file = f
        
        if skeleton_file:
            print(f"\n🍽️ VALIDATING DISH-SKELETONS FILE: {skeleton_file}")
            skeleton_data = zip_file.read(skeleton_file)
            skeleton_workbook = load_workbook(io.BytesIO(skeleton_data))
            skeleton_ws = skeleton_workbook.active
            
            # Check headers
            headers = [cell.value for cell in skeleton_ws[1]]
            print(f"📋 Headers: {headers}")
            
            expected_headers = ["Артикул", "Наименование", "Тип", "Ед. выпуска", "Выход"]
            if all(h in headers for h in expected_headers):
                print("✅ All required headers present")
            else:
                print(f"❌ Missing headers. Expected: {expected_headers}")
            
            # Check data rows
            for row_num, row in enumerate(skeleton_ws.iter_rows(min_row=2, values_only=True), 2):
                if row[0]:  # Has data
                    print(f"📊 Row {row_num}: Артикул={row[0]}, Наименование={row[1]}, Тип={row[2]}, Выход={row[4]}")
                    
                    # Validate dish code format
                    dish_code = str(row[0])
                    if len(dish_code) == 5 and dish_code.isdigit():
                        print(f"✅ Dish code format valid: {dish_code}")
                    else:
                        print(f"❌ Invalid dish code format: {dish_code}")
        
        if ttk_file:
            print(f"\n📋 VALIDATING TTK FILE: {ttk_file}")
            ttk_data = zip_file.read(ttk_file)
            ttk_workbook = load_workbook(io.BytesIO(ttk_data))
            ttk_ws = ttk_workbook.active
            
            # Check headers
            headers = [cell.value for cell in ttk_ws[1]]
            print(f"📋 Headers: {headers}")
            
            expected_headers = ["Артикул блюда", "Наименование блюда", "Артикул продукта", "Наименование продукта"]
            if all(h in headers for h in expected_headers):
                print("✅ All required headers present")
            else:
                print(f"❌ Missing headers. Expected: {expected_headers}")
            
            # Check cell formatting for code columns
            dish_code_col = headers.index("Артикул блюда") + 1 if "Артикул блюда" in headers else 1
            product_code_col = headers.index("Артикул продукта") + 1 if "Артикул продукта" in headers else 3
            
            # Check first data row formatting
            dish_code_cell = ttk_ws.cell(row=2, column=dish_code_col)
            product_code_cell = ttk_ws.cell(row=2, column=product_code_col)
            
            print(f"🔢 Dish code cell format: {dish_code_cell.number_format}")
            print(f"🔢 Product code cell format: {product_code_cell.number_format}")
            
            if dish_code_cell.number_format == '@' and product_code_cell.number_format == '@':
                print("✅ Code columns properly formatted as text (@)")
            else:
                print("❌ Code columns not formatted as text")
            
            # Check data rows
            for row_num, row in enumerate(ttk_ws.iter_rows(min_row=2, values_only=True), 2):
                if row[0]:  # Has data
                    print(f"📊 Row {row_num}: Блюдо={row[0]}, Продукт={row[2]}, Брутто={row[4]}, Нетто={row[6]}")
                    break  # Just check first row
    
    print("\n🎯 EXCEL VALIDATION COMPLETE")

if __name__ == "__main__":
    test_excel_file_validation()