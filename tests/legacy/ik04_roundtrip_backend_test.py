#!/usr/bin/env python3
"""
IK-04/03 Round-trip Golden: iiko XLSX ↔ TechCardV2 Backend Testing
Comprehensive testing of XLSX import/export round-trip workflow with fixtures
"""

import os
import sys
import json
import time
import requests
from pathlib import Path
from typing import Dict, List, Any, Optional
import tempfile

# Add backend to path for imports
sys.path.append('/app/backend')

class IK04RoundTripTester:
    """Comprehensive tester for IK-04/03 round-trip golden workflow"""
    
    def __init__(self):
        # Get backend URL from environment
        self.backend_url = os.getenv('REACT_APP_BACKEND_URL', 'http://localhost:8001')
        if not self.backend_url.startswith('http'):
            self.backend_url = f"https://{self.backend_url}"
        
        # Test fixtures paths
        self.fixtures_dir = Path('/app/tests/fixtures/iiko_xlsx')
        self.test_fixtures = ['hot.xlsx', 'cold.xlsx', 'sauce.xlsx']
        
        # Results tracking
        self.results = {
            'total_tests': 0,
            'passed_tests': 0,
            'failed_tests': 0,
            'test_details': []
        }
        
        print(f"🎯 IK-04/03 Round-trip Golden Tester initialized")
        print(f"Backend URL: {self.backend_url}")
        print(f"Fixtures: {self.test_fixtures}")
    
    def run_all_tests(self):
        """Run comprehensive round-trip testing suite"""
        print("\n" + "="*80)
        print("🚀 STARTING IK-04/03 ROUND-TRIP GOLDEN TESTS")
        print("="*80)
        
        # Test 1: XLSX Import API Endpoint
        self.test_xlsx_import_endpoint()
        
        # Test 2: XLSX Export API Endpoint  
        self.test_xlsx_export_endpoint()
        
        # Test 3: Round-trip Validation with Fixtures
        self.test_round_trip_with_fixtures()
        
        # Test 4: Technology Parsing Validation
        self.test_technology_parsing()
        
        # Test 5: Unit Conversion Validation
        self.test_unit_conversions()
        
        # Test 6: SKU Preservation Testing
        self.test_sku_preservation()
        
        # Test 7: Performance Testing
        self.test_performance()
        
        # Print final results
        self.print_final_results()
    
    def test_xlsx_import_endpoint(self):
        """Test POST /api/v1/iiko/import/ttk.xlsx endpoint"""
        print("\n📥 Testing XLSX Import API Endpoint...")
        
        for fixture_name in self.test_fixtures:
            fixture_path = self.fixtures_dir / fixture_name
            if not fixture_path.exists():
                self.record_test_result(
                    f"XLSX Import - {fixture_name}",
                    False,
                    f"Fixture file not found: {fixture_path}"
                )
                continue
            
            try:
                # Read fixture file
                with open(fixture_path, 'rb') as f:
                    file_content = f.read()
                
                # Test import endpoint
                url = f"{self.backend_url}/api/v1/iiko/import/ttk.xlsx"
                files = {'file': (fixture_name, file_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                
                start_time = time.time()
                response = requests.post(url, files=files, timeout=30)
                duration = time.time() - start_time
                
                if response.status_code == 200:
                    data = response.json()
                    
                    # Validate response structure
                    required_fields = ['status', 'techcard', 'issues', 'meta']
                    missing_fields = [field for field in required_fields if field not in data]
                    
                    if missing_fields:
                        self.record_test_result(
                            f"XLSX Import - {fixture_name}",
                            False,
                            f"Missing response fields: {missing_fields}"
                        )
                        continue
                    
                    # Validate TechCardV2 structure
                    techcard = data['techcard']
                    if not self.validate_techcard_structure(techcard):
                        self.record_test_result(
                            f"XLSX Import - {fixture_name}",
                            False,
                            "Invalid TechCardV2 structure"
                        )
                        continue
                    
                    # Check parsing statistics
                    meta = data['meta']
                    parsed_rows = meta.get('parsed_rows', 0)
                    ingredients_count = len(techcard.get('ingredients', []))
                    process_steps = len(techcard.get('process', []))
                    
                    # Validate requirements
                    success = True
                    issues = []
                    
                    if parsed_rows < 1:
                        success = False
                        issues.append("No rows parsed")
                    
                    if ingredients_count < 1:
                        success = False
                        issues.append("No ingredients found")
                    
                    if process_steps < 3:
                        success = False
                        issues.append(f"Process steps < 3 (found {process_steps})")
                    
                    if duration > 3.0:
                        success = False
                        issues.append(f"Performance: {duration:.2f}s > 3s limit")
                    
                    # Check that issues are warnings only, no errors
                    error_issues = [issue for issue in data['issues'] if issue.get('level') == 'error']
                    if error_issues:
                        success = False
                        issues.append(f"Found {len(error_issues)} error-level issues")
                    
                    result_msg = f"✅ Parsed {parsed_rows} rows, {ingredients_count} ingredients, {process_steps} steps in {duration:.2f}s"
                    if not success:
                        result_msg = f"❌ Issues: {'; '.join(issues)}"
                    
                    self.record_test_result(
                        f"XLSX Import - {fixture_name}",
                        success,
                        result_msg,
                        {
                            'parsed_rows': parsed_rows,
                            'ingredients_count': ingredients_count,
                            'process_steps': process_steps,
                            'duration': duration,
                            'status': data['status'],
                            'issues_count': len(data['issues'])
                        }
                    )
                    
                else:
                    self.record_test_result(
                        f"XLSX Import - {fixture_name}",
                        False,
                        f"HTTP {response.status_code}: {response.text[:200]}"
                    )
                    
            except Exception as e:
                self.record_test_result(
                    f"XLSX Import - {fixture_name}",
                    False,
                    f"Exception: {str(e)}"
                )
    
    def test_xlsx_export_endpoint(self):
        """Test POST /api/v1/techcards.v2/export/iiko.xlsx endpoint"""
        print("\n📤 Testing XLSX Export API Endpoint...")
        
        # First, generate a sample TechCardV2 for export testing
        sample_techcard = self.create_sample_techcard()
        
        try:
            url = f"{self.backend_url}/api/v1/techcards.v2/export/iiko.xlsx"
            headers = {'Content-Type': 'application/json'}
            
            start_time = time.time()
            response = requests.post(url, json=sample_techcard, headers=headers, timeout=30)
            duration = time.time() - start_time
            
            if response.status_code == 200:
                # Validate response headers
                content_type = response.headers.get('content-type', '')
                content_disposition = response.headers.get('content-disposition', '')
                
                success = True
                issues = []
                
                if 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' not in content_type:
                    success = False
                    issues.append(f"Wrong content-type: {content_type}")
                
                if 'attachment' not in content_disposition or 'filename' not in content_disposition:
                    success = False
                    issues.append(f"Invalid content-disposition: {content_disposition}")
                
                # Check file size (should be reasonable)
                file_size = len(response.content)
                if file_size < 1000:  # Too small
                    success = False
                    issues.append(f"File too small: {file_size} bytes")
                elif file_size > 1000000:  # Too large (>1MB)
                    success = False
                    issues.append(f"File too large: {file_size} bytes")
                
                if duration > 3.0:
                    success = False
                    issues.append(f"Performance: {duration:.2f}s > 3s limit")
                
                # Try to validate XLSX structure
                try:
                    import openpyxl
                    from io import BytesIO
                    
                    wb = openpyxl.load_workbook(BytesIO(response.content))
                    ws = wb.active
                    
                    # Check basic structure
                    if ws.max_row < 2:  # Should have header + at least 1 data row
                        success = False
                        issues.append("XLSX has insufficient rows")
                    
                    if ws.max_column < 10:  # Should have multiple columns
                        success = False
                        issues.append("XLSX has insufficient columns")
                    
                except Exception as xlsx_error:
                    success = False
                    issues.append(f"XLSX validation error: {str(xlsx_error)}")
                
                result_msg = f"✅ Generated {file_size} bytes XLSX in {duration:.2f}s"
                if not success:
                    result_msg = f"❌ Issues: {'; '.join(issues)}"
                
                self.record_test_result(
                    "XLSX Export",
                    success,
                    result_msg,
                    {
                        'file_size': file_size,
                        'duration': duration,
                        'content_type': content_type
                    }
                )
                
            else:
                self.record_test_result(
                    "XLSX Export",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}"
                )
                
        except Exception as e:
            self.record_test_result(
                "XLSX Export",
                False,
                f"Exception: {str(e)}"
            )
    
    def test_round_trip_with_fixtures(self):
        """Test complete round-trip: Import → Export → Import cycle"""
        print("\n🔄 Testing Round-trip Validation with Fixtures...")
        
        for fixture_name in self.test_fixtures:
            fixture_path = self.fixtures_dir / fixture_name
            if not fixture_path.exists():
                continue
            
            try:
                print(f"  Testing round-trip for {fixture_name}...")
                
                # Step 1: Import original XLSX
                with open(fixture_path, 'rb') as f:
                    original_content = f.read()
                
                import_url = f"{self.backend_url}/api/v1/iiko/import/ttk.xlsx"
                files = {'file': (fixture_name, original_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                
                import_response = requests.post(import_url, files=files, timeout=30)
                if import_response.status_code != 200:
                    self.record_test_result(
                        f"Round-trip {fixture_name} - Import",
                        False,
                        f"Import failed: HTTP {import_response.status_code}"
                    )
                    continue
                
                import_data = import_response.json()
                original_techcard = import_data['techcard']
                
                # Step 2: Export to XLSX
                export_url = f"{self.backend_url}/api/v1/techcards.v2/export/iiko.xlsx"
                export_response = requests.post(export_url, json=original_techcard, timeout=30)
                
                if export_response.status_code != 200:
                    self.record_test_result(
                        f"Round-trip {fixture_name} - Export",
                        False,
                        f"Export failed: HTTP {export_response.status_code}"
                    )
                    continue
                
                exported_xlsx = export_response.content
                
                # Step 3: Import exported XLSX
                files2 = {'file': (f'exported_{fixture_name}', exported_xlsx, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                reimport_response = requests.post(import_url, files=files2, timeout=30)
                
                if reimport_response.status_code != 200:
                    self.record_test_result(
                        f"Round-trip {fixture_name} - Re-import",
                        False,
                        f"Re-import failed: HTTP {reimport_response.status_code}"
                    )
                    continue
                
                reimport_data = reimport_response.json()
                reimported_techcard = reimport_data['techcard']
                
                # Step 4: Compare original vs reimported
                comparison_result = self.compare_techcards(original_techcard, reimported_techcard, fixture_name)
                
                self.record_test_result(
                    f"Round-trip {fixture_name}",
                    comparison_result['success'],
                    comparison_result['message'],
                    comparison_result['details']
                )
                
            except Exception as e:
                self.record_test_result(
                    f"Round-trip {fixture_name}",
                    False,
                    f"Exception: {str(e)}"
                )
    
    def test_technology_parsing(self):
        """Test technology parsing with temperature and time extraction"""
        print("\n🔧 Testing Technology Parsing...")
        
        # Create test XLSX with technology text
        test_cases = [
            {
                'name': 'Temperature Extraction',
                'technology': 'Разогреть духовку до 180°C. Выпекать 25 минут при 200°C.',
                'expected_temps': [180, 200],
                'expected_times': [25]
            },
            {
                'name': 'Time Extraction', 
                'technology': 'Варить 15 мин, затем тушить 30 минут на медленном огне.',
                'expected_temps': [],
                'expected_times': [15, 30]
            },
            {
                'name': 'Mixed Units',
                'technology': 'Готовить 1 час при 160°C, затем 10 мин при 220°F.',
                'expected_temps': [160, 104],  # 220°F ≈ 104°C
                'expected_times': [60, 10]  # 1 hour = 60 min
            }
        ]
        
        for test_case in test_cases:
            try:
                # Create minimal XLSX with technology text
                test_xlsx = self.create_test_xlsx_with_technology(test_case['technology'])
                
                # Import and check parsing
                import_url = f"{self.backend_url}/api/v1/iiko/import/ttk.xlsx"
                files = {'file': ('test_tech.xlsx', test_xlsx, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                
                response = requests.post(import_url, files=files, timeout=30)
                if response.status_code != 200:
                    self.record_test_result(
                        f"Technology Parsing - {test_case['name']}",
                        False,
                        f"Import failed: HTTP {response.status_code}"
                    )
                    continue
                
                data = response.json()
                techcard = data['techcard']
                process_steps = techcard.get('process', [])
                
                # Extract temperatures and times from process steps
                found_temps = []
                found_times = []
                
                for step in process_steps:
                    if step.get('temp_c'):
                        found_temps.append(int(step['temp_c']))
                    if step.get('time_min'):
                        found_times.append(int(step['time_min']))
                
                # Validate extraction
                success = True
                issues = []
                
                for expected_temp in test_case['expected_temps']:
                    if not any(abs(t - expected_temp) <= 5 for t in found_temps):  # Allow 5°C tolerance
                        success = False
                        issues.append(f"Missing temperature {expected_temp}°C")
                
                for expected_time in test_case['expected_times']:
                    if expected_time not in found_times:
                        success = False
                        issues.append(f"Missing time {expected_time} min")
                
                result_msg = f"✅ Found temps: {found_temps}, times: {found_times}"
                if not success:
                    result_msg = f"❌ Issues: {'; '.join(issues)}"
                
                self.record_test_result(
                    f"Technology Parsing - {test_case['name']}",
                    success,
                    result_msg,
                    {
                        'expected_temps': test_case['expected_temps'],
                        'found_temps': found_temps,
                        'expected_times': test_case['expected_times'],
                        'found_times': found_times
                    }
                )
                
            except Exception as e:
                self.record_test_result(
                    f"Technology Parsing - {test_case['name']}",
                    False,
                    f"Exception: {str(e)}"
                )
    
    def test_unit_conversions(self):
        """Test unit conversion functionality"""
        print("\n⚖️ Testing Unit Conversions...")
        
        conversion_tests = [
            {
                'name': 'kg→g conversion',
                'ingredient': 'Мука пшеничная',
                'brutto': 1.5,
                'unit': 'кг',
                'expected_brutto_g': 1500
            },
            {
                'name': 'ml→g with density (oil)',
                'ingredient': 'Масло подсолнечное',
                'brutto': 100,
                'unit': 'мл',
                'expected_brutto_g': 90  # density 0.9
            },
            {
                'name': 'pcs→g with piece weights',
                'ingredient': 'Яйцо куриное',
                'brutto': 2,
                'unit': 'шт',
                'expected_brutto_g': 100  # 2 * 50g
            },
            {
                'name': 'l→ml conversion',
                'ingredient': 'Молоко',
                'brutto': 0.5,
                'unit': 'л',
                'expected_brutto_g': 515  # 500ml * 1.03 density
            }
        ]
        
        for test_case in conversion_tests:
            try:
                # Create test XLSX with specific unit
                test_xlsx = self.create_test_xlsx_with_ingredient(
                    test_case['ingredient'],
                    test_case['brutto'],
                    test_case['unit']
                )
                
                # Import and check conversion
                import_url = f"{self.backend_url}/api/v1/iiko/import/ttk.xlsx"
                files = {'file': ('test_units.xlsx', test_xlsx, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                
                response = requests.post(import_url, files=files, timeout=30)
                if response.status_code != 200:
                    self.record_test_result(
                        f"Unit Conversion - {test_case['name']}",
                        False,
                        f"Import failed: HTTP {response.status_code}"
                    )
                    continue
                
                data = response.json()
                techcard = data['techcard']
                ingredients = techcard.get('ingredients', [])
                
                if not ingredients:
                    self.record_test_result(
                        f"Unit Conversion - {test_case['name']}",
                        False,
                        "No ingredients found"
                    )
                    continue
                
                ingredient = ingredients[0]  # First ingredient
                actual_brutto_g = ingredient.get('brutto_g', 0)
                
                # Allow 10% tolerance for conversions
                tolerance = test_case['expected_brutto_g'] * 0.1
                success = abs(actual_brutto_g - test_case['expected_brutto_g']) <= tolerance
                
                result_msg = f"✅ {test_case['brutto']} {test_case['unit']} → {actual_brutto_g}g (expected {test_case['expected_brutto_g']}g)"
                if not success:
                    result_msg = f"❌ {test_case['brutto']} {test_case['unit']} → {actual_brutto_g}g (expected {test_case['expected_brutto_g']}g)"
                
                self.record_test_result(
                    f"Unit Conversion - {test_case['name']}",
                    success,
                    result_msg,
                    {
                        'original_value': test_case['brutto'],
                        'original_unit': test_case['unit'],
                        'expected_g': test_case['expected_brutto_g'],
                        'actual_g': actual_brutto_g,
                        'tolerance': tolerance
                    }
                )
                
            except Exception as e:
                self.record_test_result(
                    f"Unit Conversion - {test_case['name']}",
                    False,
                    f"Exception: {str(e)}"
                )
    
    def test_sku_preservation(self):
        """Test SKU preservation through round-trip cycle"""
        print("\n🏷️ Testing SKU Preservation...")
        
        # Test with fixtures that should have SKU codes
        for fixture_name in self.test_fixtures:
            fixture_path = self.fixtures_dir / fixture_name
            if not fixture_path.exists():
                continue
            
            try:
                # Import original
                with open(fixture_path, 'rb') as f:
                    original_content = f.read()
                
                import_url = f"{self.backend_url}/api/v1/iiko/import/ttk.xlsx"
                files = {'file': (fixture_name, original_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                
                response = requests.post(import_url, files=files, timeout=30)
                if response.status_code != 200:
                    continue
                
                data = response.json()
                techcard = data['techcard']
                ingredients = techcard.get('ingredients', [])
                
                # Count SKUs
                original_skus = []
                generated_skus = []
                
                for ingredient in ingredients:
                    sku_id = ingredient.get('skuId')
                    if sku_id:
                        if sku_id.startswith('GENERATED_'):
                            generated_skus.append(sku_id)
                        else:
                            original_skus.append(sku_id)
                
                # Export and re-import
                export_url = f"{self.backend_url}/api/v1/techcards.v2/export/iiko.xlsx"
                export_response = requests.post(export_url, json=techcard, timeout=30)
                
                if export_response.status_code != 200:
                    continue
                
                # Re-import
                files2 = {'file': (f'exported_{fixture_name}', export_response.content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                reimport_response = requests.post(import_url, files=files2, timeout=30)
                
                if reimport_response.status_code != 200:
                    continue
                
                reimport_data = reimport_response.json()
                reimported_techcard = reimport_data['techcard']
                reimported_ingredients = reimported_techcard.get('ingredients', [])
                
                # Count SKUs after round-trip
                final_original_skus = []
                final_generated_skus = []
                
                for ingredient in reimported_ingredients:
                    sku_id = ingredient.get('skuId')
                    if sku_id:
                        if sku_id.startswith('GENERATED_'):
                            final_generated_skus.append(sku_id)
                        else:
                            final_original_skus.append(sku_id)
                
                # Validate preservation
                success = True
                issues = []
                
                # Original SKUs should be preserved
                for sku in original_skus:
                    if sku not in final_original_skus:
                        success = False
                        issues.append(f"Lost original SKU: {sku}")
                
                # No new GENERATED_ SKUs should appear
                new_generated = set(final_generated_skus) - set(generated_skus)
                if new_generated:
                    success = False
                    issues.append(f"New generated SKUs: {list(new_generated)}")
                
                result_msg = f"✅ Preserved {len(final_original_skus)}/{len(original_skus)} original SKUs, {len(final_generated_skus)} generated"
                if not success:
                    result_msg = f"❌ Issues: {'; '.join(issues)}"
                
                self.record_test_result(
                    f"SKU Preservation - {fixture_name}",
                    success,
                    result_msg,
                    {
                        'original_skus_count': len(original_skus),
                        'final_original_skus_count': len(final_original_skus),
                        'original_generated_count': len(generated_skus),
                        'final_generated_count': len(final_generated_skus)
                    }
                )
                
            except Exception as e:
                self.record_test_result(
                    f"SKU Preservation - {fixture_name}",
                    False,
                    f"Exception: {str(e)}"
                )
    
    def test_performance(self):
        """Test performance requirements (<3 seconds)"""
        print("\n⚡ Testing Performance Requirements...")
        
        # Test with largest fixture
        largest_fixture = None
        largest_size = 0
        
        for fixture_name in self.test_fixtures:
            fixture_path = self.fixtures_dir / fixture_name
            if fixture_path.exists():
                size = fixture_path.stat().st_size
                if size > largest_size:
                    largest_size = size
                    largest_fixture = fixture_name
        
        if not largest_fixture:
            self.record_test_result(
                "Performance Test",
                False,
                "No fixtures available for performance testing"
            )
            return
        
        try:
            fixture_path = self.fixtures_dir / largest_fixture
            with open(fixture_path, 'rb') as f:
                file_content = f.read()
            
            # Test import performance
            import_url = f"{self.backend_url}/api/v1/iiko/import/ttk.xlsx"
            files = {'file': (largest_fixture, file_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            
            start_time = time.time()
            response = requests.post(import_url, files=files, timeout=30)
            import_duration = time.time() - start_time
            
            if response.status_code != 200:
                self.record_test_result(
                    "Performance Test - Import",
                    False,
                    f"Import failed: HTTP {response.status_code}"
                )
                return
            
            data = response.json()
            techcard = data['techcard']
            
            # Test export performance
            export_url = f"{self.backend_url}/api/v1/techcards.v2/export/iiko.xlsx"
            
            start_time = time.time()
            export_response = requests.post(export_url, json=techcard, timeout=30)
            export_duration = time.time() - start_time
            
            # Validate performance
            import_success = import_duration < 3.0
            export_success = export_duration < 3.0
            overall_success = import_success and export_success
            
            result_msg = f"✅ Import: {import_duration:.2f}s, Export: {export_duration:.2f}s (both <3s)"
            if not overall_success:
                issues = []
                if not import_success:
                    issues.append(f"Import {import_duration:.2f}s > 3s")
                if not export_success:
                    issues.append(f"Export {export_duration:.2f}s > 3s")
                result_msg = f"❌ {'; '.join(issues)}"
            
            self.record_test_result(
                "Performance Test",
                overall_success,
                result_msg,
                {
                    'fixture': largest_fixture,
                    'file_size': largest_size,
                    'import_duration': import_duration,
                    'export_duration': export_duration,
                    'import_success': import_success,
                    'export_success': export_success
                }
            )
            
        except Exception as e:
            self.record_test_result(
                "Performance Test",
                False,
                f"Exception: {str(e)}"
            )
    
    def validate_techcard_structure(self, techcard: Dict[str, Any]) -> bool:
        """Validate TechCardV2 structure"""
        required_fields = ['meta', 'ingredients', 'process', 'yield', 'portions']
        
        for field in required_fields:
            if field not in techcard:
                return False
        
        # Validate meta
        meta = techcard['meta']
        if not isinstance(meta, dict) or 'title' not in meta:
            return False
        
        # Validate ingredients
        ingredients = techcard['ingredients']
        if not isinstance(ingredients, list) or len(ingredients) == 0:
            return False
        
        # Validate process
        process = techcard['process']
        if not isinstance(process, list) or len(process) < 3:
            return False
        
        return True
    
    def compare_techcards(self, original: Dict[str, Any], reimported: Dict[str, Any], fixture_name: str) -> Dict[str, Any]:
        """Compare original and reimported techcards for round-trip validation"""
        
        success = True
        issues = []
        details = {}
        
        # Compare ingredient count
        orig_ingredients = original.get('ingredients', [])
        reimp_ingredients = reimported.get('ingredients', [])
        
        if len(orig_ingredients) != len(reimp_ingredients):
            success = False
            issues.append(f"Ingredient count mismatch: {len(orig_ingredients)} vs {len(reimp_ingredients)}")
        
        details['original_ingredients'] = len(orig_ingredients)
        details['reimported_ingredients'] = len(reimp_ingredients)
        
        # Compare SKU preservation
        orig_skus = set(ing.get('skuId') for ing in orig_ingredients if ing.get('skuId') and not ing.get('skuId', '').startswith('GENERATED_'))
        reimp_skus = set(ing.get('skuId') for ing in reimp_ingredients if ing.get('skuId') and not ing.get('skuId', '').startswith('GENERATED_'))
        
        lost_skus = orig_skus - reimp_skus
        if lost_skus:
            success = False
            issues.append(f"Lost SKUs: {list(lost_skus)}")
        
        details['original_skus'] = len(orig_skus)
        details['preserved_skus'] = len(reimp_skus)
        
        # Compare process steps
        orig_process = original.get('process', [])
        reimp_process = reimported.get('process', [])
        
        if len(orig_process) != len(reimp_process):
            # Allow some variation in process steps due to parsing differences
            if abs(len(orig_process) - len(reimp_process)) > 2:
                success = False
                issues.append(f"Process steps mismatch: {len(orig_process)} vs {len(reimp_process)}")
        
        details['original_process_steps'] = len(orig_process)
        details['reimported_process_steps'] = len(reimp_process)
        
        # Compare title preservation
        orig_title = original.get('meta', {}).get('title', '')
        reimp_title = reimported.get('meta', {}).get('title', '')
        
        if orig_title != reimp_title:
            # Allow minor differences due to encoding/parsing
            if not (orig_title.lower().strip() == reimp_title.lower().strip()):
                success = False
                issues.append(f"Title mismatch: '{orig_title}' vs '{reimp_title}'")
        
        details['title_preserved'] = orig_title == reimp_title
        
        message = f"✅ Round-trip successful for {fixture_name}"
        if not success:
            message = f"❌ Round-trip issues for {fixture_name}: {'; '.join(issues)}"
        
        return {
            'success': success,
            'message': message,
            'details': details
        }
    
    def create_sample_techcard(self) -> Dict[str, Any]:
        """Create a sample TechCardV2 for testing"""
        return {
            "meta": {
                "title": "Test Dish for Export",
                "version": "2.0",
                "createdAt": "2025-01-17T10:00:00Z",
                "cuisine": "test",
                "tags": ["test"],
                "timings": {}
            },
            "portions": 4,
            "yield": {
                "perPortion_g": 200.0,
                "perBatch_g": 800.0
            },
            "ingredients": [
                {
                    "name": "Мука пшеничная",
                    "brutto_g": 500.0,
                    "netto_g": 500.0,
                    "loss_pct": 0.0,
                    "unit": "g",
                    "skuId": "FLOUR_001",
                    "allergens": [],
                    "canonical_id": None,
                    "subRecipe": None
                },
                {
                    "name": "Яйцо куриное",
                    "brutto_g": 100.0,
                    "netto_g": 90.0,
                    "loss_pct": 10.0,
                    "unit": "g",
                    "skuId": "EGG_001",
                    "allergens": [],
                    "canonical_id": None,
                    "subRecipe": None
                }
            ],
            "process": [
                {
                    "n": 1,
                    "action": "Подготовить ингредиенты",
                    "time_min": 5.0,
                    "temp_c": 20.0,
                    "equipment": None,
                    "details": None
                },
                {
                    "n": 2,
                    "action": "Смешать муку с яйцами",
                    "time_min": 10.0,
                    "temp_c": 20.0,
                    "equipment": None,
                    "details": None
                },
                {
                    "n": 3,
                    "action": "Выпекать в духовке",
                    "time_min": 25.0,
                    "temp_c": 180.0,
                    "equipment": None,
                    "details": None
                }
            ],
            "storage": {
                "conditions": "Хранить в холодильнике",
                "shelfLife_hours": 24.0,
                "servingTemp_c": 65.0
            },
            "nutrition": {
                "per100g": {"kcal": 250.0, "proteins_g": 8.0, "fats_g": 5.0, "carbs_g": 45.0, "fiber_g": 2.0, "sugar_g": 1.0, "sodium_mg": 200.0},
                "perPortion": {"kcal": 500.0, "proteins_g": 16.0, "fats_g": 10.0, "carbs_g": 90.0, "fiber_g": 4.0, "sugar_g": 2.0, "sodium_mg": 400.0}
            },
            "nutritionMeta": {
                "source": "bootstrap",
                "coveragePct": 100.0
            },
            "cost": {
                "rawCost": 150.0,
                "costPerPortion": 37.5,
                "markup_pct": None,
                "vat_pct": None
            },
            "costMeta": {
                "source": "catalog",
                "coveragePct": 100.0,
                "asOf": "2025-01-17"
            },
            "printNotes": []
        }
    
    def create_test_xlsx_with_technology(self, technology_text: str) -> bytes:
        """Create a test XLSX with specific technology text"""
        try:
            import openpyxl
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Headers
            headers = ['Наименование продукта', 'Брутто', 'Нетто', 'Ед.', 'Технология приготовления']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Sample ingredient with technology
            ws.cell(row=2, column=1, value='Тестовый ингредиент')
            ws.cell(row=2, column=2, value=100)
            ws.cell(row=2, column=3, value=90)
            ws.cell(row=2, column=4, value='г')
            ws.cell(row=2, column=5, value=technology_text)
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except ImportError:
            # Fallback: create minimal CSV-like content
            return b"Test XLSX content"
    
    def create_test_xlsx_with_ingredient(self, ingredient_name: str, brutto: float, unit: str) -> bytes:
        """Create a test XLSX with specific ingredient and unit"""
        try:
            import openpyxl
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Headers
            headers = ['Наименование продукта', 'Брутто', 'Нетто', 'Ед.']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Test ingredient
            ws.cell(row=2, column=1, value=ingredient_name)
            ws.cell(row=2, column=2, value=brutto)
            ws.cell(row=2, column=3, value=brutto * 0.9)  # 10% loss
            ws.cell(row=2, column=4, value=unit)
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except ImportError:
            return b"Test XLSX content"
    
    def record_test_result(self, test_name: str, success: bool, message: str, details: Optional[Dict] = None):
        """Record test result"""
        self.results['total_tests'] += 1
        if success:
            self.results['passed_tests'] += 1
        else:
            self.results['failed_tests'] += 1
        
        self.results['test_details'].append({
            'test_name': test_name,
            'success': success,
            'message': message,
            'details': details or {}
        })
        
        # Print result
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"  {status} {test_name}: {message}")
    
    def print_final_results(self):
        """Print comprehensive test results"""
        print("\n" + "="*80)
        print("🎯 IK-04/03 ROUND-TRIP GOLDEN TESTS RESULTS")
        print("="*80)
        
        total = self.results['total_tests']
        passed = self.results['passed_tests']
        failed = self.results['failed_tests']
        success_rate = (passed / total * 100) if total > 0 else 0
        
        print(f"📊 SUMMARY:")
        print(f"   Total Tests: {total}")
        print(f"   Passed: {passed}")
        print(f"   Failed: {failed}")
        print(f"   Success Rate: {success_rate:.1f}%")
        
        if failed > 0:
            print(f"\n❌ FAILED TESTS:")
            for result in self.results['test_details']:
                if not result['success']:
                    print(f"   • {result['test_name']}: {result['message']}")
        
        print(f"\n✅ PASSED TESTS:")
        for result in self.results['test_details']:
            if result['success']:
                print(f"   • {result['test_name']}: {result['message']}")
        
        # Overall assessment
        print(f"\n🎯 OVERALL ASSESSMENT:")
        if success_rate >= 90:
            print("   🎉 EXCELLENT: IK-04/03 round-trip golden workflow is fully functional!")
        elif success_rate >= 75:
            print("   ✅ GOOD: IK-04/03 round-trip workflow is mostly functional with minor issues.")
        elif success_rate >= 50:
            print("   ⚠️ PARTIAL: IK-04/03 round-trip workflow has significant issues requiring attention.")
        else:
            print("   🚨 CRITICAL: IK-04/03 round-trip workflow has major failures requiring immediate fixes.")
        
        return success_rate >= 75  # Consider 75%+ as overall success

def main():
    """Main test execution"""
    print("🚀 Starting IK-04/03 Round-trip Golden Backend Testing...")
    
    tester = IK04RoundTripTester()
    tester.run_all_tests()
    
    return 0

if __name__ == "__main__":
    sys.exit(main())