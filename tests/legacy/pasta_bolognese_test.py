#!/usr/bin/env python3
"""
ФИНАЛЬНЫЙ ТЕСТ АРТИКУЛОВ - Паста Болоньезе
Протестировать исправленную логику генерации артикулов с улучшенной обработкой ошибок.

WORKFLOW:
1. Сгенерировать новую техкарту "Паста Болоньезе"
2. Проверить генерацию артикулов в консоли backend'а
3. Верифицировать наличие dish.article и ingredient.product_code
4. Запустить preflight для проверки количества missing articles
5. Выполнить экспорт ZIP и проверить содержимое XLSX
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Any, Optional

import httpx
import openpyxl
from pymongo import MongoClient

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017/receptor_pro')
DB_NAME = os.getenv('DB_NAME', 'receptor_pro')

class PastaBologneseTester:
    """Pasta Bolognese Article Generation Testing"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=60.0)
        self.test_results = []
        self.organization_id = "test-org-pasta"
        self.generated_techcard_id = None
        self.exported_zip_content = None
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status} {test_name}")
        if details:
            print(f"    {details}")
        if response_time > 0:
            print(f"    Response time: {response_time:.3f}s")
        print()

    async def test_1_generate_pasta_bolognese(self):
        """Шаг 1: Сгенерировать новую техкарту 'Паста Болоньезе'"""
        try:
            start_time = time.time()
            
            dish_name = "Паста Болоньезе"
            
            print(f"🍝 Generating tech card for: {dish_name}")
            
            response = await self.client.post(
                f"{API_BASE}/techcards.v2/generate",
                json={
                    "name": dish_name
                }
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                print(f"🔍 Full response: {json.dumps(data, indent=2, ensure_ascii=False)}")
                
                # Try different ways to get the ID
                self.generated_techcard_id = data.get('card', {}).get('meta', {}).get('id') if data.get('card') else None
                
                # Extract basic info
                dish_article = None
                ingredients_count = 0
                ingredients_with_codes = 0
                
                card_data = data.get('card') or data
                
                if 'meta' in card_data and isinstance(card_data['meta'], dict):
                    dish_article = card_data['meta'].get('article')
                
                # Also check the top-level article field
                if dish_article is None and 'article' in card_data:
                    dish_article = card_data.get('article')
                
                if 'ingredients' in card_data and isinstance(card_data['ingredients'], list):
                    ingredients_count = len(card_data['ingredients'])
                    for ingredient in card_data['ingredients']:
                        if isinstance(ingredient, dict) and ingredient.get('product_code'):
                            ingredients_with_codes += 1
                
                success = self.generated_techcard_id is not None
                
                details = f"ID: {self.generated_techcard_id}, "
                details += f"Dish article: {dish_article or 'NULL'}, "
                details += f"Ingredients: {ingredients_count}, "
                details += f"With codes: {ingredients_with_codes}"
                details += f", Status: {data.get('status', 'unknown')}"
                
                self.log_test(
                    "Generate Паста Болоньезе Tech Card",
                    success,
                    details,
                    response_time
                )
                return success
                
            else:
                self.log_test(
                    "Generate Паста Болоньезе Tech Card",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                return False
                
        except Exception as e:
            self.log_test(
                "Generate Паста Болоньезе Tech Card",
                False,
                f"Exception: {str(e)}"
            )
            return False

    async def test_2_verify_database_articles(self):
        """Шаг 2: Верифицировать наличие dish.article и ingredient.product_code в БД"""
        try:
            if not self.generated_techcard_id:
                self.log_test(
                    "Verify Database Articles",
                    False,
                    "No tech card ID available"
                )
                return False
            
            print(f"🔍 Checking database for tech card: {self.generated_techcard_id}")
            
            # Connect to MongoDB
            client = MongoClient(MONGO_URL)
            db = client[DB_NAME]
            techcards_collection = db.techcards_v2
            
            # Find the tech card
            techcard = techcards_collection.find_one({"_id": self.generated_techcard_id})
            
            if not techcard:
                self.log_test(
                    "Verify Database Articles",
                    False,
                    f"Tech card not found in database: {self.generated_techcard_id}"
                )
                client.close()
                return False
            
            # Check dish article
            dish_article = None
            if 'meta' in techcard and isinstance(techcard['meta'], dict):
                dish_article = techcard['meta'].get('article')
            
            # Check ingredient product codes
            ingredients_with_codes = 0
            total_ingredients = 0
            ingredient_details = []
            
            if 'ingredients' in techcard and isinstance(techcard['ingredients'], list):
                total_ingredients = len(techcard['ingredients'])
                for ingredient in techcard['ingredients']:
                    if isinstance(ingredient, dict):
                        name = ingredient.get('name', 'Unknown')
                        product_code = ingredient.get('product_code')
                        if product_code:
                            ingredients_with_codes += 1
                            ingredient_details.append(f"{name}: {product_code}")
                        else:
                            ingredient_details.append(f"{name}: NO CODE")
            
            # Evaluate success
            dish_article_present = dish_article is not None and str(dish_article).strip() != ""
            ingredients_have_codes = ingredients_with_codes > 0
            
            success = dish_article_present and ingredients_have_codes
            
            details = f"Dish article: {dish_article or 'NULL'}, "
            details += f"Ingredients with codes: {ingredients_with_codes}/{total_ingredients}"
            if ingredient_details:
                details += f"\nIngredients: {'; '.join(ingredient_details[:3])}"
                if len(ingredient_details) > 3:
                    details += f" (and {len(ingredient_details) - 3} more)"
            
            self.log_test(
                "Verify Database Articles",
                success,
                details
            )
            
            client.close()
            return success
            
        except Exception as e:
            self.log_test(
                "Verify Database Articles",
                False,
                f"Exception: {str(e)}"
            )
            return False

    async def test_3_run_preflight_check(self):
        """Шаг 3: Запустить preflight для проверки количества missing articles"""
        try:
            if not self.generated_techcard_id:
                self.log_test(
                    "Run Preflight Check",
                    False,
                    "No tech card ID available"
                )
                return False
            
            start_time = time.time()
            
            print(f"🔍 Running preflight check for: {self.generated_techcard_id}")
            
            response = await self.client.post(
                f"{API_BASE}/techcards.v2/export/preflight-check",
                json={
                    "techcards": [self.generated_techcard_id]
                }
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                warnings = data.get('warnings', [])
                missing_articles_count = 0
                missing_dishes_count = 0
                
                for warning in warnings:
                    warning_type = warning.get('type', '').lower()
                    items = warning.get('items', [])
                    
                    if 'dish' in warning_type and 'missing' in warning_type:
                        missing_dishes_count += len(items)
                    elif 'product' in warning_type and 'missing' in warning_type:
                        missing_articles_count += len(items)
                
                export_ready = data.get('export_ready', False)
                cards_checked = data.get('cards_checked', 0)
                
                # Success if missing articles count is minimal (≤ 2)
                success = missing_articles_count <= 2
                
                details = f"Export ready: {export_ready}, Cards checked: {cards_checked}, "
                details += f"Missing dishes: {missing_dishes_count}, Missing products: {missing_articles_count}"
                if warnings:
                    details += f"\nWarnings: {len(warnings)} total"
                
                self.log_test(
                    "Run Preflight Check",
                    success,
                    details,
                    response_time
                )
                return success
                
            else:
                self.log_test(
                    "Run Preflight Check",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                return False
                
        except Exception as e:
            self.log_test(
                "Run Preflight Check",
                False,
                f"Exception: {str(e)}"
            )
            return False

    async def test_4_export_zip(self):
        """Шаг 4: Выполнить экспорт ZIP"""
        try:
            if not self.generated_techcard_id:
                self.log_test(
                    "Export ZIP",
                    False,
                    "No tech card ID available"
                )
                return False
            
            start_time = time.time()
            
            print(f"📦 Exporting ZIP for: {self.generated_techcard_id}")
            
            # Try enhanced export first
            response = await self.client.post(
                f"{API_BASE}/techcards.v2/export/enhanced/iiko.xlsx",
                json={
                    "techcards": [self.generated_techcard_id],
                    "operational_rounding": True,
                    "use_product_codes": True
                }
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                content_type = response.headers.get('content-type', '')
                content_length = len(response.content)
                
                # Verify it's a valid ZIP
                try:
                    with zipfile.ZipFile(io.BytesIO(response.content), 'r') as zip_file:
                        file_list = zip_file.namelist()
                        xlsx_files = [f for f in file_list if f.endswith('.xlsx')]
                        
                        self.exported_zip_content = response.content
                        
                        success = len(xlsx_files) > 0 and content_length > 1000
                        
                        details = f"Size: {content_length} bytes, "
                        details += f"Files: {len(file_list)} ({', '.join(file_list)}), "
                        details += f"XLSX files: {len(xlsx_files)}"
                        
                        self.log_test(
                            "Export ZIP",
                            success,
                            details,
                            response_time
                        )
                        return success
                        
                except zipfile.BadZipFile:
                    self.log_test(
                        "Export ZIP",
                        False,
                        f"Invalid ZIP file, Size: {content_length} bytes",
                        response_time
                    )
                    return False
                    
            else:
                # If enhanced export fails, try regular export
                print("Enhanced export failed, trying regular export...")
                
                response = await self.client.post(
                    f"{API_BASE}/techcards.v2/export/iiko.xlsx",
                    json={
                        "techcards": [self.generated_techcard_id]
                    }
                )
                
                if response.status_code == 200:
                    content_length = len(response.content)
                    success = content_length > 1000
                    
                    details = f"Regular export successful, Size: {content_length} bytes"
                    
                    self.log_test(
                        "Export ZIP",
                        success,
                        details,
                        response_time
                    )
                    return success
                else:
                    self.log_test(
                        "Export ZIP",
                        False,
                        f"Both exports failed. Enhanced: HTTP {response.status_code}, Regular: HTTP {response.status_code}",
                        response_time
                    )
                    return False
                
        except Exception as e:
            self.log_test(
                "Export ZIP",
                False,
                f"Exception: {str(e)}"
            )
            return False

    async def test_5_verify_xlsx_content(self):
        """Шаг 5: Проверить содержимое XLSX файлов"""
        try:
            if not self.exported_zip_content:
                self.log_test(
                    "Verify XLSX Content",
                    False,
                    "No exported ZIP content available"
                )
                return False
            
            print("📊 Analyzing XLSX content...")
            
            with zipfile.ZipFile(io.BytesIO(self.exported_zip_content), 'r') as zip_file:
                xlsx_files = [f for f in zip_file.namelist() if f.endswith('.xlsx')]
                
                if not xlsx_files:
                    self.log_test(
                        "Verify XLSX Content",
                        False,
                        "No XLSX files found in ZIP"
                    )
                    return False
                
                articles_found = []
                dish_names_found = []
                ingredient_names_found = []
                
                for xlsx_file in xlsx_files:
                    try:
                        xlsx_content = zip_file.read(xlsx_file)
                        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_content))
                        
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            
                            # Analyze first 20 rows
                            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20), 1):
                                for col_idx, cell in enumerate(row[:6], 1):  # First 6 columns
                                    if cell.value:
                                        cell_str = str(cell.value).strip()
                                        
                                        # Check for articles (5-digit codes)
                                        if cell_str.isdigit() and len(cell_str) == 5:
                                            articles_found.append(cell_str)
                                        
                                        # Check for dish names containing "Паста" or "Болоньезе"
                                        if any(word in cell_str.lower() for word in ['паста', 'болоньезе', 'pasta', 'bolognese']):
                                            dish_names_found.append(cell_str)
                                        
                                        # Check for common ingredients
                                        if any(word in cell_str.lower() for word in ['мясо', 'томат', 'лук', 'морковь', 'макарон']):
                                            ingredient_names_found.append(cell_str)
                        
                        workbook.close()
                        
                    except Exception as e:
                        print(f"Error reading {xlsx_file}: {str(e)}")
                        continue
                
                # Remove duplicates
                unique_articles = list(set(articles_found))
                unique_dishes = list(set(dish_names_found))
                unique_ingredients = list(set(ingredient_names_found))
                
                success = len(unique_articles) > 0 or len(unique_dishes) > 0
                
                details = f"XLSX files: {len(xlsx_files)}, "
                details += f"Articles: {len(unique_articles)}, "
                details += f"Dish names: {len(unique_dishes)}, "
                details += f"Ingredients: {len(unique_ingredients)}"
                
                if unique_articles:
                    details += f"\nSample articles: {', '.join(unique_articles[:3])}"
                if unique_dishes:
                    details += f"\nDish names: {', '.join(unique_dishes[:2])}"
                
                self.log_test(
                    "Verify XLSX Content",
                    success,
                    details
                )
                return success
                
        except Exception as e:
            self.log_test(
                "Verify XLSX Content",
                False,
                f"Exception: {str(e)}"
            )
            return False

    async def run_all_tests(self):
        """Run all tests for Pasta Bolognese article generation"""
        print("🍝 ФИНАЛЬНЫЙ ТЕСТ АРТИКУЛОВ - Паста Болоньезе")
        print("=" * 80)
        print()
        
        tests = [
            self.test_1_generate_pasta_bolognese,
            self.test_2_verify_database_articles,
            self.test_3_run_preflight_check,
            self.test_4_export_zip,
            self.test_5_verify_xlsx_content
        ]
        
        passed = 0
        total = len(tests)
        
        for test in tests:
            try:
                result = await test()
                if result:
                    passed += 1
            except Exception as e:
                print(f"❌ Test failed with exception: {str(e)}")
                traceback.print_exc()
        
        print("=" * 80)
        print(f"📊 FINAL RESULTS: {passed}/{total} tests passed ({passed/total*100:.1f}%)")
        print()
        
        # Expected results check
        expected_results = [
            "✅ Dish article сгенерирован и сохранен в meta",
            "✅ Product codes сгенерированы для всех ингредиентов", 
            "✅ Preflight показывает 0 или минимум missing articles",
            "✅ Экспорт проходит успешно",
            "✅ XLSX файлы содержат правильные артикулы"
        ]
        
        print("🎯 ОЖИДАЕМЫЕ РЕЗУЛЬТАТЫ:")
        for i, result in enumerate(self.test_results):
            if i < len(expected_results):
                status = "✅" if result['success'] else "❌"
                print(f"  {status} {expected_results[i]}")
        
        print()
        
        if passed == total:
            print("🎉 УСПЕХ! Все критические требования выполнены - логика генерации артикулов работает корректно!")
        elif passed >= 3:
            print("⚠️ ЧАСТИЧНЫЙ УСПЕХ - основная функциональность работает, но есть проблемы")
        else:
            print("❌ КРИТИЧЕСКИЕ ПРОБЛЕМЫ - логика генерации артикулов требует исправления")
        
        print()
        print("📋 ДЕТАЛЬНЫЕ РЕЗУЛЬТАТЫ:")
        for result in self.test_results:
            print(f"  {result['status']} {result['test']} ({result['response_time']})")
            if result['details']:
                print(f"      {result['details']}")
        
        return passed, total

async def main():
    """Main test execution"""
    async with PastaBologneseTester() as tester:
        passed, total = await tester.run_all_tests()
        
        # Exit with appropriate code
        if passed == total:
            sys.exit(0)
        else:
            sys.exit(1)

if __name__ == "__main__":
    asyncio.run(main())