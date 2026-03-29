#!/usr/bin/env python3
"""
Detailed Analysis of Reconcile TTK↔Skeletons Issues
"""

import asyncio
import json
import os
import zipfile
import io
from typing import Dict, List, Any

import httpx
import openpyxl

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"

async def analyze_dish_skeletons_issue():
    """Analyze why Dish-Skeletons.xlsx appears empty"""
    print("🔍 Analyzing Dish-Skeletons.xlsx Issue")
    
    async with httpx.AsyncClient(timeout=60.0) as client:
        # Generate tech card
        payload = {
            "name": "Борщ украинский с говядиной",
            "cuisine": "русская",
            "equipment": [],
            "budget": None,
            "dietary": []
        }
        
        response = await client.post(f"{API_BASE}/techcards.v2/generate", json=payload)
        if response.status_code != 200:
            print(f"❌ Tech card generation failed: {response.status_code}")
            return
        
        data = response.json()
        if not (data.get("status") in ["success", "draft"] and data.get("card")):
            print(f"❌ Invalid tech card response: {data.get('status')}")
            return
        
        techcard = data["card"]
        techcard_id = techcard.get("meta", {}).get("id")
        print(f"✅ Generated tech card: {techcard_id}")
        
        # Run preflight
        preflight_payload = {
            "techcardIds": [techcard_id],
            "organization_id": "test-org-reconcile"
        }
        
        preflight_response = await client.post(f"{API_BASE}/export/preflight", json=preflight_payload)
        if preflight_response.status_code != 200:
            print(f"❌ Preflight failed: {preflight_response.status_code}")
            return
        
        preflight_data = preflight_response.json()
        missing_dishes = preflight_data.get("missing", {}).get("dishes", [])
        print(f"✅ Preflight: {len(missing_dishes)} missing dishes")
        
        if missing_dishes:
            print("📋 Missing dish details:")
            for dish in missing_dishes:
                print(f"  - Name: {dish.get('name')}")
                print(f"  - Article: {dish.get('article')}")
                print(f"  - Type: {dish.get('type')}")
                print(f"  - Unit: {dish.get('unit')}")
                print(f"  - Yield: {dish.get('yield')}")
        
        # Export ZIP
        export_payload = {
            "techcardIds": [techcard_id],
            "operational_rounding": True,
            "organization_id": "test-org-reconcile",
            "preflight_result": preflight_data
        }
        
        export_response = await client.post(f"{API_BASE}/export/zip", json=export_payload)
        if export_response.status_code != 200:
            print(f"❌ Export failed: {export_response.status_code}")
            return
        
        print(f"✅ Export ZIP: {len(export_response.content)} bytes")
        
        # Analyze ZIP contents in detail
        with zipfile.ZipFile(io.BytesIO(export_response.content), 'r') as zip_file:
            for file_name in zip_file.namelist():
                print(f"\n📄 Analyzing {file_name}:")
                
                if file_name.endswith('.xlsx'):
                    with zip_file.open(file_name) as xlsx_file:
                        workbook = openpyxl.load_workbook(xlsx_file)
                        sheet = workbook.active
                        
                        print(f"  Sheet name: {sheet.title}")
                        print(f"  Max row: {sheet.max_row}")
                        print(f"  Max column: {sheet.max_column}")
                        
                        # Print first few rows
                        print("  Content preview:")
                        for row_idx, row in enumerate(sheet.iter_rows(max_row=5, values_only=True), 1):
                            print(f"    Row {row_idx}: {row}")
                        
                        # Check for article formatting
                        if 'Dish-Skeletons' in file_name:
                            print("  🔍 Dish-Skeletons specific analysis:")
                            data_rows = 0
                            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                                if row and any(cell for cell in row):
                                    data_rows += 1
                                    print(f"    Data row {row_idx}: {row}")
                            print(f"  Total data rows: {data_rows}")
                        
                        # Check article formatting
                        if file_name in ['iiko_TTK.xlsx', 'Product-Skeletons.xlsx', 'Dish-Skeletons.xlsx']:
                            print("  📊 Article formatting analysis:")
                            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=5, values_only=False), 2):
                                for col_idx, cell in enumerate(row):
                                    if col_idx in [0, 2] and cell.value:  # Article columns
                                        cell_format = cell.number_format
                                        value = str(cell.value)
                                        is_text_format = '@' in cell_format or cell_format == '@'
                                        is_5_digit = len(value) == 5 and value.isdigit()
                                        
                                        print(f"    Cell {chr(65+col_idx)}{row_idx}: value='{value}', format='{cell_format}', text_format={is_text_format}, 5_digit={is_5_digit}")
                                        
                                        if row_idx > 5:  # Limit output
                                            break
                                if row_idx > 5:
                                    break

async def test_ui_alt_legacy_removal():
    """Test if alt/legacy export options are removed from UI"""
    print("\n🔍 Testing UI Alt/Legacy Export Removal")
    
    # This would require frontend testing, but we can check if the backend endpoints exist
    async with httpx.AsyncClient(timeout=30.0) as client:
        # Check if legacy endpoints are still accessible
        legacy_endpoints = [
            f"{API_BASE}/export/legacy",
            f"{API_BASE}/export/alt",
            f"{API_BASE}/export/alternative"
        ]
        
        for endpoint in legacy_endpoints:
            try:
                response = await client.get(endpoint)
                if response.status_code == 404:
                    print(f"✅ {endpoint} - Not found (good)")
                else:
                    print(f"⚠️ {endpoint} - Still accessible ({response.status_code})")
            except Exception as e:
                print(f"✅ {endpoint} - Not accessible (good)")
        
        # Check if main export endpoints are working
        main_endpoints = [
            f"{API_BASE}/export/preflight",
            f"{API_BASE}/export/zip"
        ]
        
        for endpoint in main_endpoints:
            try:
                response = await client.post(endpoint, json={})
                if response.status_code in [200, 400, 422]:  # Any response means it exists
                    print(f"✅ {endpoint} - Available")
                else:
                    print(f"❌ {endpoint} - Unexpected status: {response.status_code}")
            except Exception as e:
                print(f"❌ {endpoint} - Error: {str(e)}")

async def main():
    print("🚀 Detailed Reconcile TTK↔Skeletons Analysis")
    print("="*60)
    
    await analyze_dish_skeletons_issue()
    await test_ui_alt_legacy_removal()

if __name__ == "__main__":
    asyncio.run(main())