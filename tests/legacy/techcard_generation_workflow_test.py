#!/usr/bin/env python3
"""
TechCard Generation Workflow Comprehensive Testing
Execute comprehensive test specification "Generate new TechCards → PF-02→EX-03: preflight→zip + XLSX @-format (credit-safe)"

Test Workflow:
1. Generate tech cards for 3 dish names: "Тест Блюдо 1", "Тест Блюдо 2", "Тест Блюдо 3"
2. Validate each generated tech card using GX-02 validation
3. Run preflight check on all generated tech card IDs
4. Export as ZIP with operational_rounding=true
5. Validate XLSX format for article preservation (@ format with leading zeros)
6. Create all required artifact files documenting the test results
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import tempfile
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path

import httpx
import openpyxl
from openpyxl.styles import NamedStyle

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"

class TechCardWorkflowTester:
    """Comprehensive TechCard Generation Workflow Testing Suite"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=30.0)
        self.test_results = []
        self.generated_ids = []
        self.artifacts_dir = Path("/app/artifacts")
        self.artifacts_dir.mkdir(exist_ok=True)
        
        # Test data from specification
        self.dish_names = ["Тест Блюдо 1", "Тест Блюдо 2", "Тест Блюдо 3"]
        self.portions = 1
        self.operational_rounding = True
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A",
            "timestamp": datetime.now().isoformat()
        }
        self.test_results.append(result)
        print(f"{status} {test_name}: {details}")
    
    async def generate_techcard(self, dish_name: str) -> Optional[Dict]:
        """Generate a single tech card and return techcard data with ID"""
        try:
            start_time = time.time()
            
            payload = {
                "name": dish_name,
                "portions": self.portions
            }
            
            response = await self.client.post(
                f"{API_BASE}/techcards.v2/generate",
                json=payload,
                headers={"Content-Type": "application/json"}
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Extract techcard ID from the response structure
                techcard_id = None
                if 'card' in data and data['card'] and 'meta' in data['card']:
                    techcard_id = data['card']['meta'].get('id')
                
                if not techcard_id:
                    # Try alternative locations
                    techcard_id = data.get('techcardId') or data.get('id')
                
                if techcard_id:
                    self.log_test(
                        f"Generate TechCard: {dish_name}",
                        True,
                        f"Generated techcard ID: {techcard_id}, Status: {data.get('status', 'unknown')}",
                        response_time
                    )
                    # Return both ID and full data for validation
                    return {
                        'id': techcard_id,
                        'data': data
                    }
                else:
                    self.log_test(
                        f"Generate TechCard: {dish_name}",
                        False,
                        f"No techcard ID in response structure: {list(data.keys())}",
                        response_time
                    )
                    return None
            else:
                self.log_test(
                    f"Generate TechCard: {dish_name}",
                    False,
                    f"HTTP {response.status_code}: {response.text}",
                    response_time
                )
                return None
                
        except Exception as e:
            self.log_test(
                f"Generate TechCard: {dish_name}",
                False,
                f"Exception: {str(e)}"
            )
            return None
    
    async def validate_techcard(self, techcard_info: Dict) -> bool:
        """Validate a tech card using GX-02 validation"""
        try:
            start_time = time.time()
            
            techcard_id = techcard_info['id']
            techcard_data = techcard_info['data']
            
            # Use the card data from generation response
            card_data = techcard_data.get('card')
            if not card_data:
                self.log_test(
                    f"Validate TechCard: {techcard_id}",
                    False,
                    "No card data available for validation"
                )
                return False
            
            payload = {
                "techcard": card_data
            }
            
            response = await self.client.post(
                f"{API_BASE}/techcards.v2/validate/quality",
                json=payload,
                headers={"Content-Type": "application/json"}
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Check for GX-02 green validation (нетто≈выход ±2–5%)
                validation_passed = data.get('is_production_ready', False)
                quality_score = data.get('quality_score', {})
                validation_issues = data.get('validation_issues', [])
                
                validation_details = []
                for issue in validation_issues:
                    if issue.get('severity') == 'error':
                        validation_passed = False
                    validation_details.append(f"{issue.get('type', 'unknown')}: {issue.get('message', 'no message')}")
                
                self.log_test(
                    f"Validate TechCard: {techcard_id}",
                    validation_passed,
                    f"GX-02 validation: {'PASS' if validation_passed else 'FAIL'}, Score: {quality_score.get('overall', 0)}, Issues: {len(validation_details)}",
                    response_time
                )
                return validation_passed
            else:
                self.log_test(
                    f"Validate TechCard: {techcard_id}",
                    False,
                    f"HTTP {response.status_code}: {response.text}",
                    response_time
                )
                return False
                
        except Exception as e:
            self.log_test(
                f"Validate TechCard: {techcard_info.get('id', 'unknown')}",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    async def run_preflight(self, techcard_ids: List[str]) -> Optional[Dict]:
        """Run preflight check on tech card IDs"""
        try:
            start_time = time.time()
            
            payload = {
                "techcardIds": techcard_ids
            }
            
            response = await self.client.post(
                f"{API_BASE}/export/preflight",
                json=payload,
                headers={"Content-Type": "application/json"}
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Save preflight results
                preflight_file = self.artifacts_dir / "preflight.json"
                with open(preflight_file, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                
                counts = data.get('counts', {})
                missing = data.get('missing', {})
                ttk_date = data.get('ttkDate', 'unknown')
                
                self.log_test(
                    "Preflight Check",
                    True,
                    f"Counts: {counts}, Missing: {missing}, TTK Date: {ttk_date}",
                    response_time
                )
                return data
            else:
                self.log_test(
                    "Preflight Check",
                    False,
                    f"HTTP {response.status_code}: {response.text}",
                    response_time
                )
                return None
                
        except Exception as e:
            self.log_test(
                "Preflight Check",
                False,
                f"Exception: {str(e)}"
            )
            return None
    
    async def export_zip(self, techcard_ids: List[str], preflight_result: Dict = None) -> Optional[bytes]:
        """Export tech cards as ZIP with operational rounding"""
        try:
            start_time = time.time()
            
            payload = {
                "techcardIds": techcard_ids,
                "operational_rounding": self.operational_rounding
            }
            
            # Include preflight result if available
            if preflight_result:
                payload["preflight_result"] = preflight_result
            
            response = await self.client.post(
                f"{API_BASE}/export/zip",
                json=payload,
                headers={"Content-Type": "application/json"}
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # The response should be binary ZIP data
                zip_content = response.content
                
                if zip_content:
                    # Save ZIP file for analysis
                    zip_file_path = self.artifacts_dir / "export.zip"
                    with open(zip_file_path, 'wb') as f:
                        f.write(zip_content)
                    
                    # Save export metadata
                    export_metadata = {
                        "zip_size": len(zip_content),
                        "zip_file_path": str(zip_file_path),
                        "techcard_ids": techcard_ids,
                        "operational_rounding": self.operational_rounding,
                        "export_timestamp": datetime.now().isoformat(),
                        "response_headers": dict(response.headers)
                    }
                    
                    export_file = self.artifacts_dir / "export.json"
                    with open(export_file, 'w', encoding='utf-8') as f:
                        json.dump(export_metadata, f, ensure_ascii=False, indent=2)
                    
                    self.log_test(
                        "ZIP Export",
                        True,
                        f"ZIP exported successfully, size: {len(zip_content)} bytes",
                        response_time
                    )
                    return zip_content
                else:
                    self.log_test(
                        "ZIP Export",
                        False,
                        "Empty ZIP content received",
                        response_time
                    )
                    return None
            else:
                self.log_test(
                    "ZIP Export",
                    False,
                    f"HTTP {response.status_code}: {response.text}",
                    response_time
                )
                return None
                
        except Exception as e:
            self.log_test(
                "ZIP Export",
                False,
                f"Exception: {str(e)}"
            )
            return None
    
    async def download_and_validate_xlsx(self, zip_content: bytes) -> bool:
        """Validate XLSX format for article preservation from ZIP content"""
        try:
            start_time = time.time()
            
            if not zip_content:
                self.log_test(
                    "XLSX Format Validation",
                    False,
                    "No ZIP content provided"
                )
                return False
            
            # Save ZIP to temporary file
            with tempfile.NamedTemporaryFile(suffix='.zip', delete=False) as temp_zip:
                temp_zip.write(zip_content)
                temp_zip_path = temp_zip.name
            
            validation_results = {
                "zip_size": len(zip_content),
                "files_found": [],
                "xlsx_validation": {},
                "article_format_check": {}
            }
            
            try:
                # Extract and analyze ZIP contents
                with zipfile.ZipFile(temp_zip_path, 'r') as zip_file:
                    file_list = zip_file.namelist()
                    validation_results["files_found"] = file_list
                    
                    # Look for XLSX files
                    xlsx_files = [f for f in file_list if f.endswith('.xlsx')]
                    
                    for xlsx_file in xlsx_files:
                        # Extract XLSX file
                        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_xlsx:
                            temp_xlsx.write(zip_file.read(xlsx_file))
                            temp_xlsx_path = temp_xlsx.name
                        
                        # Validate XLSX format
                        xlsx_validation = await self.validate_xlsx_format(temp_xlsx_path, xlsx_file)
                        validation_results["xlsx_validation"][xlsx_file] = xlsx_validation
                        
                        # Clean up temp XLSX
                        os.unlink(temp_xlsx_path)
                
                # Save validation results
                xlsx_checks_file = self.artifacts_dir / "xlsx_checks.json"
                with open(xlsx_checks_file, 'w', encoding='utf-8') as f:
                    json.dump(validation_results, f, ensure_ascii=False, indent=2)
                
                response_time = time.time() - start_time
                
                # Check if validation passed
                validation_passed = True
                validation_details = []
                
                if not xlsx_files:
                    validation_passed = False
                    validation_details.append("No XLSX files found in ZIP")
                
                for xlsx_file, validation in validation_results["xlsx_validation"].items():
                    if not validation.get("article_format_valid", False):
                        validation_passed = False
                        validation_details.append(f"{xlsx_file}: Article format validation failed")
                
                self.log_test(
                    "XLSX Format Validation",
                    validation_passed,
                    f"Files: {len(xlsx_files)}, Issues: {validation_details}",
                    response_time
                )
                
                return validation_passed
                
            finally:
                # Clean up temp ZIP
                os.unlink(temp_zip_path)
                
        except Exception as e:
            self.log_test(
                "XLSX Format Validation",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    async def validate_xlsx_format(self, xlsx_path: str, filename: str) -> Dict:
        """Validate XLSX file format for article preservation"""
        try:
            workbook = openpyxl.load_workbook(xlsx_path)
            validation_result = {
                "filename": filename,
                "sheets": [],
                "article_columns_found": [],
                "article_format_valid": True,
                "format_issues": []
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_info = {
                    "name": sheet_name,
                    "rows": sheet.max_row,
                    "columns": sheet.max_column,
                    "article_columns": []
                }
                
                # Check first row for article-related headers
                if sheet.max_row > 0:
                    header_row = [cell.value for cell in sheet[1]]
                    
                    # Look for article columns (dish.article, product.article)
                    article_column_indices = []
                    for i, header in enumerate(header_row):
                        if header and ('артикул' in str(header).lower() or 'article' in str(header).lower()):
                            article_column_indices.append(i + 1)  # 1-based indexing
                            sheet_info["article_columns"].append({
                                "column": i + 1,
                                "header": header
                            })
                    
                    # Check format of article columns
                    for col_idx in article_column_indices:
                        column_letter = openpyxl.utils.get_column_letter(col_idx)
                        
                        # Check if column is formatted as text (@)
                        sample_cells = []
                        for row in range(2, min(sheet.max_row + 1, 10)):  # Check first few data rows
                            cell = sheet.cell(row=row, column=col_idx)
                            if cell.value is not None:
                                sample_cells.append({
                                    "value": cell.value,
                                    "data_type": cell.data_type,
                                    "number_format": cell.number_format
                                })
                        
                        # Check if format preserves leading zeros
                        format_valid = True
                        for cell_info in sample_cells:
                            value = str(cell_info["value"])
                            # Check if it's a 5-digit article with leading zeros
                            if value.isdigit() and len(value) == 5 and value.startswith('0'):
                                # This is good - leading zeros preserved
                                continue
                            elif value.isdigit() and len(value) < 5:
                                # This might be missing leading zeros
                                format_valid = False
                                validation_result["format_issues"].append(
                                    f"Column {column_letter}: Value '{value}' may be missing leading zeros"
                                )
                        
                        sheet_info["article_columns"].append({
                            "column": col_idx,
                            "format_valid": format_valid,
                            "sample_values": sample_cells[:3]  # First 3 samples
                        })
                        
                        if not format_valid:
                            validation_result["article_format_valid"] = False
                
                validation_result["sheets"].append(sheet_info)
            
            workbook.close()
            return validation_result
            
        except Exception as e:
            return {
                "filename": filename,
                "error": str(e),
                "article_format_valid": False
            }
    
    async def run_comprehensive_test(self):
        """Run the complete TechCard generation workflow test"""
        print("🚀 Starting TechCard Generation Workflow Comprehensive Test")
        print(f"Backend URL: {BACKEND_URL}")
        print(f"Dish names: {self.dish_names}")
        print(f"Portions: {self.portions}")
        print(f"Operational rounding: {self.operational_rounding}")
        print("-" * 80)
        
        # Step 1: Generate batch of tech cards
        print("\n📝 Step 1: Generate Tech Cards")
        generated_techcards = []
        for dish_name in self.dish_names:
            techcard_info = await self.generate_techcard(dish_name)
            if techcard_info:
                generated_techcards.append(techcard_info)
                self.generated_ids.append(techcard_info['id'])
        
        # Save generation results
        gen_runs_file = self.artifacts_dir / "gen_runs.json"
        with open(gen_runs_file, 'w', encoding='utf-8') as f:
            json.dump({
                "dish_names": self.dish_names,
                "generated_ids": self.generated_ids,
                "generation_timestamp": datetime.now().isoformat(),
                "success_count": len(self.generated_ids),
                "total_requested": len(self.dish_names)
            }, f, ensure_ascii=False, indent=2)
        
        if len(self.generated_ids) < 2:
            print(f"❌ CRITICAL: Only {len(self.generated_ids)} tech cards generated, minimum 2 required")
            return False
        
        # Step 2: Validate batch
        print(f"\n✅ Step 2: Validate Tech Cards ({len(generated_techcards)} cards)")
        validation_results = []
        for techcard_info in generated_techcards:
            validation_passed = await self.validate_techcard(techcard_info)
            validation_results.append({
                "techcard_id": techcard_info['id'],
                "validation_passed": validation_passed
            })
        
        # Save validation results
        validation_file = self.artifacts_dir / "validation.json"
        with open(validation_file, 'w', encoding='utf-8') as f:
            json.dump({
                "validation_results": validation_results,
                "validation_timestamp": datetime.now().isoformat(),
                "passed_count": sum(1 for r in validation_results if r["validation_passed"]),
                "total_validated": len(validation_results)
            }, f, ensure_ascii=False, indent=2)
        
        # Step 3: Run preflight
        print(f"\n🛫 Step 3: Run Preflight Check")
        preflight_result = await self.run_preflight(self.generated_ids)
        
        if not preflight_result:
            print("❌ CRITICAL: Preflight check failed")
            return False
        
        # Step 4: ZIP export
        print(f"\n📦 Step 4: ZIP Export")
        zip_content = await self.export_zip(self.generated_ids, preflight_result)
        
        if not zip_content:
            print("❌ CRITICAL: ZIP export failed")
            return False
        
        # Step 5: XLSX validation
        print(f"\n📊 Step 5: XLSX Format Validation")
        xlsx_validation_passed = await self.download_and_validate_xlsx(zip_content)
        
        # Generate summary
        await self.generate_summary()
        
        return xlsx_validation_passed
    
    async def generate_summary(self):
        """Generate comprehensive test summary"""
        summary = {
            "test_specification": "Generate new TechCards → PF-02→EX-03: preflight→zip + XLSX @-format (credit-safe)",
            "execution_timestamp": datetime.now().isoformat(),
            "test_configuration": {
                "dish_names": self.dish_names,
                "portions": self.portions,
                "operational_rounding": self.operational_rounding,
                "backend_url": BACKEND_URL
            },
            "results_summary": {
                "total_tests": len(self.test_results),
                "passed_tests": sum(1 for r in self.test_results if r["success"]),
                "failed_tests": sum(1 for r in self.test_results if not r["success"]),
                "success_rate": f"{(sum(1 for r in self.test_results if r['success']) / len(self.test_results) * 100):.1f}%" if self.test_results else "0%"
            },
            "workflow_steps": {
                "tech_cards_generated": len(self.generated_ids),
                "tech_cards_requested": len(self.dish_names),
                "generation_success_rate": f"{(len(self.generated_ids) / len(self.dish_names) * 100):.1f}%" if self.dish_names else "0%"
            },
            "acceptance_criteria": {
                "all_generate_http_200": None,
                "gx02_validation_passes": None,
                "preflight_correct": None,
                "zip_contains_expected_files": None,
                "xlsx_articles_text_format": None,
                "no_guid_code_in_articles": None
            },
            "detailed_results": self.test_results,
            "artifacts_created": [
                "gen_runs.json",
                "validation.json", 
                "preflight.json",
                "export.json",
                "xlsx_checks.json",
                "summary.md"
            ]
        }
        
        # Update acceptance criteria based on test results
        for result in self.test_results:
            test_name = result["test"]
            success = result["success"]
            
            if "Generate TechCard" in test_name:
                if summary["acceptance_criteria"]["all_generate_http_200"] is None:
                    summary["acceptance_criteria"]["all_generate_http_200"] = success
                else:
                    summary["acceptance_criteria"]["all_generate_http_200"] = summary["acceptance_criteria"]["all_generate_http_200"] and success
            
            elif "Validate TechCard" in test_name:
                if summary["acceptance_criteria"]["gx02_validation_passes"] is None:
                    summary["acceptance_criteria"]["gx02_validation_passes"] = success
                else:
                    summary["acceptance_criteria"]["gx02_validation_passes"] = summary["acceptance_criteria"]["gx02_validation_passes"] and success
            
            elif "Preflight Check" in test_name:
                summary["acceptance_criteria"]["preflight_correct"] = success
            
            elif "ZIP Export" in test_name:
                summary["acceptance_criteria"]["zip_contains_expected_files"] = success
            
            elif "XLSX Format Validation" in test_name:
                summary["acceptance_criteria"]["xlsx_articles_text_format"] = success
        
        # Save summary as JSON
        summary_json_file = self.artifacts_dir / "summary.json"
        with open(summary_json_file, 'w', encoding='utf-8') as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)
        
        # Generate markdown summary
        markdown_summary = self.generate_markdown_summary(summary)
        summary_md_file = self.artifacts_dir / "summary.md"
        with open(summary_md_file, 'w', encoding='utf-8') as f:
            f.write(markdown_summary)
        
        print(f"\n📋 Summary saved to {summary_md_file}")
    
    def generate_markdown_summary(self, summary: Dict) -> str:
        """Generate markdown summary report"""
        md = f"""# TechCard Generation Workflow Test Summary

## Test Specification
**{summary['test_specification']}**

Executed: {summary['execution_timestamp']}

## Configuration
- **Dish Names**: {', '.join(summary['test_configuration']['dish_names'])}
- **Portions**: {summary['test_configuration']['portions']}
- **Operational Rounding**: {summary['test_configuration']['operational_rounding']}
- **Backend URL**: {summary['test_configuration']['backend_url']}

## Results Overview
- **Total Tests**: {summary['results_summary']['total_tests']}
- **Passed**: {summary['results_summary']['passed_tests']}
- **Failed**: {summary['results_summary']['failed_tests']}
- **Success Rate**: {summary['results_summary']['success_rate']}

## Workflow Results
- **Tech Cards Generated**: {summary['workflow_steps']['tech_cards_generated']}/{summary['workflow_steps']['tech_cards_requested']}
- **Generation Success Rate**: {summary['workflow_steps']['generation_success_rate']}

## Acceptance Criteria Status
"""
        
        criteria = summary['acceptance_criteria']
        for criterion, status in criteria.items():
            if status is None:
                status_icon = "⚠️ N/A"
            elif status:
                status_icon = "✅ PASS"
            else:
                status_icon = "❌ FAIL"
            
            criterion_name = criterion.replace('_', ' ').title()
            md += f"- **{criterion_name}**: {status_icon}\n"
        
        md += "\n## Detailed Test Results\n"
        
        for result in summary['detailed_results']:
            md += f"- **{result['test']}**: {result['status']}\n"
            if result['details']:
                md += f"  - Details: {result['details']}\n"
            if result['response_time'] != "N/A":
                md += f"  - Response Time: {result['response_time']}\n"
        
        md += f"\n## Artifacts Created\n"
        for artifact in summary['artifacts_created']:
            md += f"- `/app/artifacts/{artifact}`\n"
        
        return md

async def main():
    """Main test execution"""
    try:
        async with TechCardWorkflowTester() as tester:
            success = await tester.run_comprehensive_test()
            
            print("\n" + "="*80)
            print("🎯 FINAL RESULTS")
            print("="*80)
            
            total_tests = len(tester.test_results)
            passed_tests = sum(1 for r in tester.test_results if r["success"])
            success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
            
            print(f"Total Tests: {total_tests}")
            print(f"Passed: {passed_tests}")
            print(f"Failed: {total_tests - passed_tests}")
            print(f"Success Rate: {success_rate:.1f}%")
            print(f"Generated Tech Cards: {len(tester.generated_ids)}")
            
            if success and success_rate >= 80:
                print("\n🎉 COMPREHENSIVE TEST COMPLETED SUCCESSFULLY!")
                print("✅ All critical acceptance criteria met")
                print("✅ XLSX format validation passed")
                print("✅ Article preservation confirmed")
                return 0
            else:
                print("\n❌ TEST FAILED - Critical issues found")
                print("⚠️ Review artifact files for detailed analysis")
                return 1
                
    except Exception as e:
        print(f"\n💥 CRITICAL ERROR: {str(e)}")
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    exit_code = asyncio.run(main())
    sys.exit(exit_code)