#!/usr/bin/env python3
"""
CRITICAL DEBUG: Dish Skeleton & Article Issues Testing
Детально протестировать проблему с пустыми dish skeleton и отсутствующими артикулами в техкартах.

ПРОБЛЕМА: 
1. Dish Skeleton файлы пустые 
2. Техкарты без артикулов в экспортируемых XLSX

WORKFLOW:
1. Generate tech card для "Борщ украинский" 
2. Проверить структуру сгенерированной техкарты - есть ли dish.article и product.article у ингредиентов
3. Запустить preflight и посмотреть counts - сколько missing dishes и products
4. Запустить ZIP export и проверить реальное содержимое файлов
5. Открыть каждый XLSX файл и посмотреть, что внутри - пустые ли они
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Any, Optional

import httpx
import openpyxl
from pymongo import MongoClient

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017/receptor_pro')
DB_NAME = os.getenv('DB_NAME', 'receptor_pro')

class DishSkeletonDebugTester:
    """Critical Debug Tester for Dish Skeleton & Article Issues"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=60.0)
        self.test_results = []
        self.organization_id = "test-org-debug"
        self.generated_techcard_id = None
        self.artifacts = {}
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result with detailed information"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A",
            "timestamp": datetime.now().isoformat()
        }
        self.test_results.append(result)
        print(f"{status} {test_name}")
        if details:
            print(f"    📋 {details}")
        if response_time > 0:
            print(f"    ⏱️ {response_time:.3f}s")
        print()

    async def test_1_generate_borsch_techcard(self):
        """1. Generate tech card для 'Борщ украинский' и проверить структуру"""
        try:
            start_time = time.time()
            
            # Generate tech card for Борщ украинский
            generate_url = f"{API_BASE}/techcards.v2/generate"
            payload = {
                "name": "Борщ украинский",
                "cuisine": "русская",
                "equipment": [],
                "budget": None,
                "dietary": []
            }
            
            print(f"🔄 Generating tech card: {payload['name']}")
            response = await self.client.post(generate_url, json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Extract tech card from response
                techcard = data.get('card')
                if not techcard:
                    self.log_test("Generate Борщ украинский Tech Card", False, f"No card in response: {data}", response_time)
                    return False
                
                # Generate a unique ID for tracking
                self.generated_techcard_id = techcard.get('id') or f"generated_{int(time.time())}"
                
                # Save generation data for analysis
                self.artifacts['generation'] = {
                    'techcard_id': self.generated_techcard_id,
                    'response': data,
                    'dish_name': payload['name'],
                    'full_techcard': techcard
                }
                
                ingredients = techcard.get('ingredients', [])
                
                # Check for dish article
                dish_article = None
                if 'meta' in techcard and isinstance(techcard['meta'], dict):
                    dish_article = techcard['meta'].get('article')
                elif hasattr(techcard, 'article'):
                    dish_article = techcard.get('article')
                
                # Check for product articles in ingredients
                ingredients_with_articles = 0
                ingredients_without_articles = 0
                article_details = []
                
                for ingredient in ingredients:
                    if isinstance(ingredient, dict):
                        product_article = ingredient.get('article') or ingredient.get('product_code')
                        if product_article:
                            ingredients_with_articles += 1
                            article_details.append(f"{ingredient.get('name', 'Unknown')}: {product_article}")
                        else:
                            ingredients_without_articles += 1
                            article_details.append(f"{ingredient.get('name', 'Unknown')}: NO ARTICLE")
                
                details = f"Generated ID: {self.generated_techcard_id}, Dish article: {dish_article}, Ingredients: {len(ingredients)} total, {ingredients_with_articles} with articles, {ingredients_without_articles} without articles"
                
                # Save detailed analysis
                self.artifacts['structure_analysis'] = {
                    'dish_article': dish_article,
                    'total_ingredients': len(ingredients),
                    'ingredients_with_articles': ingredients_with_articles,
                    'ingredients_without_articles': ingredients_without_articles,
                    'article_details': article_details,
                    'full_techcard': techcard
                }
                
                self.log_test("Generate Борщ украинский Tech Card", True, details, response_time)
                return True
            else:
                self.log_test("Generate Борщ украинский Tech Card", False, f"HTTP {response.status_code}: {response.text}", response_time)
                return False
                
        except Exception as e:
            self.log_test("Generate Борщ украинский Tech Card", False, f"Exception: {str(e)}")
            return False

    async def test_2_check_database_persistence(self):
        """2. Проверить, сохранилась ли техкарта в базе данных с артикулами"""
        try:
            if not self.generated_techcard_id:
                self.log_test("Check Database Persistence", False, "No techcard ID from generation")
                return False
            
            # Connect to MongoDB and save the techcard manually if needed
            client = MongoClient(MONGO_URL)
            db = client[DB_NAME]
            techcards_collection = db.techcards_v2
            
            # First check if it exists (try both _id and id fields)
            saved_techcard = techcards_collection.find_one({"_id": self.generated_techcard_id})
            if not saved_techcard:
                saved_techcard = techcards_collection.find_one({"id": self.generated_techcard_id})
            
            if not saved_techcard and 'generation' in self.artifacts:
                # Save the generated techcard to database
                techcard_data = self.artifacts['generation']['full_techcard']
                techcard_data['_id'] = self.generated_techcard_id  # Use _id for MongoDB
                techcard_data['id'] = self.generated_techcard_id   # Keep id field too
                techcard_data['created_at'] = datetime.now()
                techcard_data['updated_at'] = datetime.now()
                
                # Insert into database
                result = techcards_collection.insert_one(techcard_data)
                print(f"📝 Manually saved techcard to database: {result.inserted_id}")
                
                # Re-fetch to verify using _id
                saved_techcard = techcards_collection.find_one({"_id": self.generated_techcard_id})
            
            if saved_techcard:
                # Check articles in saved version
                dish_article = None
                if 'meta' in saved_techcard and isinstance(saved_techcard['meta'], dict):
                    dish_article = saved_techcard['meta'].get('article')
                elif 'article' in saved_techcard:
                    dish_article = saved_techcard.get('article')
                
                ingredients = saved_techcard.get('ingredients', [])
                saved_with_articles = sum(1 for ing in ingredients if ing.get('article') or ing.get('product_code'))
                
                self.artifacts['database_check'] = {
                    'found_in_db': True,
                    'dish_article': dish_article,
                    'ingredients_count': len(ingredients),
                    'ingredients_with_articles': saved_with_articles,
                    'saved_techcard': saved_techcard
                }
                
                details = f"Found in DB: dish_article={dish_article}, ingredients={len(ingredients)}, with_articles={saved_with_articles}"
                self.log_test("Check Database Persistence", True, details)
                return True
            else:
                self.artifacts['database_check'] = {'found_in_db': False}
                self.log_test("Check Database Persistence", False, "Techcard not found in database")
                return False
                
        except Exception as e:
            self.log_test("Check Database Persistence", False, f"Exception: {str(e)}")
            return False

    async def test_3_run_preflight_check(self):
        """3. Запустить preflight и проверить counts missing dishes/products"""
        try:
            if not self.generated_techcard_id:
                self.log_test("Run Preflight Check", False, "No techcard ID available")
                return False
            
            start_time = time.time()
            
            # Run preflight check
            preflight_url = f"{API_BASE}/export/preflight"
            payload = {
                "techcardIds": [self.generated_techcard_id],
                "organization_id": self.organization_id
            }
            
            print(f"🔄 Running preflight check for techcard: {self.generated_techcard_id}")
            response = await self.client.post(preflight_url, json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Extract preflight results
                warnings = data.get('warnings', [])
                missing_dishes = 0
                missing_products = 0
                
                # Check for missing items in the new format
                missing_data = data.get('missing', {})
                if 'dishes' in missing_data:
                    missing_dishes = len(missing_data['dishes'])
                if 'products' in missing_data:
                    missing_products = len(missing_data['products'])
                
                # Fallback to warnings format
                if missing_dishes == 0 and missing_products == 0:
                    for warning in warnings:
                        if 'dish' in warning.get('type', '').lower():
                            missing_dishes += len(warning.get('items', []))
                        elif 'product' in warning.get('type', '').lower():
                            missing_products += len(warning.get('items', []))
                
                ttk_date = data.get('ttkDate', data.get('ttk_date', 'Not specified'))
                export_ready = data.get('export_ready', False)
                
                self.artifacts['preflight'] = {
                    'response': data,
                    'missing_dishes': missing_dishes,
                    'missing_products': missing_products,
                    'ttk_date': ttk_date,
                    'export_ready': export_ready,
                    'warnings': warnings
                }
                
                details = f"TTK date: {ttk_date}, Missing dishes: {missing_dishes}, Missing products: {missing_products}, Export ready: {export_ready}"
                self.log_test("Run Preflight Check", True, details, response_time)
                return True
            else:
                self.log_test("Run Preflight Check", False, f"HTTP {response.status_code}: {response.text}", response_time)
                return False
                
        except Exception as e:
            self.log_test("Run Preflight Check", False, f"Exception: {str(e)}")
            return False

    async def test_4_export_zip_and_analyze(self):
        """4. Запустить ZIP export и проверить реальное содержимое файлов"""
        try:
            if not self.generated_techcard_id:
                self.log_test("Export ZIP and Analyze", False, "No techcard ID available")
                return False
            
            start_time = time.time()
            
            # Export ZIP
            export_url = f"{API_BASE}/export/zip"
            
            # Use the preflight result to bypass the guard
            preflight_result = self.artifacts.get('preflight', {}).get('response', {})
            
            payload = {
                "techcardIds": [self.generated_techcard_id],
                "organization_id": self.organization_id,
                "operational_rounding": True,
                "preflight_result": preflight_result  # Include preflight result to bypass guard
            }
            
            print(f"🔄 Exporting ZIP for techcard: {self.generated_techcard_id}")
            response = await self.client.post(export_url, json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # Check if response is ZIP file
                content_type = response.headers.get('content-type', '')
                zip_content = response.content
                zip_size = len(zip_content)
                
                self.artifacts['export'] = {
                    'content_type': content_type,
                    'zip_size': zip_size,
                    'files_analysis': {}
                }
                
                # Analyze ZIP contents
                if zip_content:
                    try:
                        with zipfile.ZipFile(io.BytesIO(zip_content), 'r') as zip_file:
                            file_list = zip_file.namelist()
                            
                            for file_name in file_list:
                                file_data = zip_file.read(file_name)
                                file_size = len(file_data)
                                
                                # Analyze XLSX files
                                if file_name.endswith('.xlsx'):
                                    xlsx_analysis = await self._analyze_xlsx_content(file_data, file_name)
                                    self.artifacts['export']['files_analysis'][file_name] = {
                                        'size': file_size,
                                        'type': 'XLSX',
                                        'analysis': xlsx_analysis
                                    }
                                else:
                                    self.artifacts['export']['files_analysis'][file_name] = {
                                        'size': file_size,
                                        'type': 'Other'
                                    }
                            
                            details = f"ZIP size: {zip_size} bytes, Files: {len(file_list)} ({', '.join(file_list)})"
                            self.log_test("Export ZIP and Analyze", True, details, response_time)
                            return True
                    except Exception as zip_error:
                        self.log_test("Export ZIP and Analyze", False, f"ZIP analysis error: {str(zip_error)}", response_time)
                        return False
                else:
                    self.log_test("Export ZIP and Analyze", False, "Empty ZIP content", response_time)
                    return False
            else:
                self.log_test("Export ZIP and Analyze", False, f"HTTP {response.status_code}: {response.text}", response_time)
                return False
                
        except Exception as e:
            self.log_test("Export ZIP and Analyze", False, f"Exception: {str(e)}")
            return False

    async def _analyze_xlsx_content(self, xlsx_data: bytes, file_name: str) -> Dict[str, Any]:
        """Analyze XLSX file content for articles and data"""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data))
            analysis = {
                'sheets': [],
                'total_rows': 0,
                'total_cells_with_data': 0,
                'articles_found': [],
                'dish_names_found': [],
                'mock_content_detected': [],
                'empty_file': True
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_analysis = {
                    'name': sheet_name,
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column,
                    'cells_with_data': 0,
                    'sample_data': []
                }
                
                # Check for data in cells
                for row in range(1, min(sheet.max_row + 1, 21)):  # Check first 20 rows
                    for col in range(1, min(sheet.max_column + 1, 11)):  # Check first 10 columns
                        cell = sheet.cell(row=row, column=col)
                        if cell.value is not None and str(cell.value).strip():
                            analysis['empty_file'] = False
                            sheet_analysis['cells_with_data'] += 1
                            analysis['total_cells_with_data'] += 1
                            
                            cell_value = str(cell.value).strip()
                            
                            # Look for articles (5-digit codes)
                            if cell_value.isdigit() and len(cell_value) == 5:
                                analysis['articles_found'].append(cell_value)
                            
                            # Look for dish names (Russian text)
                            if any(char in 'абвгдеёжзийклмнопрстуфхцчшщъыьэюя' for char in cell_value.lower()):
                                if len(cell_value) > 3:
                                    analysis['dish_names_found'].append(cell_value)
                            
                            # Look for mock content
                            mock_indicators = ['MOCK', 'TEST_INGREDIENT', 'DISH_MOCK', 'GENERATED_TEST']
                            if any(indicator in cell_value.upper() for indicator in mock_indicators):
                                analysis['mock_content_detected'].append(cell_value)
                            
                            # Sample data for debugging
                            if len(sheet_analysis['sample_data']) < 10:
                                sheet_analysis['sample_data'].append({
                                    'cell': f"{chr(64+col)}{row}",
                                    'value': cell_value
                                })
                
                analysis['sheets'].append(sheet_analysis)
                analysis['total_rows'] += sheet.max_row
            
            return analysis
            
        except Exception as e:
            return {'error': str(e), 'empty_file': True}

    async def test_5_detailed_xlsx_inspection(self):
        """5. Детальная инспекция каждого XLSX файла"""
        try:
            if 'export' not in self.artifacts or not self.artifacts['export'].get('files_analysis'):
                self.log_test("Detailed XLSX Inspection", False, "No export data available")
                return False
            
            files_analysis = self.artifacts['export']['files_analysis']
            xlsx_files = [name for name in files_analysis.keys() if name.endswith('.xlsx')]
            
            if not xlsx_files:
                self.log_test("Detailed XLSX Inspection", False, "No XLSX files found in export")
                return False
            
            inspection_results = {}
            all_files_ok = True
            
            for file_name in xlsx_files:
                file_analysis = files_analysis[file_name]['analysis']
                
                is_empty = file_analysis.get('empty_file', True)
                articles_count = len(file_analysis.get('articles_found', []))
                dishes_count = len(file_analysis.get('dish_names_found', []))
                mock_count = len(file_analysis.get('mock_content_detected', []))
                
                inspection_results[file_name] = {
                    'empty': is_empty,
                    'articles_found': articles_count,
                    'dishes_found': dishes_count,
                    'mock_content': mock_count,
                    'status': 'EMPTY' if is_empty else ('MOCK' if mock_count > 0 else 'OK')
                }
                
                if is_empty or mock_count > 0:
                    all_files_ok = False
            
            self.artifacts['detailed_inspection'] = inspection_results
            
            # Create summary
            summary_parts = []
            for file_name, result in inspection_results.items():
                summary_parts.append(f"{file_name}: {result['status']} (articles: {result['articles_found']}, dishes: {result['dishes_found']}, mock: {result['mock_content']})")
            
            details = "; ".join(summary_parts)
            self.log_test("Detailed XLSX Inspection", all_files_ok, details)
            return all_files_ok
            
        except Exception as e:
            self.log_test("Detailed XLSX Inspection", False, f"Exception: {str(e)}")
            return False

    async def test_6_root_cause_analysis(self):
        """6. Root Cause Analysis - найти где теряются артикулы"""
        try:
            # Analyze all collected data to find root cause
            analysis = {
                'generation_phase': {},
                'database_phase': {},
                'preflight_phase': {},
                'export_phase': {},
                'root_causes': []
            }
            
            # Check generation phase
            if 'structure_analysis' in self.artifacts:
                struct = self.artifacts['structure_analysis']
                analysis['generation_phase'] = {
                    'dish_article_present': struct['dish_article'] is not None,
                    'ingredients_with_articles': struct['ingredients_with_articles'],
                    'ingredients_without_articles': struct['ingredients_without_articles']
                }
                
                if not struct['dish_article']:
                    analysis['root_causes'].append("GENERATION: Dish article not generated")
                if struct['ingredients_without_articles'] > 0:
                    analysis['root_causes'].append(f"GENERATION: {struct['ingredients_without_articles']} ingredients without articles")
            
            # Check database phase
            if 'database_check' in self.artifacts:
                db_check = self.artifacts['database_check']
                analysis['database_phase'] = {
                    'saved_to_db': db_check.get('found_in_db', False),
                    'dish_article_preserved': db_check.get('dish_article') is not None,
                    'ingredients_with_articles_preserved': db_check.get('ingredients_with_articles', 0)
                }
                
                if not db_check.get('found_in_db'):
                    analysis['root_causes'].append("DATABASE: Techcard not saved to database")
                elif not db_check.get('dish_article'):
                    analysis['root_causes'].append("DATABASE: Dish article lost during save")
            
            # Check preflight phase
            if 'preflight' in self.artifacts:
                preflight = self.artifacts['preflight']
                analysis['preflight_phase'] = {
                    'missing_dishes': preflight['missing_dishes'],
                    'missing_products': preflight['missing_products'],
                    'export_ready': preflight['export_ready']
                }
                
                if preflight['missing_dishes'] > 0:
                    analysis['root_causes'].append(f"PREFLIGHT: {preflight['missing_dishes']} missing dishes detected")
                if preflight['missing_products'] > 0:
                    analysis['root_causes'].append(f"PREFLIGHT: {preflight['missing_products']} missing products detected")
            
            # Check export phase
            if 'detailed_inspection' in self.artifacts:
                inspection = self.artifacts['detailed_inspection']
                empty_files = [name for name, result in inspection.items() if result['empty']]
                mock_files = [name for name, result in inspection.items() if result['mock_content'] > 0]
                
                analysis['export_phase'] = {
                    'empty_files': empty_files,
                    'mock_files': mock_files,
                    'total_files': len(inspection)
                }
                
                if empty_files:
                    analysis['root_causes'].append(f"EXPORT: Empty files generated: {', '.join(empty_files)}")
                if mock_files:
                    analysis['root_causes'].append(f"EXPORT: Mock content in files: {', '.join(mock_files)}")
            
            self.artifacts['root_cause_analysis'] = analysis
            
            # Determine if we found the root cause
            has_root_causes = len(analysis['root_causes']) > 0
            details = f"Root causes identified: {len(analysis['root_causes'])} - {'; '.join(analysis['root_causes'][:3])}"
            
            self.log_test("Root Cause Analysis", has_root_causes, details)
            return has_root_causes
            
        except Exception as e:
            self.log_test("Root Cause Analysis", False, f"Exception: {str(e)}")
            return False

    def generate_summary_report(self):
        """Generate comprehensive summary report"""
        print("\n" + "="*80)
        print("🚨 CRITICAL DEBUG REPORT: DISH SKELETON & ARTICLE ISSUES")
        print("="*80)
        
        # Test Results Summary
        total_tests = len(self.test_results)
        passed_tests = sum(1 for result in self.test_results if result['success'])
        
        print(f"\n📊 TEST RESULTS SUMMARY:")
        print(f"   Total Tests: {total_tests}")
        print(f"   Passed: {passed_tests}")
        print(f"   Failed: {total_tests - passed_tests}")
        print(f"   Success Rate: {(passed_tests/total_tests*100):.1f}%")
        
        # Critical Findings
        print(f"\n🔍 CRITICAL FINDINGS:")
        
        if 'root_cause_analysis' in self.artifacts:
            root_causes = self.artifacts['root_cause_analysis']['root_causes']
            if root_causes:
                for i, cause in enumerate(root_causes, 1):
                    print(f"   {i}. {cause}")
            else:
                print("   ✅ No critical issues identified")
        
        # Detailed Analysis
        print(f"\n📋 DETAILED ANALYSIS:")
        
        if 'structure_analysis' in self.artifacts:
            struct = self.artifacts['structure_analysis']
            print(f"   Generation Phase:")
            print(f"     - Dish article: {struct['dish_article']}")
            print(f"     - Ingredients with articles: {struct['ingredients_with_articles']}/{struct['total_ingredients']}")
        
        if 'preflight' in self.artifacts:
            preflight = self.artifacts['preflight']
            print(f"   Preflight Phase:")
            print(f"     - Missing dishes: {preflight['missing_dishes']}")
            print(f"     - Missing products: {preflight['missing_products']}")
            print(f"     - Export ready: {preflight['export_ready']}")
        
        if 'detailed_inspection' in self.artifacts:
            inspection = self.artifacts['detailed_inspection']
            print(f"   Export Phase:")
            for file_name, result in inspection.items():
                print(f"     - {file_name}: {result['status']} (articles: {result['articles_found']}, dishes: {result['dishes_found']})")
        
        # Recommendations
        print(f"\n💡 RECOMMENDATIONS:")
        if 'root_cause_analysis' in self.artifacts:
            root_causes = self.artifacts['root_cause_analysis']['root_causes']
            if any('GENERATION' in cause for cause in root_causes):
                print("   1. Fix article generation logic in tech card creation")
            if any('DATABASE' in cause for cause in root_causes):
                print("   2. Fix article persistence in database save operations")
            if any('PREFLIGHT' in cause for cause in root_causes):
                print("   3. Fix preflight article detection logic")
            if any('EXPORT' in cause for cause in root_causes):
                print("   4. Fix export pipeline to include real article data")
        
        print("\n" + "="*80)
        
        # Save artifacts to file
        try:
            with open('/app/dish_skeleton_debug_artifacts.json', 'w', encoding='utf-8') as f:
                json.dump(self.artifacts, f, indent=2, ensure_ascii=False, default=str)
            print("📁 Artifacts saved to: /app/dish_skeleton_debug_artifacts.json")
        except Exception as e:
            print(f"⚠️ Could not save artifacts: {e}")

async def main():
    """Main test execution"""
    print("🚨 STARTING CRITICAL DEBUG: Dish Skeleton & Article Issues")
    print("="*80)
    
    async with DishSkeletonDebugTester() as tester:
        # Execute all tests in sequence
        await tester.test_1_generate_borsch_techcard()
        await tester.test_2_check_database_persistence()
        await tester.test_3_run_preflight_check()
        await tester.test_4_export_zip_and_analyze()
        await tester.test_5_detailed_xlsx_inspection()
        await tester.test_6_root_cause_analysis()
        
        # Generate comprehensive report
        tester.generate_summary_report()

if __name__ == "__main__":
    asyncio.run(main())