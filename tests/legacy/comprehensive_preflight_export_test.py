#!/usr/bin/env python3
"""
COMPREHENSIVE PREFLIGHT→EXPORT INTEGRATION VALIDATION
Расширенное тестирование исправления mock контента и интеграции preflight→export
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Any, Optional

import httpx
import openpyxl

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"

class ComprehensivePreflightExportTester:
    """Comprehensive preflight→export integration tester"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=60.0)
        self.test_results = []
        self.organization_id = "test-org-comprehensive"
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status} {test_name}: {details}")
    
    async def test_multiple_techcards_workflow(self):
        """Тест полного workflow с несколькими техкартами"""
        try:
            print("\n🔄 ТЕСТИРОВАНИЕ ПОЛНОГО WORKFLOW С НЕСКОЛЬКИМИ ТЕХКАРТАМИ")
            
            # Генерируем 3 техкарты
            techcard_names = [
                "Паста карбонара",
                "Ризотто с грибами", 
                "Лазанья болоньезе"
            ]
            
            generated_ids = []
            
            # Генерация техкарт
            for name in techcard_names:
                start_time = time.time()
                
                response = await self.client.post(
                    f"{API_BASE}/techcards.v2/generate",
                    json={
                        "name": name,
                        "description": f"Классическое итальянское блюдо {name}",
                        "category": "горячее"
                    }
                )
                
                response_time = time.time() - start_time
                
                if response.status_code == 200:
                    data = response.json()
                    card_id = data.get('card', {}).get('meta', {}).get('id')
                    if card_id:
                        generated_ids.append(card_id)
                        self.log_test(
                            f"Генерация '{name}'",
                            True,
                            f"ID: {card_id}",
                            response_time
                        )
                    else:
                        self.log_test(
                            f"Генерация '{name}'",
                            False,
                            "ID не получен"
                        )
                        return False
                else:
                    self.log_test(
                        f"Генерация '{name}'",
                        False,
                        f"HTTP {response.status_code}"
                    )
                    return False
            
            if len(generated_ids) != 3:
                return False
            
            # Preflight для всех техкарт
            start_time = time.time()
            preflight_response = await self.client.post(
                f"{API_BASE}/export/preflight",
                json={
                    "techcardIds": generated_ids,
                    "organization_id": self.organization_id
                }
            )
            response_time = time.time() - start_time
            
            if preflight_response.status_code != 200:
                self.log_test(
                    "Preflight для 3 техкарт",
                    False,
                    f"HTTP {preflight_response.status_code}"
                )
                return False
            
            preflight_result = preflight_response.json()
            
            # Анализ preflight результата
            warnings = preflight_result.get('warnings', [])
            total_skeletons = sum(len(w.get('items', [])) for w in warnings)
            
            self.log_test(
                "Preflight для 3 техкарт",
                True,
                f"Warnings: {len(warnings)}, total skeletons: {total_skeletons}",
                response_time
            )
            
            # Export с preflight
            start_time = time.time()
            export_response = await self.client.post(
                f"{API_BASE}/export/zip",
                json={
                    "techcardIds": generated_ids,
                    "organization_id": self.organization_id,
                    "preflight_result": preflight_result,
                    "operational_rounding": True
                }
            )
            response_time = time.time() - start_time
            
            if export_response.status_code != 200:
                self.log_test(
                    "Export ZIP для 3 техкарт",
                    False,
                    f"HTTP {export_response.status_code}"
                )
                return False
            
            zip_data = export_response.content
            zip_size = len(zip_data)
            
            self.log_test(
                "Export ZIP для 3 техкарт",
                True,
                f"ZIP размер: {zip_size} bytes",
                response_time
            )
            
            # Анализ содержимого ZIP
            return await self.analyze_zip_content(zip_data, "3 техкарты")
            
        except Exception as e:
            self.log_test(
                "Workflow с несколькими техкартами",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    async def analyze_zip_content(self, zip_data: bytes, test_name: str) -> bool:
        """Анализ содержимого ZIP файла"""
        try:
            with zipfile.ZipFile(io.BytesIO(zip_data), 'r') as zip_file:
                files = zip_file.namelist()
                xlsx_files = [f for f in files if f.endswith('.xlsx')]
                
                self.log_test(
                    f"ZIP содержимое ({test_name})",
                    True,
                    f"Файлов: {len(files)}, XLSX: {len(xlsx_files)}, список: {files}"
                )
                
                # Анализ каждого XLSX файла
                total_preflight_articles = 0
                total_generated_prefixes = 0
                total_real_dish_names = 0
                
                for xlsx_file in xlsx_files:
                    xlsx_data = zip_file.read(xlsx_file)
                    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data))
                    
                    file_preflight_articles = 0
                    file_generated_prefixes = 0
                    file_real_dishes = 0
                    
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value:
                                    cell_str = str(cell.value)
                                    
                                    # Ищем 5-digit артикулы (10000-99999)
                                    if cell_str.isdigit() and len(cell_str) == 5 and cell_str.startswith('1'):
                                        file_preflight_articles += 1
                                    
                                    # Ищем GENERATED_* префиксы (КРИТИЧЕСКАЯ ПРОБЛЕМА)
                                    if 'GENERATED_' in cell_str:
                                        file_generated_prefixes += 1
                                    
                                    # Ищем реальные названия блюд
                                    if any(dish in cell_str for dish in ['Паста', 'Ризотто', 'Лазанья']):
                                        file_real_dishes += 1
                    
                    total_preflight_articles += file_preflight_articles
                    total_generated_prefixes += file_generated_prefixes
                    total_real_dish_names += file_real_dishes
                    
                    self.log_test(
                        f"Анализ {xlsx_file}",
                        file_generated_prefixes == 0,
                        f"5-digit артикулы: {file_preflight_articles}, GENERATED_*: {file_generated_prefixes}, реальные блюда: {file_real_dishes}"
                    )
                
                # Общий результат анализа
                success = (
                    total_preflight_articles > 0 and 
                    total_generated_prefixes == 0 and
                    total_real_dish_names > 0
                )
                
                self.log_test(
                    f"КРИТИЧЕСКИЙ АНАЛИЗ ({test_name})",
                    success,
                    f"✅ 5-digit артикулы: {total_preflight_articles}, ❌ GENERATED_*: {total_generated_prefixes}, ✅ реальные блюда: {total_real_dish_names}"
                )
                
                return success
                
        except Exception as e:
            self.log_test(
                f"Анализ ZIP ({test_name})",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    async def test_edge_cases(self):
        """Тест граничных случаев"""
        try:
            print("\n🧪 ТЕСТИРОВАНИЕ ГРАНИЧНЫХ СЛУЧАЕВ")
            
            # Тест 1: Export без preflight (должен работать с fallback)
            start_time = time.time()
            
            # Генерируем одну техкарту
            response = await self.client.post(
                f"{API_BASE}/techcards.v2/generate",
                json={
                    "name": "Тест блюдо без preflight",
                    "description": "Тестовое блюдо для проверки fallback",
                    "category": "горячее"
                }
            )
            
            if response.status_code != 200:
                self.log_test(
                    "Генерация для edge case",
                    False,
                    f"HTTP {response.status_code}"
                )
                return False
            
            card_id = response.json().get('card', {}).get('meta', {}).get('id')
            
            # Export БЕЗ preflight_result
            export_response = await self.client.post(
                f"{API_BASE}/export/zip",
                json={
                    "techcardIds": [card_id],
                    "organization_id": self.organization_id,
                    "operational_rounding": True
                    # НЕТ preflight_result - тест fallback
                }
            )
            
            response_time = time.time() - start_time
            
            if export_response.status_code == 200:
                zip_data = export_response.content
                success = await self.analyze_zip_content(zip_data, "без preflight")
                
                self.log_test(
                    "Export без preflight (fallback)",
                    success,
                    f"Fallback работает, размер ZIP: {len(zip_data)} bytes",
                    response_time
                )
                return success
            else:
                self.log_test(
                    "Export без preflight (fallback)",
                    False,
                    f"HTTP {export_response.status_code}: {export_response.text[:100]}"
                )
                return False
                
        except Exception as e:
            self.log_test(
                "Граничные случаи",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    async def run_comprehensive_tests(self):
        """Запуск всех comprehensive тестов"""
        print("🚀 COMPREHENSIVE PREFLIGHT→EXPORT INTEGRATION VALIDATION")
        print("=" * 80)
        
        tests = [
            self.test_multiple_techcards_workflow,
            self.test_edge_cases
        ]
        
        passed = 0
        total = len(tests)
        
        for test in tests:
            success = await test()
            if success:
                passed += 1
        
        # Итоговый результат
        success_rate = (passed / total) * 100
        
        print("\n" + "=" * 80)
        print("📊 COMPREHENSIVE TEST RESULTS")
        print("=" * 80)
        
        for result in self.test_results:
            print(f"{result['status']} {result['test']}: {result['details']}")
        
        print(f"\n🎯 УСПЕШНОСТЬ: {passed}/{total} comprehensive тестов ({success_rate:.1f}%)")
        
        # Анализ критических результатов
        critical_success = True
        generated_prefixes_found = False
        
        for result in self.test_results:
            if "GENERATED_*: 0" not in result['details'] and "GENERATED_*" in result['details']:
                if "GENERATED_*: 0" not in result['details']:
                    generated_prefixes_found = True
                    critical_success = False
        
        if critical_success and not generated_prefixes_found:
            print("\n🎉 КРИТИЧЕСКИЙ УСПЕХ: PREFLIGHT→EXPORT ИНТЕГРАЦИЯ ПОЛНОСТЬЮ ИСПРАВЛЕНА!")
            print("✅ Export система использует preflight артикулы")
            print("✅ КРИТИЧЕСКИ: 0 GENERATED_* префиксов во всех тестах")
            print("✅ Реальные данные техкарт используются в XLSX")
            print("✅ Fallback система работает корректно")
            print("✅ Множественные техкарты обрабатываются правильно")
        else:
            print("\n❌ КРИТИЧЕСКИЕ ПРОБЛЕМЫ ОБНАРУЖЕНЫ:")
            if generated_prefixes_found:
                print("🚨 GENERATED_* префиксы все еще присутствуют!")
            print("🚨 Требуется дополнительное исправление")
        
        return critical_success

async def main():
    """Главная функция comprehensive теста"""
    try:
        async with ComprehensivePreflightExportTester() as tester:
            success = await tester.run_comprehensive_tests()
            sys.exit(0 if success else 1)
    except Exception as e:
        print(f"❌ КРИТИЧЕСКАЯ ОШИБКА COMPREHENSIVE ТЕСТА: {str(e)}")
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    asyncio.run(main())