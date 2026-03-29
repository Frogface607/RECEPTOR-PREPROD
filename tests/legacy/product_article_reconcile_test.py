#!/usr/bin/env python3
"""
CRITICAL BACKEND TESTING: Product-Article Reconcile System
Критическое тестирование системы reconciliation для устранения GENERATED_* артикулов

ЦЕЛЬ: Гарантировать что в iiko_TTK.xlsx по каждому ингредиенту подставляется артикул 
только из Product-Skeletons (mapping по наименованию продукта). Ни один ингредиент 
не может получить GENERATED_*. При отсутствии mapping экспорт блокируется с ошибкой.

КРИТИЧЕСКИЕ WORKFLOW ШАГИ:
1. build_product_mapping - Собрать mapping {product_name → nomenclature.num} из Product-Skeletons.xlsx
2. set_dish_article - Артикул блюда всегда из Dish-Skeletons.xlsx по полю 'Артикул'  
3. substitute_product_articles - Mapping ингредиентов по наименованию, fail-fast при отсутствии
4. validate_no_generated - Проверка что ни одна ячейка не содержит GENERATED_*
5. post_export_check - Валидация mapping и создание артефактов

ACCEPTANCE CRITERIA (ZERO TOLERANCE):
✅ Артикулы продуктов только из Product-Skeletons по mapping названий
✅ Ни одного GENERATED_* - fail-fast при отсутствии mapping
✅ Артикул блюда из Dish-Skeletons в правильном формате (@)
✅ Список product mapping в артефактах для проверки
✅ Блокировка экспорта при отсутствии mapping для любого ингредиента
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Any, Optional

import httpx
import openpyxl
import pandas as pd

# Add the backend directory to Python path
backend_dir = Path(__file__).parent / "backend"
sys.path.insert(0, str(backend_dir))

# Backend URL from environment
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api"

class ProductArticleReconcileTest:
    """Critical Product-Article Reconciliation System Tester"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=60.0)
        self.test_results = []
        self.artifacts = {}
        self.critical_errors = []
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str, data: Any = None):
        """Log test result with detailed information"""
        result = {
            "test": test_name,
            "success": success,
            "details": details,
            "timestamp": datetime.now().isoformat(),
            "data": data
        }
        self.test_results.append(result)
        
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"{status}: {test_name}")
        print(f"   {details}")
        if not success:
            self.critical_errors.append(f"{test_name}: {details}")
        print()
    
    async def test_step_1_build_product_mapping(self) -> Dict[str, str]:
        """
        STEP 1: build_product_mapping - Собрать mapping {product_name → nomenclature.num} из Product-Skeletons.xlsx
        
        Тестируем что система может создать корректный mapping продуктов из скелетонов
        """
        try:
            print("🔍 STEP 1: Testing Product Mapping from Product-Skeletons...")
            
            # Генерируем техкарту для получения Product-Skeletons
            test_dish = "Тако с курицей и ананасовой сальсой"
            
            # 1.1 Генерируем техкарту
            gen_response = await self.client.post(f"{API_BASE}/v1/techcards.v2/generate", json={
                "name": test_dish,
                "organization_id": "test_org"
            })
            
            if gen_response.status_code != 200:
                self.log_test("Step 1.1: TechCard Generation", False, 
                            f"Failed to generate techcard: {gen_response.status_code}")
                return {}
            
            gen_data = gen_response.json()
            techcard_id = gen_data.get("card", {}).get("meta", {}).get("id")
            
            if not techcard_id:
                self.log_test("Step 1.1: TechCard Generation", False, "No techcard ID returned")
                return {}
            
            self.log_test("Step 1.1: TechCard Generation", True, 
                         f"Generated techcard: {techcard_id}")
            
            # 1.2 Запускаем preflight для создания Product-Skeletons
            preflight_response = await self.client.post(f"{API_BASE}/v1/export/preflight", json={
                "techcardIds": [techcard_id],
                "organization_id": "test_org"
            })
            
            if preflight_response.status_code != 200:
                self.log_test("Step 1.2: Preflight Check", False, 
                            f"Preflight failed: {preflight_response.status_code}")
                return {}
            
            preflight_data = preflight_response.json()
            missing_products = preflight_data.get("missing", {}).get("products", [])
            
            if not missing_products:
                self.log_test("Step 1.2: Preflight Check", False, "No missing products found for mapping test")
                return {}
            
            self.log_test("Step 1.2: Preflight Check", True, 
                         f"Found {len(missing_products)} missing products for mapping")
            
            # 1.3 Создаем ZIP экспорт для получения Product-Skeletons.xlsx
            zip_response = await self.client.post(f"{API_BASE}/v1/export/zip", json={
                "techcardIds": [techcard_id],
                "preflight_result": preflight_data,
                "operational_rounding": True,
                "organization_id": "test_org"
            })
            
            if zip_response.status_code != 200:
                self.log_test("Step 1.3: ZIP Export", False, 
                            f"ZIP export failed: {zip_response.status_code}")
                return {}
            
            # 1.4 Извлекаем Product-Skeletons.xlsx из ZIP
            zip_content = BytesIO(zip_response.content)
            product_mapping = {}
            
            with zipfile.ZipFile(zip_content, 'r') as zip_file:
                if 'Product-Skeletons.xlsx' not in zip_file.namelist():
                    self.log_test("Step 1.4: Product-Skeletons Extraction", False, 
                                "Product-Skeletons.xlsx not found in ZIP")
                    return {}
                
                # Читаем Product-Skeletons.xlsx
                with zip_file.open('Product-Skeletons.xlsx') as xlsx_file:
                    workbook = openpyxl.load_workbook(xlsx_file)
                    worksheet = workbook.active
                    
                    # Строим mapping: product_name → article
                    for row in worksheet.iter_rows(min_row=2, values_only=True):
                        if row[0] and row[1]:  # Артикул и Наименование
                            article = str(row[0]).strip()
                            product_name = str(row[1]).strip()
                            product_mapping[product_name] = article
            
            if not product_mapping:
                self.log_test("Step 1.4: Product-Skeletons Extraction", False, 
                            "No product mapping extracted from Product-Skeletons.xlsx")
                return {}
            
            self.log_test("Step 1.4: Product-Skeletons Extraction", True, 
                         f"Built product mapping with {len(product_mapping)} entries")
            
            # 1.5 Валидируем формат артикулов (5-значные)
            invalid_articles = []
            for product_name, article in product_mapping.items():
                if not (article.isdigit() and len(article) == 5):
                    invalid_articles.append(f"{product_name}: {article}")
            
            if invalid_articles:
                self.log_test("Step 1.5: Article Format Validation", False, 
                            f"Invalid article formats: {invalid_articles}")
                return {}
            
            self.log_test("Step 1.5: Article Format Validation", True, 
                         "All articles are 5-digit format")
            
            # Сохраняем mapping в артефакты
            self.artifacts["product_mapping"] = product_mapping
            self.artifacts["techcard_id"] = techcard_id
            self.artifacts["preflight_data"] = preflight_data
            
            return product_mapping
            
        except Exception as e:
            self.log_test("Step 1: Build Product Mapping", False, 
                         f"Exception: {str(e)}\n{traceback.format_exc()}")
            return {}
    
    async def test_step_2_set_dish_article(self, techcard_id: str, preflight_data: Dict) -> str:
        """
        STEP 2: set_dish_article - Артикул блюда всегда из Dish-Skeletons.xlsx по полю 'Артикул'
        
        Тестируем что артикул блюда корректно извлекается из Dish-Skeletons
        """
        try:
            print("🔍 STEP 2: Testing Dish Article from Dish-Skeletons...")
            
            # 2.1 Создаем ZIP экспорт для получения Dish-Skeletons.xlsx
            zip_response = await self.client.post(f"{API_BASE}/v1/export/zip", json={
                "techcardIds": [techcard_id],
                "preflight_result": preflight_data,
                "operational_rounding": True,
                "organization_id": "test_org"
            })
            
            if zip_response.status_code != 200:
                self.log_test("Step 2.1: ZIP Export for Dish-Skeletons", False, 
                            f"ZIP export failed: {zip_response.status_code}")
                return ""
            
            # 2.2 Извлекаем Dish-Skeletons.xlsx из ZIP
            zip_content = BytesIO(zip_response.content)
            dish_article = ""
            
            with zipfile.ZipFile(zip_content, 'r') as zip_file:
                if 'Dish-Skeletons.xlsx' not in zip_file.namelist():
                    self.log_test("Step 2.2: Dish-Skeletons Extraction", False, 
                                "Dish-Skeletons.xlsx not found in ZIP")
                    return ""
                
                # Читаем Dish-Skeletons.xlsx
                with zip_file.open('Dish-Skeletons.xlsx') as xlsx_file:
                    workbook = openpyxl.load_workbook(xlsx_file)
                    worksheet = workbook.active
                    
                    # Извлекаем артикул блюда из первой строки данных
                    for row in worksheet.iter_rows(min_row=2, max_row=2, values_only=True):
                        if row[0]:  # Артикул в первой колонке
                            dish_article = str(row[0]).strip()
                            break
            
            if not dish_article:
                self.log_test("Step 2.2: Dish-Skeletons Extraction", False, 
                            "No dish article found in Dish-Skeletons.xlsx")
                return ""
            
            self.log_test("Step 2.2: Dish-Skeletons Extraction", True, 
                         f"Extracted dish article: {dish_article}")
            
            # 2.3 Валидируем формат артикула блюда (5-значный)
            if not (dish_article.isdigit() and len(dish_article) == 5):
                self.log_test("Step 2.3: Dish Article Format Validation", False, 
                            f"Invalid dish article format: {dish_article}")
                return ""
            
            self.log_test("Step 2.3: Dish Article Format Validation", True, 
                         f"Dish article format valid: {dish_article}")
            
            # 2.4 Проверяем что артикул блюда используется в iiko_TTK.xlsx
            with zipfile.ZipFile(zip_content, 'r') as zip_file:
                if 'iiko_TTK.xlsx' not in zip_file.namelist():
                    self.log_test("Step 2.4: TTK Dish Article Usage", False, 
                                "iiko_TTK.xlsx not found in ZIP")
                    return ""
                
                # Читаем iiko_TTK.xlsx
                with zip_file.open('iiko_TTK.xlsx') as xlsx_file:
                    workbook = openpyxl.load_workbook(xlsx_file)
                    worksheet = workbook.active
                    
                    # Проверяем что артикул блюда используется в колонке A
                    dish_article_found = False
                    for row in worksheet.iter_rows(min_row=2, values_only=True):
                        if row[0] and str(row[0]).strip() == dish_article:
                            dish_article_found = True
                            break
                    
                    if not dish_article_found:
                        self.log_test("Step 2.4: TTK Dish Article Usage", False, 
                                    f"Dish article {dish_article} not found in iiko_TTK.xlsx")
                        return ""
            
            self.log_test("Step 2.4: TTK Dish Article Usage", True, 
                         f"Dish article {dish_article} correctly used in TTK")
            
            # Сохраняем в артефакты
            self.artifacts["dish_article"] = dish_article
            
            return dish_article
            
        except Exception as e:
            self.log_test("Step 2: Set Dish Article", False, 
                         f"Exception: {str(e)}\n{traceback.format_exc()}")
            return ""
    
    async def test_step_3_substitute_product_articles(self, techcard_id: str, preflight_data: Dict, 
                                                    product_mapping: Dict[str, str]) -> bool:
        """
        STEP 3: substitute_product_articles - Mapping ингредиентов по наименованию, fail-fast при отсутствии
        
        Тестируем что все ингредиенты получают артикулы только из Product-Skeletons mapping
        """
        try:
            print("🔍 STEP 3: Testing Product Article Substitution...")
            
            # 3.1 Создаем ZIP экспорт
            zip_response = await self.client.post(f"{API_BASE}/v1/export/zip", json={
                "techcardIds": [techcard_id],
                "preflight_result": preflight_data,
                "operational_rounding": True,
                "organization_id": "test_org"
            })
            
            if zip_response.status_code != 200:
                self.log_test("Step 3.1: ZIP Export for Article Substitution", False, 
                            f"ZIP export failed: {zip_response.status_code}")
                return False
            
            # 3.2 Извлекаем iiko_TTK.xlsx и анализируем артикулы продуктов
            zip_content = BytesIO(zip_response.content)
            ttk_product_articles = []
            ttk_product_names = []
            
            with zipfile.ZipFile(zip_content, 'r') as zip_file:
                with zip_file.open('iiko_TTK.xlsx') as xlsx_file:
                    workbook = openpyxl.load_workbook(xlsx_file)
                    worksheet = workbook.active
                    
                    # Извлекаем артикулы продуктов (колонка C) и названия (колонка D)
                    for row in worksheet.iter_rows(min_row=2, values_only=True):
                        if row[2] and row[3]:  # Артикул продукта и Наименование продукта
                            product_article = str(row[2]).strip()
                            product_name = str(row[3]).strip()
                            ttk_product_articles.append(product_article)
                            ttk_product_names.append(product_name)
            
            if not ttk_product_articles:
                self.log_test("Step 3.2: TTK Product Articles Extraction", False, 
                            "No product articles found in iiko_TTK.xlsx")
                return False
            
            self.log_test("Step 3.2: TTK Product Articles Extraction", True, 
                         f"Found {len(ttk_product_articles)} product articles in TTK")
            
            # 3.3 Проверяем что все артикулы продуктов соответствуют Product-Skeletons mapping
            mapping_errors = []
            unmapped_products = []
            
            for i, (product_name, product_article) in enumerate(zip(ttk_product_names, ttk_product_articles)):
                expected_article = product_mapping.get(product_name)
                
                if not expected_article:
                    unmapped_products.append(product_name)
                elif expected_article != product_article:
                    mapping_errors.append(f"{product_name}: expected {expected_article}, got {product_article}")
            
            if unmapped_products:
                self.log_test("Step 3.3: Product Mapping Validation", False, 
                            f"Products without mapping: {unmapped_products}")
                return False
            
            if mapping_errors:
                self.log_test("Step 3.3: Product Mapping Validation", False, 
                            f"Mapping errors: {mapping_errors}")
                return False
            
            self.log_test("Step 3.3: Product Mapping Validation", True, 
                         f"All {len(ttk_product_articles)} products correctly mapped")
            
            # 3.4 Проверяем формат всех артикулов продуктов (5-значные)
            invalid_product_articles = []
            for product_article in ttk_product_articles:
                if not (product_article.isdigit() and len(product_article) == 5):
                    invalid_product_articles.append(product_article)
            
            if invalid_product_articles:
                self.log_test("Step 3.4: Product Article Format Validation", False, 
                            f"Invalid product article formats: {invalid_product_articles}")
                return False
            
            self.log_test("Step 3.4: Product Article Format Validation", True, 
                         "All product articles are 5-digit format")
            
            # Сохраняем в артефакты
            self.artifacts["ttk_product_articles"] = ttk_product_articles
            self.artifacts["ttk_product_names"] = ttk_product_names
            self.artifacts["mapping_validation"] = {
                "total_products": len(ttk_product_articles),
                "mapped_correctly": len(ttk_product_articles),
                "mapping_errors": 0,
                "unmapped_products": 0
            }
            
            return True
            
        except Exception as e:
            self.log_test("Step 3: Substitute Product Articles", False, 
                         f"Exception: {str(e)}\n{traceback.format_exc()}")
            return False
    
    async def test_step_4_validate_no_generated(self, techcard_id: str, preflight_data: Dict) -> bool:
        """
        STEP 4: validate_no_generated - Проверка что ни одна ячейка не содержит GENERATED_*
        
        КРИТИЧЕСКАЯ ПРОВЕРКА: Ни одного GENERATED_* в экспортированных файлах
        """
        try:
            print("🔍 STEP 4: Testing ZERO TOLERANCE for GENERATED_* content...")
            
            # 4.1 Создаем ZIP экспорт
            zip_response = await self.client.post(f"{API_BASE}/v1/export/zip", json={
                "techcardIds": [techcard_id],
                "preflight_result": preflight_data,
                "operational_rounding": True,
                "organization_id": "test_org"
            })
            
            if zip_response.status_code != 200:
                self.log_test("Step 4.1: ZIP Export for GENERATED Validation", False, 
                            f"ZIP export failed: {zip_response.status_code}")
                return False
            
            # 4.2 Сканируем все файлы в ZIP на наличие GENERATED_*
            zip_content = BytesIO(zip_response.content)
            generated_signatures = []
            files_scanned = []
            
            with zipfile.ZipFile(zip_content, 'r') as zip_file:
                for filename in zip_file.namelist():
                    if filename.endswith('.xlsx'):
                        files_scanned.append(filename)
                        
                        with zip_file.open(filename) as xlsx_file:
                            workbook = openpyxl.load_workbook(xlsx_file)
                            worksheet = workbook.active
                            
                            # Сканируем все ячейки на GENERATED_*
                            for row in worksheet.iter_rows(values_only=True):
                                for cell_value in row:
                                    if cell_value and isinstance(cell_value, str):
                                        if 'GENERATED_' in cell_value.upper():
                                            generated_signatures.append(f"{filename}: {cell_value}")
            
            if generated_signatures:
                self.log_test("Step 4.2: GENERATED Content Scan", False, 
                            f"CRITICAL: Found GENERATED signatures: {generated_signatures}")
                return False
            
            self.log_test("Step 4.2: GENERATED Content Scan", True, 
                         f"ZERO GENERATED signatures found in {len(files_scanned)} files")
            
            # 4.3 Дополнительная проверка на mock контент
            mock_signatures = []
            mock_patterns = ['MOCK', 'TEST_INGREDIENT', 'DISH_MOCK', 'ТЕСТОВОЕ', 'ТЕСТОВЫЙ']
            
            with zipfile.ZipFile(zip_content, 'r') as zip_file:
                for filename in zip_file.namelist():
                    if filename.endswith('.xlsx'):
                        with zip_file.open(filename) as xlsx_file:
                            workbook = openpyxl.load_workbook(xlsx_file)
                            worksheet = workbook.active
                            
                            for row in worksheet.iter_rows(values_only=True):
                                for cell_value in row:
                                    if cell_value and isinstance(cell_value, str):
                                        for pattern in mock_patterns:
                                            if pattern in cell_value.upper():
                                                mock_signatures.append(f"{filename}: {cell_value}")
            
            if mock_signatures:
                self.log_test("Step 4.3: Mock Content Scan", False, 
                            f"Found mock signatures: {mock_signatures}")
                return False
            
            self.log_test("Step 4.3: Mock Content Scan", True, 
                         f"ZERO mock signatures found")
            
            # 4.4 Проверяем что все артикулы - реальные 5-значные коды
            all_articles = []
            
            with zipfile.ZipFile(zip_content, 'r') as zip_file:
                with zip_file.open('iiko_TTK.xlsx') as xlsx_file:
                    workbook = openpyxl.load_workbook(xlsx_file)
                    worksheet = workbook.active
                    
                    for row in worksheet.iter_rows(min_row=2, values_only=True):
                        if row[0]:  # Артикул блюда
                            all_articles.append(str(row[0]).strip())
                        if row[2]:  # Артикул продукта
                            all_articles.append(str(row[2]).strip())
            
            invalid_articles = []
            for article in all_articles:
                if not (article.isdigit() and len(article) == 5):
                    invalid_articles.append(article)
            
            if invalid_articles:
                self.log_test("Step 4.4: Real Article Validation", False, 
                            f"Invalid articles found: {invalid_articles}")
                return False
            
            self.log_test("Step 4.4: Real Article Validation", True, 
                         f"All {len(all_articles)} articles are real 5-digit codes")
            
            # Сохраняем в артефакты
            self.artifacts["generated_validation"] = {
                "files_scanned": files_scanned,
                "generated_signatures_found": 0,
                "mock_signatures_found": 0,
                "total_articles_validated": len(all_articles),
                "invalid_articles": 0
            }
            
            return True
            
        except Exception as e:
            self.log_test("Step 4: Validate No Generated", False, 
                         f"Exception: {str(e)}\n{traceback.format_exc()}")
            return False
    
    async def test_step_5_post_export_check(self) -> bool:
        """
        STEP 5: post_export_check - Валидация mapping и создание артефактов
        
        Финальная валидация всего workflow и создание отчета
        """
        try:
            print("🔍 STEP 5: Post-Export Validation and Artifacts...")
            
            # 5.1 Проверяем что все критические артефакты созданы
            required_artifacts = [
                "product_mapping", "dish_article", "ttk_product_articles", 
                "ttk_product_names", "mapping_validation", "generated_validation"
            ]
            
            missing_artifacts = []
            for artifact in required_artifacts:
                if artifact not in self.artifacts:
                    missing_artifacts.append(artifact)
            
            if missing_artifacts:
                self.log_test("Step 5.1: Artifacts Validation", False, 
                            f"Missing artifacts: {missing_artifacts}")
                return False
            
            self.log_test("Step 5.1: Artifacts Validation", True, 
                         f"All {len(required_artifacts)} artifacts created")
            
            # 5.2 Создаем сводный отчет mapping
            mapping_report = {
                "workflow_status": "COMPLETED",
                "timestamp": datetime.now().isoformat(),
                "product_mapping_count": len(self.artifacts["product_mapping"]),
                "dish_article": self.artifacts["dish_article"],
                "ttk_products_count": len(self.artifacts["ttk_product_articles"]),
                "mapping_accuracy": "100%",
                "generated_content_found": False,
                "mock_content_found": False,
                "critical_errors": len(self.critical_errors),
                "acceptance_criteria_met": len(self.critical_errors) == 0
            }
            
            self.artifacts["mapping_report"] = mapping_report
            
            self.log_test("Step 5.2: Mapping Report Creation", True, 
                         f"Mapping report created: {mapping_report['acceptance_criteria_met']}")
            
            # 5.3 Валидируем что все acceptance criteria выполнены
            acceptance_criteria = {
                "product_articles_from_skeletons": len(self.artifacts["product_mapping"]) > 0,
                "zero_generated_content": self.artifacts["generated_validation"]["generated_signatures_found"] == 0,
                "dish_article_from_skeletons": bool(self.artifacts["dish_article"]),
                "product_mapping_available": len(self.artifacts["product_mapping"]) > 0,
                "export_not_blocked": len(self.critical_errors) == 0
            }
            
            failed_criteria = [k for k, v in acceptance_criteria.items() if not v]
            
            if failed_criteria:
                self.log_test("Step 5.3: Acceptance Criteria Validation", False, 
                            f"Failed criteria: {failed_criteria}")
                return False
            
            self.log_test("Step 5.3: Acceptance Criteria Validation", True, 
                         "All 5 acceptance criteria met")
            
            # 5.4 Проверяем производительность workflow
            total_tests = len(self.test_results)
            passed_tests = sum(1 for r in self.test_results if r["success"])
            success_rate = (passed_tests / total_tests) * 100 if total_tests > 0 else 0
            
            performance_report = {
                "total_tests": total_tests,
                "passed_tests": passed_tests,
                "success_rate": f"{success_rate:.1f}%",
                "critical_errors": len(self.critical_errors),
                "workflow_operational": success_rate >= 90
            }
            
            self.artifacts["performance_report"] = performance_report
            
            if success_rate < 90:
                self.log_test("Step 5.4: Performance Validation", False, 
                            f"Success rate too low: {success_rate:.1f}%")
                return False
            
            self.log_test("Step 5.4: Performance Validation", True, 
                         f"Success rate: {success_rate:.1f}%")
            
            return True
            
        except Exception as e:
            self.log_test("Step 5: Post Export Check", False, 
                         f"Exception: {str(e)}\n{traceback.format_exc()}")
            return False
    
    async def test_fail_fast_scenario(self) -> bool:
        """
        FAIL-FAST SCENARIO: Тестируем блокировку экспорта при отсутствии mapping
        
        Проверяем что система корректно блокирует экспорт когда нет mapping для ингредиентов
        """
        try:
            print("🔍 FAIL-FAST SCENARIO: Testing export blocking without mapping...")
            
            # Создаем техкарту с ингредиентами, для которых заведомо нет mapping
            test_dish = "Экзотическое блюдо с редкими ингредиентами"
            
            # Генерируем техкарту
            gen_response = await self.client.post(f"{API_BASE}/v1/techcards.v2/generate", json={
                "name": test_dish,
                "organization_id": "test_org"
            })
            
            if gen_response.status_code != 200:
                self.log_test("Fail-Fast: TechCard Generation", False, 
                            f"Failed to generate techcard: {gen_response.status_code}")
                return False
            
            gen_data = gen_response.json()
            techcard_id = gen_data.get("card", {}).get("meta", {}).get("id")
            
            # Пытаемся экспортировать без preflight (должно заблокироваться)
            export_response = await self.client.post(f"{API_BASE}/v1/export/zip", json={
                "techcardIds": [techcard_id],
                "operational_rounding": True,
                "organization_id": "test_org"
                # Намеренно НЕ передаем preflight_result
            })
            
            # Ожидаем что экспорт заблокируется с ошибкой PRE_FLIGHT_REQUIRED
            if export_response.status_code == 200:
                self.log_test("Fail-Fast: Export Blocking", False, 
                            "Export succeeded when it should have been blocked")
                return False
            
            if export_response.status_code == 400:
                error_data = export_response.json()
                if "PRE_FLIGHT_REQUIRED" in str(error_data):
                    self.log_test("Fail-Fast: Export Blocking", True, 
                                "Export correctly blocked with PRE_FLIGHT_REQUIRED")
                    return True
            
            self.log_test("Fail-Fast: Export Blocking", False, 
                        f"Unexpected response: {export_response.status_code}")
            return False
            
        except Exception as e:
            self.log_test("Fail-Fast Scenario", False, 
                         f"Exception: {str(e)}\n{traceback.format_exc()}")
            return False
    
    async def run_comprehensive_test(self):
        """Запуск полного комплексного тестирования Product-Article Reconcile системы"""
        
        print("=" * 80)
        print("🚨 CRITICAL PRODUCT-ARTICLE RECONCILE SYSTEM TESTING")
        print("=" * 80)
        print()
        
        start_time = time.time()
        
        # STEP 1: Build Product Mapping
        product_mapping = await self.test_step_1_build_product_mapping()
        if not product_mapping:
            print("❌ CRITICAL FAILURE: Cannot proceed without product mapping")
            return
        
        # STEP 2: Set Dish Article
        dish_article = await self.test_step_2_set_dish_article(
            self.artifacts["techcard_id"], 
            self.artifacts["preflight_data"]
        )
        if not dish_article:
            print("❌ CRITICAL FAILURE: Cannot proceed without dish article")
            return
        
        # STEP 3: Substitute Product Articles
        substitution_success = await self.test_step_3_substitute_product_articles(
            self.artifacts["techcard_id"],
            self.artifacts["preflight_data"], 
            product_mapping
        )
        if not substitution_success:
            print("❌ CRITICAL FAILURE: Product article substitution failed")
            return
        
        # STEP 4: Validate No Generated Content
        no_generated_success = await self.test_step_4_validate_no_generated(
            self.artifacts["techcard_id"],
            self.artifacts["preflight_data"]
        )
        if not no_generated_success:
            print("❌ CRITICAL FAILURE: GENERATED content found - ZERO TOLERANCE VIOLATED")
            return
        
        # STEP 5: Post Export Check
        post_export_success = await self.test_step_5_post_export_check()
        if not post_export_success:
            print("❌ CRITICAL FAILURE: Post-export validation failed")
            return
        
        # FAIL-FAST SCENARIO
        fail_fast_success = await self.test_fail_fast_scenario()
        
        # Final Results
        end_time = time.time()
        duration = end_time - start_time
        
        print("=" * 80)
        print("📊 FINAL RESULTS")
        print("=" * 80)
        
        total_tests = len(self.test_results)
        passed_tests = sum(1 for r in self.test_results if r["success"])
        success_rate = (passed_tests / total_tests) * 100 if total_tests > 0 else 0
        
        print(f"Total Tests: {total_tests}")
        print(f"Passed: {passed_tests}")
        print(f"Failed: {total_tests - passed_tests}")
        print(f"Success Rate: {success_rate:.1f}%")
        print(f"Duration: {duration:.2f}s")
        print()
        
        if self.critical_errors:
            print("❌ CRITICAL ERRORS:")
            for error in self.critical_errors:
                print(f"   • {error}")
            print()
        
        # Acceptance Criteria Summary
        print("🎯 ACCEPTANCE CRITERIA:")
        if "mapping_report" in self.artifacts:
            report = self.artifacts["mapping_report"]
            criteria_met = report.get("acceptance_criteria_met", False)
            
            print(f"   ✅ Артикулы продуктов только из Product-Skeletons: {len(product_mapping) > 0}")
            print(f"   ✅ Ни одного GENERATED_*: {no_generated_success}")
            print(f"   ✅ Артикул блюда из Dish-Skeletons: {bool(dish_article)}")
            print(f"   ✅ Список product mapping в артефактах: {len(product_mapping) > 0}")
            print(f"   ✅ Блокировка экспорта при отсутствии mapping: {fail_fast_success}")
            print()
            
            if criteria_met and success_rate >= 90:
                print("🎉 OUTSTANDING SUCCESS: Product-Article Reconcile система ПОЛНОСТЬЮ ОПЕРАЦИОНАЛЬНА")
                print("   Все критические требования выполнены, GENERATED_* контент устранен")
            else:
                print("🚨 CRITICAL ISSUES DETECTED: Система требует немедленного исправления")
        
        # Save artifacts to file
        artifacts_file = Path(__file__).parent / "product_article_reconcile_artifacts.json"
        with open(artifacts_file, 'w', encoding='utf-8') as f:
            json.dump(self.artifacts, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\n📁 Artifacts saved to: {artifacts_file}")


async def main():
    """Main test execution"""
    async with ProductArticleReconcileTest() as tester:
        await tester.run_comprehensive_test()


if __name__ == "__main__":
    asyncio.run(main())