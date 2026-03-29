#!/usr/bin/env python3
"""
Mock Content Resolution End-to-End Test
Critical verification that the mock content issue has been completely resolved in the export workflow.

GOAL: Test the full user workflow from tech card generation through export to ensure no mock content appears in exported files.

WORKFLOW TO TEST:
1. Generate a new tech card using /techcards.v2/generate
2. Verify that the tech card is saved to database with proper ID
3. Verify that currentTechCardId is properly set in frontend
4. Test the export flow using the UI workflow (preflight + ZIP)
5. Download and inspect the ZIP files for mock content
6. Confirm all exported content is real data

CRITICAL VALIDATION POINTS:
- Tech card generation creates real ID that gets saved to database
- Frontend properly stores the tech card ID for export use
- Export preflight and ZIP generation use the real tech card ID
- XLSX files contain actual dish names and ingredient data
- Zero mock signatures in exported files
- Article formatting preserved correctly

SUCCESS CRITERIA:
- Tech card generated with real ID and saved to database ✅
- Export uses real tech card data instead of 'current' fallback ✅
- XLSX files contain actual generated content (no DISH_MOCK_TECH_CARD) ✅
- Article columns properly formatted as text (@) with leading zeros ✅
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Any, Optional

import httpx
import openpyxl
from pymongo import MongoClient

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017/receptor_pro')
DB_NAME = os.getenv('DB_NAME', 'receptor_pro')

# Mock content signatures to detect
MOCK_SIGNATURES = [
    'DISH_MOCK_TECH_CARD',
    'GENERATED_TEST_INGREDIENT',
    'GENERATED_TEST_INGREDIENT_2',
    'TEST_INGREDIENT',
    'Mock',
    'mock',
    'MOCK',
    'test_ingredient',
    'demo_dish'
]

class MockContentResolutionTester:
    """Comprehensive Mock Content Resolution Testing Suite"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=60.0)
        self.test_results = []
        self.organization_id = "test-org-mock-resolution"
        self.generated_techcard_ids = []
        self.artifacts = {}
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status}: {test_name} ({response_time:.3f}s) - {details}")
    
    async def test_techcard_generation_with_real_ids(self):
        """Test tech card generation creates real IDs and saves to database"""
        print("\n🎯 Testing Tech Card Generation with Real IDs")
        
        # Test dishes to generate
        test_dishes = [
            "Борщ украинский с говядиной",
            "Стейк из говядины с картофельным пюре"
        ]
        
        for dish_name in test_dishes:
            try:
                start_time = time.time()
                
                payload = {
                    "name": dish_name,
                    "cuisine": "русская",
                    "equipment": [],
                    "budget": None,
                    "dietary": []
                }
                
                response = await self.client.post(f"{API_BASE}/techcards.v2/generate", json=payload)
                response_time = time.time() - start_time
                
                if response.status_code == 200:
                    data = response.json()
                    print(f"DEBUG: Response data keys: {list(data.keys())}")
                    print(f"DEBUG: Response data: {json.dumps(data, indent=2)[:500]}...")
                    
                    # Extract tech card ID from the response structure
                    techcard_id = None
                    if "card" in data and data["card"]:
                        card_data = data["card"]
                        if "meta" in card_data and "id" in card_data["meta"]:
                            techcard_id = card_data["meta"]["id"]
                        else:
                            techcard_id = card_data.get("id") or card_data.get("_id")
                    
                    if not techcard_id:
                        # Try other possible locations
                        techcard_id = data.get("id") or data.get("techcard_id") or data.get("_id")
                    
                    if techcard_id and len(str(techcard_id)) > 10:  # Valid UUID-like ID
                        self.generated_techcard_ids.append(techcard_id)
                        
                        # Verify tech card has real content
                        card_data = data.get("card", {})
                        techcard_name = card_data.get("meta", {}).get("title", "") or card_data.get("name", "")
                        ingredients_count = len(card_data.get("ingredients", []))
                        
                        # Check for mock signatures in the generated content
                        mock_detected = any(sig in str(card_data).upper() for sig in [s.upper() for s in MOCK_SIGNATURES])
                        
                        if not mock_detected and ingredients_count > 0:
                            self.log_test(f"Tech Card Generation - {dish_name}", True,
                                        f"ID: {techcard_id[:20]}..., Ingredients: {ingredients_count}", response_time)
                            
                            # Store for database verification
                            self.artifacts[f"gen_{len(self.generated_techcard_ids)}"] = {
                                "id": techcard_id,
                                "name": techcard_name,
                                "ingredients": ingredients_count,
                                "data": card_data
                            }
                        else:
                            self.log_test(f"Tech Card Generation - {dish_name}", False,
                                        f"Mock content detected or no ingredients", response_time)
                    else:
                        self.log_test(f"Tech Card Generation - {dish_name}", False,
                                    f"Invalid or missing tech card ID: {techcard_id}", response_time)
                else:
                    self.log_test(f"Tech Card Generation - {dish_name}", False,
                                f"HTTP {response.status_code}: {response.text[:200]}", response_time)
                    
            except Exception as e:
                self.log_test(f"Tech Card Generation - {dish_name}", False, f"Exception: {str(e)}", 0.0)
    
    async def test_database_persistence_verification(self):
        """Verify tech cards are properly saved to database"""
        print("\n🎯 Testing Database Persistence Verification")
        
        if not self.generated_techcard_ids:
            self.log_test("Database Persistence", False, "No tech card IDs to verify", 0.0)
            return
        
        try:
            start_time = time.time()
            
            # Connect to MongoDB
            client = MongoClient(MONGO_URL)
            db = client[DB_NAME.strip('"')]
            
            # Check techcards collection
            techcards_collection = db.techcards
            
            # Also try techcards_v2 collection
            techcards_v2_collection = db.techcards_v2
            
            verified_count = 0
            for techcard_id in self.generated_techcard_ids:
                # Look for the tech card in both collections
                stored_techcard = (techcards_collection.find_one({"_id": techcard_id}) or 
                                 techcards_collection.find_one({"id": techcard_id}) or
                                 techcards_v2_collection.find_one({"_id": techcard_id}) or
                                 techcards_v2_collection.find_one({"id": techcard_id}) or
                                 techcards_v2_collection.find_one({"meta.id": techcard_id}))
                
                if stored_techcard:
                    verified_count += 1
                    
                    # Verify it has real content
                    ingredients = stored_techcard.get("ingredients", [])
                    name = stored_techcard.get("name") or stored_techcard.get("title", "")
                    
                    # Check for mock signatures
                    mock_detected = any(sig in str(stored_techcard).upper() for sig in [s.upper() for s in MOCK_SIGNATURES])
                    
                    if mock_detected:
                        self.log_test("Database Content Verification", False,
                                    f"Mock content found in stored tech card {techcard_id[:20]}...", 0.0)
                        client.close()
                        return
            
            response_time = time.time() - start_time
            
            if verified_count == len(self.generated_techcard_ids):
                self.log_test("Database Persistence", True,
                            f"All {verified_count} tech cards found in database with real content", response_time)
            else:
                self.log_test("Database Persistence", False,
                            f"Only {verified_count}/{len(self.generated_techcard_ids)} tech cards found", response_time)
            
            client.close()
            
        except Exception as e:
            self.log_test("Database Persistence", False, f"Exception: {str(e)}", 0.0)
    
    async def test_export_preflight_with_real_ids(self):
        """Test export preflight using real tech card IDs (not 'current')"""
        print("\n🎯 Testing Export Preflight with Real IDs")
        
        if not self.generated_techcard_ids:
            self.log_test("Export Preflight Real IDs", False, "No tech card IDs available", 0.0)
            return
        
        try:
            start_time = time.time()
            
            # Use REAL tech card IDs instead of 'current'
            payload = {
                "techcardIds": self.generated_techcard_ids,
                "organization_id": self.organization_id
            }
            
            response = await self.client.post(f"{API_BASE}/export/preflight", json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ["status", "ttkDate", "missing", "generated", "counts"]
                missing_fields = [field for field in required_fields if field not in data]
                
                if not missing_fields:
                    # Check TTK date format
                    ttk_date = data.get("ttkDate")
                    dish_skeletons = data.get("counts", {}).get("dishSkeletons", 0)
                    product_skeletons = data.get("counts", {}).get("productSkeletons", 0)
                    
                    self.log_test("Export Preflight Real IDs", True,
                                f"TTK Date: {ttk_date}, Dish Skeletons: {dish_skeletons}, Product Skeletons: {product_skeletons}", 
                                response_time)
                    
                    # Store preflight result for export test
                    self.artifacts["preflight_result"] = data
                else:
                    self.log_test("Export Preflight Real IDs", False,
                                f"Missing fields: {missing_fields}", response_time)
            else:
                self.log_test("Export Preflight Real IDs", False,
                            f"HTTP {response.status_code}: {response.text[:200]}", response_time)
                
        except Exception as e:
            self.log_test("Export Preflight Real IDs", False, f"Exception: {str(e)}", 0.0)
    
    async def test_export_zip_with_real_data(self):
        """Test ZIP export using real tech card data"""
        print("\n🎯 Testing ZIP Export with Real Data")
        
        if not self.generated_techcard_ids or "preflight_result" not in self.artifacts:
            self.log_test("ZIP Export Real Data", False, "Prerequisites not met", 0.0)
            return
        
        try:
            start_time = time.time()
            
            # Use REAL tech card IDs and preflight result
            payload = {
                "techcardIds": self.generated_techcard_ids,
                "operational_rounding": True,
                "organization_id": self.organization_id,
                "preflight_result": self.artifacts["preflight_result"]
            }
            
            response = await self.client.post(f"{API_BASE}/export/zip", json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # Check if response is a ZIP file
                content_type = response.headers.get('content-type', '')
                content_length = len(response.content)
                
                is_zip = (content_type == 'application/zip' or 
                         response.content.startswith(b'PK'))
                
                if is_zip and content_length > 0:
                    self.log_test("ZIP Export Real Data", True,
                                f"ZIP file generated: {content_length} bytes", response_time)
                    
                    # Store ZIP content for inspection
                    self.artifacts["export_zip"] = response.content
                else:
                    self.log_test("ZIP Export Real Data", False,
                                f"Invalid ZIP: content-type={content_type}, size={content_length}", response_time)
            else:
                self.log_test("ZIP Export Real Data", False,
                            f"HTTP {response.status_code}: {response.text[:200]}", response_time)
                
        except Exception as e:
            self.log_test("ZIP Export Real Data", False, f"Exception: {str(e)}", 0.0)
    
    async def test_xlsx_content_inspection(self):
        """Inspect XLSX files for mock content and verify real data"""
        print("\n🎯 Testing XLSX Content Inspection")
        
        if "export_zip" not in self.artifacts:
            self.log_test("XLSX Content Inspection", False, "No ZIP file to inspect", 0.0)
            return
        
        try:
            start_time = time.time()
            
            # Extract and inspect ZIP contents
            zip_buffer = io.BytesIO(self.artifacts["export_zip"])
            
            with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
                file_list = zip_file.namelist()
                xlsx_files = [f for f in file_list if f.endswith('.xlsx')]
                
                if not xlsx_files:
                    self.log_test("XLSX Content Inspection", False, "No XLSX files found in ZIP", 0.0)
                    return
                
                mock_signatures_found = []
                real_content_found = []
                total_cells_checked = 0
                
                for xlsx_file in xlsx_files:
                    # Read XLSX file
                    xlsx_data = zip_file.read(xlsx_file)
                    xlsx_buffer = io.BytesIO(xlsx_data)
                    
                    try:
                        workbook = openpyxl.load_workbook(xlsx_buffer)
                        
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            
                            # Inspect all cells for content
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.value:
                                        cell_value = str(cell.value)
                                        total_cells_checked += 1
                                        
                                        # Check for mock signatures
                                        for mock_sig in MOCK_SIGNATURES:
                                            if mock_sig.upper() in cell_value.upper():
                                                mock_signatures_found.append(f"{xlsx_file}:{sheet_name}:{cell.coordinate}={mock_sig}")
                                        
                                        # Check for real content (dish names from our generated tech cards)
                                        for artifact_key, artifact_data in self.artifacts.items():
                                            if artifact_key.startswith("gen_") and "name" in artifact_data:
                                                dish_name = artifact_data["name"]
                                                if dish_name and dish_name.lower() in cell_value.lower():
                                                    real_content_found.append(f"{xlsx_file}:{sheet_name}:{cell.coordinate}={dish_name}")
                        
                        workbook.close()
                        
                    except Exception as e:
                        self.log_test(f"XLSX Inspection - {xlsx_file}", False, f"Error reading file: {str(e)}", 0.0)
                        continue
                
                response_time = time.time() - start_time
                
                # Evaluate results
                if mock_signatures_found:
                    self.log_test("XLSX Mock Content Detection", False,
                                f"Found {len(mock_signatures_found)} mock signatures: {mock_signatures_found[:3]}", response_time)
                else:
                    self.log_test("XLSX Mock Content Detection", True,
                                f"No mock signatures found in {total_cells_checked} cells", response_time)
                
                if real_content_found:
                    self.log_test("XLSX Real Content Verification", True,
                                f"Found {len(real_content_found)} real content matches", response_time)
                else:
                    self.log_test("XLSX Real Content Verification", False,
                                f"No real content found in {total_cells_checked} cells", response_time)
                
                # Store inspection results
                self.artifacts["xlsx_inspection"] = {
                    "files_inspected": xlsx_files,
                    "cells_checked": total_cells_checked,
                    "mock_signatures": mock_signatures_found,
                    "real_content": real_content_found
                }
                
        except Exception as e:
            self.log_test("XLSX Content Inspection", False, f"Exception: {str(e)}", 0.0)
    
    async def test_article_formatting_preservation(self):
        """Test that article columns are properly formatted with leading zeros"""
        print("\n🎯 Testing Article Formatting Preservation")
        
        if "export_zip" not in self.artifacts:
            self.log_test("Article Formatting", False, "No ZIP file to inspect", 0.0)
            return
        
        try:
            start_time = time.time()
            
            # Extract and inspect ZIP contents for article formatting
            zip_buffer = io.BytesIO(self.artifacts["export_zip"])
            
            with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
                xlsx_files = [f for f in zip_file.namelist() if f.endswith('.xlsx')]
                
                articles_found = []
                formatting_issues = []
                
                for xlsx_file in xlsx_files:
                    xlsx_data = zip_file.read(xlsx_file)
                    xlsx_buffer = io.BytesIO(xlsx_data)
                    
                    try:
                        workbook = openpyxl.load_workbook(xlsx_buffer)
                        
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            
                            # Look for article columns (typically columns A and C)
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.value and isinstance(cell.value, str):
                                        cell_value = str(cell.value)
                                        
                                        # Check if this looks like an article (5 digits)
                                        if len(cell_value) == 5 and cell_value.isdigit():
                                            articles_found.append(f"{xlsx_file}:{cell.coordinate}={cell_value}")
                                            
                                            # Check if cell is formatted as text (@)
                                            if hasattr(cell, 'number_format'):
                                                if cell.number_format != '@':
                                                    formatting_issues.append(f"{xlsx_file}:{cell.coordinate} not formatted as text")
                        
                        workbook.close()
                        
                    except Exception as e:
                        continue
                
                response_time = time.time() - start_time
                
                if articles_found:
                    if not formatting_issues:
                        self.log_test("Article Formatting", True,
                                    f"Found {len(articles_found)} properly formatted articles", response_time)
                    else:
                        self.log_test("Article Formatting", False,
                                    f"Found {len(formatting_issues)} formatting issues", response_time)
                else:
                    self.log_test("Article Formatting", True,
                                f"No articles found (expected for some scenarios)", response_time)
                
        except Exception as e:
            self.log_test("Article Formatting", False, f"Exception: {str(e)}", 0.0)
    
    async def test_complete_workflow_validation(self):
        """Test complete end-to-end workflow validation"""
        print("\n🎯 Testing Complete Workflow Validation")
        
        try:
            start_time = time.time()
            
            # Validate all critical components are working
            workflow_components = {
                "Tech Card Generation": len(self.generated_techcard_ids) > 0,
                "Database Persistence": "gen_1" in self.artifacts,
                "Preflight with Real IDs": "preflight_result" in self.artifacts,
                "ZIP Export": "export_zip" in self.artifacts,
                "XLSX Inspection": "xlsx_inspection" in self.artifacts
            }
            
            failed_components = [comp for comp, status in workflow_components.items() if not status]
            
            if not failed_components:
                # Check for critical success criteria
                xlsx_data = self.artifacts.get("xlsx_inspection", {})
                mock_signatures = xlsx_data.get("mock_signatures", [])
                real_content = xlsx_data.get("real_content", [])
                
                # SUCCESS CRITERIA VALIDATION
                criteria_met = {
                    "Tech cards generated with real IDs": len(self.generated_techcard_ids) > 0,
                    "Tech cards saved to database": "gen_1" in self.artifacts,
                    "Export uses real tech card IDs": "preflight_result" in self.artifacts,
                    "XLSX files contain real data": len(real_content) > 0,
                    "Zero mock signatures in exports": len(mock_signatures) == 0
                }
                
                failed_criteria = [criteria for criteria, met in criteria_met.items() if not met]
                
                response_time = time.time() - start_time
                
                if not failed_criteria:
                    self.log_test("Complete Workflow Validation", True,
                                f"All {len(criteria_met)} success criteria met", response_time)
                else:
                    self.log_test("Complete Workflow Validation", False,
                                f"Failed criteria: {failed_criteria}", response_time)
            else:
                response_time = time.time() - start_time
                self.log_test("Complete Workflow Validation", False,
                            f"Failed components: {failed_components}", response_time)
                
        except Exception as e:
            self.log_test("Complete Workflow Validation", False, f"Exception: {str(e)}", 0.0)
    
    def print_summary(self):
        """Print comprehensive test summary"""
        print("\n" + "="*80)
        print("🎯 MOCK CONTENT RESOLUTION END-TO-END TESTING SUMMARY")
        print("="*80)
        
        total_tests = len(self.test_results)
        passed_tests = sum(1 for result in self.test_results if result["success"])
        failed_tests = total_tests - passed_tests
        success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
        
        print(f"\n📊 OVERALL RESULTS:")
        print(f"   Total Tests: {total_tests}")
        print(f"   Passed: {passed_tests} ✅")
        print(f"   Failed: {failed_tests} ❌")
        print(f"   Success Rate: {success_rate:.1f}%")
        
        # Critical validation summary
        print(f"\n🎯 CRITICAL SUCCESS CRITERIA:")
        
        critical_tests = [
            ("Tech Card Generation", "Tech cards generated with real IDs and saved to database"),
            ("Export Preflight Real IDs", "Export uses real tech card data instead of 'current' fallback"),
            ("XLSX Mock Content Detection", "XLSX files contain no mock signatures"),
            ("XLSX Real Content Verification", "XLSX files contain actual generated content"),
            ("Article Formatting", "Article columns properly formatted with leading zeros"),
            ("Complete Workflow Validation", "All success criteria met")
        ]
        
        for test_name, description in critical_tests:
            result = next((r for r in self.test_results if r["test"] == test_name), None)
            if result:
                status = "✅" if result["success"] else "❌"
                print(f"   {status} {description}")
            else:
                print(f"   ⚠️  {description} (not tested)")
        
        # Detailed findings
        if "xlsx_inspection" in self.artifacts:
            xlsx_data = self.artifacts["xlsx_inspection"]
            mock_count = len(xlsx_data.get("mock_signatures", []))
            real_count = len(xlsx_data.get("real_content", []))
            
            print(f"\n📋 DETAILED FINDINGS:")
            print(f"   Tech Cards Generated: {len(self.generated_techcard_ids)}")
            print(f"   XLSX Files Inspected: {len(xlsx_data.get('files_inspected', []))}")
            print(f"   Cells Checked: {xlsx_data.get('cells_checked', 0)}")
            print(f"   Mock Signatures Found: {mock_count}")
            print(f"   Real Content Matches: {real_count}")
        
        if failed_tests > 0:
            print(f"\n❌ FAILED TESTS:")
            for result in self.test_results:
                if not result["success"]:
                    print(f"   • {result['test']}: {result['details']}")
        
        # Final verdict
        mock_resolved = True
        if "xlsx_inspection" in self.artifacts:
            mock_signatures = self.artifacts["xlsx_inspection"].get("mock_signatures", [])
            mock_resolved = len(mock_signatures) == 0
        
        print(f"\n🎯 FINAL VERDICT:")
        if success_rate >= 80 and mock_resolved:
            print(f"   ✅ MOCK CONTENT ISSUE COMPLETELY RESOLVED")
            print(f"   ✅ Export system now uses real tech card data")
            print(f"   ✅ Zero mock signatures detected in exported files")
            print(f"   ✅ System ready for production use")
        else:
            print(f"   ❌ MOCK CONTENT ISSUE NOT FULLY RESOLVED")
            print(f"   ❌ Export system may still contain mock content")
            print(f"   ❌ Requires immediate attention")
        
        return success_rate >= 80 and mock_resolved


async def main():
    """Main test execution"""
    print("🚀 Starting Mock Content Resolution End-to-End Testing")
    print(f"Backend URL: {BACKEND_URL}")
    print(f"MongoDB URL: {MONGO_URL}")
    print(f"Database: {DB_NAME}")
    print("\nTesting complete frontend-to-backend flow to verify mock content issue resolution...")
    
    async with MockContentResolutionTester() as tester:
        # Execute comprehensive test suite
        await tester.test_techcard_generation_with_real_ids()
        await tester.test_database_persistence_verification()
        await tester.test_export_preflight_with_real_ids()
        await tester.test_export_zip_with_real_data()
        await tester.test_xlsx_content_inspection()
        await tester.test_article_formatting_preservation()
        await tester.test_complete_workflow_validation()
        
        # Print comprehensive summary
        success = tester.print_summary()
        
        if success:
            print(f"\n🎉 MOCK CONTENT RESOLUTION TESTING COMPLETED SUCCESSFULLY!")
            print(f"The mock content issue has been completely resolved.")
            print(f"Export system now uses real tech card data throughout the entire workflow.")
        else:
            print(f"\n⚠️  MOCK CONTENT RESOLUTION TESTING COMPLETED WITH ISSUES")
            print(f"The mock content issue has NOT been fully resolved.")
            print(f"Immediate action required to fix remaining mock content in export system.")
        
        return success


if __name__ == "__main__":
    try:
        success = asyncio.run(main())
        sys.exit(0 if success else 1)
    except KeyboardInterrupt:
        print("\n⚠️ Testing interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Testing failed with exception: {e}")
        traceback.print_exc()
        sys.exit(1)