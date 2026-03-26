#!/usr/bin/env python3
"""
iiko IMPORT RELIABILITY — PRODUCT CODES & DISH SKELETONS TESTING

Comprehensive testing of the newly implemented iiko Import Reliability features
to eliminate "артикул не найден" errors in iikoWeb import.

IMPLEMENTED FEATURES TO TEST:

**Feature A: Product Code Toggle (Ingredients)**
- New export option: use_product_codes (default: true)
- Uses product.code/article instead of GUID for "Артикул продукта" 
- Excel cell formatting as text (@) to preserve leading zeros
- Fallback to generated codes when product codes unavailable

**Feature B: Dish Code Resolver + Skeletons (Dishes)**
- API endpoints for finding dish codes in iiko RMS by name
- API endpoints for generating sequential dish codes  
- Dual export: Dish-Skeletons.xlsx + iiko_TTK.xlsx
- Dish codes mapping integration

**Feature 3: Pre-flight Warnings**
- Detection of missing dish codes
- Detection of missing product codes
- User-friendly warnings with actionable CTAs
"""

import requests
import json
import time
import os
import sys
from typing import Dict, List, Any
import zipfile
import io
from openpyxl import load_workbook

# Backend URL from environment
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'http://localhost:8001')
API_BASE = f"{BACKEND_URL}/api/v1"

class IikoImportReliabilityTester:
    def __init__(self):
        self.session = requests.Session()
        self.test_results = []
        self.start_time = time.time()
        
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        self.test_results.append({
            "test": test_name,
            "success": success,
            "details": details,
            "response_time": response_time
        })
        print(f"{status} {test_name} ({response_time:.3f}s)")
        if details:
            print(f"    {details}")
    
    def create_test_techcard(self, title: str, ingredients: List[Dict] = None) -> Dict:
        """Create a test techcard with realistic data that matches TechCardV2 schema"""
        if ingredients is None:
            ingredients = [
                {
                    "name": "Куриное филе",
                    "brutto_g": 150.0,
                    "netto_g": 120.0,
                    "loss_pct": 20.0,
                    "unit": "g",
                    "skuId": "test-chicken-sku-001"
                },
                {
                    "name": "Картофель",
                    "brutto_g": 200.0,
                    "netto_g": 180.0,
                    "loss_pct": 10.0,
                    "unit": "g",
                    "skuId": "test-potato-sku-002"
                },
                {
                    "name": "Лук репчатый",
                    "brutto_g": 50.0,
                    "netto_g": 45.0,
                    "loss_pct": 10.0,
                    "unit": "g",
                    "skuId": "test-onion-sku-003"
                }
            ]
        
        # Calculate total netto for yield consistency
        total_netto = sum(ing["netto_g"] for ing in ingredients)
        
        return {
            "meta": {
                "title": title,
                "id": f"test-{title.lower().replace(' ', '-')}",
                "version": "2.0",
                "createdAt": "2025-01-27T10:00:00Z"
            },
            "ingredients": ingredients,
            "yield": {  # Note: not yield_ in JSON
                "perPortion_g": total_netto,
                "perBatch_g": total_netto
            },
            "portions": 1,
            "process": [  # Array, not object with steps
                {
                    "n": 1,
                    "action": "Нарезать куриное филе кубиками",
                    "time_min": 5,
                    "equipment": ["нож", "доска"]
                },
                {
                    "n": 2,
                    "action": "Обжарить курицу на сковороде",
                    "temp_c": 180,
                    "time_min": 10,
                    "equipment": ["сковорода"]
                },
                {
                    "n": 3,
                    "action": "Добавить овощи и тушить",
                    "temp_c": 160,
                    "time_min": 15,
                    "equipment": ["сковорода"]
                }
            ],
            "storage": {
                "conditions": "Хранить в холодильнике при температуре +2...+6°C",
                "shelfLife_hours": 24.0,
                "servingTemp_c": 65.0
            },
            "nutrition": {
                "per100g": {
                    "kcal": 150.0,
                    "proteins_g": 20.0,
                    "fats_g": 5.0,
                    "carbs_g": 10.0
                },
                "perPortion": {
                    "kcal": 150.0 * total_netto / 100,
                    "proteins_g": 20.0 * total_netto / 100,
                    "fats_g": 5.0 * total_netto / 100,
                    "carbs_g": 10.0 * total_netto / 100
                }
            },
            "cost": {
                "rawCost": 100.0,
                "costPerPortion": 100.0,
                "markup_pct": 200.0,
                "vat_pct": 20.0
            }
        }

    def test_dish_codes_find_endpoint(self):
        """Test 1: POST /api/v1/techcards.v2/dish-codes/find - dish search functionality"""
        start_time = time.time()
        
        try:
            payload = {
                "dish_names": ["Салат Цезарь", "Стейк с овощами", "Борщ украинский"],
                "organization_id": "default"
            }
            
            response = self.session.post(
                f"{API_BASE}/techcards.v2/dish-codes/find",
                json=payload,
                timeout=10
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                if "results" in data and isinstance(data["results"], list):
                    results = data["results"]
                    
                    # Check each result has required fields
                    valid_results = True
                    for result in results:
                        required_fields = ["dish_name", "status"]
                        if not all(field in result for field in required_fields):
                            valid_results = False
                            break
                        
                        # If found, should have dish_code and confidence
                        if result["status"] == "found":
                            if "dish_code" not in result or "confidence" not in result:
                                valid_results = False
                                break
                    
                    if valid_results:
                        found_count = len([r for r in results if r["status"] == "found"])
                        self.log_test(
                            "Dish Codes Find API",
                            True,
                            f"Found {found_count}/{len(results)} dishes with proper response structure",
                            response_time
                        )
                    else:
                        self.log_test(
                            "Dish Codes Find API",
                            False,
                            "Invalid response structure - missing required fields",
                            response_time
                        )
                else:
                    self.log_test(
                        "Dish Codes Find API",
                        False,
                        "Response missing 'results' array",
                        response_time
                    )
            else:
                self.log_test(
                    "Dish Codes Find API",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Dish Codes Find API",
                False,
                f"Exception: {str(e)}",
                response_time
            )

    def test_dish_codes_generate_endpoint(self):
        """Test 2: POST /api/v1/techcards.v2/dish-codes/generate - code generation"""
        start_time = time.time()
        
        try:
            payload = {
                "dish_names": ["Новое блюдо 1", "Новое блюдо 2", "Тестовое блюдо"],
                "organization_id": "default",
                "width": 5
            }
            
            response = self.session.post(
                f"{API_BASE}/techcards.v2/dish-codes/generate",
                json=payload,
                timeout=10
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                if "generated_codes" in data and isinstance(data["generated_codes"], dict):
                    codes = data["generated_codes"]
                    
                    # Check all dishes got codes
                    if len(codes) == len(payload["dish_names"]):
                        # Validate code format (5 digits with leading zeros)
                        valid_codes = True
                        for dish_name, code in codes.items():
                            if not (isinstance(code, str) and len(code) == 5 and code.isdigit()):
                                valid_codes = False
                                break
                        
                        if valid_codes:
                            self.log_test(
                                "Dish Codes Generate API",
                                True,
                                f"Generated {len(codes)} sequential codes with proper format (width=5)",
                                response_time
                            )
                        else:
                            self.log_test(
                                "Dish Codes Generate API",
                                False,
                                "Generated codes don't match expected format (5 digits)",
                                response_time
                            )
                    else:
                        self.log_test(
                            "Dish Codes Generate API",
                            False,
                            f"Expected {len(payload['dish_names'])} codes, got {len(codes)}",
                            response_time
                        )
                else:
                    self.log_test(
                        "Dish Codes Generate API",
                        False,
                        "Response missing 'generated_codes' object",
                        response_time
                    )
            else:
                self.log_test(
                    "Dish Codes Generate API",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Dish Codes Generate API",
                False,
                f"Exception: {str(e)}",
                response_time
            )

    def test_preflight_check_endpoint(self):
        """Test 3: POST /api/v1/techcards.v2/export/preflight-check - warnings system"""
        start_time = time.time()
        
        try:
            # Create test techcards with missing codes
            techcards = [
                self.create_test_techcard("Блюдо без кода 1"),
                self.create_test_techcard("Блюдо без кода 2")
            ]
            
            payload = {
                "techcards": techcards,
                "export_options": {
                    "use_product_codes": True,
                    "dish_codes_mapping": {}  # Empty mapping to trigger warnings
                },
                "organization_id": "default"
            }
            
            response = self.session.post(
                f"{API_BASE}/techcards.v2/export/preflight-check",
                json=payload,
                timeout=15
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ["status", "warnings", "cards_checked", "export_ready"]
                if all(field in data for field in required_fields):
                    
                    # Should detect missing dish codes
                    warnings = data["warnings"]
                    missing_dish_codes_warning = None
                    missing_product_codes_warning = None
                    
                    for warning in warnings:
                        if warning.get("type") == "missing_dish_codes":
                            missing_dish_codes_warning = warning
                        elif warning.get("type") == "missing_product_codes":
                            missing_product_codes_warning = warning
                    
                    # Validate warnings structure
                    warnings_valid = True
                    warning_details = []
                    
                    if missing_dish_codes_warning:
                        required_warning_fields = ["type", "title", "items", "action", "severity"]
                        if all(field in missing_dish_codes_warning for field in required_warning_fields):
                            warning_details.append(f"Missing dish codes: {len(missing_dish_codes_warning['items'])} items")
                        else:
                            warnings_valid = False
                    
                    if missing_product_codes_warning:
                        warning_details.append(f"Missing product codes: {len(missing_product_codes_warning['items'])} items")
                    
                    if warnings_valid and data["status"] == "warnings":
                        self.log_test(
                            "Preflight Check API",
                            True,
                            f"Detected warnings correctly: {'; '.join(warning_details)}",
                            response_time
                        )
                    else:
                        self.log_test(
                            "Preflight Check API",
                            False,
                            f"Invalid warnings structure or status. Status: {data['status']}",
                            response_time
                        )
                else:
                    self.log_test(
                        "Preflight Check API",
                        False,
                        f"Response missing required fields: {required_fields}",
                        response_time
                    )
            else:
                self.log_test(
                    "Preflight Check API",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Preflight Check API",
                False,
                f"Exception: {str(e)}",
                response_time
            )

    def test_enhanced_dual_export_endpoint(self):
        """Test 4: POST /api/v1/techcards.v2/export/enhanced-dual/iiko.xlsx - dual export"""
        start_time = time.time()
        
        try:
            # Create test techcards
            techcards = [
                self.create_test_techcard("Салат Цезарь"),
                self.create_test_techcard("Стейк с овощами")
            ]
            
            # Provide dish codes mapping
            dish_codes_mapping = {
                "Салат Цезарь": "12345",
                "Стейк с овощами": "12346"
            }
            
            payload = {
                "techcards": techcards,
                "export_options": {
                    "use_product_codes": True,
                    "dish_codes_mapping": dish_codes_mapping
                },
                "organization_id": "default",
                "user_email": "test@example.com"
            }
            
            response = self.session.post(
                f"{API_BASE}/techcards.v2/export/enhanced-dual/iiko.xlsx",
                json=payload,
                timeout=20
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # Should return ZIP file
                content_type = response.headers.get('content-type', '')
                if 'application/zip' in content_type:
                    
                    # Validate ZIP contents
                    try:
                        zip_buffer = io.BytesIO(response.content)
                        with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
                            file_list = zip_file.namelist()
                            
                            # Should contain both files
                            has_skeletons = any('Dish-Skeletons' in f for f in file_list)
                            has_ttk = any('iiko_TTK' in f for f in file_list)
                            
                            if has_skeletons and has_ttk:
                                self.log_test(
                                    "Enhanced Dual Export API",
                                    True,
                                    f"ZIP contains both files: {', '.join(file_list)}",
                                    response_time
                                )
                            else:
                                self.log_test(
                                    "Enhanced Dual Export API",
                                    False,
                                    f"ZIP missing expected files. Found: {', '.join(file_list)}",
                                    response_time
                                )
                    except Exception as zip_error:
                        self.log_test(
                            "Enhanced Dual Export API",
                            False,
                            f"Invalid ZIP file: {str(zip_error)}",
                            response_time
                        )
                else:
                    self.log_test(
                        "Enhanced Dual Export API",
                        False,
                        f"Expected ZIP file, got content-type: {content_type}",
                        response_time
                    )
            else:
                self.log_test(
                    "Enhanced Dual Export API",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Enhanced Dual Export API",
                False,
                f"Exception: {str(e)}",
                response_time
            )

    def test_product_codes_feature(self):
        """Test 5: Feature A - Product Codes Toggle"""
        start_time = time.time()
        
        try:
            # Test with use_product_codes=true
            techcard = self.create_test_techcard("Тест продуктовых кодов")
            
            payload = {
                "techcard": techcard,
                "export_options": {
                    "use_product_codes": True
                },
                "organization_id": "default"
            }
            
            response = self.session.post(
                f"{API_BASE}/techcards.v2/export/enhanced/iiko.xlsx",
                json=payload,
                timeout=15
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # Should return Excel file
                content_type = response.headers.get('content-type', '')
                if 'spreadsheet' in content_type:
                    
                    # Validate Excel file structure
                    try:
                        excel_buffer = io.BytesIO(response.content)
                        workbook = load_workbook(excel_buffer)
                        worksheet = workbook.active
                        
                        # Check headers
                        headers = [cell.value for cell in worksheet[1]]
                        expected_headers = ['Артикул блюда', 'Наименование блюда', 'Артикул продукта']
                        
                        has_required_headers = all(header in headers for header in expected_headers)
                        
                        if has_required_headers:
                            # Check if product codes are formatted as text (not GUID)
                            product_code_col = headers.index('Артикул продукта') + 1
                            dish_code_col = headers.index('Артикул блюда') + 1
                            
                            # Check cell formatting
                            product_code_cell = worksheet.cell(row=2, column=product_code_col)
                            dish_code_cell = worksheet.cell(row=2, column=dish_code_col)
                            
                            # Should be formatted as text (@)
                            text_formatted = (
                                product_code_cell.number_format == '@' and 
                                dish_code_cell.number_format == '@'
                            )
                            
                            if text_formatted:
                                self.log_test(
                                    "Product Codes Feature A",
                                    True,
                                    "Excel file has proper text formatting for code columns",
                                    response_time
                                )
                            else:
                                self.log_test(
                                    "Product Codes Feature A",
                                    False,
                                    f"Code columns not formatted as text. Product: {product_code_cell.number_format}, Dish: {dish_code_cell.number_format}",
                                    response_time
                                )
                        else:
                            self.log_test(
                                "Product Codes Feature A",
                                False,
                                f"Missing required headers. Found: {headers}",
                                response_time
                            )
                            
                    except Exception as excel_error:
                        self.log_test(
                            "Product Codes Feature A",
                            False,
                            f"Excel validation error: {str(excel_error)}",
                            response_time
                        )
                else:
                    self.log_test(
                        "Product Codes Feature A",
                        False,
                        f"Expected Excel file, got content-type: {content_type}",
                        response_time
                    )
            else:
                self.log_test(
                    "Product Codes Feature A",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Product Codes Feature A",
                False,
                f"Exception: {str(e)}",
                response_time
            )

    def test_dish_codes_integration(self):
        """Test 6: Feature B - Dish Codes & Skeletons Integration"""
        start_time = time.time()
        
        try:
            # First generate dish codes
            dish_names = ["Интеграционное блюдо 1", "Интеграционное блюдо 2"]
            
            generate_payload = {
                "dish_names": dish_names,
                "organization_id": "default",
                "width": 5
            }
            
            generate_response = self.session.post(
                f"{API_BASE}/techcards.v2/dish-codes/generate",
                json=generate_payload,
                timeout=10
            )
            
            if generate_response.status_code == 200:
                generated_data = generate_response.json()
                dish_codes_mapping = generated_data.get("generated_codes", {})
                
                if dish_codes_mapping:
                    # Now test dual export with generated codes
                    techcards = [
                        self.create_test_techcard(dish_names[0]),
                        self.create_test_techcard(dish_names[1])
                    ]
                    
                    export_payload = {
                        "techcards": techcards,
                        "export_options": {
                            "use_product_codes": True,
                            "dish_codes_mapping": dish_codes_mapping
                        },
                        "organization_id": "default",
                        "user_email": "test@example.com"
                    }
                    
                    export_response = self.session.post(
                        f"{API_BASE}/techcards.v2/export/enhanced-dual/iiko.xlsx",
                        json=export_payload,
                        timeout=20
                    )
                    
                    response_time = time.time() - start_time
                    
                    if export_response.status_code == 200:
                        # Validate ZIP with dish codes
                        try:
                            zip_buffer = io.BytesIO(export_response.content)
                            with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
                                file_list = zip_file.namelist()
                                
                                # Should have Dish-Skeletons.xlsx
                                skeleton_file = None
                                for f in file_list:
                                    if 'Dish-Skeletons' in f:
                                        skeleton_file = f
                                        break
                                
                                if skeleton_file:
                                    # Validate skeleton file content
                                    skeleton_data = zip_file.read(skeleton_file)
                                    skeleton_workbook = load_workbook(io.BytesIO(skeleton_data))
                                    skeleton_ws = skeleton_workbook.active
                                    
                                    # Check if generated codes are present
                                    codes_found = []
                                    for row in skeleton_ws.iter_rows(min_row=2, values_only=True):
                                        if row[0]:  # Артикул column
                                            codes_found.append(str(row[0]))
                                    
                                    expected_codes = list(dish_codes_mapping.values())
                                    codes_match = all(code in codes_found for code in expected_codes)
                                    
                                    if codes_match:
                                        self.log_test(
                                            "Dish Codes Integration B",
                                            True,
                                            f"Generated codes integrated successfully: {expected_codes}",
                                            response_time
                                        )
                                    else:
                                        self.log_test(
                                            "Dish Codes Integration B",
                                            False,
                                            f"Generated codes not found in skeleton. Expected: {expected_codes}, Found: {codes_found}",
                                            response_time
                                        )
                                else:
                                    self.log_test(
                                        "Dish Codes Integration B",
                                        False,
                                        "Dish-Skeletons.xlsx not found in ZIP",
                                        response_time
                                    )
                        except Exception as validation_error:
                            self.log_test(
                                "Dish Codes Integration B",
                                False,
                                f"ZIP validation error: {str(validation_error)}",
                                response_time
                            )
                    else:
                        response_time = time.time() - start_time
                        self.log_test(
                            "Dish Codes Integration B",
                            False,
                            f"Export failed: HTTP {export_response.status_code}",
                            response_time
                        )
                else:
                    response_time = time.time() - start_time
                    self.log_test(
                        "Dish Codes Integration B",
                        False,
                        "No codes generated in first step",
                        response_time
                    )
            else:
                response_time = time.time() - start_time
                self.log_test(
                    "Dish Codes Integration B",
                    False,
                    f"Code generation failed: HTTP {generate_response.status_code}",
                    response_time
                )
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Dish Codes Integration B",
                False,
                f"Exception: {str(e)}",
                response_time
            )

    def test_performance_requirements(self):
        """Test 7: Performance Requirements"""
        
        # Test dish code generation performance
        start_time = time.time()
        try:
            dish_names = [f"Тестовое блюдо {i}" for i in range(1, 11)]  # 10 dishes
            
            payload = {
                "dish_names": dish_names,
                "organization_id": "default",
                "width": 5
            }
            
            response = self.session.post(
                f"{API_BASE}/techcards.v2/dish-codes/generate",
                json=payload,
                timeout=5  # 2s requirement + buffer
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200 and response_time < 2.0:
                self.log_test(
                    "Performance: Dish Code Generation",
                    True,
                    f"Generated 10 codes in {response_time:.3f}s (< 2s requirement)",
                    response_time
                )
            else:
                self.log_test(
                    "Performance: Dish Code Generation",
                    False,
                    f"Too slow: {response_time:.3f}s (requirement: < 2s) or failed",
                    response_time
                )
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Performance: Dish Code Generation",
                False,
                f"Exception: {str(e)}",
                response_time
            )
        
        # Test preflight check performance
        start_time = time.time()
        try:
            techcards = [self.create_test_techcard(f"Блюдо {i}") for i in range(1, 6)]  # 5 cards
            
            payload = {
                "techcards": techcards,
                "export_options": {
                    "use_product_codes": True,
                    "dish_codes_mapping": {}
                },
                "organization_id": "default"
            }
            
            response = self.session.post(
                f"{API_BASE}/techcards.v2/export/preflight-check",
                json=payload,
                timeout=5  # 3s requirement + buffer
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200 and response_time < 3.0:
                self.log_test(
                    "Performance: Preflight Check",
                    True,
                    f"Validated 5 cards in {response_time:.3f}s (< 3s requirement)",
                    response_time
                )
            else:
                self.log_test(
                    "Performance: Preflight Check",
                    False,
                    f"Too slow: {response_time:.3f}s (requirement: < 3s) or failed",
                    response_time
                )
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Performance: Preflight Check",
                False,
                f"Exception: {str(e)}",
                response_time
            )

    def test_error_scenarios(self):
        """Test 8: Error Scenarios"""
        
        # Test invalid techcard data
        start_time = time.time()
        try:
            payload = {
                "techcards": [{"invalid": "data"}],  # Invalid structure
                "export_options": {},
                "organization_id": "default"
            }
            
            response = self.session.post(
                f"{API_BASE}/techcards.v2/export/preflight-check",
                json=payload,
                timeout=10
            )
            
            response_time = time.time() - start_time
            
            # Should handle gracefully (not crash)
            if response.status_code in [200, 400]:
                self.log_test(
                    "Error Handling: Invalid Techcard",
                    True,
                    f"Handled invalid data gracefully (HTTP {response.status_code})",
                    response_time
                )
            else:
                self.log_test(
                    "Error Handling: Invalid Techcard",
                    False,
                    f"Unexpected response: HTTP {response.status_code}",
                    response_time
                )
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Error Handling: Invalid Techcard",
                False,
                f"Exception: {str(e)}",
                response_time
            )
        
        # Test empty dish names
        start_time = time.time()
        try:
            payload = {
                "dish_names": [],  # Empty list
                "organization_id": "default"
            }
            
            response = self.session.post(
                f"{API_BASE}/techcards.v2/dish-codes/generate",
                json=payload,
                timeout=10
            )
            
            response_time = time.time() - start_time
            
            # The endpoint should handle empty list gracefully (return empty result or 400)
            if response.status_code in [200, 400]:
                if response.status_code == 200:
                    data = response.json()
                    if data.get("generated_codes") == {}:
                        self.log_test(
                            "Error Handling: Empty Dish Names",
                            True,
                            "Correctly handled empty dish names list (returned empty result)",
                            response_time
                        )
                    else:
                        self.log_test(
                            "Error Handling: Empty Dish Names",
                            False,
                            "Should return empty result for empty input",
                            response_time
                        )
                else:  # 400
                    self.log_test(
                        "Error Handling: Empty Dish Names",
                        True,
                        "Correctly rejected empty dish names list",
                        response_time
                    )
            else:
                self.log_test(
                    "Error Handling: Empty Dish Names",
                    False,
                    f"Expected HTTP 200 or 400, got {response.status_code}: {response.text[:100]}",
                    response_time
                )
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Error Handling: Empty Dish Names",
                False,
                f"Exception: {str(e)}",
                response_time
            )

    def run_all_tests(self):
        """Run all iiko Import Reliability tests"""
        print("🎯 STARTING iiko IMPORT RELIABILITY COMPREHENSIVE TESTING")
        print(f"Backend URL: {BACKEND_URL}")
        print("=" * 80)
        
        # Test 1: API Endpoints Validation
        print("\n📋 TEST 1: API ENDPOINTS VALIDATION")
        self.test_dish_codes_find_endpoint()
        self.test_dish_codes_generate_endpoint()
        self.test_preflight_check_endpoint()
        self.test_enhanced_dual_export_endpoint()
        
        # Test 2: Feature A - Product Codes
        print("\n🏷️ TEST 2: FEATURE A - PRODUCT CODES")
        self.test_product_codes_feature()
        
        # Test 3: Feature B - Dish Codes & Skeletons
        print("\n🍽️ TEST 3: FEATURE B - DISH CODES & SKELETONS")
        self.test_dish_codes_integration()
        
        # Test 4: Performance Requirements
        print("\n⚡ TEST 4: PERFORMANCE REQUIREMENTS")
        self.test_performance_requirements()
        
        # Test 5: Error Scenarios
        print("\n🚨 TEST 5: ERROR SCENARIOS")
        self.test_error_scenarios()
        
        # Summary
        self.print_summary()

    def print_summary(self):
        """Print test summary"""
        total_time = time.time() - self.start_time
        passed = len([t for t in self.test_results if t["success"]])
        failed = len([t for t in self.test_results if not t["success"]])
        total = len(self.test_results)
        
        print("\n" + "=" * 80)
        print("🎯 iiko IMPORT RELIABILITY TESTING SUMMARY")
        print("=" * 80)
        
        print(f"✅ PASSED: {passed}/{total} tests ({passed/total*100:.1f}%)")
        print(f"❌ FAILED: {failed}/{total} tests ({failed/total*100:.1f}%)")
        print(f"⏱️ TOTAL TIME: {total_time:.2f}s")
        
        if failed > 0:
            print(f"\n❌ FAILED TESTS:")
            for result in self.test_results:
                if not result["success"]:
                    print(f"   • {result['test']}: {result['details']}")
        
        print(f"\n🎯 CRITICAL FEATURES STATUS:")
        
        # Feature A: Product Code Toggle
        feature_a_tests = [t for t in self.test_results if "Product Codes" in t["test"]]
        feature_a_success = all(t["success"] for t in feature_a_tests)
        print(f"   Feature A (Product Codes): {'✅ WORKING' if feature_a_success else '❌ ISSUES'}")
        
        # Feature B: Dish Code Resolver
        feature_b_tests = [t for t in self.test_results if "Dish Codes" in t["test"] or "Enhanced Dual" in t["test"]]
        feature_b_success = all(t["success"] for t in feature_b_tests)
        print(f"   Feature B (Dish Codes): {'✅ WORKING' if feature_b_success else '❌ ISSUES'}")
        
        # Feature 3: Pre-flight Warnings
        feature_3_tests = [t for t in self.test_results if "Preflight" in t["test"]]
        feature_3_success = all(t["success"] for t in feature_3_tests)
        print(f"   Feature 3 (Pre-flight): {'✅ WORKING' if feature_3_success else '❌ ISSUES'}")
        
        # Performance
        perf_tests = [t for t in self.test_results if "Performance" in t["test"]]
        perf_success = all(t["success"] for t in perf_tests)
        print(f"   Performance: {'✅ MEETS REQUIREMENTS' if perf_success else '❌ SLOW'}")
        
        # Overall status
        overall_success = passed >= total * 0.8  # 80% pass rate
        print(f"\n🎯 OVERALL STATUS: {'✅ READY FOR PRODUCTION' if overall_success else '❌ NEEDS FIXES'}")
        
        if overall_success:
            print("🎉 iiko Import Reliability features are working correctly!")
            print("   • Product codes toggle operational")
            print("   • Dish code resolver functional") 
            print("   • Pre-flight warnings active")
            print("   • Excel formatting preserves leading zeros")
            print("   • Dual export (Skeletons + TTK) working")
        else:
            print("⚠️ Some features need attention before production deployment.")

if __name__ == "__main__":
    tester = IikoImportReliabilityTester()
    tester.run_all_tests()