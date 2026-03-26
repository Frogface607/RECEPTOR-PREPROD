#!/usr/bin/env python3
"""
Alt XLSX Export Article Generation Analysis Test
Testing the existing Alt XLSX Export functionality to understand article generation logic.

ЦЕЛЬ: Проверить, как работает существующий "Alt XLSX Export" и почему там артикулы генерируются корректно для ингредиентов.
ENDPOINT: POST /api/v1/techcards.v2/export/iiko.xlsx
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Any, Optional

import httpx
import openpyxl
from pymongo import MongoClient

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017/receptor_pro')
DB_NAME = os.getenv('DB_NAME', 'receptor_pro')

class AltXLSXExportTester:
    """Alt XLSX Export Article Generation Analysis"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=60.0)
        self.test_results = []
        self.organization_id = "test-org-alt-export"
        self.generated_techcard_id = None
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status} {test_name}")
        if details:
            print(f"    {details}")
        if response_time > 0:
            print(f"    Response time: {response_time:.3f}s")
        print()

    async def generate_greek_salad_techcard(self) -> Optional[Dict[str, Any]]:
        """Generate 'Салат Греческий' tech card for testing"""
        try:
            start_time = time.time()
            
            # Generate Greek Salad tech card
            generate_url = f"{API_BASE}/techcards.v2/generate"
            
            payload = {
                "name": "Салат Греческий",
                "description": "Классический греческий салат с оливками, фетой и овощами",
                "organization_id": self.organization_id
            }
            
            response = await self.client.post(generate_url, json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Handle different response structures
                techcard_id = None
                techcard_data = None
                
                if 'id' in data:
                    techcard_id = data.get('id')
                    techcard_data = data
                elif 'card' in data and 'meta' in data['card']:
                    techcard_id = data['card']['meta'].get('id')
                    techcard_data = data['card']  # Use the card data directly
                
                if techcard_id and techcard_data:
                    self.generated_techcard_id = techcard_id
                    
                    # Get ingredient count for analysis
                    ingredients = techcard_data.get('ingredients', [])
                    ingredient_count = len(ingredients)
                    
                    # Check article generation status
                    article_status = "No articles"
                    dish_article = techcard_data.get('meta', {}).get('article') or techcard_data.get('article')
                    ingredient_articles = [ing.get('product_code') for ing in ingredients if ing.get('product_code')]
                    article_status = f"Dish article: {dish_article}, Ingredient articles: {len(ingredient_articles)}"
                    
                    self.log_test(
                        "Generate Greek Salad TechCard",
                        True,
                        f"Generated 'Салат Греческий' with ID: {techcard_id}, {ingredient_count} ingredients. {article_status}",
                        response_time
                    )
                    return techcard_data  # Return the tech card data directly
                else:
                    self.log_test(
                        "Generate Greek Salad TechCard",
                        False,
                        f"No techcard ID found in response structure",
                        response_time
                    )
                    return None
            else:
                self.log_test(
                    "Generate Greek Salad TechCard",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                return None
                
        except Exception as e:
            self.log_test(
                "Generate Greek Salad TechCard",
                False,
                f"Exception: {str(e)}"
            )
            return None

    async def test_alt_xlsx_export(self, techcard_data: Dict[str, Any]) -> Optional[bytes]:
        """Test Alt XLSX Export endpoint"""
        try:
            start_time = time.time()
            
            # Use Alt XLSX Export endpoint with TechCardV2 structure
            export_url = f"{API_BASE}/techcards.v2/export/iiko.xlsx"
            
            # The endpoint expects a TechCardV2 object directly
            response = await self.client.post(export_url, json=techcard_data)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # Check if response is Excel file
                content_type = response.headers.get('content-type', '')
                content_length = len(response.content)
                
                if 'spreadsheet' in content_type or 'excel' in content_type or content_length > 1000:
                    self.log_test(
                        "Alt XLSX Export Request",
                        True,
                        f"Successfully exported XLSX file ({content_length} bytes, content-type: {content_type})",
                        response_time
                    )
                    return response.content
                else:
                    self.log_test(
                        "Alt XLSX Export Request",
                        False,
                        f"Response not Excel file - content-type: {content_type}, size: {content_length} bytes, content: {response.text[:200]}",
                        response_time
                    )
                    return None
            else:
                self.log_test(
                    "Alt XLSX Export Request",
                    False,
                    f"HTTP {response.status_code}: {response.text[:300]}",
                    response_time
                )
                return None
                
        except Exception as e:
            self.log_test(
                "Alt XLSX Export Request",
                False,
                f"Exception: {str(e)}"
            )
            return None

    def analyze_xlsx_articles(self, xlsx_content: bytes) -> Dict[str, Any]:
        """Analyze XLSX file for article generation logic"""
        try:
            # Load Excel file
            workbook = openpyxl.load_workbook(io.BytesIO(xlsx_content))
            
            analysis = {
                "worksheets": [],
                "dish_articles": [],
                "ingredient_articles": [],
                "article_patterns": [],
                "total_rows": 0,
                "articles_found": 0,
                "raw_content": []  # Store raw content for debugging
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_analysis = {
                    "name": sheet_name,
                    "rows": 0,
                    "articles": [],
                    "dish_names": [],
                    "ingredient_names": [],
                    "all_content": []  # Store all content for debugging
                }
                
                # Analyze each row
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if not any(row):  # Skip empty rows
                        continue
                        
                    sheet_analysis["rows"] += 1
                    analysis["total_rows"] += 1
                    
                    # Store raw row content for debugging
                    row_content = [str(cell) if cell is not None else "" for cell in row]
                    sheet_analysis["all_content"].append({
                        "row": row_idx,
                        "content": row_content
                    })
                    
                    # Look for articles in each cell
                    for col_idx, cell_value in enumerate(row):
                        if cell_value is None:
                            continue
                            
                        cell_str = str(cell_value).strip()
                        
                        # Check if this looks like an article (5-digit number or any number)
                        if cell_str.isdigit():
                            article_info = {
                                "article": cell_str,
                                "row": row_idx,
                                "col": col_idx + 1,
                                "sheet": sheet_name,
                                "length": len(cell_str)
                            }
                            sheet_analysis["articles"].append(article_info)
                            
                            # Only count 5-digit articles as proper articles
                            if len(cell_str) == 5:
                                analysis["articles_found"] += 1
                            
                            # Try to find associated name in same row
                            name_cell = None
                            for name_col_idx, name_cell_value in enumerate(row):
                                if name_cell_value and isinstance(name_cell_value, str) and len(name_cell_value) > 3:
                                    if name_col_idx != col_idx:  # Different column
                                        name_cell = str(name_cell_value).strip()
                                        break
                            
                            article_info["associated_name"] = name_cell
                            
                            # Categorize as dish or ingredient article
                            if "Салат Греческий" in str(row):
                                analysis["dish_articles"].append(article_info)
                            else:
                                analysis["ingredient_articles"].append(article_info)
                        
                        # Collect dish names
                        elif "Салат Греческий" in cell_str:
                            sheet_analysis["dish_names"].append({
                                "name": cell_str,
                                "row": row_idx,
                                "col": col_idx + 1
                            })
                        
                        # Collect potential ingredient names (Russian words)
                        elif (len(cell_str) > 2 and 
                              any(ord(c) >= 1040 and ord(c) <= 1103 for c in cell_str) and  # Cyrillic
                              not cell_str.isdigit()):
                            sheet_analysis["ingredient_names"].append({
                                "name": cell_str,
                                "row": row_idx,
                                "col": col_idx + 1
                            })
                
                analysis["worksheets"].append(sheet_analysis)
            
            # Analyze article patterns
            all_articles = [art["article"] for art in analysis["dish_articles"] + analysis["ingredient_articles"]]
            if all_articles:
                analysis["article_patterns"] = {
                    "min_article": min(all_articles),
                    "max_article": max(all_articles),
                    "unique_count": len(set(all_articles)),
                    "total_count": len(all_articles),
                    "sequential": self._check_sequential_pattern(all_articles)
                }
            
            # Store first few rows of content for debugging
            for sheet_data in analysis["worksheets"]:
                analysis["raw_content"].extend(sheet_data["all_content"][:10])  # First 10 rows per sheet
            
            self.log_test(
                "XLSX Article Analysis",
                True,
                f"Found {analysis['articles_found']} proper articles (5-digit): {len(analysis['dish_articles'])} dish, {len(analysis['ingredient_articles'])} ingredient. Total numeric values: {len(all_articles)}"
            )
            
            return analysis
            
        except Exception as e:
            self.log_test(
                "XLSX Article Analysis",
                False,
                f"Exception analyzing XLSX: {str(e)}"
            )
            return {"error": str(e)}

    def _check_sequential_pattern(self, articles: List[str]) -> Dict[str, Any]:
        """Check if articles follow sequential pattern"""
        try:
            numbers = [int(art) for art in articles]
            numbers.sort()
            
            sequential = True
            gaps = []
            
            for i in range(1, len(numbers)):
                gap = numbers[i] - numbers[i-1]
                if gap != 1:
                    sequential = False
                    gaps.append(gap)
            
            return {
                "is_sequential": sequential,
                "gaps": gaps,
                "range": f"{min(numbers)}-{max(numbers)}" if numbers else "none"
            }
        except:
            return {"error": "Could not analyze pattern"}

    async def compare_with_new_zip_export(self, techcard_data: Dict[str, Any]) -> Dict[str, Any]:
        """Compare Alt Export with new ZIP export"""
        try:
            start_time = time.time()
            
            # Try new ZIP export endpoint
            zip_export_url = f"{API_BASE}/export/zip"
            
            techcard_id = techcard_data.get('meta', {}).get('id') or techcard_data.get('id')
            
            payload = {
                "techcard_ids": [techcard_id],
                "organization_id": self.organization_id
            }
            
            response = await self.client.post(zip_export_url, json=payload)
            response_time = time.time() - start_time
            
            comparison = {
                "zip_export_available": False,
                "zip_content_analysis": None,
                "differences": []
            }
            
            if response.status_code == 200:
                content_type = response.headers.get('content-type', '')
                
                if 'zip' in content_type or 'application/zip' in content_type:
                    comparison["zip_export_available"] = True
                    
                    # Analyze ZIP content
                    try:
                        with zipfile.ZipFile(io.BytesIO(response.content)) as zip_file:
                            files = zip_file.namelist()
                            comparison["zip_files"] = files
                            
                            # Look for XLSX files in ZIP
                            xlsx_files = [f for f in files if f.endswith('.xlsx')]
                            if xlsx_files:
                                # Analyze first XLSX file
                                xlsx_content = zip_file.read(xlsx_files[0])
                                comparison["zip_content_analysis"] = self.analyze_xlsx_articles(xlsx_content)
                    except Exception as e:
                        comparison["zip_analysis_error"] = str(e)
                    
                    self.log_test(
                        "New ZIP Export Comparison",
                        True,
                        f"ZIP export available with {len(comparison.get('zip_files', []))} files",
                        response_time
                    )
                else:
                    self.log_test(
                        "New ZIP Export Comparison",
                        False,
                        f"ZIP export returned non-ZIP content: {content_type}",
                        response_time
                    )
            else:
                self.log_test(
                    "New ZIP Export Comparison",
                    False,
                    f"ZIP export failed: HTTP {response.status_code}",
                    response_time
                )
            
            return comparison
            
        except Exception as e:
            self.log_test(
                "New ZIP Export Comparison",
                False,
                f"Exception: {str(e)}"
            )
            return {"error": str(e)}

    async def run_comprehensive_test(self):
        """Run comprehensive Alt XLSX Export analysis"""
        print("🎯 ALT XLSX EXPORT ARTICLE GENERATION ANALYSIS")
        print("=" * 60)
        print()
        
        # Step 1: Generate Greek Salad tech card
        print("📝 STEP 1: Generate 'Салат Греческий' Tech Card")
        techcard_data = await self.generate_greek_salad_techcard()
        
        if not techcard_data:
            print("❌ Cannot proceed without tech card")
            return
        
        print()
        
        # Step 2: Test Alt XLSX Export
        print("📊 STEP 2: Test Alt XLSX Export")
        xlsx_content = await self.test_alt_xlsx_export(techcard_data)
        
        if not xlsx_content:
            print("❌ Cannot proceed without XLSX export")
            return
        
        print()
        
        # Step 3: Analyze XLSX for articles
        print("🔍 STEP 3: Analyze XLSX Articles")
        article_analysis = self.analyze_xlsx_articles(xlsx_content)
        
        print()
        
        # Step 4: Compare with new ZIP export
        print("⚖️ STEP 4: Compare with New ZIP Export")
        comparison = await self.compare_with_new_zip_export(techcard_data)
        
        print()
        
        # Generate comprehensive report
        await self.generate_analysis_report(article_analysis, comparison)

    async def generate_analysis_report(self, article_analysis: Dict[str, Any], comparison: Dict[str, Any]):
        """Generate comprehensive analysis report"""
        print("📋 COMPREHENSIVE ANALYSIS REPORT")
        print("=" * 60)
        print()
        
        # Article Generation Summary
        print("🎯 ARTICLE GENERATION FINDINGS:")
        if article_analysis.get("articles_found", 0) > 0:
            print(f"✅ Articles found: {article_analysis['articles_found']}")
            print(f"   - Dish articles: {len(article_analysis.get('dish_articles', []))}")
            print(f"   - Ingredient articles: {len(article_analysis.get('ingredient_articles', []))}")
            
            # Show article patterns
            patterns = article_analysis.get("article_patterns", {})
            if patterns:
                print(f"   - Article range: {patterns.get('range', 'unknown')}")
                print(f"   - Sequential: {patterns.get('sequential', {}).get('is_sequential', False)}")
        else:
            print("❌ No articles found in Alt XLSX Export")
        
        print()
        
        # Detailed Article List
        if article_analysis.get("dish_articles"):
            print("🍽️ DISH ARTICLES:")
            for art in article_analysis["dish_articles"]:
                print(f"   - {art['article']} (Row {art['row']}, Col {art['col']}) - {art.get('associated_name', 'No name')}")
        
        if article_analysis.get("ingredient_articles"):
            print("🥗 INGREDIENT ARTICLES:")
            for art in article_analysis["ingredient_articles"][:10]:  # Show first 10
                print(f"   - {art['article']} (Row {art['row']}, Col {art['col']}) - {art.get('associated_name', 'No name')}")
            if len(article_analysis["ingredient_articles"]) > 10:
                print(f"   ... and {len(article_analysis['ingredient_articles']) - 10} more")
        
        # Show raw XLSX content for debugging
        if article_analysis.get("raw_content"):
            print("📄 RAW XLSX CONTENT (First 10 rows):")
            for row_data in article_analysis["raw_content"][:10]:
                row_str = " | ".join([cell[:20] + "..." if len(cell) > 20 else cell for cell in row_data["content"] if cell])
                print(f"   Row {row_data['row']}: {row_str}")
        
        print()
        
        # Comparison with ZIP Export
        print("⚖️ COMPARISON WITH NEW ZIP EXPORT:")
        if comparison.get("zip_export_available"):
            zip_analysis = comparison.get("zip_content_analysis", {})
            if zip_analysis and not zip_analysis.get("error"):
                alt_articles = article_analysis.get("articles_found", 0)
                zip_articles = zip_analysis.get("articles_found", 0)
                
                print(f"   - Alt Export articles: {alt_articles}")
                print(f"   - ZIP Export articles: {zip_articles}")
                
                if alt_articles > 0 and zip_articles == 0:
                    print("   🎯 KEY FINDING: Alt Export generates articles, ZIP Export does not!")
                elif alt_articles > 0 and zip_articles > 0:
                    print("   ✅ Both exports generate articles")
                else:
                    print("   ❌ Neither export generates articles")
            else:
                print("   ❌ Could not analyze ZIP export content")
        else:
            print("   ❌ ZIP export not available for comparison")
        
        print()
        
        # Critical Questions Answers
        print("❓ CRITICAL QUESTIONS ANSWERED:")
        
        dish_articles = len(article_analysis.get("dish_articles", []))
        ingredient_articles = len(article_analysis.get("ingredient_articles", []))
        
        # Check for placeholder articles
        placeholder_dish_articles = 0
        placeholder_ingredient_articles = 0
        numeric_dish_articles = 0
        numeric_ingredient_articles = 0
        
        # Look at raw content to find placeholder articles
        for row_data in article_analysis.get("raw_content", []):
            row_content = row_data["content"]
            for cell in row_content:
                if cell.startswith("DISH_"):
                    placeholder_dish_articles += 1
                elif cell.startswith("GENERATED_"):
                    placeholder_ingredient_articles += 1
        
        # Also check the articles list for numeric ones
        for art in article_analysis.get("dish_articles", []):
            if art["article"].isdigit() and len(art["article"]) == 5:
                numeric_dish_articles += 1
        
        for art in article_analysis.get("ingredient_articles", []):
            if art["article"].isdigit() and len(art["article"]) == 5:
                numeric_ingredient_articles += 1
        
        print(f"   1. Генерируются ли артикулы для ингредиентов в Alt Export?")
        if numeric_ingredient_articles > 0:
            print(f"      ✅ ДА - найдено {numeric_ingredient_articles} числовых артикулов ингредиентов")
        elif placeholder_ingredient_articles > 0:
            print(f"      ⚠️ ЧАСТИЧНО - найдено {placeholder_ingredient_articles} placeholder артикулов (GENERATED_*)")
            print(f"         🔍 КЛЮЧЕВОЕ ОТКРЫТИЕ: Alt Export использует placeholder'ы вместо реальных артикулов!")
        else:
            print(f"      ❌ НЕТ - артикулы ингредиентов не найдены")
        
        print(f"   2. Есть ли артикул блюда в Alt Export?")
        if numeric_dish_articles > 0:
            print(f"      ✅ ДА - найдено {numeric_dish_articles} числовых артикулов блюд")
        elif placeholder_dish_articles > 0:
            print(f"      ⚠️ ЧАСТИЧНО - найдено {placeholder_dish_articles} placeholder артикулов (DISH_*)")
            print(f"         🔍 КЛЮЧЕВОЕ ОТКРЫТИЕ: Alt Export использует placeholder'ы вместо реальных артикулов!")
        else:
            print(f"      ❌ НЕТ - артикулы блюд не найдены")
        
        print(f"   3. Какая логика используется для генерации артикулов?")
        if placeholder_dish_articles > 0 or placeholder_ingredient_articles > 0:
            print(f"      🎯 ОБНАРУЖЕНА ЛОГИКА PLACEHOLDER'ОВ:")
            print(f"         - Блюда: DISH_{{НАЗВАНИЕ_БЛЮДА}} (найдено {placeholder_dish_articles})")
            print(f"         - Ингредиенты: GENERATED_{{НАЗВАНИЕ_ИНГРЕДИЕНТА}} (найдено {placeholder_ingredient_articles})")
            print(f"         - Это fallback логика когда реальные артикулы недоступны")
        else:
            patterns = article_analysis.get("article_patterns", {})
            if patterns and not patterns.get("error"):
                if patterns.get("sequential", {}).get("is_sequential"):
                    print(f"      📊 Последовательная нумерация: {patterns.get('range')}")
                else:
                    print(f"      📊 Непоследовательная нумерация: {patterns.get('range')}")
            else:
                print(f"      ❓ Логика не определена - недостаточно данных")
        
        print(f"   4. В чем разница между Alt Export и новым ZIP export?")
        if comparison.get("zip_export_available"):
            zip_analysis = comparison.get("zip_content_analysis", {})
            if zip_analysis and not zip_analysis.get("error"):
                alt_articles = article_analysis.get("articles_found", 0)
                zip_articles = zip_analysis.get("articles_found", 0)
                
                print(f"      📊 Alt Export: {alt_articles} реальных + {placeholder_dish_articles + placeholder_ingredient_articles} placeholder артикулов")
                print(f"      📊 ZIP Export: {zip_articles} артикулов")
                
                if placeholder_dish_articles + placeholder_ingredient_articles > 0 and zip_articles == 0:
                    print(f"      🎯 КЛЮЧЕВАЯ РАЗНИЦА: Alt Export имеет fallback логику placeholder'ов, ZIP Export - нет!")
                elif alt_articles > 0 and zip_articles == 0:
                    print(f"      🎯 РАЗНИЦА: Alt Export генерирует артикулы, ZIP Export - нет")
                else:
                    print(f"      ✅ Оба экспорта работают похоже")
            else:
                print(f"      ❌ Не удалось проанализировать ZIP export для сравнения")
        else:
            print(f"      ❌ ZIP export недоступен для сравнения")
        
        print()
        
        # Success Summary
        passed_tests = sum(1 for result in self.test_results if result["success"])
        total_tests = len(self.test_results)
        success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
        
        print(f"📊 TEST SUMMARY: {passed_tests}/{total_tests} tests passed ({success_rate:.1f}% success rate)")
        
        # Key findings and recommendations
        if placeholder_ingredient_articles > 0 or placeholder_dish_articles > 0:
            print("🎉 SUCCESS: Alt Export показывает рабочую логику placeholder артикулов!")
            print("💡 КЛЮЧЕВЫЕ НАХОДКИ:")
            print("   - Alt Export использует fallback систему placeholder'ов")
            print("   - Формат: DISH_{название} для блюд, GENERATED_{название} для ингредиентов")
            print("   - Это объясняет почему Alt Export 'работает' - он не генерирует реальные артикулы")
            print("💡 РЕКОМЕНДАЦИИ:")
            print("   1. Новый ZIP export должен реализовать аналогичную fallback логику")
            print("   2. Или исправить генерацию реальных числовых артикулов в источнике")
            print("   3. Alt Export НЕ решает проблему - он её маскирует placeholder'ами")
        else:
            print("⚠️ НЕОЖИДАННО: Alt Export также не генерирует артикулы")
            print("💡 РЕКОМЕНДАЦИЯ: Проблема глубже - нужно исследовать генерацию артикулов на уровне источника")

async def main():
    """Main test execution"""
    try:
        async with AltXLSXExportTester() as tester:
            await tester.run_comprehensive_test()
            
    except KeyboardInterrupt:
        print("\n⚠️ Test interrupted by user")
    except Exception as e:
        print(f"\n❌ Critical error: {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    asyncio.run(main())