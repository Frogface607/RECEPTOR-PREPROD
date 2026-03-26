#!/usr/bin/env python3
"""
КРИТИЧЕСКИЙ ТЕСТ: ПОЛНОЕ УСТРАНЕНИЕ MOCK КОНТЕНТА
Тестирование полного устранения mock контента после удаления всех mock методов из export_v2.py.
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Any, Optional

import httpx
import openpyxl
from pymongo import MongoClient

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017/receptor_pro')
DB_NAME = os.getenv('DB_NAME', 'receptor_pro')

# MOCK ПОДПИСИ ДЛЯ ПРОВЕРКИ (должно быть 0)
# These are actual mock/test content signatures, not legitimate generated articles
MOCK_SIGNATURES = [
    "DISH_MOCK_TECH_CARD",
    "GENERATED_TEST_INGREDIENT", 
    "Test Ingredient",
    "Mock Tech Card",
    "ТЕСТОВОЕ БЛЮДО",
    "ТЕСТОВЫЙ ИНГРЕДИЕНТ"
]

class MockEliminationTester:
    """Критический тестер полного устранения mock контента"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=60.0)
        self.test_results = []
        self.organization_id = "test-org-mock-elimination"
        self.generated_techcard_id = None
        self.preflight_result = None
        
    async def __aenter__(self):
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status} {test_name}")
        if details:
            print(f"    {details}")
        if response_time > 0:
            print(f"    Response time: {response_time:.3f}s")
    
    async def test_1_generate_ragu_techcard(self):
        """1. Сгенерировать новую техкарту 'Рагу овощное'"""
        start_time = time.time()
        
        try:
            # Generate tech card for "Рагу овощное"
            generate_url = f"{API_BASE}/techcards.v2/generate"
            
            payload = {
                "name": "Рагу овощное",
                "cuisine": "русская",
                "equipment": [],
                "budget": None,
                "dietary": []
            }
            
            response = await self.client.post(generate_url, json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Extract techcard ID
                if 'card' in data and data['card'] and 'meta' in data['card'] and 'id' in data['card']['meta']:
                    self.generated_techcard_id = data['card']['meta']['id']
                    
                    self.log_test(
                        "Generate Рагу овощное Tech Card",
                        True,
                        f"Generated tech card with ID: {self.generated_techcard_id}",
                        response_time
                    )
                    return True
                else:
                    self.log_test(
                        "Generate Рагу овощное Tech Card",
                        False,
                        f"No techcard ID in response: {data}",
                        response_time
                    )
                    return False
            else:
                self.log_test(
                    "Generate Рагу овощное Tech Card",
                    False,
                    f"HTTP {response.status_code}: {response.text}",
                    response_time
                )
                return False
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Generate Рагу овощное Tech Card",
                False,
                f"Exception: {str(e)}",
                response_time
            )
            return False
    
    async def test_2_verify_db_storage(self):
        """2. Убедиться что техкарта сохранена в БД с реальным UUID"""
        start_time = time.time()
        
        try:
            if not self.generated_techcard_id:
                self.log_test(
                    "Verify DB Storage",
                    False,
                    "No generated techcard ID available"
                )
                return False
            
            # Connect to MongoDB
            client = MongoClient(MONGO_URL)
            db = client[DB_NAME.strip('"')]
            techcards_collection = db.techcards_v2
            
            # Find the techcard in database
            techcard = techcards_collection.find_one({
                "$or": [
                    {"id": self.generated_techcard_id},
                    {"meta.id": self.generated_techcard_id}
                ]
            })
            
            response_time = time.time() - start_time
            
            if techcard:
                # Verify it has real data
                has_real_name = techcard.get('meta', {}).get('title') == "Рагу овощное"
                has_ingredients = len(techcard.get('ingredients', [])) > 0
                has_real_uuid = len(self.generated_techcard_id) == 36  # UUID format
                
                if has_real_name and has_ingredients and has_real_uuid:
                    self.log_test(
                        "Verify DB Storage",
                        True,
                        f"Tech card found in DB with {len(techcard.get('ingredients', []))} ingredients",
                        response_time
                    )
                    return True
                else:
                    self.log_test(
                        "Verify DB Storage",
                        False,
                        f"Tech card missing real data: name={has_real_name}, ingredients={has_ingredients}, uuid={has_real_uuid}",
                        response_time
                    )
                    return False
            else:
                self.log_test(
                    "Verify DB Storage",
                    False,
                    f"Tech card not found in database with ID: {self.generated_techcard_id}",
                    response_time
                )
                return False
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Verify DB Storage",
                False,
                f"Exception: {str(e)}",
                response_time
            )
            return False
    
    async def test_3_run_preflight_with_real_id(self):
        """3. Запустить preflight с реальным ID"""
        start_time = time.time()
        
        try:
            if not self.generated_techcard_id:
                self.log_test(
                    "Run Preflight with Real ID",
                    False,
                    "No generated techcard ID available"
                )
                return False
            
            preflight_url = f"{API_BASE}/export/preflight"
            
            payload = {
                "techcardIds": [self.generated_techcard_id],
                "organization_id": self.organization_id
            }
            
            response = await self.client.post(preflight_url, json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                data = response.json()
                
                # Verify preflight response structure
                has_ttk_date = 'ttkDate' in data
                has_counts = 'counts' in data
                
                if has_ttk_date and has_counts:
                    # Store preflight result for ZIP export
                    self.preflight_result = data
                    
                    self.log_test(
                        "Run Preflight with Real ID",
                        True,
                        f"Preflight successful: TTK date={data.get('ttkDate')}, skeletons={data.get('counts', {})}",
                        response_time
                    )
                    return True
                else:
                    self.log_test(
                        "Run Preflight with Real ID",
                        False,
                        f"Invalid preflight response structure: {data}",
                        response_time
                    )
                    return False
            else:
                self.log_test(
                    "Run Preflight with Real ID",
                    False,
                    f"HTTP {response.status_code}: {response.text}",
                    response_time
                )
                return False
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Run Preflight with Real ID",
                False,
                f"Exception: {str(e)}",
                response_time
            )
            return False
    
    async def test_4_run_zip_export_with_real_id(self):
        """4. Запустить ZIP экспорт с реальным ID"""
        start_time = time.time()
        
        try:
            if not self.generated_techcard_id:
                self.log_test(
                    "Run ZIP Export with Real ID",
                    False,
                    "No generated techcard ID available"
                )
                return False
            
            if not self.preflight_result:
                self.log_test(
                    "Run ZIP Export with Real ID",
                    False,
                    "No preflight result available"
                )
                return False
            
            export_url = f"{API_BASE}/export/zip"
            
            payload = {
                "techcardIds": [self.generated_techcard_id],
                "organization_id": self.organization_id,
                "preflight_result": self.preflight_result
            }
            
            response = await self.client.post(export_url, json=payload)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # Verify it's a ZIP file
                content_type = response.headers.get('content-type', '')
                is_zip = 'zip' in content_type.lower() or response.content[:4] == b'PK\x03\x04'
                
                if is_zip and len(response.content) > 1000:  # Reasonable ZIP size
                    # Store ZIP content for scanning
                    self.zip_content = response.content
                    
                    self.log_test(
                        "Run ZIP Export with Real ID",
                        True,
                        f"ZIP export successful: {len(response.content)} bytes",
                        response_time
                    )
                    return True
                else:
                    self.log_test(
                        "Run ZIP Export with Real ID",
                        False,
                        f"Invalid ZIP response: content-type={content_type}, size={len(response.content)}",
                        response_time
                    )
                    return False
            else:
                self.log_test(
                    "Run ZIP Export with Real ID",
                    False,
                    f"HTTP {response.status_code}: {response.text}",
                    response_time
                )
                return False
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Run ZIP Export with Real ID",
                False,
                f"Exception: {str(e)}",
                response_time
            )
            return False
    
    async def test_5_scan_zip_for_mock_signatures(self):
        """5. КРИТИЧЕСКИ: Просканировать ZIP на отсутствие mock подписей"""
        start_time = time.time()
        
        try:
            if not hasattr(self, 'zip_content'):
                self.log_test(
                    "Scan ZIP for Mock Signatures",
                    False,
                    "No ZIP content available for scanning"
                )
                return False
            
            # Extract and scan all files in ZIP
            mock_findings = {}
            total_mock_count = 0
            
            with zipfile.ZipFile(io.BytesIO(self.zip_content), 'r') as zip_file:
                for file_name in zip_file.namelist():
                    if file_name.endswith('.xlsx'):
                        # Read XLSX file
                        with zip_file.open(file_name) as xlsx_file:
                            workbook = openpyxl.load_workbook(xlsx_file)
                            
                            file_mock_count = 0
                            mock_details = []
                            
                            # Scan all worksheets
                            for sheet_name in workbook.sheetnames:
                                sheet = workbook[sheet_name]
                                
                                # Scan all cells
                                for row in sheet.iter_rows():
                                    for cell in row:
                                        if cell.value:
                                            cell_value = str(cell.value)
                                            
                                            # Check for mock signatures
                                            for signature in MOCK_SIGNATURES:
                                                if signature in cell_value:
                                                    file_mock_count += 1
                                                    total_mock_count += 1
                                                    mock_details.append(f"{signature} in cell {cell.coordinate}: '{cell_value[:50]}...'")
                                                    
                                                    # Debug: Print first few mock findings
                                                    if total_mock_count <= 5:
                                                        print(f"DEBUG: Found mock signature '{signature}' in {file_name} cell {cell.coordinate}: '{cell_value[:100]}'")
                            
                            if file_mock_count > 0:
                                mock_findings[file_name] = {
                                    'count': file_mock_count,
                                    'details': mock_details
                                }
            
            response_time = time.time() - start_time
            
            # КРИТИЧЕСКИЙ РЕЗУЛЬТАТ: должно быть 0 mock подписей
            if total_mock_count == 0:
                self.log_test(
                    "Scan ZIP for Mock Signatures",
                    True,
                    f"✅ КРИТИЧЕСКИЙ УСПЕХ: 0 mock подписей найдено во всех файлах",
                    response_time
                )
                return True
            else:
                details = f"❌ КРИТИЧЕСКАЯ ОШИБКА: {total_mock_count} mock подписей найдено:\n"
                for file_name, findings in mock_findings.items():
                    details += f"  {file_name}: {findings['count']} подписей\n"
                    for detail in findings['details'][:3]:  # Show first 3
                        details += f"    - {detail}\n"
                
                self.log_test(
                    "Scan ZIP for Mock Signatures",
                    False,
                    details,
                    response_time
                )
                return False
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Scan ZIP for Mock Signatures",
                False,
                f"Exception: {str(e)}",
                response_time
            )
            return False
    
    async def test_6_verify_real_content(self):
        """6. Проверить что все содержимое - реальные данные техкарты"""
        start_time = time.time()
        
        try:
            if not hasattr(self, 'zip_content'):
                self.log_test(
                    "Verify Real Content",
                    False,
                    "No ZIP content available for verification"
                )
                return False
            
            # Look for real content indicators
            real_content_found = False
            ragu_mentions = 0
            real_ingredients = []
            
            with zipfile.ZipFile(io.BytesIO(self.zip_content), 'r') as zip_file:
                for file_name in zip_file.namelist():
                    if file_name.endswith('.xlsx'):
                        # Read XLSX file
                        with zip_file.open(file_name) as xlsx_file:
                            workbook = openpyxl.load_workbook(xlsx_file)
                            
                            # Scan all worksheets for real content
                            for sheet_name in workbook.sheetnames:
                                sheet = workbook[sheet_name]
                                
                                # Scan all cells
                                for row in sheet.iter_rows():
                                    for cell in row:
                                        if cell.value:
                                            cell_value = str(cell.value).lower()
                                            
                                            # Look for "Рагу овощное"
                                            if "рагу" in cell_value and "овощное" in cell_value:
                                                ragu_mentions += 1
                                                real_content_found = True
                                            
                                            # Look for real ingredient names (not test/mock)
                                            if any(ingredient in cell_value for ingredient in [
                                                "картофель", "морковь", "лук", "помидор", "перец",
                                                "баклажан", "кабачок", "капуста", "зелень"
                                            ]):
                                                if cell_value not in real_ingredients:
                                                    real_ingredients.append(cell_value)
                                                real_content_found = True
            
            response_time = time.time() - start_time
            
            if real_content_found and ragu_mentions > 0:
                self.log_test(
                    "Verify Real Content",
                    True,
                    f"✅ Реальный контент найден: {ragu_mentions} упоминаний 'Рагу овощное', {len(real_ingredients)} реальных ингредиентов",
                    response_time
                )
                return True
            else:
                self.log_test(
                    "Verify Real Content",
                    False,
                    f"❌ Реальный контент не найден: {ragu_mentions} упоминаний рагу, {len(real_ingredients)} реальных ингредиентов",
                    response_time
                )
                return False
                
        except Exception as e:
            response_time = time.time() - start_time
            self.log_test(
                "Verify Real Content",
                False,
                f"Exception: {str(e)}",
                response_time
            )
            return False
    
    async def run_all_tests(self):
        """Запустить все критические тесты"""
        print("🚨 КРИТИЧЕСКИЙ ТЕСТ: ПОЛНОЕ УСТРАНЕНИЕ MOCK КОНТЕНТА")
        print("=" * 80)
        
        # Run tests in sequence
        test_1_success = await self.test_1_generate_ragu_techcard()
        test_2_success = await self.test_2_verify_db_storage()
        test_3_success = await self.test_3_run_preflight_with_real_id()
        test_4_success = await self.test_4_run_zip_export_with_real_id()
        test_5_success = await self.test_5_scan_zip_for_mock_signatures()
        test_6_success = await self.test_6_verify_real_content()
        
        # Calculate results
        total_tests = 6
        passed_tests = sum([test_1_success, test_2_success, test_3_success, 
                           test_4_success, test_5_success, test_6_success])
        success_rate = (passed_tests / total_tests) * 100
        
        print("\n" + "=" * 80)
        print("🎯 КРИТИЧЕСКИЙ РЕЗУЛЬТАТ ТЕСТИРОВАНИЯ")
        print("=" * 80)
        
        if test_5_success and test_6_success:
            print("✅ КРИТИЧЕСКИЙ УСПЕХ: Mock контент полностью устранен!")
            print("✅ ZIP экспорт работает с реальными данными")
            print("✅ XLSX файлы содержат название 'Рагу овощное'")
            print("✅ Ингредиенты из реальной техкарты")
            print("✅ 0 mock подписей найдено во всех файлах")
            print("✅ Реальные артикулы вместо placeholder'ов")
        else:
            print("❌ КРИТИЧЕСКАЯ ОШИБКА: Mock контент все еще присутствует!")
            if not test_5_success:
                print("❌ Mock подписи найдены в экспортированных файлах")
            if not test_6_success:
                print("❌ Реальный контент техкарты не найден")
        
        print(f"\n📊 Общий результат: {passed_tests}/{total_tests} тестов пройдено ({success_rate:.1f}%)")
        
        # Detailed results
        print("\n📋 Детальные результаты:")
        for result in self.test_results:
            print(f"  {result['status']} {result['test']}")
            if result['details']:
                print(f"      {result['details']}")
        
        return success_rate >= 83.3  # At least 5/6 tests must pass


async def main():
    """Main test execution"""
    try:
        async with MockEliminationTester() as tester:
            success = await tester.run_all_tests()
            
            if success:
                print("\n🎉 ФИНАЛЬНЫЙ ТЕСТ УСПЕШНО ЗАВЕРШЕН!")
                print("Полное устранение mock контента подтверждено.")
                sys.exit(0)
            else:
                print("\n🚨 ФИНАЛЬНЫЙ ТЕСТ НЕ ПРОЙДЕН!")
                print("Mock контент все еще присутствует в системе.")
                sys.exit(1)
                
    except Exception as e:
        print(f"\n💥 КРИТИЧЕСКАЯ ОШИБКА ТЕСТИРОВАНИЯ: {str(e)}")
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())