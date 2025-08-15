"""
Unit тесты для iiko XLSX Export (ТТК)
Проверка корректности экспорта технологических карт в формат iiko
"""

import pytest
import io
from pathlib import Path
import json
from openpyxl import load_workbook

# Импортируем тестируемые модули
from receptor_agent.exports.iiko_xlsx import (
    create_iiko_ttk_xlsx,
    generate_dish_slug,
    normalize_unit_to_grams,
    generate_technology_text,
    IIKO_HEADERS
)
from receptor_agent.techcards_v2.schemas import TechCardV2

class TestIikoXlsxExport:
    """Тесты для экспорта в iiko XLSX"""
    
    @pytest.fixture
    def sample_tech_card_data(self):
        """Создает образец данных TechCardV2 для тестов"""
        return {
            "meta": {
                "title": "Говядина тушеная с овощами",
                "description": "Тестовое блюдо",
                "category": "Горячие блюда",
                "dish_code": "BEEF_STEW_001"
            },
            "ingredients": [
                {
                    "name": "говядина",
                    "netto_g": 400.0,
                    "brutto_g": 450.0,
                    "loss_pct": 11.1,
                    "unit": "g",
                    "skuId": "BEEF_001",
                    "canonical_id": "beef"
                },
                {
                    "name": "лук репчатый",
                    "netto_g": 100.0,
                    "brutto_g": 120.0,
                    "loss_pct": 16.7,
                    "unit": "g",
                    "skuId": "ONION_001",
                    "canonical_id": "onion"
                },
                {
                    "name": "морковь",
                    "netto_g": 80.0,
                    "brutto_g": 100.0,
                    "loss_pct": 20.0,
                    "unit": "g",
                    "skuId": None,  # Тест отсутствующего SKU
                    "canonical_id": None
                },
                {
                    "name": "растительное масло",
                    "netto_g": 30.0,
                    "brutto_g": 30.0,
                    "loss_pct": 0.0,
                    "unit": "ml",  # Тест конвертации единиц
                    "skuId": "OIL_001",
                    "canonical_id": "oil"
                }
            ],
            "portions": 4,
            "yield": {
                "perPortion_g": 150.0,
                "perBatch_g": 600.0
            },
            "process": [
                {
                    "n": 1,
                    "action": "Подготовить ингредиенты",
                    "time_min": 10,
                    "temp_c": None
                },
                {
                    "n": 2,
                    "action": "Обжарить говядину",
                    "time_min": 15,
                    "temp_c": 180
                },
                {
                    "n": 3,
                    "action": "Добавить овощи и тушить",
                    "time_min": 25,
                    "temp_c": 160,
                    "equipment": "сковорода"
                }
            ]
        }
    
    @pytest.fixture
    def tech_card(self, sample_tech_card_data):
        """Создает объект TechCardV2 из данных"""
        return TechCardV2.model_validate(sample_tech_card_data)
    
    def test_generate_dish_slug(self):
        """Тест генерации slug из названия блюда"""
        # Тест обычного названия
        assert generate_dish_slug("Борщ классический") == "БОРЩ_КЛАССИЧЕСКИЙ"
        
        # Тест с специальными символами
        assert generate_dish_slug("Суп-пюре №1 (особый)") == "СУП_ПЮРЕ_1_ОСОБЫЙ"
        
        # Тест длинного названия (должно обрезаться)
        long_title = "Очень длинное название блюда которое должно быть обрезано"
        slug = generate_dish_slug(long_title)
        assert len(slug) <= 20
        assert slug.startswith("ОЧЕНЬ_ДЛИННОЕ")
        
        # Тест пустого названия
        assert generate_dish_slug("") == ""
    
    def test_normalize_unit_to_grams(self):
        """Тест нормализации единиц измерения в граммы"""
        # Граммы остаются граммами
        assert normalize_unit_to_grams(100, "г", "тест") == 100
        assert normalize_unit_to_grams(100, "g", "тест") == 100
        
        # Килограммы в граммы
        assert normalize_unit_to_grams(1, "кг", "тест") == 1000
        assert normalize_unit_to_grams(1, "kg", "тест") == 1000
        
        # Миллилитры в граммы (плотность = 1 для большинства)
        assert normalize_unit_to_grams(100, "мл", "вода") == 100
        assert normalize_unit_to_grams(100, "ml", "вода") == 100
        
        # Масло имеет плотность 0.9
        assert normalize_unit_to_grams(100, "мл", "растительное масло") == 90
        
        # Литры в граммы
        assert normalize_unit_to_grams(1, "л", "вода") == 1000
        
        # Штуки в граммы (известные продукты)
        assert normalize_unit_to_grams(1, "шт", "яйцо") == 50
        assert normalize_unit_to_grams(2, "шт", "луковица") == 300  # 2 * 150
        
        # Штуки неизвестного продукта (средний вес 100г)
        assert normalize_unit_to_grams(1, "шт", "неизвестный продукт") == 100
        
        # Неизвестная единица возвращается как есть
        assert normalize_unit_to_grams(100, "неизвестная", "тест") == 100
    
    def test_generate_technology_text(self):
        """Тест генерации текста технологии приготовления"""
        # Тест с полными данными
        process = [
            {
                "n": 1,
                "action": "Подготовить ингредиенты",
                "time_min": 10,
                "temp_c": None
            },
            {
                "n": 2,
                "action": "Обжарить мясо",
                "time_min": 15,
                "temp_c": 180,
                "equipment": "сковорода"
            }
        ]
        
        result = generate_technology_text(process)
        assert "#1. Подготовить ингредиенты [10 мин]" in result
        assert "#2. Обжарить мясо [t=180°C] [15 мин] [сковорода]" in result
        
        # Тест пустого процесса
        result = generate_technology_text([])
        assert result == "Технология приготовления не указана"
        
        # Тест обрезки длинного текста (>1000 символов)
        long_process = []
        for i in range(50):
            long_process.append({
                "n": i + 1,
                "action": f"Очень длинное действие номер {i+1} с подробным описанием процесса приготовления",
                "time_min": 5
            })
        
        result = generate_technology_text(long_process)
        assert len(result) <= 1000
        assert result.endswith("...")
    
    def test_create_iiko_ttk_xlsx_structure(self, tech_card):
        """Тест создания XLSX файла и проверки структуры"""
        # Генерируем файл
        excel_buffer, issues = create_iiko_ttk_xlsx(tech_card)
        
        # Проверяем, что файл создался
        assert isinstance(excel_buffer, io.BytesIO)
        assert excel_buffer.getvalue()  # Файл не пустой
        
        # Загружаем и проверяем содержимое
        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        # Проверяем заголовки (первая строка)
        expected_headers = [
            'Артикул блюда', 'Наименование блюда', 'Артикул продукта', 
            'Наименование продукта', 'Брутто', 'Потери, %', 'Нетто',
            'Ед.', 'Выход готового продукта', 'Норма закладки', 
            'Метод списания', 'Технология приготовления'
        ]
        
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = ws.cell(row=1, column=col).value
            assert actual_header == expected_header, f"Header mismatch at column {col}: expected {expected_header}, got {actual_header}"
    
    def test_create_iiko_ttk_xlsx_data_mapping(self, tech_card):
        """Тест корректности маппинга данных в XLSX"""
        excel_buffer, issues = create_iiko_ttk_xlsx(tech_card)
        
        # Загружаем файл
        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        # Проверяем данные первого ингредиента (говядина)
        row = 2  # Первая строка данных
        
        # Артикул блюда
        assert ws.cell(row=row, column=1).value == "BEEF_STEW_001"
        
        # Наименование блюда
        assert ws.cell(row=row, column=2).value == "Говядина тушеная с овощами"
        
        # Артикул продукта
        assert ws.cell(row=row, column=3).value == "BEEF_001"
        
        # Наименование продукта
        assert ws.cell(row=row, column=4).value == "говядина"
        
        # Брутто (округлено до 0.1)
        assert ws.cell(row=row, column=5).value == 450.0
        
        # Потери, % (округлено до 0.1)
        assert ws.cell(row=row, column=6).value == 11.1
        
        # Нетто (округлено до 0.1)
        assert ws.cell(row=row, column=7).value == 400.0
        
        # Единица (всегда граммы)
        assert ws.cell(row=row, column=8).value == "г"
        
        # Выход готового продукта
        assert ws.cell(row=row, column=9).value == 600.0
        
        # Норма закладки
        assert ws.cell(row=row, column=10).value == 1
        
        # Метод списания
        assert ws.cell(row=row, column=11).value == 1
        
        # Технология (только в первой строке)
        technology = ws.cell(row=row, column=12).value
        assert technology is not None
        assert "Подготовить ингредиенты" in technology
        assert "Обжарить говядину" in technology
    
    def test_create_iiko_ttk_xlsx_unit_conversion(self, tech_card):
        """Тест конвертации единиц измерения"""
        excel_buffer, issues = create_iiko_ttk_xlsx(tech_card)
        
        # Загружаем файл
        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        # Проверяем масло (должно быть сконвертировано из мл в г)
        oil_row = 5  # Четвертый ингредиент (растительное масло)
        
        # Наименование продукта
        assert ws.cell(row=oil_row, column=4).value == "растительное масло"
        
        # Единица должна быть "г"
        assert ws.cell(row=oil_row, column=8).value == "г"
        
        # Нетто: 30 мл масла ≈ 27г (плотность 0.9)
        netto_value = ws.cell(row=oil_row, column=7).value
        assert netto_value == 27.0  # 30 * 0.9 = 27
    
    def test_create_iiko_ttk_xlsx_missing_sku_handling(self, tech_card):
        """Тест обработки отсутствующих SKU"""
        excel_buffer, issues = create_iiko_ttk_xlsx(tech_card)
        
        # Проверяем issues
        assert len(issues) == 1
        assert issues[0]["type"] == "noSku"
        assert issues[0]["name"] == "морковь"
        assert "GENERATED_" in issues[0]["hint"]
        
        # Загружаем файл и проверяем сгенерированный SKU
        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        # Морковь - третий ингредиент
        carrot_row = 4
        product_code = ws.cell(row=carrot_row, column=3).value
        assert product_code.startswith("GENERATED_")
        assert "МОРКОВЬ" in product_code.upper()
    
    def test_create_iiko_ttk_xlsx_ingredients_count(self, tech_card):
        """Тест количества ингредиентов в экспорте"""
        excel_buffer, issues = create_iiko_ttk_xlsx(tech_card)
        
        # Загружаем файл
        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        # Считаем количество строк с данными
        row_count = 0
        for row in range(2, ws.max_row + 1):  # Начиная со второй строки
            if ws.cell(row=row, column=1).value:  # Если есть артикул блюда
                row_count += 1
        
        # Должно быть 4 ингредиента
        assert row_count == 4
    
    def test_create_iiko_ttk_xlsx_decimal_precision(self, tech_card):
        """Тест точности десятичных значений"""
        excel_buffer, issues = create_iiko_ttk_xlsx(tech_card)
        
        # Загружаем файл
        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        # Проверяем округление до 0.1
        for row in range(2, 6):  # Строки 2-5 (4 ингредиента)
            brutto = ws.cell(row=row, column=5).value
            loss_pct = ws.cell(row=row, column=6).value
            netto = ws.cell(row=row, column=7).value
            output_qty = ws.cell(row=row, column=9).value
            
            # Проверяем, что значения округлены до 1 знака после запятой
            if brutto is not None:
                assert brutto == round(float(brutto), 1)
            if loss_pct is not None:
                assert loss_pct == round(float(loss_pct), 1)
            if netto is not None:
                assert netto == round(float(netto), 1)
            if output_qty is not None:
                assert output_qty == round(float(output_qty), 1)

if __name__ == "__main__":
    pytest.main([__file__, "-v"])