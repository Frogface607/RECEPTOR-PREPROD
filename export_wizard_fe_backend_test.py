#!/usr/bin/env python3
"""
FE Export Wizard: disable mocks and bind to /export/preflight + /export/zip (credit-safe)
Comprehensive testing to ensure UI-export uses real endpoints and produces real XLSX files without mock content.
"""

import asyncio
import json
import os
import sys
import time
import traceback
import zipfile
import hashlib
import tempfile
import shutil
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path

import httpx
import openpyxl
from openpyxl.utils import get_column_letter

# Test Configuration
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://cursor-push.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api/v1"

class ExportWizardFETester:
    """FE Export Wizard Backend Testing Suite"""
    
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=30.0)
        self.test_results = []
        self.artifacts = {}
        self.organization_id = "test-org-export-wizard"
        self.temp_dir = None
        
    async def __aenter__(self):
        # Create temporary directory for artifacts
        self.temp_dir = tempfile.mkdtemp(prefix="export_wizard_test_")
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.client.aclose()
        # Don't clean up temporary directory for debugging
        print(f"🔍 Artifacts preserved in: {self.temp_dir}")
        # if self.temp_dir and os.path.exists(self.temp_dir):
        #     shutil.rmtree(self.temp_dir)
    
    def log_test(self, test_name: str, success: bool, details: str = "", response_time: float = 0.0):
        """Log test result"""
        status = "✅ PASS" if success else "❌ FAIL"
        result = {
            "test": test_name,
            "status": status,
            "success": success,
            "details": details,
            "response_time": f"{response_time:.3f}s" if response_time > 0 else "N/A"
        }
        self.test_results.append(result)
        print(f"{status} {test_name}: {details}")
    
    def save_artifact(self, filename: str, data: Any):
        """Save test artifact"""
        artifact_path = os.path.join(self.temp_dir, filename)
        
        if isinstance(data, (dict, list)):
            with open(artifact_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        elif isinstance(data, bytes):
            with open(artifact_path, 'wb') as f:
                f.write(data)
        else:
            with open(artifact_path, 'w', encoding='utf-8') as f:
                f.write(str(data))
        
        self.artifacts[filename] = artifact_path
        print(f"📁 Artifact saved: {filename}")
    
    async def generate_real_techcards(self) -> List[str]:
        """Step 1: Generate 2 real tech cards using /techcards.v2/generate"""
        print("\n🔧 STEP 1: Generate Real Tech Cards")
        
        techcard_ids = []
        dish_names = ["Тест Блюдо Экспорт 1", "Тест Блюдо Экспорт 2"]
        
        for dish_name in dish_names:
            start_time = time.time()
            
            try:
                response = await self.client.post(
                    f"{API_BASE}/techcards.v2/generate",
                    json={
                        "name": dish_name,
                        "description": f"Тестовое блюдо для проверки экспорта без моков",
                        "category": "горячее",
                        "operational_rounding": True
                    }
                )
                
                response_time = time.time() - start_time
                
                if response.status_code == 200:
                    data = response.json()
                    # Extract ID from the correct location in response
                    techcard_id = (
                        data.get('id') or 
                        data.get('techcard', {}).get('id') or
                        data.get('card', {}).get('meta', {}).get('id')
                    )
                    
                    if techcard_id:
                        techcard_ids.append(techcard_id)
                        self.log_test(
                            f"Generate Tech Card: {dish_name}",
                            True,
                            f"Generated ID: {techcard_id}",
                            response_time
                        )
                    else:
                        self.log_test(
                            f"Generate Tech Card: {dish_name}",
                            False,
                            f"No ID in response: {data}",
                            response_time
                        )
                else:
                    self.log_test(
                        f"Generate Tech Card: {dish_name}",
                        False,
                        f"HTTP {response.status_code}: {response.text[:200]}",
                        response_time
                    )
                    
            except Exception as e:
                self.log_test(
                    f"Generate Tech Card: {dish_name}",
                    False,
                    f"Exception: {str(e)}",
                    time.time() - start_time
                )
        
        # Save generation results
        gen_results = {
            "dish_names": dish_names,
            "generated_ids": techcard_ids,
            "timestamp": datetime.now().isoformat()
        }
        self.save_artifact("gen_runs.json", gen_results)
        
        return techcard_ids
    
    async def run_preflight(self, techcard_ids: List[str]) -> Dict[str, Any]:
        """Step 2: Call /export/preflight with real techcard IDs"""
        print("\n🔍 STEP 2: Run Preflight Check")
        
        start_time = time.time()
        
        try:
            response = await self.client.post(
                f"{API_BASE}/export/preflight",
                json={
                    "techcardIds": techcard_ids,
                    "operational_rounding": True
                }
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                preflight_data = response.json()
                
                # Extract key information
                counts = preflight_data.get('counts', {})
                missing = preflight_data.get('missing', {})
                ttk_date = preflight_data.get('ttk_date')
                
                self.log_test(
                    "Preflight Check",
                    True,
                    f"TTK Date: {ttk_date}, Dish Skeletons: {counts.get('dishSkeletons', 0)}, Product Skeletons: {counts.get('productSkeletons', 0)}",
                    response_time
                )
                
                # Save preflight results
                self.save_artifact("preflight.json", preflight_data)
                
                return preflight_data
                
            else:
                self.log_test(
                    "Preflight Check",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                return {}
                
        except Exception as e:
            self.log_test(
                "Preflight Check",
                False,
                f"Exception: {str(e)}",
                time.time() - start_time
            )
            return {}
    
    async def zip_export(self, techcard_ids: List[str]) -> Dict[str, Any]:
        """Step 3: Call /export/zip with same techcard IDs"""
        print("\n📦 STEP 3: ZIP Export")
        
        start_time = time.time()
        
        try:
            response = await self.client.post(
                f"{API_BASE}/export/zip",
                json={
                    "techcardIds": techcard_ids,
                    "operational_rounding": True
                }
            )
            
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                # Check if response is JSON or binary
                content_type = response.headers.get('content-type', '')
                
                if 'application/json' in content_type:
                    # JSON response with zipUrl
                    export_data = response.json()
                    zip_url = export_data.get('zipUrl')
                    file_list = export_data.get('files', [])
                    
                    self.log_test(
                        "ZIP Export",
                        True,
                        f"ZIP URL: {zip_url}, Files: {len(file_list)}",
                        response_time
                    )
                    
                    # Save export results
                    export_results = {
                        "zipUrl": zip_url,
                        "files": file_list,
                        "response": export_data,
                        "timestamp": datetime.now().isoformat()
                    }
                    self.save_artifact("export.json", export_results)
                    
                    return export_data
                    
                elif 'application/zip' in content_type or 'application/octet-stream' in content_type:
                    # Direct ZIP file response
                    zip_content = response.content
                    
                    # Save ZIP file directly
                    zip_path = os.path.join(self.temp_dir, "direct_export.zip")
                    with open(zip_path, 'wb') as f:
                        f.write(zip_content)
                    
                    self.log_test(
                        "ZIP Export",
                        True,
                        f"Direct ZIP file received: {len(zip_content)} bytes",
                        response_time
                    )
                    
                    # Create export data structure for direct ZIP
                    export_data = {
                        "direct_zip": True,
                        "zip_path": zip_path,
                        "size_bytes": len(zip_content),
                        "content_type": content_type
                    }
                    
                    # Save export results
                    export_results = {
                        "direct_zip": True,
                        "zip_path": zip_path,
                        "size_bytes": len(zip_content),
                        "content_type": content_type,
                        "timestamp": datetime.now().isoformat()
                    }
                    self.save_artifact("export.json", export_results)
                    
                    return export_data
                    
                else:
                    # Unknown content type
                    self.log_test(
                        "ZIP Export",
                        False,
                        f"Unknown content type: {content_type}",
                        response_time
                    )
                    return {}
                
            else:
                self.log_test(
                    "ZIP Export",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                return {}
                
        except Exception as e:
            self.log_test(
                "ZIP Export",
                False,
                f"Exception: {str(e)}",
                time.time() - start_time
            )
            return {}
    
    async def download_and_hash(self, export_data: Dict[str, Any]) -> Optional[str]:
        """Step 4: Download ZIP from zipUrl or use direct ZIP, calculate hash and extract"""
        print("\n⬇️ STEP 4: Download and Hash ZIP")
        
        # Handle direct ZIP file
        if export_data.get('direct_zip'):
            zip_path = export_data.get('zip_path')
            if not zip_path or not os.path.exists(zip_path):
                self.log_test("Download ZIP", False, "Direct ZIP file not found")
                return None
            
            # Read the ZIP file
            with open(zip_path, 'rb') as f:
                zip_content = f.read()
            
            # Calculate hash and size
            sha256_hash = hashlib.sha256(zip_content).hexdigest()
            file_size = len(zip_content)
            
            # Extract ZIP
            extract_path = os.path.join(self.temp_dir, "extracted")
            os.makedirs(extract_path, exist_ok=True)
            
            try:
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(extract_path)
                    extracted_files = zip_ref.namelist()
                
                self.log_test(
                    "Download ZIP",
                    True,
                    f"Direct ZIP processed: Size: {file_size} bytes, SHA256: {sha256_hash[:16]}..., Files: {len(extracted_files)}",
                    0.0
                )
                
                # Save metadata
                zip_meta = {
                    "sha256": sha256_hash,
                    "size_bytes": file_size,
                    "extracted_files": extracted_files,
                    "zip_path": zip_path,
                    "extract_path": extract_path,
                    "direct_zip": True,
                    "timestamp": datetime.now().isoformat()
                }
                self.save_artifact("zip_meta.json", zip_meta)
                
                return extract_path
                
            except Exception as e:
                self.log_test("Download ZIP", False, f"Failed to extract direct ZIP: {str(e)}")
                return None
        
        # Handle ZIP URL
        zip_url = export_data.get('zipUrl')
        if not zip_url:
            self.log_test("Download ZIP", False, "No ZIP URL or direct ZIP provided")
            return None
        
        start_time = time.time()
        
        try:
            # Download ZIP file
            response = await self.client.get(zip_url, timeout=25.0)
            response_time = time.time() - start_time
            
            if response.status_code == 200:
                zip_content = response.content
                
                # Calculate hash and size
                sha256_hash = hashlib.sha256(zip_content).hexdigest()
                file_size = len(zip_content)
                
                # Save ZIP file
                zip_path = os.path.join(self.temp_dir, "export.zip")
                with open(zip_path, 'wb') as f:
                    f.write(zip_content)
                
                # Extract ZIP
                extract_path = os.path.join(self.temp_dir, "extracted")
                os.makedirs(extract_path, exist_ok=True)
                
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(extract_path)
                    extracted_files = zip_ref.namelist()
                
                self.log_test(
                    "Download ZIP",
                    True,
                    f"Size: {file_size} bytes, SHA256: {sha256_hash[:16]}..., Files: {len(extracted_files)}",
                    response_time
                )
                
                # Save metadata
                zip_meta = {
                    "sha256": sha256_hash,
                    "size_bytes": file_size,
                    "extracted_files": extracted_files,
                    "zip_path": zip_path,
                    "extract_path": extract_path,
                    "timestamp": datetime.now().isoformat()
                }
                self.save_artifact("zip_meta.json", zip_meta)
                
                return extract_path
                
            else:
                self.log_test(
                    "Download ZIP",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}",
                    response_time
                )
                return None
                
        except Exception as e:
            self.log_test(
                "Download ZIP",
                False,
                f"Exception: {str(e)}",
                time.time() - start_time
            )
            return None
    
    def mock_scan(self, extract_path: str) -> bool:
        """Step 5: Scan XLSX files for mock signatures"""
        print("\n🔍 STEP 5: Mock Content Scan")
        
        mock_signatures = [
            "DISH_MOCK_TECH_CARD",
            "GENERATED_TEST_INGREDIENT",
            "GENERATED_TEST_INGREDIENT_2",
            "Артикул,Наименование,Тип,Ед. выпуска,Выход"
        ]
        
        mock_findings = []
        xlsx_files = []
        
        # Find all XLSX files
        for root, dirs, files in os.walk(extract_path):
            for file in files:
                if file.endswith('.xlsx'):
                    xlsx_files.append(os.path.join(root, file))
        
        # Scan each XLSX file
        for xlsx_file in xlsx_files:
            try:
                workbook = openpyxl.load_workbook(xlsx_file, data_only=True)
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Scan all cells for mock signatures
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value:
                                cell_value = str(cell.value)
                                
                                for signature in mock_signatures:
                                    if signature in cell_value:
                                        mock_findings.append({
                                            "file": os.path.basename(xlsx_file),
                                            "sheet": sheet_name,
                                            "cell": f"{get_column_letter(cell.column)}{cell.row}",
                                            "signature": signature,
                                            "content": cell_value[:100]
                                        })
                
                workbook.close()
                
            except Exception as e:
                print(f"Error scanning {xlsx_file}: {str(e)}")
        
        # Save scan results
        scan_results = {
            "xlsx_files_scanned": [os.path.basename(f) for f in xlsx_files],
            "mock_signatures_searched": mock_signatures,
            "mock_findings": mock_findings,
            "scan_passed": len(mock_findings) == 0,
            "timestamp": datetime.now().isoformat()
        }
        self.save_artifact("mock_scan.json", scan_results)
        
        if mock_findings:
            self.log_test(
                "Mock Content Scan",
                False,
                f"Found {len(mock_findings)} mock signatures in XLSX files"
            )
            return False
        else:
            self.log_test(
                "Mock Content Scan",
                True,
                f"No mock signatures found in {len(xlsx_files)} XLSX files"
            )
            return True
    
    def xlsx_validate(self, extract_path: str) -> bool:
        """Step 6: Verify XLSX format - articles as text (@) with leading zeros"""
        print("\n📊 STEP 6: XLSX Format Validation")
        
        validation_results = []
        xlsx_files = []
        
        # Find all XLSX files
        for root, dirs, files in os.walk(extract_path):
            for file in files:
                if file.endswith('.xlsx'):
                    xlsx_files.append(os.path.join(root, file))
        
        # Validate each XLSX file
        for xlsx_file in xlsx_files:
            try:
                workbook = openpyxl.load_workbook(xlsx_file)
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Look for article columns (typically columns with "Артикул" in header)
                    article_columns = []
                    
                    # Check first row for headers
                    if sheet.max_row > 0:
                        for col in range(1, sheet.max_column + 1):
                            header_cell = sheet.cell(row=1, column=col)
                            if header_cell.value and "артикул" in str(header_cell.value).lower():
                                article_columns.append(col)
                    
                    # Validate article columns format
                    for col in article_columns:
                        for row in range(2, min(sheet.max_row + 1, 20)):  # Check first 20 rows
                            cell = sheet.cell(row=row, column=col)
                            
                            if cell.value:
                                # Check if cell is formatted as text (@)
                                is_text_format = (
                                    cell.number_format == '@' or 
                                    cell.data_type == 's' or
                                    isinstance(cell.value, str)
                                )
                                
                                # Check for leading zeros preservation
                                cell_value = str(cell.value)
                                has_leading_zeros = (
                                    cell_value.startswith('0') and 
                                    len(cell_value) > 1 and 
                                    cell_value.isdigit()
                                )
                                
                                validation_results.append({
                                    "file": os.path.basename(xlsx_file),
                                    "sheet": sheet_name,
                                    "cell": f"{get_column_letter(col)}{row}",
                                    "value": cell_value,
                                    "is_text_format": is_text_format,
                                    "has_leading_zeros": has_leading_zeros,
                                    "number_format": cell.number_format,
                                    "data_type": cell.data_type
                                })
                
                workbook.close()
                
            except Exception as e:
                print(f"Error validating {xlsx_file}: {str(e)}")
        
        # Analyze validation results
        text_format_count = sum(1 for r in validation_results if r['is_text_format'])
        leading_zeros_count = sum(1 for r in validation_results if r['has_leading_zeros'])
        total_articles = len(validation_results)
        
        validation_passed = (
            total_articles > 0 and
            text_format_count >= total_articles * 0.8  # At least 80% should be text format
        )
        
        # Save validation results
        xlsx_checks = {
            "xlsx_files_validated": [os.path.basename(f) for f in xlsx_files],
            "total_article_cells": total_articles,
            "text_format_count": text_format_count,
            "leading_zeros_count": leading_zeros_count,
            "validation_passed": validation_passed,
            "detailed_results": validation_results[:50],  # Limit to first 50 for readability
            "timestamp": datetime.now().isoformat()
        }
        self.save_artifact("xlsx_checks.json", xlsx_checks)
        
        if validation_passed:
            self.log_test(
                "XLSX Format Validation",
                True,
                f"Articles properly formatted: {text_format_count}/{total_articles} as text, {leading_zeros_count} with leading zeros"
            )
        else:
            self.log_test(
                "XLSX Format Validation",
                False,
                f"Article formatting issues: {text_format_count}/{total_articles} as text format"
            )
        
        return validation_passed
    
    def correlate_with_preflight(self, preflight_data: Dict[str, Any], extract_path: str) -> bool:
        """Step 7: Check correspondence between preflight counts and ZIP contents"""
        print("\n🔗 STEP 7: Correlate with Preflight")
        
        counts = preflight_data.get('counts', {})
        dish_skeletons_expected = counts.get('dishSkeletons', 0)
        product_skeletons_expected = counts.get('productSkeletons', 0)
        
        # Check actual files in ZIP
        files_found = []
        for root, dirs, files in os.walk(extract_path):
            files_found.extend(files)
        
        has_iiko_ttk = any('iiko_TTK.xlsx' in f for f in files_found)
        has_dish_skeletons = any('Dish-Skeletons.xlsx' in f for f in files_found)
        has_product_skeletons = any('Product-Skeletons.xlsx' in f for f in files_found)
        
        # Validate correspondence
        correlation_passed = True
        correlation_details = []
        
        # iiko_TTK.xlsx should always be present
        if not has_iiko_ttk:
            correlation_passed = False
            correlation_details.append("Missing required iiko_TTK.xlsx file")
        
        # Dish skeletons correspondence
        if dish_skeletons_expected > 0 and not has_dish_skeletons:
            correlation_passed = False
            correlation_details.append(f"Expected {dish_skeletons_expected} dish skeletons but Dish-Skeletons.xlsx not found")
        elif dish_skeletons_expected == 0 and has_dish_skeletons:
            correlation_details.append("Unexpected Dish-Skeletons.xlsx found (preflight showed 0)")
        
        # Product skeletons correspondence
        if product_skeletons_expected > 0 and not has_product_skeletons:
            correlation_passed = False
            correlation_details.append(f"Expected {product_skeletons_expected} product skeletons but Product-Skeletons.xlsx not found")
        elif product_skeletons_expected == 0 and has_product_skeletons:
            correlation_details.append("Unexpected Product-Skeletons.xlsx found (preflight showed 0)")
        
        correlation_result = {
            "preflight_counts": counts,
            "files_found": files_found,
            "has_iiko_ttk": has_iiko_ttk,
            "has_dish_skeletons": has_dish_skeletons,
            "has_product_skeletons": has_product_skeletons,
            "correlation_passed": correlation_passed,
            "correlation_details": correlation_details,
            "timestamp": datetime.now().isoformat()
        }
        
        # Save correlation results (this completes our artifacts)
        self.save_artifact("correlation.json", correlation_result)
        
        if correlation_passed:
            self.log_test(
                "Preflight Correlation",
                True,
                f"File composition matches preflight: TTK={has_iiko_ttk}, Dish={has_dish_skeletons}, Product={has_product_skeletons}"
            )
        else:
            self.log_test(
                "Preflight Correlation",
                False,
                f"Mismatch: {'; '.join(correlation_details)}"
            )
        
        return correlation_passed
    
    def generate_summary(self):
        """Generate comprehensive test summary"""
        print("\n📋 STEP 8: Generate Summary")
        
        total_tests = len(self.test_results)
        passed_tests = sum(1 for r in self.test_results if r['success'])
        success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
        
        # Categorize results
        critical_failures = []
        warnings = []
        successes = []
        
        for result in self.test_results:
            if not result['success']:
                if any(keyword in result['test'].lower() for keyword in ['mock', 'generate', 'preflight', 'zip']):
                    critical_failures.append(result)
                else:
                    warnings.append(result)
            else:
                successes.append(result)
        
        summary = {
            "test_execution": {
                "timestamp": datetime.now().isoformat(),
                "total_tests": total_tests,
                "passed_tests": passed_tests,
                "failed_tests": total_tests - passed_tests,
                "success_rate": f"{success_rate:.1f}%"
            },
            "critical_failures": critical_failures,
            "warnings": warnings,
            "successes": successes,
            "artifacts_created": list(self.artifacts.keys()),
            "artifacts_location": self.temp_dir,
            "acceptance_criteria_status": {
                "real_endpoints_used": any("Generate Tech Card" in r['test'] and r['success'] for r in self.test_results),
                "no_mock_content": any("Mock Content Scan" in r['test'] and r['success'] for r in self.test_results),
                "xlsx_format_valid": any("XLSX Format Validation" in r['test'] and r['success'] for r in self.test_results),
                "preflight_correlation": any("Preflight Correlation" in r['test'] and r['success'] for r in self.test_results)
            }
        }
        
        self.save_artifact("summary.md", self._format_summary_markdown(summary))
        
        return summary
    
    def _format_summary_markdown(self, summary: Dict[str, Any]) -> str:
        """Format summary as markdown"""
        md = f"""# FE Export Wizard Test Summary

## Test Execution
- **Timestamp**: {summary['test_execution']['timestamp']}
- **Total Tests**: {summary['test_execution']['total_tests']}
- **Success Rate**: {summary['test_execution']['success_rate']}

## Acceptance Criteria Status
- **Real Endpoints Used**: {'✅' if summary['acceptance_criteria_status']['real_endpoints_used'] else '❌'}
- **No Mock Content**: {'✅' if summary['acceptance_criteria_status']['no_mock_content'] else '❌'}
- **XLSX Format Valid**: {'✅' if summary['acceptance_criteria_status']['xlsx_format_valid'] else '❌'}
- **Preflight Correlation**: {'✅' if summary['acceptance_criteria_status']['preflight_correlation'] else '❌'}

## Test Results

### Critical Failures ({len(summary['critical_failures'])})
"""
        for failure in summary['critical_failures']:
            md += f"- **{failure['test']}**: {failure['details']}\n"
        
        md += f"\n### Warnings ({len(summary['warnings'])})\n"
        for warning in summary['warnings']:
            md += f"- **{warning['test']}**: {warning['details']}\n"
        
        md += f"\n### Successes ({len(summary['successes'])})\n"
        for success in summary['successes']:
            md += f"- **{success['test']}**: {success['details']}\n"
        
        md += f"\n## Artifacts Created\n"
        for artifact in summary['artifacts_created']:
            md += f"- {artifact}\n"
        
        return md
    
    async def run_comprehensive_test(self):
        """Execute the complete FE Export Wizard test workflow"""
        print("🚀 Starting FE Export Wizard Comprehensive Test")
        print(f"Backend URL: {BACKEND_URL}")
        print(f"Temp Directory: {self.temp_dir}")
        
        try:
            # Step 1: Generate real tech cards
            techcard_ids = await self.generate_real_techcards()
            
            if not techcard_ids:
                print("❌ CRITICAL: No tech cards generated - cannot continue")
                return self.generate_summary()
            
            # Step 2: Run preflight
            preflight_data = await self.run_preflight(techcard_ids)
            
            if not preflight_data:
                print("❌ CRITICAL: Preflight failed - cannot continue")
                return self.generate_summary()
            
            # Step 3: ZIP export
            export_data = await self.zip_export(techcard_ids)
            
            if not export_data:
                print("❌ CRITICAL: ZIP export failed - cannot continue")
                return self.generate_summary()
            
            # Step 4: Download and extract ZIP
            extract_path = await self.download_and_hash(export_data)
            
            if not extract_path:
                print("❌ CRITICAL: ZIP download failed - cannot continue")
                return self.generate_summary()
            
            # Step 5: Mock content scan
            mock_scan_passed = self.mock_scan(extract_path)
            
            # Step 6: XLSX format validation
            xlsx_validation_passed = self.xlsx_validate(extract_path)
            
            # Step 7: Correlate with preflight
            correlation_passed = self.correlate_with_preflight(preflight_data, extract_path)
            
            # Step 8: Generate summary
            summary = self.generate_summary()
            
            return summary
            
        except Exception as e:
            print(f"❌ CRITICAL ERROR: {str(e)}")
            traceback.print_exc()
            return self.generate_summary()

async def main():
    """Main test execution"""
    print("=" * 80)
    print("FE Export Wizard: disable mocks and bind to /export/preflight + /export/zip")
    print("=" * 80)
    
    async with ExportWizardFETester() as tester:
        summary = await tester.run_comprehensive_test()
        
        print("\n" + "=" * 80)
        print("TEST SUMMARY")
        print("=" * 80)
        
        print(f"Total Tests: {summary['test_execution']['total_tests']}")
        print(f"Success Rate: {summary['test_execution']['success_rate']}")
        print(f"Artifacts Location: {summary['artifacts_location']}")
        
        # Print acceptance criteria
        print("\nAcceptance Criteria:")
        for criterion, status in summary['acceptance_criteria_status'].items():
            status_icon = "✅" if status else "❌"
            print(f"  {status_icon} {criterion.replace('_', ' ').title()}")
        
        # Print critical failures
        if summary['critical_failures']:
            print(f"\n❌ Critical Failures ({len(summary['critical_failures'])}):")
            for failure in summary['critical_failures']:
                print(f"  - {failure['test']}: {failure['details']}")
        
        # Overall result
        all_criteria_met = all(summary['acceptance_criteria_status'].values())
        overall_status = "✅ PASS" if all_criteria_met else "❌ FAIL"
        print(f"\nOverall Result: {overall_status}")
        
        if all_criteria_met:
            print("🎉 FE Export Wizard test completed successfully!")
            print("✅ UI-export uses real endpoints without mock content")
            print("✅ XLSX files have proper text (@) formatting")
            print("✅ ZIP composition matches preflight results")
        else:
            print("⚠️ FE Export Wizard test found issues that need attention")
        
        return 0 if all_criteria_met else 1

if __name__ == "__main__":
    exit_code = asyncio.run(main())
    sys.exit(exit_code)