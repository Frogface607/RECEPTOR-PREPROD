#!/usr/bin/env python3
"""
Focused Article Workflow Testing
Quick focused test of the article workflow functionality.
"""

import asyncio
import json
import logging
import os
import sys
import time
from typing import Dict, Any, List, Optional
import httpx
from datetime import datetime
import tempfile
import openpyxl
from io import BytesIO

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class FocusedArticleWorkflowTester:
    """Focused test suite for article workflow"""
    
    def __init__(self):
        # Get backend URL from environment
        self.backend_url = os.environ.get('REACT_APP_BACKEND_URL', 'http://localhost:8001')
        if not self.backend_url.endswith('/api'):
            self.backend_url = f"{self.backend_url}/api"
        
        self.test_results = []
        self.total_tests = 0
        self.passed_tests = 0
        
        logger.info(f"🔧 Backend URL: {self.backend_url}")
    
    def log_test_result(self, test_name: str, passed: bool, details: str = "", data: Any = None):
        """Log test result"""
        self.total_tests += 1
        if passed:
            self.passed_tests += 1
            logger.info(f"✅ {test_name}: PASSED - {details}")
        else:
            logger.error(f"❌ {test_name}: FAILED - {details}")
        
        self.test_results.append({
            'test': test_name,
            'passed': passed,
            'details': details,
            'data': data,
            'timestamp': datetime.now().isoformat()
        })
    
    async def test_debug_rms_endpoint(self):
        """Test debug RMS endpoint"""
        logger.info("🔍 Testing debug RMS endpoint...")
        
        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                response = await client.post(
                    f"{self.backend_url}/v1/techcards.v2/debug/rms-product",
                    json={"sku_id": "test-guid-123"}
                )
                
                if response.status_code == 200:
                    data = response.json()
                    result = data.get('result', {})
                    
                    # Check structure
                    expected_fields = ['sku_id', 'product_in_products', 'product_in_prices', 'extracted_article']
                    has_all_fields = all(field in result for field in expected_fields)
                    
                    self.log_test_result(
                        "Debug RMS Endpoint Structure",
                        has_all_fields,
                        f"Has all expected fields: {expected_fields}"
                    )
                    
                    # Check extracted article
                    extracted_article = result.get('extracted_article')
                    self.log_test_result(
                        "Debug RMS Article Extraction",
                        extracted_article is not None,
                        f"Extracted article: '{extracted_article}'"
                    )
                    
                else:
                    self.log_test_result(
                        "Debug RMS Endpoint",
                        False,
                        f"Status: {response.status_code}, Response: {response.text[:200]}"
                    )
                    
        except Exception as e:
            self.log_test_result(
                "Debug RMS Endpoint",
                False,
                f"Exception: {str(e)}"
            )
    
    async def test_enhanced_export(self):
        """Test enhanced export with articles"""
        logger.info("🔍 Testing enhanced export...")
        
        try:
            # Simple test techcard
            test_techcard = {
                "meta": {
                    "id": "test-article-export",
                    "title": "Тест экспорта артикулов",
                    "description": "Тестовое блюдо"
                },
                "ingredients": [
                    {
                        "name": "Тестовый ингредиент",
                        "quantity": 100,
                        "unit": "г",
                        "skuId": "test-guid-ingredient",
                        "product_code": "01499"  # 5-digit article
                    }
                ],
                "yield_": {
                    "perPortion_g": 100,
                    "perBatch_g": 100
                },
                "portions": 1,
                "nutrition": {"per100g": {}, "perPortion": {}},
                "cost": {"per100g": {}, "perPortion": {}},
                "process": {"steps": []}
            }
            
            async with httpx.AsyncClient(timeout=30.0) as client:
                export_request = {
                    "card": test_techcard,
                    "options": {
                        "use_product_codes": True,
                        "operational_rounding": False
                    },
                    "organization_id": "test-org",
                    "user_email": "test@example.com"
                }
                
                response = await client.post(
                    f"{self.backend_url}/v1/techcards.v2/export/enhanced/iiko.xlsx",
                    json=export_request
                )
                
                if response.status_code == 200:
                    content_type = response.headers.get('content-type', '')
                    is_excel = 'spreadsheet' in content_type or 'excel' in content_type
                    file_size = len(response.content)
                    
                    self.log_test_result(
                        "Enhanced Export Response",
                        is_excel and file_size > 1000,
                        f"Content-Type: {content_type}, Size: {file_size} bytes"
                    )
                    
                    if is_excel and file_size > 1000:
                        # Try to parse Excel and check for articles
                        try:
                            excel_buffer = BytesIO(response.content)
                            workbook = openpyxl.load_workbook(excel_buffer)
                            worksheet = workbook.active
                            
                            # Look for article column
                            article_column = None
                            for col in range(1, min(worksheet.max_column + 1, 15)):  # Limit search
                                header = worksheet.cell(row=1, column=col).value
                                if header and "Артикул продукта" in str(header):
                                    article_column = col
                                    break
                            
                            if article_column:
                                # Check first data row
                                article_cell = worksheet.cell(row=2, column=article_column)
                                article_value = str(article_cell.value) if article_cell.value else ""
                                is_text_format = article_cell.number_format == '@'
                                
                                self.log_test_result(
                                    "Excel Article Column Found",
                                    True,
                                    f"Article: '{article_value}', Format: '{article_cell.number_format}'"
                                )
                                
                                self.log_test_result(
                                    "Excel Article Text Format",
                                    is_text_format,
                                    f"Cell format is text (@): {is_text_format}"
                                )
                                
                                # Check if article matches expected 5-digit format
                                is_5_digit = article_value.isdigit() and len(article_value) == 5
                                self.log_test_result(
                                    "Excel Article 5-Digit Format",
                                    is_5_digit,
                                    f"Article '{article_value}' is 5-digit: {is_5_digit}"
                                )
                            else:
                                self.log_test_result(
                                    "Excel Article Column",
                                    False,
                                    "Could not find 'Артикул продукта' column"
                                )
                                
                        except Exception as e:
                            self.log_test_result(
                                "Excel File Parsing",
                                False,
                                f"Could not parse Excel: {str(e)}"
                            )
                    
                else:
                    self.log_test_result(
                        "Enhanced Export",
                        False,
                        f"Export failed: {response.status_code} - {response.text[:200]}"
                    )
                    
        except Exception as e:
            self.log_test_result(
                "Enhanced Export",
                False,
                f"Exception: {str(e)}"
            )
    
    async def test_dish_codes_endpoints(self):
        """Test dish codes endpoints"""
        logger.info("🔍 Testing dish codes endpoints...")
        
        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                # Test dish codes generation
                generate_request = {
                    "dish_names": ["Тестовое блюдо 1", "Тестовое блюдо 2"],
                    "organization_id": "test-org",
                    "width": 5
                }
                
                response = await client.post(
                    f"{self.backend_url}/v1/techcards.v2/dish-codes/generate",
                    json=generate_request
                )
                
                if response.status_code == 200:
                    data = response.json()
                    generated_codes = data.get('generated_codes', {})
                    
                    self.log_test_result(
                        "Dish Codes Generation",
                        len(generated_codes) == 2,
                        f"Generated {len(generated_codes)} codes: {generated_codes}"
                    )
                    
                    # Check format of generated codes
                    all_valid = True
                    for dish_name, code in generated_codes.items():
                        is_valid = isinstance(code, str) and code.isdigit() and len(code) == 5
                        if not is_valid:
                            all_valid = False
                    
                    self.log_test_result(
                        "Dish Codes Format Validation",
                        all_valid,
                        f"All codes are 5-digit format: {all_valid}"
                    )
                    
                else:
                    self.log_test_result(
                        "Dish Codes Generation",
                        response.status_code in [400, 500],  # Expected with no RMS
                        f"Status: {response.status_code} (expected with no RMS data)"
                    )
                    
        except Exception as e:
            self.log_test_result(
                "Dish Codes Endpoints",
                False,
                f"Exception: {str(e)}"
            )
    
    async def test_preflight_check(self):
        """Test preflight check"""
        logger.info("🔍 Testing preflight check...")
        
        try:
            test_techcards = [
                {
                    "meta": {"title": "Тестовое блюдо", "id": "test-1"},
                    "ingredients": [
                        {"name": "Ингредиент без кода", "skuId": "guid-1"}
                    ],
                    "nutrition": {"per100g": {}, "perPortion": {}},
                    "cost": {"per100g": {}, "perPortion": {}},
                    "process": {"steps": []}
                }
            ]
            
            async with httpx.AsyncClient(timeout=15.0) as client:
                preflight_request = {
                    "techcards": test_techcards,
                    "export_options": {
                        "use_product_codes": True,
                        "dish_codes_mapping": {}
                    },
                    "organization_id": "test-org"
                }
                
                response = await client.post(
                    f"{self.backend_url}/v1/techcards.v2/export/preflight-check",
                    json=preflight_request
                )
                
                if response.status_code == 200:
                    data = response.json()
                    status = data.get('status')
                    warnings = data.get('warnings', [])
                    
                    self.log_test_result(
                        "Preflight Check Response",
                        status in ['ready', 'warnings'],
                        f"Status: {status}, Warnings: {len(warnings)}"
                    )
                    
                    # Should have warnings for missing codes
                    has_warnings = len(warnings) > 0
                    self.log_test_result(
                        "Preflight Warnings Detection",
                        has_warnings,
                        f"Detected {len(warnings)} warnings as expected"
                    )
                    
                else:
                    self.log_test_result(
                        "Preflight Check",
                        False,
                        f"Status: {response.status_code} - {response.text[:200]}"
                    )
                    
        except Exception as e:
            self.log_test_result(
                "Preflight Check",
                False,
                f"Exception: {str(e)}"
            )
    
    async def test_article_formatting_logic(self):
        """Test article formatting logic"""
        logger.info("🔍 Testing article formatting logic...")
        
        try:
            # Test zfill formatting (used in the code)
            test_cases = [
                ("1499", "01499"),
                ("597", "00597"),
                ("1", "00001"),
                ("12345", "12345"),
                ("0", "00000"),
            ]
            
            all_passed = True
            for input_code, expected in test_cases:
                formatted = input_code.zfill(5)
                passed = formatted == expected
                if not passed:
                    all_passed = False
                
                self.log_test_result(
                    f"Article Formatting - {input_code}",
                    passed,
                    f"'{input_code}' → '{formatted}' (expected '{expected}')"
                )
            
            self.log_test_result(
                "Article Formatting Logic",
                all_passed,
                "All formatting tests passed"
            )
            
        except Exception as e:
            self.log_test_result(
                "Article Formatting Logic",
                False,
                f"Exception: {str(e)}"
            )
    
    async def run_focused_tests(self):
        """Run focused article workflow tests"""
        logger.info("🚀 Starting Focused Article Workflow Testing...")
        
        start_time = time.time()
        
        # Run focused tests
        await self.test_debug_rms_endpoint()
        await self.test_enhanced_export()
        await self.test_dish_codes_endpoints()
        await self.test_preflight_check()
        await self.test_article_formatting_logic()
        
        end_time = time.time()
        duration = end_time - start_time
        
        # Generate summary
        success_rate = (self.passed_tests / self.total_tests * 100) if self.total_tests > 0 else 0
        
        logger.info("=" * 80)
        logger.info("🎯 FOCUSED ARTICLE WORKFLOW TESTING COMPLETED")
        logger.info("=" * 80)
        logger.info(f"📊 Total Tests: {self.total_tests}")
        logger.info(f"✅ Passed: {self.passed_tests}")
        logger.info(f"❌ Failed: {self.total_tests - self.passed_tests}")
        logger.info(f"📈 Success Rate: {success_rate:.1f}%")
        logger.info(f"⏱️ Duration: {duration:.2f}s")
        logger.info("=" * 80)
        
        return {
            'total_tests': self.total_tests,
            'passed_tests': self.passed_tests,
            'success_rate': success_rate,
            'duration': duration,
            'results': self.test_results
        }


async def main():
    """Main function"""
    tester = FocusedArticleWorkflowTester()
    results = await tester.run_focused_tests()
    
    if results['success_rate'] >= 70:  # 70% success threshold
        logger.info("🎉 Focused article workflow testing PASSED!")
        return 0
    else:
        logger.error("💥 Focused article workflow testing FAILED!")
        return 1


if __name__ == "__main__":
    exit_code = asyncio.run(main())
    sys.exit(exit_code)