"""
iiko XLSX Export Module
WITH iiko Import Reliability — Product Codes & Dish Skeletons
"""

from __future__ import annotations
from io import BytesIO
from pathlib import Path
from typing import List, Dict, Any, Tuple
import re
import os
import logging

# Используем openpyxl для работы с Excel
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    # Fallback если openpyxl не установлен
    Workbook = None
    load_workbook = None

from receptor_agent.techcards_v2.schemas import TechCardV2

logger = logging.getLogger(__name__)

def get_product_code_from_rms(sku_id: str, rms_service=None) -> str:
    """
    Feature A: Получить числовой код продукта из iiko RMS вместо GUID
    
    Args:
        sku_id: GUID продукта из iiko
        rms_service: Сервис для работы с iiko RMS (опционально)
    
    Returns:
        Числовой код продукта или исходный sku_id если код не найден
    """
    if not sku_id or not rms_service:
        return sku_id or ""
    
    try:
        # Поиск в коллекции products
        product = rms_service.products.find_one({"_id": sku_id})
        if product and product.get('article'):
            # Возвращаем article (числовой код) с сохранением ведущих нулей
            return str(product['article']).zfill(5)  # Минимум 5 цифр с ведущими нулями
        
        # Поиск в коллекции prices если нет в products
        pricing = rms_service.prices.find_one({"skuId": sku_id})
        if pricing and pricing.get('article'):
            return str(pricing['article']).zfill(5)
            
    except Exception as e:
        print(f"Error getting product code for {sku_id}: {e}")
    
    return sku_id or ""


def find_dish_in_iiko_rms(dish_name: str, rms_service=None) -> Dict[str, Any]:
    """
    Feature B: Поиск блюда в iiko RMS по имени с нормализацией
    
    Args:
        dish_name: Название блюда для поиска
        rms_service: Сервис для работы с iiko RMS
    
    Returns:
        {
            'status': 'found' | 'not_found' | 'error',
            'dish_code': str,
            'dish_name': str,
            'confidence': float
        }
    """
    if not dish_name or not rms_service:
        return {'status': 'error', 'message': 'Missing dish name or RMS service'}
    
    try:
        # Нормализация названия для поиска (убираем лишние символы, приводим к нижнему регистру)
        normalized_search = dish_name.lower().strip()
        normalized_search = re.sub(r'[^\w\s]', '', normalized_search)  # Убираем пунктуацию
        normalized_search = re.sub(r'\s+', ' ', normalized_search)     # Нормализуем пробелы
        
        # Поиск в коллекции products по типу "Блюдо"
        dishes = list(rms_service.products.find({
            "type": {"$in": ["Блюдо", "блюдо", "dish", "DISH"]},
            "active": True
        }))
        
        best_match = None
        best_confidence = 0
        
        for dish in dishes:
            dish_name_db = dish.get('name', '').lower().strip()
            dish_name_db = re.sub(r'[^\w\s]', '', dish_name_db)
            dish_name_db = re.sub(r'\s+', ' ', dish_name_db)
            
            # Простое совпадение по подстроке
            if normalized_search in dish_name_db or dish_name_db in normalized_search:
                confidence = 0.9
            # Проверка на совпадение слов
            elif len(set(normalized_search.split()) & set(dish_name_db.split())) > 0:
                confidence = 0.7
            else:
                continue
            
            if confidence > best_confidence:
                best_confidence = confidence
                best_match = dish
        
        if best_match and best_confidence >= 0.7:
            return {
                'status': 'found',
                'dish_code': str(best_match.get('article', best_match.get('code', ''))).zfill(5),
                'dish_name': best_match.get('name', ''),
                'confidence': best_confidence
            }
        
        return {'status': 'not_found'}
        
    except Exception as e:
        return {'status': 'error', 'message': str(e)}


def generate_dish_codes(dish_names: List[str], rms_service=None, width: int = 5) -> Dict[str, str]:
    """
    Feature B: Генерация свободных числовых кодов для новых блюд
    
    Args:
        dish_names: Список названий блюд для генерации кодов
        rms_service: Сервис для работы с iiko RMS (для проверки занятых кодов)  
        width: Ширина кода (количество цифр с ведущими нулями)
    
    Returns:
        Dict[dish_name, generated_code]
    """
    if not dish_names:
        return {}
    
    # Получаем список занятых кодов из базы
    used_codes = set()
    
    if rms_service:
        try:
            # Собираем все коды из products
            products = rms_service.products.find({}, {"article": 1, "code": 1})
            for product in products:
                code = product.get('article') or product.get('code')
                if code and str(code).isdigit():
                    used_codes.add(int(code))
        except Exception as e:
            print(f"Error getting used codes: {e}")
    
    # Генерируем свободные коды
    generated_codes = {}
    start_code = 10**(width-1)  # Минимальный код для заданной ширины (например, 10000 для width=5)
    current_code = start_code
    
    for dish_name in dish_names:
        # Ищем первый свободный код
        while current_code in used_codes:
            current_code += 1
        
        # Форматируем с ведущими нулями
        formatted_code = str(current_code).zfill(width)
        generated_codes[dish_name] = formatted_code
        
        # Помечаем как занятый
        used_codes.add(current_code)
        current_code += 1
    
    return generated_codes


def create_dish_skeletons_xlsx(dish_codes_mapping: Dict[str, str], 
                              cards: List[TechCardV2]) -> BytesIO:
    """
    Feature B: Создание файла Dish-Skeletons.xlsx для предварительного импорта блюд в iiko
    
    Args:
        dish_codes_mapping: Маппинг {dish_name: dish_code}
        cards: Список техкарт для извлечения данных о блюдах
    
    Returns:
        BytesIO buffer с Excel файлом
    """
    if not Workbook:
        raise ImportError("openpyxl is required for Excel export")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Блюда"
    
    # Заголовки согласно требованиям
    headers = [
        "Артикул",           # Числовой код блюда
        "Наименование",      # Название блюда  
        "Тип",               # Всегда "Блюдо"
        "Ед. выпуска",       # Единица измерения (г/мл)
        "Выход"              # Выход готового продукта
    ]
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Заполняем данные по блюдам
    row = 2
    for card in cards:
        dish_name = card.meta.title
        dish_code = dish_codes_mapping.get(dish_name, '')
        
        if dish_code:
            # Определяем выход и единицу
            yield_g = card.yield_.perPortion_g if hasattr(card, 'yield_') else 200
            unit = 'г'  # По умолчанию граммы
            
            row_data = [
                dish_code,      # Артикул
                dish_name,      # Наименование
                "Блюдо",        # Тип
                unit,           # Ед. выпуска
                yield_g         # Выход
            ]
            
            # Записываем данные
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                
                # Форматируем артикул как текст
                if col == 1:  # Колонка "Артикул"
                    cell.number_format = '@'
            
            row += 1
    
    # Автоширина колонок
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row_cells in ws[f"{column_letter}1:{column_letter}{row-1}"]:
            for cell in row_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

# Путь к шаблону iiko
TEMPLATE_PATH = Path(__file__).parent.parent / "data" / "iiko_templates" / "ttk.xlsx"

# Маппинг заголовков iiko (поддержка синонимов)
IIKO_HEADERS = {
    'dish_code': ['Артикул блюда', 'Код блюда'],
    'dish_name': ['Наименование блюда', 'Название блюда'],
    'product_code': ['Артикул продукта', 'Код продукта'],
    'product_name': ['Наименование продукта', 'Название продукта'],
    'brutto': ['Брутто', 'Брутто, г'],
    'loss_pct': ['Потери, %', 'Потери'],
    'netto': ['Нетто', 'Нетто, г'],
    'unit': ['Ед.', 'Единица', 'Ед. изм.'],
    'output_qty': ['Выход готового продукта', 'Выход'],
    'norm': ['Норма закладки'],
    'write_off_method': ['Метод списания'],
    'technology': ['Технология приготовления', 'Технология']
}

def generate_dish_slug(title: str) -> str:
    """Генерирует slug из названия блюда"""
    # Убираем специальные символы и заменяем пробелы на подчеркивания
    slug = re.sub(r'[^\w\s-]', '', title).strip()
    slug = re.sub(r'[-\s]+', '_', slug)
    return slug.upper()[:20]  # Ограничиваем длину

def normalize_unit_to_grams(value: float, unit: str, ingredient_name: str = "") -> float:
    """
    Конвертирует единицы в граммы для iiko
    """
    unit = unit.lower().strip()
    
    # Уже граммы
    if unit in ['г', 'g', 'gram', 'грамм']:
        return value
    
    # Килограммы -> граммы
    elif unit in ['кг', 'kg', 'kilogram', 'килограмм']:
        return value * 1000
    
    # Миллилитры -> граммы (принимаем плотность = 1 для большинства жидкостей)
    elif unit in ['мл', 'ml', 'milliliter', 'миллилитр']:
        # Для масла плотность ≈ 0.9
        if 'масло' in ingredient_name.lower():
            return value * 0.9
        # Для воды и большинства жидкостей плотность = 1
        return value
    
    # Литры -> граммы
    elif unit in ['л', 'l', 'liter', 'литр']:
        return value * 1000  # Принимаем плотность = 1
    
    # Штуки -> граммы (средние веса)
    elif unit in ['шт', 'pcs', 'piece', 'штука']:
        # Стандартные веса продуктов
        piece_weights = {
            'яйцо': 50,      # 50г за яйцо
            'луковица': 150,  # 150г за луковицу
            'морковь': 100,   # 100г за морковь
            'картофель': 120, # 120г за картофелину
            'помидор': 180,   # 180г за помидор
            'огурец': 150,    # 150г за огурец
            'яблоко': 180,    # 180г за яблоко
            'лимон': 130      # 130г за лимон
        }
        
        # Пытаемся найти подходящий вес
        for product, weight in piece_weights.items():
            if product in ingredient_name.lower():
                return value * weight
        
        # Если не нашли - используем средний вес 100г
        return value * 100
    
    # По умолчанию возвращаем как есть
    return value

def generate_technology_text(process_steps) -> str:
    """
    Генерирует текст технологии из шагов процесса
    Формат: #{n}. {action} [t={temp_c}°C] [{time_min} мин] [{equipment}]
    """
    if not process_steps:
        return "Технология приготовления не указана"
    
    technology_lines = []
    
    for step in process_steps:
        # Handle both dict and object types
        if hasattr(step, 'n'):
            step_num = step.n
            action = getattr(step, 'action', 'Действие не указано')
            temp_c = getattr(step, 'temp_c', None)
            time_min = getattr(step, 'time_min', None)
            equipment = getattr(step, 'equipment', None)
        else:
            step_num = step.get('n', len(technology_lines) + 1)
            action = step.get('action', 'Действие не указано')
            temp_c = step.get('temp_c')
            time_min = step.get('time_min')
            equipment = step.get('equipment')
        
        # Базовая строка
        line = f"#{step_num}. {action}"
        
        # Добавляем температуру если есть
        if temp_c:
            line += f" [t={temp_c}°C]"
        
        # Добавляем время если есть
        if time_min:
            line += f" [{time_min} мин]"
        
        # Добавляем оборудование если есть
        if equipment:
            if isinstance(equipment, list):
                equipment_str = ', '.join(equipment)
            else:
                equipment_str = str(equipment)
            line += f" [{equipment_str}]"
        
        technology_lines.append(line)
    
    # Объединяем в один текст, ограничиваем 1000 символами
    technology_text = '\n'.join(technology_lines)
    if len(technology_text) > 1000:
        technology_text = technology_text[:997] + "..."
    
    return technology_text

def create_iiko_ttk_xlsx(card: TechCardV2, 
                        export_options: Dict[str, Any] = None) -> Tuple[BytesIO, List[Dict[str, Any]]]:
    """
    Создать iiko XLSX файл для импорта технологических карт
    WITH iiko Import Reliability — Product Codes & Dish Skeletons
    WITH Operational Rounding v1 (Export & Kitchen View)
    
    Args:
        card: Техкарта в формате TechCardV2
        export_options: Опции экспорта:
            - use_product_codes: bool = True - использовать коды товаров вместо GUID
            - dish_codes_mapping: Dict[str, str] - маппинг блюд к кодам
            - operational_rounding: bool = True - применять операционное округление
    
    Returns:
        (BytesIO buffer, issues list)
    """
    if not export_options:
        export_options = {}
    
    # Feature A: Product Code toggle (по умолчанию включен)
    use_product_codes = export_options.get('use_product_codes', True)
    dish_codes_mapping = export_options.get('dish_codes_mapping', {})
    rms_service = export_options.get('rms_service')  # Для получения кодов продуктов
    
    # Operational Rounding v1: операционное округление (по умолчанию включено)
    operational_rounding_enabled = export_options.get('operational_rounding', True)
    
    if not Workbook:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    issues = []
    
    # Operational Rounding v1: Применяем операционное округление если включено
    working_card = card
    rounding_metadata = None
    
    if operational_rounding_enabled:
        try:
            from receptor_agent.techcards_v2.operational_rounding import get_operational_rounder
            logger.info("Applying operational rounding...")
            rounder = get_operational_rounder()
            
            # Конвертируем TechCardV2 в dict для округления
            card_dict = card.model_dump()
            logger.info(f"Card dict keys: {list(card_dict.keys())}")
            logger.info(f"Meta dict: {card_dict.get('meta', {})}")
            
            rounding_result = rounder.round_techcard_ingredients(card_dict)
            
            # DEBUG: Проверяем результат округления
            rounded_dict = rounding_result['rounded_techcard']
            logger.info(f"Rounded card keys: {list(rounded_dict.keys())}")
            logger.info(f"Rounded meta: {rounded_dict.get('meta', {})}")
            
            # Создаем новый TechCardV2 объект с округленными данными
            working_card = TechCardV2.model_validate(rounded_dict)
            rounding_metadata = rounding_result['rounding_metadata']
            
            logger.info(f"Working card meta type: {type(working_card.meta)}")
            logger.info(f"Working card meta: {working_card.meta}")
            
            # Логируем применение округления
            if rounding_metadata and rounding_metadata.get('items'):
                logger.info(f"Operational rounding applied to {len(rounding_metadata['items'])} ingredients, "
                           f"delta: {rounding_metadata.get('delta_g', 0)}g")
                
                # Добавляем информацию в issues для отчета
                if abs(rounding_metadata.get('delta_g', 0)) > 0.1:
                    issues.append({
                        "type": "operationalRounding",
                        "message": f"Применено операционное округление: {len(rounding_metadata['items'])} ингредиентов, "
                                 f"дельта: {rounding_metadata.get('delta_g', 0):+.1f}г",
                        "severity": "info",
                        "details": rounding_metadata
                    })
            
        except Exception as e:
            logger.error(f"Operational rounding error: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            # Продолжаем с исходной техкартой если округление не удалось
            issues.append({
                "type": "roundingError",
                "message": f"Ошибка операционного округления: {str(e)}",
                "severity": "warning"
            })
    
    # Пытаемся загрузить шаблон, если не получается - создаем новый
    wb = None
    if TEMPLATE_PATH.exists():
        try:
            wb = load_workbook(TEMPLATE_PATH)
            ws = wb.active
        except Exception as e:
            print(f"Failed to load template {TEMPLATE_PATH}: {e}")
            wb = None
    
    if wb is None:
        # Создаем новую книгу с правильными заголовками
        wb = Workbook()
        ws = wb.active
        ws.title = "ТТК"
        
        # Заголовки в правильном порядке
        headers = [
            'Артикул блюда',
            'Наименование блюда', 
            'Артикул продукта',
            'Наименование продукта',
            'Брутто',
            'Потери, %',
            'Нетто',
            'Ед.',
            'Выход готового продукта',
            'Норма закладки',
            'Метод списания',
            'Технология приготовления'
        ]
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    else:
        ws = wb.active
    
    # Генерируем данные для экспорта
    # Feature A: Use dish_codes_mapping if provided
    # DEBUG: Детальное логирование meta объекта для поиска проблемы с 'str' object has no attribute 'get'
    logger.info(f"DEBUG working_card type: {type(working_card)}")
    logger.info(f"DEBUG working_card.meta type: {type(working_card.meta)}")
    logger.info(f"DEBUG working_card.meta repr: {repr(working_card.meta)}")
    
    try:
        if hasattr(working_card.meta, 'title'):
            dish_title = working_card.meta.title
        elif isinstance(working_card.meta, dict):
            dish_title = working_card.meta.get('title', 'Unknown Dish')
        elif isinstance(working_card.meta, str):
            # Если meta является строкой, используем её как название
            dish_title = working_card.meta
            logger.warning(f"working_card.meta is a string: {working_card.meta}")
        else:
            dish_title = str(working_card.meta)
            logger.warning(f"working_card.meta is unexpected type {type(working_card.meta)}: {working_card.meta}")
    except Exception as e:
        logger.error(f"Error accessing working_card.meta.title: {e}")
        logger.error(f"working_card.meta type: {type(working_card.meta)}")
        logger.error(f"working_card.meta: {working_card.meta}")
        dish_title = 'Unknown Dish'
    
    dish_code = dish_codes_mapping.get(dish_title)
    if not dish_code:
        meta_dict = working_card.meta.model_dump() if hasattr(working_card.meta, 'model_dump') else working_card.meta
        dish_code = meta_dict.get('dish_code') if isinstance(meta_dict, dict) else None
    if not dish_code:
        dish_slug = generate_dish_slug(dish_title)
        dish_code = f"DISH_{dish_slug}"
    
    dish_name = dish_title
    
    # Рассчитываем выход готового продукта
    output_qty = getattr(working_card.yield_, 'perBatch_g', 0) if hasattr(working_card, 'yield_') and working_card.yield_ else 0
    if not output_qty and hasattr(working_card, 'yield_') and working_card.yield_:
        # Если нет perBatch_g, рассчитываем из порций
        portions = getattr(working_card, 'portions', 1)
        per_portion = getattr(working_card.yield_, 'perPortion_g', 0)
        output_qty = portions * per_portion if per_portion else 0
    
    # Генерируем технологию приготовления
    technology_text = ""
    if hasattr(working_card, 'process') and working_card.process:
        technology_text = generate_technology_text(working_card.process)
    
    # Заполняем данные по ингредиентам
    row = 2  # Начинаем со второй строки (первая - заголовки)
    
    for ingredient in working_card.ingredients:
        # Артикул продукта
        # Feature A: Product Code toggle - использовать реальные коды товаров
        if use_product_codes:
            # A. Hotfix & Migration: Сначала пытаемся использовать уже сохраненный product_code
            product_code = getattr(ingredient, 'product_code', None)
            
            # Если нет сохраненного кода, получаем из RMS по skuId
            if not product_code and ingredient.skuId:
                product_code = get_product_code_from_rms(ingredient.skuId, rms_service)
            
            # Если это все еще GUID или нет кода, генерируем
            if not product_code or product_code == ingredient.skuId:
                # Генерируем артикул если отсутствует код в iiko
                ingredient_slug = generate_dish_slug(ingredient.name)
                product_code = f"GENERATED_{ingredient_slug}"
                issues.append({
                    "type": "noProductCode",
                    "name": ingredient.name,
                    "hint": f"Продукт не найден в iiko RMS или отсутствует код товара. Сгенерирован: {product_code}",
                    "dish": dish_title
                })
        else:
            # Используем GUID как есть (legacy режим)
            product_code = ingredient.skuId or f"GUID_{generate_dish_slug(ingredient.name)}"
        
        # Обработка подрецептов
        if hasattr(ingredient, 'subRecipe') and ingredient.subRecipe:
            # Для подрецептов используем их dish_code как артикул продукта
            sub_recipe_code = getattr(ingredient.subRecipe, 'dish_code', None)
            if sub_recipe_code and use_product_codes:
                product_code = sub_recipe_code
            elif not use_product_codes:
                # Используем GUID для подрецептов
                product_code = getattr(ingredient.subRecipe, 'guid', sub_recipe_code or f"SUBGUID_{generate_dish_slug(ingredient.name)}")
            else:
                issues.append({
                    "type": "subRecipeNoCode",
                    "name": ingredient.name,
                    "hint": "Sub-recipe missing dish_code",
                    "dish": dish_title
                })
        
        # Конвертируем единицы в граммы
        brutto_g = round(normalize_unit_to_grams(
            getattr(ingredient, 'brutto_g', 0), 
            getattr(ingredient, 'unit', 'г'), 
            ingredient.name
        ), 1)
        
        netto_g = round(normalize_unit_to_grams(
            getattr(ingredient, 'netto_g', 0), 
            getattr(ingredient, 'unit', 'г'), 
            ingredient.name
        ), 1)
        
        loss_pct = round(getattr(ingredient, 'loss_pct', 0), 1)
        
        # Заполняем строку
        row_data = [
            dish_code,                    # Артикул блюда
            dish_name,                    # Наименование блюда
            product_code,                 # Артикул продукта
            ingredient.name,              # Наименование продукта
            brutto_g,                     # Брутто
            loss_pct,                     # Потери, %
            netto_g,                      # Нетто
            'г',                          # Ед. (всегда граммы)
            round(output_qty, 1),         # Выход готового продукта
            1,                            # Норма закладки
            1,                            # Метод списания
            technology_text if row == 2 else ""  # Технология (только для первой строки)
        ]
        
        # Записываем данные в ячейки
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            
            # Feature A: Форматирование кодов продуктов и блюд как текст для сохранения ведущих нулей
            if col in [1, 3]:  # Колонки "Артикул блюда" и "Артикул продукта"
                cell.number_format = '@'  # Текстовый формат для сохранения ведущих нулей
        
        row += 1
    
    # Автоширина колонок
    for col in range(1, 13):  # 12 колонок
        column_letter = get_column_letter(col)
        max_length = 0
        for row_cells in ws[f"{column_letter}1:{column_letter}{row-1}"]:
            for cell in row_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer, issues


def export_techcard_to_iiko_xlsx(card: TechCardV2, export_options: Dict[str, Any] = None) -> bytes:
    """
    Wrapper function for round-trip tests compatibility
    Returns just the Excel bytes
    """
    buffer, issues = create_iiko_ttk_xlsx(card, export_options)
    return buffer.getvalue()