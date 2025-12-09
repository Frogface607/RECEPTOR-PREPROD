"""
IK-04/01: iiko XLSX → TechCardV2 Parser
Backend parser for importing iiko template tech cards (XLSX) and converting to valid TechCardV2
"""

import os
import re
import logging
from io import BytesIO
from typing import Dict, List, Any, Optional, Tuple, Union
from datetime import datetime, timezone
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

class IikoXlsxParseError(Exception):
    """Custom exception for iiko XLSX parsing errors"""
    pass

class IikoXlsxParser:
    """Parser for iiko XLSX tech card templates"""
    
    # Column mapping for flexible header recognition (RU/EN support)
    COLUMN_MAPPINGS = {
        # Meta information
        "dish_code": ["артикул блюда", "dish code", "dish article", "код блюда"],
        "dish_name": ["название", "name", "dish name", "наименование блюда"],
        "category": ["категория", "category", "группа", "group"],
        "yield_total": ["выход готового продукта", "total yield", "выход", "yield", "готовый продукт"],
        "technology": ["технология приготовления", "technology", "cooking technology", "процесс приготовления", "приготовление"],
        
        # Ingredient information
        "ingredient_code": ["артикул продукта", "product code", "ingredient code", "sku", "код продукта"],
        "ingredient_name": ["наименование продукта", "product name", "ingredient name", "продукт", "ингредиент"],
        "brutto": ["брутто", "gross", "brutto weight", "масса брутто"],
        "netto": ["нетто", "net", "netto weight", "масса нетто"],
        "unit": ["единица измерения", "unit", "measure", "ед. изм.", "ед.изм", "ед."],
        "loss": ["потери", "loss", "потери %", "loss %", "процент потерь"],
        "portion_norm": ["норма закладки", "portion norm", "норма", "закладка", "на порцию"]
    }
    
    # Density mapping for ml→g conversion (kg/L)
    DENSITY_MAP = {
        # Oils and fats
        "масло": 0.9, "oil": 0.9, "жир": 0.9, "fat": 0.9,
        "подсолнечное масло": 0.92, "оливковое масло": 0.91,
        
        # Dairy products  
        "сливки": 1.03, "cream": 1.03, "сметана": 1.02, "sour cream": 1.02,
        "молоко": 1.03, "milk": 1.03, "кефир": 1.03, "yogurt": 1.04,
        
        # Syrups and honey
        "сироп": 1.3, "syrup": 1.3, "мёд": 1.4, "honey": 1.4,
        "патока": 1.35, "molasses": 1.35,
        
        # Alcoholic beverages
        "вино": 0.99, "wine": 0.99, "водка": 0.94, "vodka": 0.94,
        
        # Default water density
        "вода": 1.0, "water": 1.0
    }
    
    # Standard piece weights (g) for common ingredients
    PIECE_WEIGHTS = {
        "яйцо": 50, "egg": 50, "яйца": 50, "eggs": 50,
        "лук": 150, "onion": 150, "луковица": 150,
        "морковь": 100, "carrot": 100, "морковка": 100,
        "картофель": 120, "potato": 120, "картошка": 120,
        "помидор": 120, "tomato": 120, "томат": 120,
        "огурец": 100, "cucumber": 100,
        "лимон": 120, "lemon": 120, "лайм": 60, "lime": 60,
        "чеснок": 5, "garlic": 5, "зубчик чеснока": 5,
        "перец болгарский": 150, "bell pepper": 150
    }

    def __init__(self):
        self.issues = []
        self.parsed_rows = 0
        
    def parse_xlsx_to_techcard(self, file_bytes: bytes, filename: str = "import.xlsx") -> Dict[str, Any]:
        """
        Main parsing method: converts iiko XLSX bytes to TechCardV2 structure
        
        Args:
            file_bytes: XLSX file content as bytes
            filename: Original filename for reference
            
        Returns:
            Dict with parsed TechCardV2 structure and issues
        """
        self.issues = []
        self.parsed_rows = 0
        
        try:
            # Load workbook from bytes
            workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
            
            # Use the first worksheet
            if not workbook.worksheets:
                raise IikoXlsxParseError("XLSX file contains no worksheets")
                
            worksheet = workbook.worksheets[0]
            
            # Parse the worksheet
            techcard_data = self._parse_worksheet(worksheet, filename)
            
            return {
                "techcard": techcard_data,
                "issues": self.issues,
                "meta": {
                    "source": "iiko-xlsx",
                    "parsed_rows": self.parsed_rows,
                    "filename": filename
                }
            }
            
        except Exception as e:
            logger.error(f"Error parsing XLSX file {filename}: {str(e)}")
            raise IikoXlsxParseError(f"Failed to parse XLSX: {str(e)}")
    
    def _parse_worksheet(self, worksheet: Worksheet, filename: str) -> Dict[str, Any]:
        """Parse worksheet content and extract tech card data"""
        
        # Step 1: Find and map column headers
        header_row, column_map = self._find_headers(worksheet)
        if not column_map:
            raise IikoXlsxParseError("Could not identify required columns in XLSX")
        
        # Step 2: Extract meta information (usually in top rows)
        meta_info = self._extract_meta_info(worksheet, column_map, header_row)
        
        # Step 3: Parse ingredients data
        ingredients = self._parse_ingredients(worksheet, column_map, header_row)
        
        # Step 4: Process technology/process steps
        process_steps = self._parse_technology(meta_info.get("technology", ""))
        
        # Step 5: Calculate yield and portions
        yield_info = self._calculate_yield_portions(ingredients, meta_info)
        
        # Step 6: Build TechCardV2 structure
        techcard = self._build_techcard_structure(meta_info, ingredients, process_steps, yield_info)
        
        return techcard
    
    def _find_headers(self, worksheet: Worksheet) -> Tuple[int, Dict[str, int]]:
        """
        Find header row and create column mapping
        Returns: (header_row_index, {field_name: column_index})
        """
        column_map = {}
        
        # Search first 10 rows for headers
        for row_idx in range(1, min(11, worksheet.max_row + 1)):
            row_values = [str(cell.value or "").strip().lower() for cell in worksheet[row_idx]]
            
            # Check if this row contains header-like content
            potential_map = {}
            for field, aliases in self.COLUMN_MAPPINGS.items():
                for col_idx, value in enumerate(row_values):
                    if any(alias.lower() in value for alias in aliases):
                        potential_map[field] = col_idx
                        break
            
            # If we found enough columns, this is likely the header row
            if len(potential_map) >= 3:  # Need at least 3 key columns
                column_map = potential_map
                logger.info(f"Found headers in row {row_idx}: {list(potential_map.keys())}")
                return row_idx, column_map
        
        # If no clear header row found, assume row 1 and try basic mapping
        logger.warning("No clear header row found, using row 1 with basic mapping")
        return 1, self._create_basic_column_map(worksheet)
    
    def _create_basic_column_map(self, worksheet: Worksheet) -> Dict[str, int]:
        """Create basic column mapping when headers are unclear"""
        # Fallback mapping based on typical iiko XLSX structure
        return {
            "ingredient_name": 0,  # Usually first column
            "brutto": 1,           # Second column
            "netto": 2,            # Third column  
            "unit": 3,             # Fourth column
            "ingredient_code": 4   # Fifth column if exists
        }
    
    def _extract_meta_info(self, worksheet: Worksheet, column_map: Dict[str, int], header_row: int) -> Dict[str, Any]:
        """Extract meta information from worksheet (dish name, code, etc.)"""
        meta = {}
        
        # Look for meta info in rows before header 
        for row_idx in range(1, header_row):
            for cell in worksheet[row_idx]:
                value = str(cell.value or "").strip()
                if not value:
                    continue
                    
                # Try to identify meta information by patterns
                if any(keyword in value.lower() for keyword in ["название", "name", "блюд"]):
                    # Next cell might contain the dish name
                    next_cell = worksheet.cell(row_idx, cell.column + 1)
                    if next_cell.value:
                        meta["dish_name"] = str(next_cell.value).strip()
                        
                elif any(keyword in value.lower() for keyword in ["артикул", "код", "code"]):
                    # Next cell might contain the dish code
                    next_cell = worksheet.cell(row_idx, cell.column + 1)
                    if next_cell.value:
                        meta["dish_code"] = str(next_cell.value).strip()
                        
                elif any(keyword in value.lower() for keyword in ["технология", "technology", "приготовления"]):
                    # Next cell might contain the technology
                    next_cell = worksheet.cell(row_idx, cell.column + 1)
                    if next_cell.value:
                        meta["technology"] = str(next_cell.value).strip()
        
        # Also look for meta info in rows AFTER header but BEFORE ingredients start
        # Find first row with actual ingredient data
        first_ingredient_row = None
        for row_idx in range(header_row + 1, min(header_row + 10, worksheet.max_row + 1)):  # Search max 10 rows
            row = worksheet[row_idx]
            ingredient_name_col = column_map.get("ingredient_name", 0)
            if ingredient_name_col < len(row) and row[ingredient_name_col].value:
                ingredient_name = str(row[ingredient_name_col].value).strip()
                if ingredient_name and not self._is_meta_row(ingredient_name, row, column_map):
                    first_ingredient_row = row_idx
                    break
        
        # Search for meta in rows between header and first ingredient
        if first_ingredient_row:
            for row_idx in range(header_row + 1, first_ingredient_row):
                for cell in worksheet[row_idx]:
                    value = str(cell.value or "").strip()
                    if not value:
                        continue
                        
                    # Try to identify meta information by patterns
                    if any(keyword in value.lower() for keyword in ["название", "name", "блюд"]):
                        # Next cell might contain the dish name
                        next_cell = worksheet.cell(row_idx, cell.column + 1)
                        if next_cell.value:
                            meta["dish_name"] = str(next_cell.value).strip()
                            
                    elif any(keyword in value.lower() for keyword in ["артикул", "код", "code"]):
                        # Next cell might contain the dish code
                        next_cell = worksheet.cell(row_idx, cell.column + 1)
                        if next_cell.value:
                            meta["dish_code"] = str(next_cell.value).strip()
                            
                    elif any(keyword in value.lower() for keyword in ["технология", "technology", "приготовления"]):
                        # Next cell might contain the technology
                        next_cell = worksheet.cell(row_idx, cell.column + 1)
                        if next_cell.value:
                            meta["technology"] = str(next_cell.value).strip()
        
        # Also check if meta info is in the same row as ingredients (common in some templates)
        if "dish_name" in column_map:
            first_data_row = header_row + 1
            if first_data_row <= worksheet.max_row:
                dish_name_cell = worksheet.cell(first_data_row, column_map["dish_name"] + 1)
                if dish_name_cell.value:
                    meta["dish_name"] = str(dish_name_cell.value).strip()
                    
        if "dish_code" in column_map:
            first_data_row = header_row + 1
            if first_data_row <= worksheet.max_row:
                dish_code_cell = worksheet.cell(first_data_row, column_map["dish_code"] + 1)
                if dish_code_cell.value:
                    meta["dish_code"] = str(dish_code_cell.value).strip()
        
        # Extract technology text from the technology column
        if "technology" in column_map:
            first_data_row = header_row + 1
            if first_data_row <= worksheet.max_row:
                tech_cell = worksheet.cell(first_data_row, column_map["technology"] + 1)
                if tech_cell.value:
                    meta["technology"] = str(tech_cell.value).strip()
        
        # Set defaults if not found
        if "dish_name" not in meta:
            meta["dish_name"] = "Imported Dish"
            self.issues.append({
                "code": "dishNameMissing",
                "level": "warning", 
                "msg": "Название блюда не найдено, используется значение по умолчанию"
            })
        
        return meta
    
    def _parse_ingredients(self, worksheet: Worksheet, column_map: Dict[str, int], header_row: int) -> List[Dict[str, Any]]:
        """Parse ingredients data from worksheet"""
        ingredients = []
        
        # Start from row after header
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            row = worksheet[row_idx]
            
            # Check if row has data (at least ingredient name)
            ingredient_name_col = column_map.get("ingredient_name", 0)
            if ingredient_name_col >= len(row) or not row[ingredient_name_col].value:
                continue
                
            ingredient_name = str(row[ingredient_name_col].value).strip()
            if not ingredient_name:
                continue
            
            # Skip meta-information rows (dish info mixed with ingredients)
            if self._is_meta_row(ingredient_name, row, column_map):
                continue
            
            self.parsed_rows += 1
            
            # Parse ingredient data
            ingredient = self._parse_single_ingredient(row, column_map, row_idx)
            if ingredient:
                ingredients.append(ingredient)
        
        if not ingredients:
            self.issues.append({
                "code": "noIngredients",
                "level": "error",
                "msg": "Не найдено ингредиентов в файле"
            })
        
        return ingredients
    
    def _is_meta_row(self, ingredient_name: str, row, column_map: Dict[str, int]) -> bool:
        """Check if this row contains meta information instead of ingredient data"""
        # Check if ingredient name contains meta keywords
        meta_keywords = [
            "название блюда", "dish name", "наименование блюда",
            "артикул блюда", "dish code", "код блюда", 
            "выход готового", "total yield", "готовый продукт",
            "технология", "technology", "приготовление",
            "порций", "portions", "норма закладки"
        ]
        
        ingredient_lower = ingredient_name.lower()
        for keyword in meta_keywords:
            if keyword in ingredient_lower:
                return True
        
        # Check if brutto/netto values look like meta data (very large or text)
        brutto_col = column_map.get("brutto")
        if brutto_col is not None and brutto_col < len(row):
            brutto_val = row[brutto_col].value
            if brutto_val:
                brutto_str = str(brutto_val).strip()
                # If it contains non-numeric text like "800 г", it's likely meta
                if any(unit in brutto_str.lower() for unit in ["г", "g", "кг", "kg", "л", "l"]):
                    if not brutto_str.replace(".", "").replace(",", "").replace(" ", "").replace("г", "").replace("g", "").replace("кг", "").replace("kg", "").replace("л", "").replace("l", "").isdigit():
                        return True
        
        return False
    
    def _parse_single_ingredient(self, row, column_map: Dict[str, int], row_idx: int) -> Optional[Dict[str, Any]]:
        """Parse single ingredient row"""
        try:
            ingredient = {}
            
            # Extract basic fields
            ingredient["name"] = self._get_cell_value(row, column_map.get("ingredient_name", 0), "")
            if not ingredient["name"]:
                return None
            
            # SKU/Article
            ingredient["skuId"] = self._get_cell_value(row, column_map.get("ingredient_code"), None)
            
            # Quantities
            brutto_raw = self._get_cell_value(row, column_map.get("brutto"), 0)
            netto_raw = self._get_cell_value(row, column_map.get("netto"), 0)
            unit_raw = self._get_cell_value(row, column_map.get("unit"), "g")
            
            # Convert to float and validate
            brutto_g = self._parse_number_with_validation(brutto_raw, f"брутто в строке {row_idx}")
            netto_g = self._parse_number_with_validation(netto_raw, f"нетто в строке {row_idx}")
            
            # Normalize unit and quantities
            unit, brutto_g, netto_g = self._normalize_unit_and_quantities(
                unit_raw, brutto_g, netto_g, ingredient["name"]
            )
            
            ingredient["unit"] = unit
            ingredient["brutto_g"] = brutto_g
            ingredient["netto_g"] = netto_g
            
            # Calculate loss percentage
            if brutto_g > 0 and netto_g <= brutto_g:
                ingredient["loss_pct"] = round(((brutto_g - netto_g) / brutto_g) * 100, 1)
            else:
                ingredient["loss_pct"] = 0.0
                if brutto_g > 0 and netto_g > brutto_g:
                    self.issues.append({
                        "code": "nettoGreaterThanBrutto", 
                        "level": "warning",
                        "msg": f"Нетто больше брутто для '{ingredient['name']}'"
                    })
            
            # Validate loss bounds
            if ingredient["loss_pct"] > 50:
                self.issues.append({
                    "code": "lossOutOfBounds",
                    "level": "warning", 
                    "msg": f"Высокие потери {ingredient['loss_pct']}% для '{ingredient['name']}'"
                })
            
            # Additional fields
            ingredient["allergens"] = []  # Could be extracted from additional columns
            ingredient["canonical_id"] = None
            ingredient["subRecipe"] = None
            
            return ingredient
            
        except Exception as e:
            logger.warning(f"Error parsing ingredient in row {row_idx}: {str(e)}")
            return None
    
    def _get_cell_value(self, row, col_idx: Optional[int], default: Any = None) -> Any:
        """Safely get cell value with default"""
        if col_idx is None or col_idx >= len(row):
            return default
        
        cell_value = row[col_idx].value
        if cell_value is None:
            return default
            
        if isinstance(cell_value, str):
            return cell_value.strip()
        return cell_value
    
    def _parse_number_with_validation(self, value: Any, context: str) -> float:
        """Parse and validate numeric value"""
        if value is None or value == "":
            return 0.0
        
        try:
            # Handle string numbers with different decimal separators
            if isinstance(value, str):
                value = value.replace(",", ".").strip()
            
            num = float(value)
            
            if num < 0:
                self.issues.append({
                    "code": "negativeNumber",
                    "level": "error",
                    "msg": f"Отрицательное значение {num} для {context}"
                })
                return 0.0
                
            return num
            
        except (ValueError, TypeError):
            self.issues.append({
                "code": "invalidNumber",
                "level": "warning", 
                "msg": f"Некорректное число '{value}' для {context}"
            })
            return 0.0
    
    def _normalize_unit_and_quantities(self, unit_str: str, brutto: float, netto: float, 
                                     ingredient_name: str) -> Tuple[str, float, float]:
        """
        Normalize unit and convert quantities accordingly
        Returns: (normalized_unit, normalized_brutto, normalized_netto)
        """
        if not unit_str:
            unit_str = "g"
        
        unit_str = unit_str.lower().strip()
        
        # Kilogram to gram conversion
        if unit_str in ["кг", "kg", "kilogram", "килограмм"]:
            return "g", brutto * 1000, netto * 1000
        
        # Liter to milliliter conversion  
        elif unit_str in ["л", "l", "liter", "литр"]:
            return "ml", brutto * 1000, netto * 1000
        
        # Milliliter to gram conversion (using density)
        elif unit_str in ["мл", "ml", "milliliter", "миллилитр"]:
            density = self._get_density_for_ingredient(ingredient_name)
            return "g", brutto * density, netto * density
        
        # Gram (no conversion needed)
        elif unit_str in ["г", "g", "gram", "грамм"]:
            return "g", brutto, netto
        
        # Pieces (keep as pcs, try to estimate weight)
        elif unit_str in ["шт", "pcs", "piece", "штука", "pieces"]:
            piece_weight = self._get_piece_weight(ingredient_name)
            if piece_weight:
                return "g", brutto * piece_weight, netto * piece_weight
            else:
                self.issues.append({
                    "code": "massPerPieceMissing",
                    "level": "warning",
                    "msg": f"Ед. изм. 'шт.' без mass_per_piece для '{ingredient_name}'"
                })
                return "pcs", brutto, netto
        
        # Unknown unit
        else:
            self.issues.append({
                "code": "unitUnknown", 
                "level": "warning",
                "msg": f"Неизвестная единица измерения '{unit_str}' для '{ingredient_name}'"
            })
            return "g", brutto, netto  # Assume grams as fallback
    
    def _get_density_for_ingredient(self, ingredient_name: str) -> float:
        """Get density for ml→g conversion"""
        ingredient_lower = ingredient_name.lower()
        
        for material, density in self.DENSITY_MAP.items():
            if material in ingredient_lower:
                if density != 1.0:  # Non-water density
                    self.issues.append({
                        "code": f"densityUsed:{material}",
                        "level": "info", 
                        "msg": f"Применена плотность {density} для '{ingredient_name}'"
                    })
                return density
        
        # Default water density
        self.issues.append({
            "code": "densityAssumedWater",
            "level": "warning",
            "msg": f"ml→g по плотности 1.0 для '{ingredient_name}'"
        })
        return 1.0
    
    def _get_piece_weight(self, ingredient_name: str) -> Optional[float]:
        """Get standard weight for piece-based ingredients"""
        ingredient_lower = ingredient_name.lower()
        
        for piece_type, weight in self.PIECE_WEIGHTS.items():
            if piece_type in ingredient_lower:
                return float(weight)
        
        return None
    
    def _parse_technology(self, technology_text: str) -> List[Dict[str, Any]]:
        """Parse technology text into process steps (minimum 3 steps required)"""
        if not technology_text or not technology_text.strip():
            # IK-04/01: Generate minimum 3 steps as required by TechCardV2 schema
            return [
                {
                    "n": 1,
                    "action": "Подготовка ингредиентов и оборудования",
                    "time_min": None,
                    "temp_c": None,
                    "equipment": None,
                    "details": None
                },
                {
                    "n": 2,
                    "action": "Обработка основных ингредиентов",
                    "time_min": None,
                    "temp_c": None,
                    "equipment": None,
                    "details": None
                },
                {
                    "n": 3,
                    "action": "Окончательное приготовление и подача",
                    "time_min": None,
                    "temp_c": None,
                    "equipment": None,
                    "details": None
                }
            ]
        
        # Split by line breaks and periods
        steps_raw = re.split(r'[.\n\r]+', technology_text.strip())
        steps_raw = [s.strip() for s in steps_raw if s.strip()]
        
        if not steps_raw:
            steps_raw = [technology_text.strip()]
        
        # Ensure we have between 3 and 8 steps (TechCardV2 requirement)
        if len(steps_raw) < 3:
            # Expand to 3 steps by splitting the existing content
            if len(steps_raw) == 1:
                # Single step - split into preparation, main process, finishing
                main_step = steps_raw[0]
                steps_raw = [
                    "Подготовка ингредиентов",
                    main_step,
                    "Оформление и подача"
                ]
            elif len(steps_raw) == 2:
                # Two steps - add preparation step
                steps_raw.insert(0, "Подготовка ингредиентов")
        elif len(steps_raw) > 8:
            # Combine some steps to fit within 8 step limit
            combined_steps = []
            chunk_size = max(1, len(steps_raw) // 6)  # Target ~6 steps, min chunk_size=1
            for i in range(0, len(steps_raw), chunk_size):
                chunk = steps_raw[i:i + chunk_size]
                combined_steps.append(". ".join(chunk))
            steps_raw = combined_steps[:8]
        
        processes = []
        default_actions = [
            "Подготовка ингредиентов и оборудования",
            "Обработка основных ингредиентов", 
            "Окончательное приготовление и подача"
        ]
        
        for idx, step_text in enumerate(steps_raw, 1):
            # Handle empty or very short steps
            action = step_text.strip()
            if not action or len(action) <= 2:
                # Use default action for empty/very short steps
                if idx <= len(default_actions):
                    action = default_actions[idx - 1]
                else:
                    action = f"Дополнительный этап приготовления №{idx}"
            
            # Extract time and temperature using regex
            time_min = self._extract_time_from_text(step_text)
            temp_c = self._extract_temperature_from_text(step_text)
            
            # Ensure each step has either time_min or temp_c (TechCardV2 requirement)
            # Check if action contains thermal words
            thermal_words = ['жарить', 'варить', 'готовить', 'тушить', 'запекать', 'кипятить']
            is_thermal = any(word in action.lower() for word in thermal_words)
            
            if is_thermal and time_min is None and temp_c is None:
                # For thermal steps without time/temp, provide both defaults to avoid validation issues
                if idx == 1 or "подготов" in action.lower() or "приправ" in action.lower():
                    temp_c = 20.0  # Room temperature for preparation
                    time_min = 5.0  # 5 minutes for preparation
                else:
                    # For cooking steps, use default cooking temperature and time
                    temp_c = 60.0  # Default cooking temperature
                    time_min = 10.0  # Default cooking time
            elif is_thermal and time_min is None:
                # If we have temp_c but no time_min, set default time
                time_min = 1.0  # Minimal default time
            elif is_thermal and temp_c is None:
                # If we have time_min but no temp_c, set default temperature
                temp_c = 60.0  # Default cooking temperature
            
            processes.append({
                "n": idx,
                "action": action,
                "time_min": time_min,
                "temp_c": temp_c,
                "equipment": None,  # Could be extracted with more sophisticated parsing
                "details": None
            })
        
        return processes
    
    def _extract_time_from_text(self, text: str) -> Optional[float]:
        """Extract cooking time from text using regex"""
        # Enhanced patterns for time with support for ranges and complex formats
        time_patterns = [
            # Range patterns (take minimum value)
            r'(\d{1,3})\s*[-–]\s*(\d{1,3})\s*мин',        # "3-5 мин" → 3.0
            r'(\d{1,3})\s*[-–]\s*(\d{1,3})\s*минут',      # "3-5 минут" → 3.0  
            r'(\d{1,3})\s*[-–]\s*(\d{1,3})\s*min',        # "3-5 min" → 3.0
            # Single value patterns
            r'(\d{1,3})\s*мин',                           # "20 мин" → 20.0
            r'(\d{1,3})\s*минут',                         # "20 минут" → 20.0
            r'(\d{1,3})\s*min',                           # "20 min" → 20.0
            # Hour patterns (range and single)
            r'(\d{1,2})\s*[-–]\s*(\d{1,2})\s*час',        # "1-2 часа" → 60.0 (first * 60)
            r'(\d{1,2})\s*час',                           # "1 час" → 60.0
            r'(\d{1,2})\s*hour'                           # "1 hour" → 60.0
        ]
        
        for pattern in time_patterns:
            match = re.search(pattern, text.lower())
            if match:
                # For range patterns, take the first (minimum) value
                time_val = int(match.group(1))
                
                # Convert hours to minutes if needed
                if "час" in pattern or "hour" in pattern:
                    time_val *= 60
                    
                return float(time_val)
        
        return None
    
    def _extract_temperature_from_text(self, text: str) -> Optional[float]:
        """Extract cooking temperature from text using regex"""
        # Enhanced patterns for temperature with support for ranges and complex formats
        temp_patterns = [
            # Fahrenheit conversion (must be first to avoid °F being caught by °)
            r'(\d{2,3})\s*°\s*[FfФф]',                      # "350°F" → 176.7 (converted)
            # Range patterns (take minimum value)
            r'(\d{2,3})\s*[-–]\s*(\d{2,3})\s*°\s*[CcСс]',   # "170-180°C" → 170.0
            r'(\d{2,3})\s*[-–]\s*(\d{2,3})\s*°',            # "170-180°" → 170.0
            r'(\d{2,3})\s*[-–]\s*(\d{2,3})\s*градус',       # "170-180 градусов" → 170.0
            # Complex t= format
            r't\s*=\s*(\d{2,3})\s*°?\s*[CcСс]?',            # "t=85°C" → 85.0
            # Single value patterns  
            r'(\d{2,3})\s*°\s*[CcСс]',                      # "200°C" → 200.0
            r'(\d{2,3})\s*°',                               # "200°" → 200.0
            r'(\d{2,3})\s*градус',                          # "200 градусов" → 200.0
        ]
        
        for pattern in temp_patterns:
            match = re.search(pattern, text.lower())
            if match:
                # For range patterns, take the first (minimum) value
                temp_val = float(match.group(1))
                
                # Convert Fahrenheit to Celsius if needed
                if "°\s*[FfФф]" in pattern:
                    temp_val = round((temp_val - 32) * 5.0 / 9.0, 1)
                
                return temp_val
        
        return None
    
    def _calculate_yield_portions(self, ingredients: List[Dict[str, Any]], meta_info: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate yield and portions using heuristics"""
        
        # Calculate total netto weight
        total_netto = sum(ing.get("netto_g", 0) for ing in ingredients if ing.get("unit") == "g")
        
        # Default values
        result = {
            "perBatch_g": total_netto,
            "portions": 1,
            "perPortion_g": total_netto
        }
        
        # Look for yield information in meta or first row
        yield_info = meta_info.get("yield_total")
        portion_norm = meta_info.get("portion_norm")
        
        if yield_info:
            # Parse yield (e.g., "800 г", "1.2 кг")
            yield_match = re.search(r'(\d+(?:\.\d+)?)\s*(г|g|кг|kg)', str(yield_info).lower())
            if yield_match:
                yield_val = float(yield_match.group(1))
                yield_unit = yield_match.group(2)
                
                if yield_unit in ["кг", "kg"]:
                    yield_val *= 1000  # Convert to grams
                
                result["perBatch_g"] = yield_val
        
        if portion_norm:
            # Parse portion norm
            norm_str = str(portion_norm).strip()
            
            # If it's a simple number ≤ 20, treat as portions count
            try:
                norm_val = float(norm_str)
                if norm_val <= 20 and norm_val.is_integer():
                    result["portions"] = int(norm_val)
                    result["perPortion_g"] = result["perBatch_g"] / result["portions"]
                else:
                    # Parse as weight (e.g., "200 г")
                    weight_match = re.search(r'(\d+(?:\.\d+)?)\s*(г|g)', norm_str.lower())
                    if weight_match:
                        portion_weight = float(weight_match.group(1))
                        result["perPortion_g"] = portion_weight
                        result["portions"] = max(1, int(result["perBatch_g"] / portion_weight))
            except ValueError:
                pass
        
        # Add warning if yield was estimated
        if not yield_info and not portion_norm:
            self.issues.append({
                "code": "yieldEstimated",
                "level": "warning",
                "msg": f"Выход оценён как сумма нетто: {result['perBatch_g']}г"
            })
        
        return result
    
    def _build_techcard_structure(self, meta_info: Dict[str, Any], ingredients: List[Dict[str, Any]], 
                                process_steps: List[Dict[str, Any]], yield_info: Dict[str, Any]) -> Dict[str, Any]:
        """Build final TechCardV2 structure"""
        
        techcard = {
            "meta": {
                "title": meta_info.get("dish_name", "Импортированное блюдо"),
                "version": "2.0",
                "createdAt": datetime.now(timezone.utc).isoformat(),
                "cuisine": "импортированная",
                "tags": ["импорт", "iiko"],
                # IK-04/01: Store dish code in tags instead of meta.code (not in schema)
                "timings": {}
            },
            "portions": yield_info["portions"],
            "yield": {
                "perPortion_g": round(yield_info["perPortion_g"], 1),
                "perBatch_g": round(yield_info["perBatch_g"], 1)
            },
            "ingredients": ingredients,
            "process": process_steps,
            "storage": {
                "conditions": "Согласно требованиям ГОСТ",
                "shelfLife_hours": 24.0,
                "servingTemp_c": 65.0
            },
            "nutrition": {
                "per100g": {"kcal": 0.0, "proteins_g": 0.0, "fats_g": 0.0, "carbs_g": 0.0, "fiber_g": 0.0, "sugar_g": 0.0, "sodium_mg": 0.0},
                "perPortion": {"kcal": 0.0, "proteins_g": 0.0, "fats_g": 0.0, "carbs_g": 0.0, "fiber_g": 0.0, "sugar_g": 0.0, "sodium_mg": 0.0}
            },
            "nutritionMeta": {
                "source": "bootstrap",  # Valid enum value for nutrition
                "coveragePct": 0.0
            },
            "cost": {
                "rawCost": None,
                "costPerPortion": None,
                "markup_pct": None,
                "vat_pct": None
            },
            "costMeta": {
                "source": "none",  # Valid enum value for cost (bootstrap not allowed)
                "coveragePct": 0.0,
                "asOf": datetime.now().strftime("%Y-%m-%d")
            },
            "printNotes": []
        }
        
        # IK-04/01: Store dish code in printNotes if provided (workaround for schema limitation)
        if meta_info.get("dish_code"):
            techcard["printNotes"].append(f"Артикул блюда: {meta_info['dish_code']}")
        
        return techcard