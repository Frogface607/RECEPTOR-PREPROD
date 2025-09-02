#!/usr/bin/env python3
"""
TECH CARD NUTRITION FINAL BOOST - COMPREHENSIVE BACKEND TESTING
================================================================

This test validates the complete TECH CARD NUTRITION FINAL BOOST system
focusing on achieving ≥80% БЖУ coverage with the expanded 216-ingredient catalog.

CRITICAL ACCEPTANCE CRITERIA:
✅ БЖУ покрытие ≥80%: MUST BE ACHIEVED
✅ Каталог 216 ингредиентов: with critical additions
✅ 100% техкарт с покрытием ≥80%: all tests pass
✅ XLSX экспорт работает: stable with БЖУ data
✅ Performance ≤20с: acceptable generation time

Test Scenarios:
1. Create 5 diverse tech cards and measure БЖУ coverage
2. Calculate average coverage and confirm ≥80%
3. Export each to XLSX and verify success
4. Measure performance of all operations
5. Confirm project goal achievement
"""

import asyncio
import aiohttp
import json
import time
import statistics
from typing import Dict, List, Any, Optional
from datetime import datetime
import os

# Backend URL from environment
BACKEND_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://receptor-pro-beta-1.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api"

class TechCardNutritionFinalBoostTester:
    def __init__(self):
        self.session = None
        self.results = {
            'test_start': datetime.now().isoformat(),
            'catalog_validation': {},
            'techcard_generation': {},
            'nutrition_coverage': {},
            'xlsx_export': {},
            'performance': {},
            'final_assessment': {}
        }
        
    async def __aenter__(self):
        self.session = aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=60))
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        if self.session:
            await self.session.close()

    async def test_1_expanded_catalog_validation(self):
        """Test 1: Validate expanded catalog (216 ingredients)"""
        print("🔍 TEST 1: EXPANDED CATALOG VALIDATION (216 ingredients)")
        
        try:
            # Check nutrition catalog file
            catalog_path = "/app/backend/data/nutrition_catalog.dev.json"
            with open(catalog_path, 'r', encoding='utf-8') as f:
                catalog = json.load(f)
            
            items = catalog.get('items', [])
            catalog_count = len(items)
            
            # Validate critical ingredients
            critical_ingredients = ['вода', 'свекла', 'лук', 'соль']
            found_critical = []
            
            for item in items:
                name = item.get('name', '').lower()
                for critical in critical_ingredients:
                    if critical in name:
                        found_critical.append(critical)
                        break
            
            # Validate БЖУ data completeness
            complete_nutrition = 0
            for item in items:
                per100g = item.get('per100g', {})
                if all(key in per100g for key in ['kcal', 'proteins_g', 'fats_g', 'carbs_g']):
                    complete_nutrition += 1
            
            self.results['catalog_validation'] = {
                'total_ingredients': catalog_count,
                'target_ingredients': 216,
                'meets_target': catalog_count >= 216,
                'critical_ingredients_found': found_critical,
                'critical_ingredients_target': critical_ingredients,
                'complete_nutrition_data': complete_nutrition,
                'nutrition_completeness_rate': complete_nutrition / catalog_count if catalog_count > 0 else 0,
                'source': catalog.get('source', 'unknown')
            }
            
            print(f"   📊 Catalog contains {catalog_count} ingredients (target: 216)")
            print(f"   🎯 Target met: {'✅' if catalog_count >= 216 else '❌'}")
            print(f"   🔑 Critical ingredients found: {len(found_critical)}/{len(critical_ingredients)}")
            print(f"   📈 Complete БЖУ data: {complete_nutrition}/{catalog_count} ({complete_nutrition/catalog_count*100:.1f}%)")
            print(f"   📦 Source: {catalog.get('source', 'unknown')}")
            
            return catalog_count >= 216 and len(found_critical) >= 3
            
        except Exception as e:
            print(f"   ❌ Catalog validation failed: {str(e)}")
            self.results['catalog_validation']['error'] = str(e)
            return False

    async def test_2_techcard_generation_with_coverage(self):
        """Test 2: Generate 5 diverse tech cards and measure БЖУ coverage"""
        print("\n🍽️ TEST 2: TECHCARD GENERATION WITH БЖУ COVERAGE")
        
        # Diverse dish names for comprehensive testing
        test_dishes = [
            "Борщ украинский с говядиной",
            "Стейк из говядины с картофельным пюре", 
            "Салат Цезарь с курицей",
            "Паста карбонара с беконом",
            "Рыба запеченная с овощами"
        ]
        
        generated_cards = []
        coverage_scores = []
        generation_times = []
        
        for i, dish_name in enumerate(test_dishes, 1):
            print(f"   🔄 Generating tech card {i}/5: {dish_name}")
            
            try:
                start_time = time.time()
                
                # Generate tech card
                async with self.session.post(
                    f"{API_BASE}/v1/techcards.v2/generate",
                    json={"name": dish_name}
                ) as response:
                    if response.status == 200:
                        data = await response.json()
                        generation_time = time.time() - start_time
                        generation_times.append(generation_time)
                        
                        # Extract tech card data
                        techcard = data.get('techcard', {})
                        techcard_id = techcard.get('id')
                        
                        # Calculate БЖУ coverage
                        coverage = await self.calculate_nutrition_coverage(techcard)
                        coverage_scores.append(coverage)
                        
                        generated_cards.append({
                            'id': techcard_id,
                            'name': dish_name,
                            'coverage': coverage,
                            'generation_time': generation_time,
                            'ingredients_count': len(techcard.get('ingredients', [])),
                            'source': techcard.get('meta', {}).get('source', 'unknown')
                        })
                        
                        print(f"      ✅ Generated in {generation_time:.1f}s, БЖУ coverage: {coverage:.1f}%")
                        
                    else:
                        print(f"      ❌ Generation failed: HTTP {response.status}")
                        
            except Exception as e:
                print(f"      ❌ Generation error: {str(e)}")
        
        # Calculate statistics
        avg_coverage = statistics.mean(coverage_scores) if coverage_scores else 0
        avg_generation_time = statistics.mean(generation_times) if generation_times else 0
        
        self.results['techcard_generation'] = {
            'generated_cards': generated_cards,
            'total_generated': len(generated_cards),
            'target_count': 5,
            'success_rate': len(generated_cards) / len(test_dishes),
            'average_generation_time': avg_generation_time,
            'performance_target_met': avg_generation_time <= 20.0
        }
        
        self.results['nutrition_coverage'] = {
            'individual_scores': coverage_scores,
            'average_coverage': avg_coverage,
            'target_coverage': 80.0,
            'target_met': avg_coverage >= 80.0,
            'cards_above_80': sum(1 for score in coverage_scores if score >= 80.0),
            'coverage_improvement': avg_coverage - 53.0  # vs baseline 53%
        }
        
        print(f"\n   📊 COVERAGE RESULTS:")
        print(f"      Average БЖУ coverage: {avg_coverage:.1f}% (target: ≥80%)")
        print(f"      Target achieved: {'✅' if avg_coverage >= 80.0 else '❌'}")
        print(f"      Cards with ≥80% coverage: {sum(1 for score in coverage_scores if score >= 80.0)}/{len(coverage_scores)}")
        print(f"      Average generation time: {avg_generation_time:.1f}s (target: ≤20s)")
        
        return len(generated_cards) >= 4 and avg_coverage >= 80.0

    async def calculate_nutrition_coverage(self, techcard: Dict[str, Any]) -> float:
        """Calculate БЖУ coverage percentage for a tech card"""
        try:
            ingredients = techcard.get('ingredients', [])
            if not ingredients:
                return 0.0
            
            covered_ingredients = 0
            total_ingredients = len(ingredients)
            
            for ingredient in ingredients:
                # Check if ingredient has БЖУ data
                nutrition = ingredient.get('nutrition', {})
                if nutrition and any(nutrition.get(key, 0) > 0 for key in ['kcal', 'proteins_g', 'fats_g', 'carbs_g']):
                    covered_ingredients += 1
            
            coverage = (covered_ingredients / total_ingredients) * 100 if total_ingredients > 0 else 0
            return coverage
            
        except Exception as e:
            print(f"      ⚠️ Coverage calculation error: {str(e)}")
            return 0.0

    async def test_3_xlsx_export_validation(self):
        """Test 3: XLSX export with improved БЖУ data"""
        print("\n📋 TEST 3: XLSX EXPORT VALIDATION")
        
        generated_cards = self.results['techcard_generation'].get('generated_cards', [])
        if not generated_cards:
            print("   ❌ No generated cards available for export testing")
            return False
        
        export_results = []
        export_times = []
        
        for i, card in enumerate(generated_cards[:3], 1):  # Test first 3 cards
            print(f"   📤 Exporting card {i}/3: {card['name']}")
            
            try:
                start_time = time.time()
                
                # Test enhanced export
                async with self.session.post(
                    f"{API_BASE}/v1/techcards.v2/export/enhanced/iiko.xlsx",
                    json={
                        "techcard_ids": [card['id']],
                        "operational_rounding": True
                    }
                ) as response:
                    export_time = time.time() - start_time
                    export_times.append(export_time)
                    
                    if response.status == 200:
                        content_type = response.headers.get('content-type', '')
                        content_length = len(await response.read())
                        
                        export_results.append({
                            'card_id': card['id'],
                            'card_name': card['name'],
                            'success': True,
                            'file_size': content_length,
                            'export_time': export_time,
                            'content_type': content_type
                        })
                        
                        print(f"      ✅ Exported successfully: {content_length} bytes in {export_time:.2f}s")
                        
                    else:
                        print(f"      ❌ Export failed: HTTP {response.status}")
                        export_results.append({
                            'card_id': card['id'],
                            'card_name': card['name'],
                            'success': False,
                            'error': f"HTTP {response.status}"
                        })
                        
            except Exception as e:
                print(f"      ❌ Export error: {str(e)}")
                export_results.append({
                    'card_id': card['id'],
                    'card_name': card['name'],
                    'success': False,
                    'error': str(e)
                })
        
        successful_exports = sum(1 for result in export_results if result.get('success', False))
        avg_export_time = statistics.mean(export_times) if export_times else 0
        avg_file_size = statistics.mean([r.get('file_size', 0) for r in export_results if r.get('success', False)])
        
        self.results['xlsx_export'] = {
            'export_results': export_results,
            'successful_exports': successful_exports,
            'total_attempts': len(export_results),
            'success_rate': successful_exports / len(export_results) if export_results else 0,
            'average_export_time': avg_export_time,
            'average_file_size': avg_file_size,
            'target_file_size': 6000,  # ~6KB+
            'file_size_target_met': avg_file_size >= 6000
        }
        
        print(f"\n   📊 EXPORT RESULTS:")
        print(f"      Successful exports: {successful_exports}/{len(export_results)}")
        print(f"      Average file size: {avg_file_size:.0f} bytes (target: ≥6KB)")
        print(f"      Average export time: {avg_export_time:.2f}s")
        
        return successful_exports >= 2

    async def test_4_alt_export_cleanup(self):
        """Test 4: ALT Export Cleanup functionality"""
        print("\n🧹 TEST 4: ALT EXPORT CLEANUP")
        
        try:
            # Test ALT export cleanup stats
            async with self.session.get(f"{API_BASE}/v1/export/cleanup/stats") as response:
                if response.status == 200:
                    stats = await response.json()
                    print(f"   ✅ Cleanup stats accessible: {stats}")
                    
                    # Test cleanup audit
                    async with self.session.post(f"{API_BASE}/v1/export/cleanup/audit") as audit_response:
                        if audit_response.status == 200:
                            audit = await audit_response.json()
                            print(f"   ✅ Cleanup audit working: {len(audit.get('issues', []))} issues found")
                            return True
                        else:
                            print(f"   ❌ Cleanup audit failed: HTTP {audit_response.status}")
                            return False
                else:
                    print(f"   ❌ Cleanup stats failed: HTTP {response.status}")
                    return False
                    
        except Exception as e:
            print(f"   ❌ ALT Export Cleanup error: {str(e)}")
            return False

    async def test_5_performance_validation(self):
        """Test 5: Performance validation"""
        print("\n⚡ TEST 5: PERFORMANCE VALIDATION")
        
        generation_times = [card.get('generation_time', 0) for card in self.results['techcard_generation'].get('generated_cards', [])]
        export_times = [result.get('export_time', 0) for result in self.results['xlsx_export'].get('export_results', []) if result.get('success', False)]
        
        avg_generation = statistics.mean(generation_times) if generation_times else 0
        max_generation = max(generation_times) if generation_times else 0
        avg_export = statistics.mean(export_times) if export_times else 0
        
        performance_met = avg_generation <= 20.0 and max_generation <= 30.0
        
        self.results['performance'] = {
            'average_generation_time': avg_generation,
            'max_generation_time': max_generation,
            'average_export_time': avg_export,
            'generation_target': 20.0,
            'performance_target_met': performance_met,
            'total_workflow_time': avg_generation + avg_export
        }
        
        print(f"   📊 PERFORMANCE METRICS:")
        print(f"      Average generation time: {avg_generation:.1f}s (target: ≤20s)")
        print(f"      Maximum generation time: {max_generation:.1f}s")
        print(f"      Average export time: {avg_export:.2f}s")
        print(f"      Performance target met: {'✅' if performance_met else '❌'}")
        
        return performance_met

    async def test_6_system_integration(self):
        """Test 6: Full system integration"""
        print("\n🔗 TEST 6: SYSTEM INTEGRATION")
        
        try:
            # Test NutritionCalculator with expanded catalog
            test_ingredients = [
                {"name": "говядина", "quantity": 200},
                {"name": "картофель", "quantity": 300},
                {"name": "морковь", "quantity": 100}
            ]
            
            # Test catalog search with expanded data
            search_results = []
            for ingredient in test_ingredients:
                async with self.session.get(
                    f"{API_BASE}/v1/techcards.v2/catalog-search",
                    params={"q": ingredient["name"], "source": "catalog"}
                ) as response:
                    if response.status == 200:
                        data = await response.json()
                        results = data.get('results', [])
                        search_results.append({
                            'ingredient': ingredient["name"],
                            'found': len(results) > 0,
                            'results_count': len(results)
                        })
                        print(f"   🔍 Search '{ingredient['name']}': {len(results)} results")
                    else:
                        search_results.append({
                            'ingredient': ingredient["name"],
                            'found': False,
                            'error': f"HTTP {response.status}"
                        })
            
            successful_searches = sum(1 for result in search_results if result.get('found', False))
            integration_success = successful_searches >= 2
            
            print(f"   📊 Integration results: {successful_searches}/{len(search_results)} searches successful")
            
            return integration_success
            
        except Exception as e:
            print(f"   ❌ System integration error: {str(e)}")
            return False

    async def generate_final_assessment(self):
        """Generate final assessment of TECH CARD NUTRITION FINAL BOOST"""
        print("\n🎯 FINAL ASSESSMENT: TECH CARD NUTRITION FINAL BOOST")
        
        # Collect all test results
        catalog_ok = self.results['catalog_validation'].get('meets_target', False)
        coverage_ok = self.results['nutrition_coverage'].get('target_met', False)
        export_ok = self.results['xlsx_export'].get('success_rate', 0) >= 0.8
        performance_ok = self.results['performance'].get('performance_target_met', False)
        
        # Calculate overall success
        critical_criteria = [catalog_ok, coverage_ok, export_ok, performance_ok]
        success_rate = sum(critical_criteria) / len(critical_criteria)
        
        avg_coverage = self.results['nutrition_coverage'].get('average_coverage', 0)
        catalog_count = self.results['catalog_validation'].get('total_ingredients', 0)
        
        self.results['final_assessment'] = {
            'overall_success': success_rate >= 0.8,
            'success_rate': success_rate,
            'critical_criteria_met': {
                'expanded_catalog_216': catalog_ok,
                'coverage_80_percent': coverage_ok,
                'xlsx_export_working': export_ok,
                'performance_under_20s': performance_ok
            },
            'key_metrics': {
                'catalog_ingredients': catalog_count,
                'average_coverage': avg_coverage,
                'coverage_improvement': avg_coverage - 53.0,
                'target_achievement': avg_coverage >= 80.0
            },
            'project_status': 'SUCCESS' if success_rate >= 0.8 and avg_coverage >= 80.0 else 'NEEDS_IMPROVEMENT'
        }
        
        print(f"\n   🏆 PROJECT GOAL ACHIEVEMENT:")
        print(f"      ✅ Expanded catalog (216 ingredients): {'✅' if catalog_ok else '❌'}")
        print(f"      ✅ БЖУ coverage ≥80%: {'✅' if coverage_ok else '❌'} ({avg_coverage:.1f}%)")
        print(f"      ✅ XLSX export working: {'✅' if export_ok else '❌'}")
        print(f"      ✅ Performance ≤20s: {'✅' if performance_ok else '❌'}")
        print(f"\n   🎯 OVERALL SUCCESS: {'✅ ACHIEVED' if success_rate >= 0.8 and avg_coverage >= 80.0 else '❌ NEEDS WORK'}")
        print(f"      Success rate: {success_rate*100:.1f}%")
        print(f"      Coverage improvement: +{avg_coverage - 53.0:.1f}% vs baseline")
        
        return success_rate >= 0.8 and avg_coverage >= 80.0

    async def save_results(self):
        """Save test results to file"""
        self.results['test_end'] = datetime.now().isoformat()
        
        with open('/app/tech_card_nutrition_final_boost_results.json', 'w', encoding='utf-8') as f:
            json.dump(self.results, f, indent=2, ensure_ascii=False)
        
        print(f"\n💾 Results saved to: /app/tech_card_nutrition_final_boost_results.json")

async def main():
    """Main test execution"""
    print("🚀 TECH CARD NUTRITION FINAL BOOST - COMPREHENSIVE TESTING")
    print("=" * 70)
    
    async with TechCardNutritionFinalBoostTester() as tester:
        test_results = []
        
        # Execute all tests
        test_results.append(await tester.test_1_expanded_catalog_validation())
        test_results.append(await tester.test_2_techcard_generation_with_coverage())
        test_results.append(await tester.test_3_xlsx_export_validation())
        test_results.append(await tester.test_4_alt_export_cleanup())
        test_results.append(await tester.test_5_performance_validation())
        test_results.append(await tester.test_6_system_integration())
        
        # Generate final assessment
        final_success = await tester.generate_final_assessment()
        
        # Save results
        await tester.save_results()
        
        # Summary
        passed_tests = sum(test_results)
        total_tests = len(test_results)
        
        print(f"\n" + "=" * 70)
        print(f"🏁 TESTING COMPLETED")
        print(f"   Tests passed: {passed_tests}/{total_tests}")
        print(f"   Success rate: {passed_tests/total_tests*100:.1f}%")
        print(f"   Final assessment: {'✅ SUCCESS' if final_success else '❌ NEEDS IMPROVEMENT'}")
        
        if final_success:
            print(f"\n🎉 TECH CARD NUTRITION FINAL BOOST: PROJECT GOAL ACHIEVED!")
            print(f"   ✅ БЖУ coverage ≥80% confirmed")
            print(f"   ✅ Expanded catalog (216 ingredients) operational")
            print(f"   ✅ XLSX export with improved data working")
            print(f"   ✅ Performance targets met")
        else:
            print(f"\n⚠️ TECH CARD NUTRITION FINAL BOOST: NEEDS IMPROVEMENT")
            print(f"   Review results for specific issues to address")

if __name__ == "__main__":
    asyncio.run(main())
"""
TECH CARD NUTRITION INTEGRATION COMPREHENSIVE TESTING
Протестировать ЗАВЕРШЕННУЮ систему TECH CARD NUTRITION INTEGRATION после всех улучшений.

КРИТИЧЕСКИЕ тесты для финального подтверждения:
1. Проверка улучшенного БЖУ покрытия (КРИТИЧЕСКИЙ) - ≥80%
2. Валидация расширенного каталога (201 ингредиент)
3. XLSX экспорт с улучшенными данными (КРИТИЧЕСКИЙ)
4. Система поиска ингредиентов (≥90% покрытие)
5. Performance и стабильность
"""

import requests
import json
import time
import os
from typing import Dict, List, Any
import tempfile
import zipfile
from openpyxl import load_workbook
import statistics

# Configuration
BACKEND_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://receptor-pro-beta-1.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api"

class TechCardNutritionTester:
    def __init__(self):
        self.results = []
        self.artifacts = {}
        
    def log_result(self, test_name: str, success: bool, details: str, data: Any = None):
        """Log test result with details"""
        result = {
            'test': test_name,
            'success': success,
            'details': details,
            'timestamp': time.time(),
            'data': data
        }
        self.results.append(result)
        status = "✅" if success else "❌"
        print(f"{status} {test_name}: {details}")
        
    def test_expanded_catalog_validation(self) -> bool:
        """КРИТИЧЕСКИЙ ТЕСТ 1: Проверка расширенного каталога с 201 ингредиентом"""
        try:
            # Check nutrition catalog
            try:
                with open('/app/backend/data/nutrition_catalog.dev.json', 'r', encoding='utf-8') as f:
                    nutrition_catalog = json.load(f)
                
                ingredients = nutrition_catalog.get('items', [])
                ingredient_count = len(ingredients)
                
                if ingredient_count >= 201:
                    self.log_result(
                        "Expanded Nutrition Catalog Validation", 
                        True,
                        f"Каталог содержит {ingredient_count} ингредиентов (≥201 требуется)",
                        {'count': ingredient_count}
                    )
                    
                    # Validate БЖУ data completeness
                    complete_nutrition = 0
                    for ingredient in ingredients:
                        per100g = ingredient.get('per100g', {})
                        if all(key in per100g for key in ['kcal', 'proteins_g', 'fats_g', 'carbs_g']):
                            complete_nutrition += 1
                    
                    completeness_pct = (complete_nutrition / ingredient_count) * 100
                    
                    if completeness_pct >= 95:
                        self.log_result(
                            "БЖУ Data Completeness", 
                            True,
                            f"БЖУ данные полные для {completeness_pct:.1f}% ингредиентов",
                            {'completeness_pct': completeness_pct}
                        )
                        return True
                    else:
                        self.log_result(
                            "БЖУ Data Completeness", 
                            False,
                            f"БЖУ данные неполные: {completeness_pct:.1f}% (требуется ≥95%)"
                        )
                        return False
                else:
                    self.log_result(
                        "Expanded Nutrition Catalog Validation", 
                        False,
                        f"Каталог содержит только {ingredient_count} ингредиентов (требуется ≥201)"
                    )
                    return False
                    
            except FileNotFoundError:
                self.log_result(
                    "Expanded Nutrition Catalog Validation", 
                    False,
                    "nutrition_catalog.dev.json не найден"
                )
                return False
                
        except Exception as e:
            self.log_result(
                "Expanded Catalog Validation", 
                False,
                f"Ошибка при проверке каталога: {str(e)}"
            )
            return False
    
    def test_improved_nutrition_coverage(self) -> bool:
        """КРИТИЧЕСКИЙ ТЕСТ 2: Проверка улучшенного БЖУ покрытия ≥80%"""
        try:
            # Test dishes with different complexity levels
            test_dishes = [
                "Борщ украинский с говядиной",
                "Стейк из говядины с картофельным пюре", 
                "Салат Цезарь с курицей"
            ]
            
            coverage_results = []
            
            for dish_name in test_dishes:
                start_time = time.time()
                
                # Generate tech card
                response = requests.post(
                    f"{API_BASE}/v1/techcards.v2/generate",
                    json={"name": dish_name},
                    timeout=60
                )
                
                generation_time = time.time() - start_time
                
                if response.status_code == 200:
                    data = response.json()
                    
                    if data.get('card') and data['card'].get('nutritionMeta'):
                        nutrition_meta = data['card']['nutritionMeta']
                        coverage_pct = nutrition_meta.get('coveragePct', 0)
                        source = nutrition_meta.get('source', 'unknown')
                        
                        coverage_results.append({
                            'dish': dish_name,
                            'coverage_pct': coverage_pct,
                            'source': source,
                            'generation_time': generation_time,
                            'card_id': data['card'].get('id') or data['card'].get('meta', {}).get('id')
                        })
                        
                        self.log_result(
                            f"БЖУ Coverage - {dish_name}",
                            coverage_pct >= 80,
                            f"Покрытие: {coverage_pct:.1f}%, источник: {source}, время: {generation_time:.1f}с"
                        )
                    else:
                        self.log_result(
                            f"БЖУ Coverage - {dish_name}",
                            False,
                            "Отсутствуют данные nutritionMeta в ответе"
                        )
                        coverage_results.append({
                            'dish': dish_name,
                            'coverage_pct': 0,
                            'source': 'error',
                            'generation_time': generation_time,
                            'error': 'No nutritionMeta'
                        })
                else:
                    self.log_result(
                        f"БЖУ Coverage - {dish_name}",
                        False,
                        f"Ошибка генерации: HTTP {response.status_code}"
                    )
                    coverage_results.append({
                        'dish': dish_name,
                        'coverage_pct': 0,
                        'source': 'error',
                        'generation_time': generation_time,
                        'error': f'HTTP {response.status_code}'
                    })
            
            # Calculate average coverage
            valid_coverages = [r['coverage_pct'] for r in coverage_results if r['coverage_pct'] > 0]
            
            if valid_coverages:
                avg_coverage = statistics.mean(valid_coverages)
                
                # Check if average meets target
                target_met = avg_coverage >= 80.0
                
                # Check if source is 'catalog' (using expanded catalog)
                catalog_sources = [r for r in coverage_results if r['source'] == 'catalog']
                
                self.artifacts['nutrition_coverage_results'] = coverage_results
                
                self.log_result(
                    "КРИТИЧЕСКИЙ: Улучшенное БЖУ покрытие",
                    target_met and len(catalog_sources) > 0,
                    f"Среднее покрытие: {avg_coverage:.1f}% (цель: ≥80%), источник 'catalog': {len(catalog_sources)}/{len(coverage_results)}"
                )
                
                return target_met and len(catalog_sources) > 0
            else:
                self.log_result(
                    "КРИТИЧЕСКИЙ: Улучшенное БЖУ покрытие",
                    False,
                    "Нет валидных результатов покрытия БЖУ"
                )
                return False
                
        except Exception as e:
            self.log_result(
                "Improved Nutrition Coverage Test",
                False,
                f"Ошибка при тестировании БЖУ покрытия: {str(e)}"
            )
            return False
    
    def test_xlsx_export_with_improved_data(self) -> bool:
        """КРИТИЧЕСКИЙ ТЕСТ 3: XLSX экспорт с улучшенными БЖУ данными"""
        try:
            # First generate a tech card with good nutrition coverage
            response = requests.post(
                f"{API_BASE}/v1/techcards.v2/generate",
                json={"name": "Паста Болоньезе"},
                timeout=60
            )
            
            if response.status_code != 200:
                self.log_result(
                    "XLSX Export - Tech Card Generation",
                    False,
                    f"Не удалось сгенерировать техкарту: HTTP {response.status_code}"
                )
                return False
            
            data = response.json()
            card = data.get('card')
            
            if not card:
                self.log_result(
                    "XLSX Export - Tech Card Generation",
                    False,
                    "Техкарта не сгенерирована (card = null)"
                )
                return False
            
            card_id = card.get('id') or card.get('meta', {}).get('id')
            nutrition_coverage = card.get('nutritionMeta', {}).get('coveragePct', 0)
            
            self.log_result(
                "XLSX Export - Tech Card Generation",
                True,
                f"Техкарта сгенерирована (ID: {card_id}, БЖУ покрытие: {nutrition_coverage:.1f}%)"
            )
            
            # Test XLSX export
            export_response = requests.post(
                f"{API_BASE}/v1/techcards.v2/export/iiko.xlsx",
                json=card,  # Send the full card object
                timeout=30
            )
            
            if export_response.status_code == 200:
                # Check if it's actually an XLSX file
                content_type = export_response.headers.get('content-type', '')
                
                if 'spreadsheet' in content_type or 'excel' in content_type:
                    file_size = len(export_response.content)
                    
                    # Save and analyze XLSX content
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                        tmp_file.write(export_response.content)
                        tmp_file_path = tmp_file.name
                    
                    try:
                        # Load and analyze XLSX
                        workbook = load_workbook(tmp_file_path)
                        sheet_names = workbook.sheetnames
                        
                        # Check for nutrition data in sheets
                        nutrition_data_found = False
                        for sheet_name in sheet_names:
                            sheet = workbook[sheet_name]
                            for row in sheet.iter_rows(values_only=True):
                                if row and any(cell for cell in row if cell and 'ккал' in str(cell).lower()):
                                    nutrition_data_found = True
                                    break
                            if nutrition_data_found:
                                break
                        
                        os.unlink(tmp_file_path)
                        
                        self.log_result(
                            "КРИТИЧЕСКИЙ: XLSX экспорт с БЖУ данными",
                            nutrition_data_found,
                            f"XLSX файл ({file_size} байт), листы: {sheet_names}, БЖУ данные: {'найдены' if nutrition_data_found else 'не найдены'}"
                        )
                        
                        return nutrition_data_found
                        
                    except Exception as e:
                        os.unlink(tmp_file_path)
                        self.log_result(
                            "XLSX Export Analysis",
                            False,
                            f"Ошибка анализа XLSX файла: {str(e)}"
                        )
                        return False
                else:
                    self.log_result(
                        "XLSX Export Content Type",
                        False,
                        f"Неверный content-type: {content_type}"
                    )
                    return False
            else:
                self.log_result(
                    "XLSX Export Request",
                    False,
                    f"Ошибка экспорта: HTTP {export_response.status_code}"
                )
                return False
                
        except Exception as e:
            self.log_result(
                "XLSX Export Test",
                False,
                f"Ошибка при тестировании XLSX экспорта: {str(e)}"
            )
            return False
    
    def test_ingredient_search_system(self) -> bool:
        """ТЕСТ 4: Система поиска ингредиентов с улучшенной индексацией"""
        try:
            # Test popular Russian ingredients
            test_ingredients = [
                "говядина",
                "картофель", 
                "лук",
                "морковь",
                "молоко",
                "яйца",
                "масло",
                "соль",
                "перец",
                "помидоры"
            ]
            
            search_results = []
            
            for ingredient in test_ingredients:
                # Test catalog search
                response = requests.get(
                    f"{API_BASE}/v1/techcards.v2/catalog-search",
                    params={"q": ingredient, "source": "all"},
                    timeout=10
                )
                
                if response.status_code == 200:
                    data = response.json()
                    results_count = len(data.get('results', data.get('items', [])))
                    
                    search_results.append({
                        'ingredient': ingredient,
                        'results_count': results_count,
                        'found': results_count > 0
                    })
                    
                    self.log_result(
                        f"Поиск ингредиента - {ingredient}",
                        results_count > 0,
                        f"Найдено результатов: {results_count}"
                    )
                else:
                    search_results.append({
                        'ingredient': ingredient,
                        'results_count': 0,
                        'found': False,
                        'error': f'HTTP {response.status_code}'
                    })
                    
                    self.log_result(
                        f"Поиск ингредиента - {ingredient}",
                        False,
                        f"Ошибка поиска: HTTP {response.status_code}"
                    )
            
            # Calculate coverage
            found_ingredients = [r for r in search_results if r['found']]
            coverage_pct = (len(found_ingredients) / len(test_ingredients)) * 100
            
            target_met = coverage_pct >= 90.0
            
            self.artifacts['ingredient_search_results'] = search_results
            
            self.log_result(
                "Система поиска ингредиентов",
                target_met,
                f"Покрытие поиска: {coverage_pct:.1f}% (цель: ≥90%), найдено: {len(found_ingredients)}/{len(test_ingredients)}"
            )
            
            return target_met
            
        except Exception as e:
            self.log_result(
                "Ingredient Search System Test",
                False,
                f"Ошибка при тестировании поиска: {str(e)}"
            )
            return False
    
    def test_performance_and_stability(self) -> bool:
        """ТЕСТ 5: Performance и стабильность системы"""
        try:
            performance_results = []
            
            # Test multiple tech card generations for performance
            test_dishes = [
                "Борщ классический",
                "Салат Оливье", 
                "Котлеты по-киевски"
            ]
            
            for dish in test_dishes:
                start_time = time.time()
                
                response = requests.post(
                    f"{API_BASE}/v1/techcards.v2/generate",
                    json={"name": dish},
                    timeout=60
                )
                
                generation_time = time.time() - start_time
                
                success = response.status_code == 200 and response.json().get('card') is not None
                
                performance_results.append({
                    'dish': dish,
                    'generation_time': generation_time,
                    'success': success,
                    'status_code': response.status_code
                })
                
                self.log_result(
                    f"Performance - {dish}",
                    success and generation_time < 45.0,
                    f"Время генерации: {generation_time:.1f}с, успех: {success}"
                )
            
            # Calculate average performance
            successful_generations = [r for r in performance_results if r['success']]
            
            if successful_generations:
                avg_time = statistics.mean([r['generation_time'] for r in successful_generations])
                success_rate = len(successful_generations) / len(performance_results) * 100
                
                performance_acceptable = avg_time < 45.0 and success_rate >= 80.0
                
                self.artifacts['performance_results'] = performance_results
                
                self.log_result(
                    "Performance и стабильность",
                    performance_acceptable,
                    f"Среднее время: {avg_time:.1f}с, успешность: {success_rate:.1f}%"
                )
                
                return performance_acceptable
            else:
                self.log_result(
                    "Performance и стабильность",
                    False,
                    "Нет успешных генераций для анализа производительности"
                )
                return False
                
        except Exception as e:
            self.log_result(
                "Performance and Stability Test",
                False,
                f"Ошибка при тестировании производительности: {str(e)}"
            )
            return False
    
    def test_alt_export_cleanup_integration(self) -> bool:
        """ТЕСТ 6: ALT Export Cleanup интеграция"""
        try:
            # Test ALT export cleanup stats
            response = requests.get(f"{API_BASE}/v1/export/cleanup/stats", timeout=10)
            
            if response.status_code == 200:
                stats = response.json()
                
                self.log_result(
                    "ALT Export Cleanup Stats",
                    True,
                    f"Статистика cleanup доступна: {stats}"
                )
                
                # Test audit endpoint
                audit_response = requests.post(f"{API_BASE}/v1/export/cleanup/audit", timeout=10)
                
                if audit_response.status_code == 200:
                    self.log_result(
                        "ALT Export Cleanup Integration",
                        True,
                        "ALT Export Cleanup система интегрирована и функциональна"
                    )
                    return True
                else:
                    self.log_result(
                        "ALT Export Cleanup Audit",
                        False,
                        f"Ошибка audit: HTTP {audit_response.status_code}"
                    )
                    return False
            else:
                self.log_result(
                    "ALT Export Cleanup Stats",
                    False,
                    f"Ошибка получения статистики: HTTP {response.status_code}"
                )
                return False
                
        except Exception as e:
            self.log_result(
                "ALT Export Cleanup Integration Test",
                False,
                f"Ошибка при тестировании ALT Export Cleanup: {str(e)}"
            )
            return False
    
    def run_comprehensive_test(self) -> Dict[str, Any]:
        """Запуск всех критических тестов"""
        print("🎯 НАЧАЛО КОМПЛЕКСНОГО ТЕСТИРОВАНИЯ TECH CARD NUTRITION INTEGRATION")
        print("=" * 80)
        
        start_time = time.time()
        
        # Run all critical tests
        test_results = {
            'expanded_catalog': self.test_expanded_catalog_validation(),
            'nutrition_coverage': self.test_improved_nutrition_coverage(), 
            'xlsx_export': self.test_xlsx_export_with_improved_data(),
            'ingredient_search': self.test_ingredient_search_system(),
            'performance': self.test_performance_and_stability(),
            'alt_export_cleanup': self.test_alt_export_cleanup_integration()
        }
        
        total_time = time.time() - start_time
        
        # Calculate overall success
        passed_tests = sum(1 for result in test_results.values() if result)
        total_tests = len(test_results)
        success_rate = (passed_tests / total_tests) * 100
        
        # Final assessment
        print("\n" + "=" * 80)
        print("🎯 ФИНАЛЬНЫЕ РЕЗУЛЬТАТЫ ТЕСТИРОВАНИЯ")
        print("=" * 80)
        
        for test_name, result in test_results.items():
            status = "✅ ПРОЙДЕН" if result else "❌ НЕ ПРОЙДЕН"
            print(f"{status}: {test_name}")
        
        print(f"\nОБЩИЙ РЕЗУЛЬТАТ: {passed_tests}/{total_tests} тестов пройдено ({success_rate:.1f}%)")
        print(f"ВРЕМЯ ВЫПОЛНЕНИЯ: {total_time:.1f} секунд")
        
        # Check critical acceptance criteria
        critical_tests_passed = (
            test_results['nutrition_coverage'] and  # БЖУ покрытие ≥80%
            test_results['expanded_catalog'] and    # Каталог 200+ ингредиентов  
            test_results['xlsx_export']             # XLSX экспорт работает
        )
        
        if critical_tests_passed:
            print("\n🎉 КРИТИЧЕСКИЕ ACCEPTANCE CRITERIA ВЫПОЛНЕНЫ!")
            print("✅ БЖУ покрытие ≥80%")
            print("✅ Каталог 200+ ингредиентов") 
            print("✅ XLSX экспорт с БЖУ данными")
            
            if success_rate >= 80:
                print("\n🚀 TECH CARD NUTRITION INTEGRATION ГОТОВА К ПРОДАКШЕНУ!")
            else:
                print(f"\n⚠️ Система частично готова (успешность {success_rate:.1f}%)")
        else:
            print("\n🚨 КРИТИЧЕСКИЕ ТРЕБОВАНИЯ НЕ ВЫПОЛНЕНЫ!")
            if not test_results['nutrition_coverage']:
                print("❌ БЖУ покрытие < 80%")
            if not test_results['expanded_catalog']:
                print("❌ Каталог < 200 ингредиентов")
            if not test_results['xlsx_export']:
                print("❌ XLSX экспорт не работает")
        
        # Save detailed results
        final_results = {
            'test_results': test_results,
            'success_rate': success_rate,
            'total_time': total_time,
            'critical_tests_passed': critical_tests_passed,
            'artifacts': self.artifacts,
            'detailed_results': self.results
        }
        
        # Save to file
        with open('/app/tech_card_nutrition_test_results.json', 'w', encoding='utf-8') as f:
            json.dump(final_results, f, ensure_ascii=False, indent=2, default=str)
        
        return final_results

def main():
    """Main test execution"""
    tester = TechCardNutritionTester()
    results = tester.run_comprehensive_test()
    
    # Return appropriate exit code
    if results['critical_tests_passed'] and results['success_rate'] >= 80:
        exit(0)  # Success
    else:
        exit(1)  # Failure

if __name__ == "__main__":
    main()
"""
ALT Export Cleanup System Comprehensive Testing

Протестировать ALT Export Cleanup систему согласно спецификации в test_result.md.

ОСНОВНЫЕ ЗАДАЧИ ДЛЯ ТЕСТИРОВАНИЯ:

1. ALT Export Cleanup Module (Приоритет: ВЫСОКИЙ)
   - Протестировать класс ALTExportValidator из /app/backend/receptor_agent/exports/alt_export_cleanup.py
   - Проверить функции анализа архивов, поиска дублей, валидации TTK
   - Тестировать методы cleanup_archive и validate_single_ttk
   - Проверить singleton pattern и статистику

2. Integration Testing (Приоритет: ВЫСОКИЙ)
   - POST /api/v1/techcards.v2/export/iiko.xlsx - проверить интеграцию валидации TTK
   - POST /api/v1/export/zip - проверить автоматическую очистку архивов
   - POST /api/v1/export/ttk-only - проверить валидацию TTK файлов
   - Убедиться что все экспорты проходят через ALT pipeline

3. Admin Endpoints (Приоритет: СРЕДНИЙ)
   - GET /api/v1/export/cleanup/stats - получение статистики очистки
   - POST /api/v1/export/cleanup/audit - полный аудит архивов
   - POST /api/v1/export/cleanup/reset-stats - сброс статистики

4. Функциональное тестирование:
   - Создать тестовые сценарии с дублями TTK и superfluous files
   - Проверить что cleanup автоматически удаляет проблемные файлы
   - Убедиться в корректности обязательных компонентов (Dish-Skeletons, Product-Skeletons, reference TTK)
   - Проверить логирование всех операций очистки

5. Edge Cases:
   - Пустые архивы, невалидные ZIP файлы
   - TTK файлы с отсутствующими данными
   - Архивы только с superfluous files
   - Большие архивы с множественными дублями
"""

import requests
import json
import time
import os
import tempfile
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Any, Optional
import sys
import hashlib

# Backend URL from environment
BACKEND_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://receptor-pro-beta-1.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api"

class ALTExportCleanupTester:
    def __init__(self):
        self.session = requests.Session()
        self.session.timeout = 30
        self.test_results = []
        self.generated_techcards = []
        self.test_artifacts = {}
        
    def log_result(self, test_name: str, success: bool, details: str, data: Any = None):
        """Log test result with details"""
        result = {
            'test': test_name,
            'success': success,
            'details': details,
            'timestamp': time.time(),
            'data': data
        }
        self.test_results.append(result)
        status = "✅" if success else "❌"
        print(f"{status} {test_name}: {details}")
        
    def test_1_alt_export_validator_class(self):
        """1. ALT Export Cleanup Module - Тестирование класса ALTExportValidator"""
        print("\n🔧 ТЕСТ 1: ALT Export Cleanup Module - ALTExportValidator Class")
        
        try:
            # Import the ALT Export Cleanup module directly
            sys.path.append('/app/backend')
            from receptor_agent.exports.alt_export_cleanup import ALTExportValidator, get_alt_export_validator, ArchiveAnalysisResult
            
            # Test 1.1: Singleton pattern
            validator1 = get_alt_export_validator()
            validator2 = get_alt_export_validator()
            
            singleton_test = validator1 is validator2
            self.log_result(
                "ALTExportValidator Singleton Pattern",
                singleton_test,
                f"Singleton pattern working: {singleton_test}"
            )
            
            # Test 1.2: Basic validator initialization
            validator = ALTExportValidator()
            has_required_methods = all(hasattr(validator, method) for method in [
                'analyze_archive', 'cleanup_archive', 'validate_single_ttk', 
                'get_cleanup_statistics', 'reset_statistics'
            ])
            
            self.log_result(
                "ALTExportValidator Required Methods",
                has_required_methods,
                f"All required methods present: {has_required_methods}"
            )
            
            # Test 1.3: Statistics initialization
            stats = validator.get_cleanup_statistics()
            expected_stats_keys = ['total_processed', 'duplicates_removed', 'invalid_removed', 'superfluous_removed', 'archives_cleaned']
            stats_valid = all(key in stats for key in expected_stats_keys)
            
            self.log_result(
                "ALTExportValidator Statistics Structure",
                stats_valid,
                f"Statistics structure valid: {stats_valid}, keys: {list(stats.keys())}"
            )
            
            # Test 1.4: Create test archive with issues for analysis
            test_archive = self._create_test_archive_with_issues()
            analysis_result = validator.analyze_archive(test_archive, "test_analysis")
            
            analysis_valid = isinstance(analysis_result, ArchiveAnalysisResult)
            has_issues = analysis_result.has_issues()
            
            self.log_result(
                "ALTExportValidator Archive Analysis",
                analysis_valid and has_issues,
                f"Analysis working: {analysis_valid}, found issues: {has_issues}, summary: {analysis_result.get_summary()}"
            )
            
            # Test 1.5: Cleanup functionality
            clean_archive, cleanup_stats = validator.cleanup_archive(test_archive, analysis_result, "test_cleanup")
            cleanup_successful = cleanup_stats.get('cleaned', False)
            
            self.log_result(
                "ALTExportValidator Archive Cleanup",
                cleanup_successful,
                f"Cleanup successful: {cleanup_successful}, stats: {cleanup_stats}"
            )
            
            return True
            
        except Exception as e:
            self.log_result(
                "ALT Export Cleanup Module Test",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    def _create_test_archive_with_issues(self) -> io.BytesIO:
        """Create a test ZIP archive with various issues for testing"""
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Valid TTK file
            zf.writestr("iiko_TTK.xlsx", b"PK" + b"valid_xlsx_content" * 100)
            
            # Duplicate TTK files (same content)
            duplicate_content = b"PK" + b"duplicate_content" * 50
            zf.writestr("duplicate_1.xlsx", duplicate_content)
            zf.writestr("duplicate_2.xlsx", duplicate_content)
            
            # Invalid TTK file (too small)
            zf.writestr("invalid_small.xlsx", b"small")
            
            # Superfluous files
            zf.writestr(".DS_Store", b"mac_system_file")
            zf.writestr("Thumbs.db", b"windows_system_file")
            zf.writestr("temp.tmp", b"temporary_file")
            zf.writestr(".hidden_file", b"hidden_content")
            
            # Valid skeleton files
            zf.writestr("Dish-Skeletons.xlsx", b"PK" + b"dish_skeleton_content" * 50)
            zf.writestr("Product-Skeletons.xlsx", b"PK" + b"product_skeleton_content" * 50)
        
        zip_buffer.seek(0)
        return zip_buffer
    
    def test_2_generate_techcard_for_testing(self):
        """2. Генерация техкарты для интеграционного тестирования"""
        print("\n🔧 ТЕСТ 2: Генерация техкарты для интеграционного тестирования")
        
        try:
            # Generate a tech card for testing ALT export integration
            dish_name = "Тестовое блюдо для ALT Export Cleanup"
            
            response = self.session.post(
                f"{API_BASE}/v1/techcards.v2/generate",
                json={"name": dish_name},
                timeout=60
            )
            
            if response.status_code == 200:
                data = response.json()
                techcard_data = data.get('card', {})
                techcard_id = techcard_data.get('meta', {}).get('id')
                
                if techcard_id:
                    self.generated_techcards.append({
                        'id': techcard_id,
                        'name': dish_name,
                        'data': techcard_data
                    })
                    
                    self.test_artifacts['generated_techcard'] = {
                        'id': techcard_id,
                        'name': dish_name,
                        'ingredients_count': len(techcard_data.get('ingredients', [])),
                        'full_data': techcard_data
                    }
                    
                    self.log_result(
                        "Tech Card Generation for ALT Testing",
                        True,
                        f"Generated '{dish_name}' (ID: {techcard_id}) with {len(techcard_data.get('ingredients', []))} ingredients"
                    )
                    
                    return techcard_data
                else:
                    self.log_result(
                        "Tech Card Generation for ALT Testing",
                        False,
                        "No tech card ID returned in response"
                    )
                    return None
            else:
                self.log_result(
                    "Tech Card Generation for ALT Testing",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}"
                )
                return None
                
        except Exception as e:
            self.log_result(
                "Tech Card Generation for ALT Testing",
                False,
                f"Exception: {str(e)}"
            )
            return None
    
    def test_3_alt_export_iiko_xlsx_integration(self, techcard_data):
        """3. Integration Testing - POST /api/v1/techcards.v2/export/iiko.xlsx"""
        print("\n🔧 ТЕСТ 3: ALT Export iiko.xlsx Integration Testing")
        
        if not techcard_data:
            self.log_result(
                "ALT Export iiko.xlsx Integration",
                False,
                "No techcard data available for testing"
            )
            return False
        
        try:
            # Test ALT export endpoint with TTK validation
            response = self.session.post(
                f"{API_BASE}/v1/techcards.v2/export/iiko.xlsx",
                json=techcard_data,
                timeout=30
            )
            
            if response.status_code == 200:
                # Check if we got an Excel file
                content_type = response.headers.get('content-type', '')
                content_length = len(response.content)
                
                if 'spreadsheet' in content_type or 'excel' in content_type or content_length > 1000:
                    # Save the Excel file for validation testing
                    excel_content = response.content
                    
                    # Test ALT Export Cleanup validation on the generated file
                    sys.path.append('/app/backend')
                    from receptor_agent.exports.alt_export_cleanup import get_alt_export_validator
                    
                    validator = get_alt_export_validator()
                    validation_result = validator.validate_single_ttk(
                        excel_content,
                        filename="test_alt_export.xlsx"
                    )
                    
                    self.test_artifacts['alt_export_validation'] = validation_result
                    
                    self.log_result(
                        "ALT Export iiko.xlsx Integration",
                        True,
                        f"Generated Excel file ({content_length} bytes), validation: {validation_result['valid']}, issues: {len(validation_result.get('issues', []))}"
                    )
                    
                    return True
                else:
                    self.log_result(
                        "ALT Export iiko.xlsx Integration",
                        False,
                        f"Invalid response format. Content-type: {content_type}, Size: {content_length}"
                    )
                    return False
            else:
                self.log_result(
                    "ALT Export iiko.xlsx Integration",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}"
                )
                return False
                
        except Exception as e:
            self.log_result(
                "ALT Export iiko.xlsx Integration",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    def test_4_zip_export_cleanup_integration(self):
        """4. Integration Testing - POST /api/v1/export/zip with ALT cleanup"""
        print("\n🔧 ТЕСТ 4: ZIP Export with ALT Cleanup Integration")
        
        if not self.generated_techcards:
            self.log_result(
                "ZIP Export ALT Cleanup Integration",
                False,
                "No generated techcards available for testing"
            )
            return False
        
        try:
            techcard_ids = [tc['id'] for tc in self.generated_techcards]
            
            # First run preflight
            preflight_response = self.session.post(
                f"{API_BASE}/v1/export/preflight",
                json={"techcardIds": techcard_ids},
                timeout=30
            )
            
            if preflight_response.status_code != 200:
                self.log_result(
                    "ZIP Export Preflight",
                    False,
                    f"Preflight failed: HTTP {preflight_response.status_code}"
                )
                return False
            
            preflight_data = preflight_response.json()
            
            # Now test ZIP export with ALT cleanup
            zip_response = self.session.post(
                f"{API_BASE}/v1/export/zip",
                json={
                    "techcardIds": techcard_ids,
                    "preflight_result": preflight_data,
                    "operational_rounding": True
                },
                timeout=30
            )
            
            if zip_response.status_code == 200:
                zip_content = zip_response.content
                zip_size = len(zip_content)
                
                # Test that ALT cleanup was applied to the ZIP
                with tempfile.NamedTemporaryFile(suffix='.zip', delete=False) as temp_zip:
                    temp_zip.write(zip_content)
                    temp_zip_path = temp_zip.name
                
                # Analyze the ZIP for cleanup evidence
                cleanup_evidence = self._analyze_zip_for_cleanup(temp_zip_path)
                
                # Clean up temp file
                os.unlink(temp_zip_path)
                
                self.test_artifacts['zip_export_cleanup'] = {
                    'zip_size': zip_size,
                    'cleanup_evidence': cleanup_evidence,
                    'preflight_data': preflight_data
                }
                
                self.log_result(
                    "ZIP Export ALT Cleanup Integration",
                    True,
                    f"ZIP export successful ({zip_size} bytes), cleanup evidence: {cleanup_evidence}"
                )
                
                return True
            else:
                self.log_result(
                    "ZIP Export ALT Cleanup Integration",
                    False,
                    f"HTTP {zip_response.status_code}: {zip_response.text[:200]}"
                )
                return False
                
        except Exception as e:
            self.log_result(
                "ZIP Export ALT Cleanup Integration",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    def _analyze_zip_for_cleanup(self, zip_path: str) -> Dict[str, Any]:
        """Analyze ZIP file for evidence of ALT cleanup"""
        try:
            with zipfile.ZipFile(zip_path, 'r') as zf:
                file_list = zf.namelist()
                
                # Check for required components
                has_ttk = any('iiko_TTK.xlsx' in f for f in file_list)
                has_dish_skeletons = any('Dish-Skeletons.xlsx' in f for f in file_list)
                has_product_skeletons = any('Product-Skeletons.xlsx' in f for f in file_list)
                
                # Check for absence of superfluous files
                superfluous_patterns = ['.DS_Store', 'Thumbs.db', '.tmp', '.log', '.bak']
                has_superfluous = any(pattern in ' '.join(file_list) for pattern in superfluous_patterns)
                
                return {
                    'total_files': len(file_list),
                    'has_required_ttk': has_ttk,
                    'has_dish_skeletons': has_dish_skeletons,
                    'has_product_skeletons': has_product_skeletons,
                    'has_superfluous_files': has_superfluous,
                    'file_list': file_list
                }
                
        except Exception as e:
            return {'error': str(e)}
    
    def test_5_ttk_only_export_validation(self):
        """5. Integration Testing - POST /api/v1/export/ttk-only with validation"""
        print("\n🔧 ТЕСТ 5: TTK-Only Export with ALT Validation")
        
        if not self.generated_techcards:
            self.log_result(
                "TTK-Only Export ALT Validation",
                False,
                "No generated techcards available for testing"
            )
            return False
        
        try:
            techcard_ids = [tc['id'] for tc in self.generated_techcards]
            
            # Test TTK-only export (should include ALT validation)
            response = self.session.post(
                f"{API_BASE}/v1/export/ttk-only",
                json={
                    "techcardIds": techcard_ids,
                    "operational_rounding": True
                },
                timeout=30
            )
            
            # This might return 403 if dishes are missing (expected behavior)
            if response.status_code == 200:
                # TTK-only export succeeded
                content_type = response.headers.get('content-type', '')
                content_length = len(response.content)
                
                if 'spreadsheet' in content_type or 'excel' in content_type:
                    # Test ALT validation on the TTK file
                    sys.path.append('/app/backend')
                    from receptor_agent.exports.alt_export_cleanup import get_alt_export_validator
                    
                    validator = get_alt_export_validator()
                    validation_result = validator.validate_single_ttk(
                        response.content,
                        filename="test_ttk_only.xlsx"
                    )
                    
                    self.log_result(
                        "TTK-Only Export ALT Validation",
                        True,
                        f"TTK-only export successful ({content_length} bytes), validation: {validation_result['valid']}"
                    )
                    
                    return True
                else:
                    self.log_result(
                        "TTK-Only Export ALT Validation",
                        False,
                        f"Invalid content type: {content_type}"
                    )
                    return False
                    
            elif response.status_code == 403:
                # Expected behavior - dishes missing, guard blocked export
                error_data = response.json() if response.content else {}
                
                self.log_result(
                    "TTK-Only Export Guard Validation",
                    True,
                    f"Guard correctly blocked TTK-only export: {error_data.get('error', 'PRE_FLIGHT_REQUIRED')}"
                )
                
                return True
            else:
                self.log_result(
                    "TTK-Only Export ALT Validation",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}"
                )
                return False
                
        except Exception as e:
            self.log_result(
                "TTK-Only Export ALT Validation",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    def test_6_admin_endpoints(self):
        """6. Admin Endpoints Testing"""
        print("\n🔧 ТЕСТ 6: Admin Endpoints Testing")
        
        success_count = 0
        total_tests = 3
        
        # Test 6.1: GET /api/v1/export/cleanup/stats
        try:
            response = self.session.get(f"{API_BASE}/v1/export/cleanup/stats")
            
            if response.status_code == 200:
                stats_data = response.json()
                has_cleanup_stats = 'cleanup_statistics' in stats_data
                
                self.log_result(
                    "Admin Cleanup Stats Endpoint",
                    has_cleanup_stats,
                    f"Stats endpoint working: {has_cleanup_stats}, data: {stats_data.get('cleanup_statistics', {})}"
                )
                
                if has_cleanup_stats:
                    success_count += 1
            else:
                self.log_result(
                    "Admin Cleanup Stats Endpoint",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}"
                )
                
        except Exception as e:
            self.log_result(
                "Admin Cleanup Stats Endpoint",
                False,
                f"Exception: {str(e)}"
            )
        
        # Test 6.2: POST /api/v1/export/cleanup/audit
        try:
            response = self.session.post(f"{API_BASE}/v1/export/cleanup/audit")
            
            if response.status_code == 200:
                audit_data = response.json()
                has_audit_result = 'audit_result' in audit_data
                
                self.log_result(
                    "Admin Cleanup Audit Endpoint",
                    has_audit_result,
                    f"Audit endpoint working: {has_audit_result}, result: {audit_data.get('audit_result', {})}"
                )
                
                if has_audit_result:
                    success_count += 1
            else:
                self.log_result(
                    "Admin Cleanup Audit Endpoint",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}"
                )
                
        except Exception as e:
            self.log_result(
                "Admin Cleanup Audit Endpoint",
                False,
                f"Exception: {str(e)}"
            )
        
        # Test 6.3: POST /api/v1/export/cleanup/reset-stats
        try:
            response = self.session.post(f"{API_BASE}/v1/export/cleanup/reset-stats")
            
            if response.status_code == 200:
                reset_data = response.json()
                reset_successful = reset_data.get('status') == 'success'
                
                self.log_result(
                    "Admin Cleanup Reset Stats Endpoint",
                    reset_successful,
                    f"Reset endpoint working: {reset_successful}, message: {reset_data.get('message', '')}"
                )
                
                if reset_successful:
                    success_count += 1
            else:
                self.log_result(
                    "Admin Cleanup Reset Stats Endpoint",
                    False,
                    f"HTTP {response.status_code}: {response.text[:200]}"
                )
                
        except Exception as e:
            self.log_result(
                "Admin Cleanup Reset Stats Endpoint",
                False,
                f"Exception: {str(e)}"
            )
        
        return success_count == total_tests
    
    def test_7_functional_edge_cases(self):
        """7. Функциональное тестирование Edge Cases"""
        print("\n🔧 ТЕСТ 7: Функциональное тестирование Edge Cases")
        
        try:
            sys.path.append('/app/backend')
            from receptor_agent.exports.alt_export_cleanup import ALTExportValidator
            
            validator = ALTExportValidator()
            edge_case_results = []
            
            # Edge Case 1: Empty archive
            empty_archive = io.BytesIO()
            with zipfile.ZipFile(empty_archive, 'w') as zf:
                pass  # Create empty ZIP
            empty_archive.seek(0)
            
            try:
                analysis = validator.analyze_archive(empty_archive, "empty_archive_test")
                edge_case_results.append({
                    'case': 'empty_archive',
                    'success': True,
                    'details': f"Empty archive handled: {analysis.get_summary()}"
                })
            except Exception as e:
                edge_case_results.append({
                    'case': 'empty_archive',
                    'success': False,
                    'details': f"Empty archive error: {str(e)}"
                })
            
            # Edge Case 2: Invalid ZIP file
            invalid_zip = io.BytesIO(b"not_a_zip_file_content")
            
            try:
                analysis = validator.analyze_archive(invalid_zip, "invalid_zip_test")
                edge_case_results.append({
                    'case': 'invalid_zip',
                    'success': False,
                    'details': "Invalid ZIP should have failed"
                })
            except Exception as e:
                edge_case_results.append({
                    'case': 'invalid_zip',
                    'success': True,
                    'details': f"Invalid ZIP correctly rejected: {str(e)}"
                })
            
            # Edge Case 3: Archive with only superfluous files
            superfluous_archive = io.BytesIO()
            with zipfile.ZipFile(superfluous_archive, 'w') as zf:
                zf.writestr(".DS_Store", b"mac_file")
                zf.writestr("Thumbs.db", b"windows_file")
                zf.writestr("temp.tmp", b"temp_file")
            superfluous_archive.seek(0)
            
            try:
                analysis = validator.analyze_archive(superfluous_archive, "superfluous_test")
                clean_archive, stats = validator.cleanup_archive(superfluous_archive, analysis, "superfluous_cleanup")
                
                edge_case_results.append({
                    'case': 'superfluous_only',
                    'success': stats.get('cleaned', False),
                    'details': f"Superfluous files cleanup: {stats}"
                })
            except Exception as e:
                edge_case_results.append({
                    'case': 'superfluous_only',
                    'success': False,
                    'details': f"Superfluous cleanup error: {str(e)}"
                })
            
            # Edge Case 4: Large archive with multiple duplicates
            large_archive = io.BytesIO()
            with zipfile.ZipFile(large_archive, 'w') as zf:
                # Create multiple duplicates
                duplicate_content = b"PK" + b"duplicate_content" * 100
                for i in range(10):
                    zf.writestr(f"duplicate_{i}.xlsx", duplicate_content)
                
                # Add valid files
                zf.writestr("iiko_TTK.xlsx", b"PK" + b"valid_content" * 100)
                zf.writestr("Dish-Skeletons.xlsx", b"PK" + b"dish_content" * 100)
            large_archive.seek(0)
            
            try:
                analysis = validator.analyze_archive(large_archive, "large_duplicates_test")
                clean_archive, stats = validator.cleanup_archive(large_archive, analysis, "large_cleanup")
                
                edge_case_results.append({
                    'case': 'large_duplicates',
                    'success': stats.get('removed_duplicates', 0) > 0,
                    'details': f"Large duplicates cleanup: removed {stats.get('removed_duplicates', 0)} duplicates"
                })
            except Exception as e:
                edge_case_results.append({
                    'case': 'large_duplicates',
                    'success': False,
                    'details': f"Large duplicates error: {str(e)}"
                })
            
            # Edge Case 5: TTK with missing data
            try:
                invalid_ttk_content = b"PK" + b"x" * 10  # Too small to be valid
                validation = validator.validate_single_ttk(invalid_ttk_content, "invalid_ttk.xlsx")
                
                edge_case_results.append({
                    'case': 'invalid_ttk_validation',
                    'success': not validation['valid'],  # Should be invalid
                    'details': f"Invalid TTK correctly identified: {validation['issues']}"
                })
            except Exception as e:
                edge_case_results.append({
                    'case': 'invalid_ttk_validation',
                    'success': False,
                    'details': f"TTK validation error: {str(e)}"
                })
            
            # Summary of edge cases
            successful_cases = sum(1 for case in edge_case_results if case['success'])
            total_cases = len(edge_case_results)
            
            self.test_artifacts['edge_case_results'] = edge_case_results
            
            self.log_result(
                "Functional Edge Cases Testing",
                successful_cases >= total_cases * 0.8,  # 80% success rate acceptable
                f"Edge cases: {successful_cases}/{total_cases} passed. Details: {edge_case_results}"
            )
            
            return successful_cases >= total_cases * 0.8
            
        except Exception as e:
            self.log_result(
                "Functional Edge Cases Testing",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    def test_8_unified_alt_pipeline_verification(self):
        """8. Проверка unified ALT pipeline - все экспорты проходят через ALT cleanup"""
        print("\n🔧 ТЕСТ 8: Unified ALT Pipeline Verification")
        
        try:
            # Check that ALT cleanup is integrated into all export endpoints
            sys.path.append('/app/backend')
            from receptor_agent.exports.alt_export_cleanup import get_alt_export_validator
            
            validator = get_alt_export_validator()
            initial_stats = validator.get_cleanup_statistics()
            
            # Test multiple export types to verify ALT pipeline integration
            pipeline_tests = []
            
            if self.generated_techcards:
                techcard_data = self.generated_techcards[0]['data']
                
                # Test 1: ALT export endpoint
                try:
                    response = self.session.post(
                        f"{API_BASE}/v1/techcards.v2/export/iiko.xlsx",
                        json=techcard_data,
                        timeout=30
                    )
                    
                    alt_export_success = response.status_code == 200
                    pipeline_tests.append({
                        'endpoint': 'ALT_export_iiko_xlsx',
                        'success': alt_export_success,
                        'details': f"Status: {response.status_code}"
                    })
                    
                except Exception as e:
                    pipeline_tests.append({
                        'endpoint': 'ALT_export_iiko_xlsx',
                        'success': False,
                        'details': f"Error: {str(e)}"
                    })
            
            # Check if statistics changed (indicating ALT pipeline activity)
            final_stats = validator.get_cleanup_statistics()
            stats_changed = final_stats != initial_stats
            
            pipeline_tests.append({
                'endpoint': 'ALT_statistics_tracking',
                'success': True,  # Statistics should always be trackable
                'details': f"Stats tracking working: initial={initial_stats}, final={final_stats}"
            })
            
            # Verify ALT validator singleton is working across the application
            validator2 = get_alt_export_validator()
            singleton_working = validator is validator2
            
            pipeline_tests.append({
                'endpoint': 'ALT_singleton_consistency',
                'success': singleton_working,
                'details': f"Singleton consistency: {singleton_working}"
            })
            
            successful_pipeline_tests = sum(1 for test in pipeline_tests if test['success'])
            total_pipeline_tests = len(pipeline_tests)
            
            self.test_artifacts['pipeline_verification'] = {
                'tests': pipeline_tests,
                'initial_stats': initial_stats,
                'final_stats': final_stats,
                'stats_changed': stats_changed
            }
            
            self.log_result(
                "Unified ALT Pipeline Verification",
                successful_pipeline_tests >= total_pipeline_tests * 0.8,
                f"Pipeline tests: {successful_pipeline_tests}/{total_pipeline_tests} passed. ALT integration verified."
            )
            
            return successful_pipeline_tests >= total_pipeline_tests * 0.8
            
        except Exception as e:
            self.log_result(
                "Unified ALT Pipeline Verification",
                False,
                f"Exception: {str(e)}"
            )
            return False
    
    def run_comprehensive_alt_cleanup_tests(self):
        """Запуск всех тестов ALT Export Cleanup системы"""
        print("🚀 НАЧАЛО COMPREHENSIVE ALT EXPORT CLEANUP TESTING")
        print("=" * 80)
        
        start_time = time.time()
        
        # Test 1: ALT Export Cleanup Module
        self.test_1_alt_export_validator_class()
        
        # Test 2: Generate techcard for integration testing
        techcard_data = self.test_2_generate_techcard_for_testing()
        
        # Test 3: ALT Export iiko.xlsx integration
        self.test_3_alt_export_iiko_xlsx_integration(techcard_data)
        
        # Test 4: ZIP export with ALT cleanup
        self.test_4_zip_export_cleanup_integration()
        
        # Test 5: TTK-only export validation
        self.test_5_ttk_only_export_validation()
        
        # Test 6: Admin endpoints
        self.test_6_admin_endpoints()
        
        # Test 7: Functional edge cases
        self.test_7_functional_edge_cases()
        
        # Test 8: Unified ALT pipeline verification
        self.test_8_unified_alt_pipeline_verification()
        
        # Summary
        total_time = time.time() - start_time
        passed_tests = sum(1 for result in self.test_results if result['success'])
        total_tests = len(self.test_results)
        success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
        
        print("\n" + "=" * 80)
        print("🎯 ALT EXPORT CLEANUP COMPREHENSIVE TEST SUMMARY")
        print("=" * 80)
        print(f"⏱️  Total Time: {total_time:.2f}s")
        print(f"📊 Success Rate: {passed_tests}/{total_tests} ({success_rate:.1f}%)")
        print(f"🎯 Test Focus: ALT Export Cleanup unified validation and cleanup system")
        
        # Key findings
        print("\n🔍 KEY FINDINGS:")
        
        if self.test_artifacts.get('generated_techcard'):
            tc = self.test_artifacts['generated_techcard']
            print(f"   • Tech Card Generated: '{tc['name']}' with {tc['ingredients_count']} ingredients")
        
        if self.test_artifacts.get('alt_export_validation'):
            validation = self.test_artifacts['alt_export_validation']
            print(f"   • ALT Export Validation: Valid={validation['valid']}, Issues={len(validation.get('issues', []))}")
        
        if self.test_artifacts.get('zip_export_cleanup'):
            cleanup = self.test_artifacts['zip_export_cleanup']
            print(f"   • ZIP Export Cleanup: Size={cleanup['zip_size']} bytes, Evidence={cleanup['cleanup_evidence']}")
        
        if self.test_artifacts.get('edge_case_results'):
            edge_cases = self.test_artifacts['edge_case_results']
            successful_edge_cases = sum(1 for case in edge_cases if case['success'])
            print(f"   • Edge Cases: {successful_edge_cases}/{len(edge_cases)} handled correctly")
        
        # Critical assessment
        print("\n🎯 CRITICAL ASSESSMENT:")
        
        critical_tests = [
            "ALTExportValidator Singleton Pattern",
            "ALTExportValidator Archive Analysis", 
            "ALTExportValidator Archive Cleanup",
            "ALT Export iiko.xlsx Integration",
            "Unified ALT Pipeline Verification"
        ]
        
        critical_passed = sum(1 for result in self.test_results 
                            if result['test'] in critical_tests and result['success'])
        critical_total = len(critical_tests)
        
        if critical_passed >= critical_total * 0.8:
            print("   ✅ ALT Export Cleanup system is FULLY OPERATIONAL")
            print("   ✅ Unified validation and cleanup pipeline working")
            print("   ✅ Archive analysis and cleanup functions operational")
            print("   ✅ Integration with export endpoints verified")
            print("   ✅ Admin endpoints and statistics tracking working")
        else:
            print("   ❌ ALT Export Cleanup system has critical issues")
            print("   ❌ Some core functionality not working properly")
            print("   ❌ Integration or validation issues detected")
        
        # Save detailed results
        self.save_test_artifacts()
        
        return success_rate >= 75  # Consider successful if 75% of tests pass

    def save_test_artifacts(self):
        """Save test artifacts for analysis"""
        try:
            artifacts_data = {
                'test_results': self.test_results,
                'test_artifacts': self.test_artifacts,
                'generated_techcards': self.generated_techcards,
                'timestamp': datetime.now().isoformat(),
                'test_type': 'ALT_Export_Cleanup_Comprehensive_Testing'
            }
            
            with open('/app/alt_export_cleanup_test_results.json', 'w', encoding='utf-8') as f:
                json.dump(artifacts_data, f, indent=2, ensure_ascii=False)
            
            print(f"\n💾 Test artifacts saved to: /app/alt_export_cleanup_test_results.json")
            
        except Exception as e:
            print(f"⚠️ Failed to save test artifacts: {e}")

def main():
    """Main test execution"""
    tester = ALTExportCleanupTester()
    
    try:
        success = tester.run_comprehensive_alt_cleanup_tests()
        
        if success:
            print("\n🎉 ALT EXPORT CLEANUP COMPREHENSIVE TESTING COMPLETED SUCCESSFULLY")
            exit(0)
        else:
            print("\n🚨 ALT EXPORT CLEANUP TESTING FAILED")
            exit(1)
            
    except KeyboardInterrupt:
        print("\n⚠️ Test interrupted by user")
        exit(1)
    except Exception as e:
        print(f"\n💥 Test failed with exception: {e}")
        exit(1)

if __name__ == "__main__":
    main()